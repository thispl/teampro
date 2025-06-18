from collections import defaultdict
import math
import frappe
from frappe.utils.csvutils import read_csv_content
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,datetime,get_first_day,get_last_day,today)
from datetime import datetime
from datetime import timedelta
import locale
from datetime import date
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from erpnext.accounts.utils import (
    unlink_ref_doc_from_payment_entries
)

@frappe.whitelist()
def get_user_task_summary_by_type(sprint_id, from_date=None, to_date=None, team=None):
    if not sprint_id:
        frappe.throw(_("Sprint is required"))

    completed_statuses = ["Pending Review", "Client Review", "Completed"]
    filters = {"custom_sprint": sprint_id}
    if team:
        filters["custom_dev_team"] = team
    if to_date:
        filters["exp_end_date"] = ["<=", to_date]

    tasks = frappe.get_all("Task", filters=filters,
        fields=["name", "custom_allocated_to", "status", "custom_spot_task", "rt", "custom_dev_team"])

    user_employee_map = {
        emp.user_id: emp.name for emp in frappe.get_all("Employee", fields=["name", "user_id"]) if emp.user_id
    }

    timesheet_rows = frappe.db.sql(f"""
        SELECT td.task, t.employee, SUM(td.hours) AS hours, SUM(t.total_hours) AS total_hours
        FROM `tabTimesheet` t
        JOIN `tabTimesheet Detail` td ON td.parent = t.name
        WHERE t.docstatus = 1 AND td.task IS NOT NULL
        {f"AND t.start_date >= %(from_date)s" if from_date else ""}
        {f"AND t.end_date <= %(to_date)s" if to_date else ""}
        GROUP BY td.task, t.employee
    """, {"from_date": from_date, "to_date": to_date}, as_dict=True)
    
    task_used_in_timesheet = set((row.task, row.employee) for row in timesheet_rows)
    task_hours_map = {(row.task, row.employee): row.hours for row in timesheet_rows}
    timesheet_total_map = defaultdict(float)
    for row in timesheet_rows:
        timesheet_total_map[row.employee] += row.total_hours or 0.0

    # Initialize summary
    user_summary = {}
    totals = defaultdict(float)

    for task in tasks:
        user = task.custom_allocated_to
        if not user:
            continue
        employee_id = user_employee_map.get(user)
        key = (task.name, employee_id)
        if user not in user_summary:
            user_summary[user] = {
                "sprint_planned": 0, "sprint_completed": 0, "sprint_working": 0, "sprint_not_taken": 0,
                "spot_planned": 0, "spot_completed": 0, "spot_working": 0, "spot_not_taken": 0,
                "sprint_planned_hours": 0.0, "sprint_completed_hours": 0.0, "sprint_working_hours": 0.0, "sprint_not_taken_hours": 0.0,
                "spot_planned_hours": 0.0, "spot_completed_hours": 0.0, "spot_working_hours": 0.0, "spot_not_taken_hours": 0.0,
                "biometric_hours": 0.0, "timesheet_hours": 0.0
            }

        is_spot = task.custom_spot_task == 1
        status = task.status
        is_completed = status in completed_statuses
        is_working = status == "Working"
        is_not_taken = employee_id and key not in task_used_in_timesheet
        rt = task.rt or 0.0

        type_prefix = "spot" if is_spot else "sprint"

        user_summary[user][f"{type_prefix}_planned"] += 1
        user_summary[user][f"{type_prefix}_planned_hours"] += rt
        totals[f"{type_prefix}_planned"] += 1
        totals[f"{type_prefix}_planned_hours"] += rt

        if is_completed:
            user_summary[user][f"{type_prefix}_completed"] += 1
            user_summary[user][f"{type_prefix}_completed_hours"] += rt
            totals[f"{type_prefix}_completed"] += 1
            totals[f"{type_prefix}_completed_hours"] += rt

        if is_working:
            user_summary[user][f"{type_prefix}_working"] += 1
            user_summary[user][f"{type_prefix}_working_hours"] += rt
            totals[f"{type_prefix}_working"] += 1
            totals[f"{type_prefix}_working_hours"] += rt

        if is_not_taken:
            user_summary[user][f"{type_prefix}_not_taken"] += 1
            user_summary[user][f"{type_prefix}_not_taken_hours"] += rt
            totals[f"{type_prefix}_not_taken"] += 1
            totals[f"{type_prefix}_not_taken_hours"] += rt

        # Ensure employee_id is available and task not logged in timesheet
        if employee_id and (task.name, employee_id) not in task_used_in_timesheet:
            if rt:
                user_summary[user][f"{type_prefix}_not_taken_hours"] += rt
                totals[f"{type_prefix}_not_taken_hours"] += rt
            user_summary[user][f"{type_prefix}_not_taken"] += 1
            totals[f"{type_prefix}_not_taken"] += 1


    # Add biometric hours and timesheet total hours
    for user, summary in user_summary.items():
        employee_id = user_employee_map.get(user)
        if not employee_id:
            continue

        # Attendance Biometric Time
        # biometric_hours = frappe.db.sql("""
        #     SELECT SUM(bt_difference) FROM `tabAttendance`
        #     WHERE employee = %s
        #     {from_date_clause}
        #     {to_date_clause}
        # """.format(
        #     from_date_clause="AND attendance_date >= %(from_date)s" if from_date else "",
        #     to_date_clause="AND attendance_date <= %(to_date)s" if to_date else ""
        # ), {"from_date": from_date, "to_date": to_date, "employee": employee_id})

        # bt_total = biometric_hours[0][0] or 0.0
        # summary["biometric_hours"] = bt_total
        # totals["biometric_hours"] += bt_total

        # Timesheet Hours
        ts_total = timesheet_total_map.get(employee_id, 0.0)
        summary["timesheet_hours"] = ts_total
        totals["timesheet_hours"] += ts_total

    result = []

    employee_short_code_map = {
        emp.user_id: emp.short_code for emp in frappe.get_all("Employee", fields=["name", "user_id","short_code"]) if emp.user_id
    }
    for user, summary in user_summary.items():
        employee_id = employee_short_code_map.get(user)
        result.append({"user": employee_id, **summary})

    result.append({"user": "Total", **totals})
    return result

@frappe.whitelist()
def update_subject_from_description(doc,method):
    if not doc.custom_subject:
        desc = re.sub(r"<.*?>", "", doc.description)
        doc.custom_subject = desc[:140]

@frappe.whitelist()
def update_task_subject_from_description(doc,method):
    if not doc.subject:
        desc = re.sub(r"<.*?>", "", doc.description)
        doc.subject = desc[:140]

@frappe.whitelist()
def set_creation_date(doc,method):
    if not doc.created_on:
        doc.created_on = today()

@frappe.whitelist()
def si_on_trash(doc,method):
    unlink_ref_doc_from_payment_entries(doc, doc.name)

@frappe.whitelist()
def attendance_calc(from_date,to_date):
    employees = frappe.get_all("Employee",{"status":"Active",'date_of_joining':['<=',from_date]},["*"],order_by='name asc')
    for emp in employees:
        user_id = frappe.get_value('Employee',{'employee':emp.name},['user_id'])
        hod = frappe.get_value('User',{'email':user_id},['name'])
        role = "HOD"
        hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
        if hod:
            late_list = frappe.db.sql("""select count(name) as count from `tabAttendance` where employee = '%s' and time(in_time) > '09:45:00' and leave_application IS NULL and attendance_request IS NULL and attendance_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0
        else:
            late_list = frappe.db.sql("""select count(name) as count from `tabAttendance` where employee = '%s' and time(in_time) > '09:30:00' and leave_application IS NULL and attendance_request IS NULL and attendance_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0
        attendance_perm = frappe.db.sql("""select count(*) as count from `tabAttendance Permission` where employee = '%s' and status in ('Approved','Open') and permission_date between '%s' and '%s' and session = "First Half" """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        on_duty = frappe.db.sql("""select count(*) as count from `tabAttendance Request` where employee = '%s' and docstatus=1 and from_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        leave = frappe.db.sql("""select count(*) as count from `tabLeave Application` where employee = '%s' and docstatus=1 and from_date between '%s' and '%s' and half_day = 0 """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        allowed_late = 3
        late = 0
        actual_late = late_list - (allowed_late + attendance_perm + on_duty + leave)
        
        if actual_late >= 0:
            at = actual_late
        else:
            at = 0
        if at >= 3:
            if at <= 5 :
                late = 0.5
            elif at <= 8 :
                late = 1
            elif at <= 11 :
                late = 1.5
            elif at <= 14 :
                late = 2
            elif at <= 17 :
                late = 2.5
            elif at <= 20 :
                late = 3
            elif at <= 23 :
                late = 3.5
            elif at <= 26 :
                late = 41
            elif at <= 29 :
                late = 4.5
        else:
            late = 0
        salary_st=frappe.db.get_value("Salary Structure Assignment",{"employee":emp.name,"docstatus":1},["name"])
        if salary_st:
            ad = frappe.db.sql("""select * from `tabSalary Structure Assignment` where employee = '%s' and docstatus = 1 ORDER BY from_date DESC LIMIT 1  """%(emp.name),as_dict=True)[0]
            if ad:
                days = date_diff(to_date,from_date) + 1
                if frappe.db.exists('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date}):
                    adsl = frappe.get_doc('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date})
                    adsl.emp_name = emp.name
                    adsl.deduction_days = late
                    adsl.actual_late = at
                    adsl.late_days = late_list 
                    adsl.on_duty= on_duty
                    adsl.permissions = attendance_perm
                    adsl.from_date = from_date
                    adsl.to_date = to_date
                    adsl.late_penalty = (late * (int(ad.base + ad.variable)/(days)))
                    adsl.save()
                else:
                    adsl = frappe.new_doc("Late Penalty")
                    adsl.emp_name = emp.name
                    adsl.deduction_days = late
                    adsl.actual_late = at
                    adsl.late_days = late_list 
                    adsl.on_duty= on_duty
                    adsl.permissions = attendance_perm
                    adsl.from_date = from_date
                    adsl.to_date = to_date
                    adsl.late_penalty = (late * (int(ad.base+ ad.variable)/(days)))
                    adsl.save()
    # 
    employee = frappe.get_all("Employee",{"status":"Left",'relieving_date':['>=',from_date]},["*"],order_by='name asc')
    for emp in employee:
        user_id = frappe.get_value('Employee',{'employee':emp.name},['user_id'])
        hod = frappe.get_value('User',{'email':user_id},['name'])
        role = "HOD"
        hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
        if hod:
            late_list = frappe.db.sql("""select count(name) as count from `tabAttendance` where employee = '%s' and time(in_time) > '09:45:00' and leave_application IS NULL and attendance_request IS NULL and attendance_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0
        else:
            late_list = frappe.db.sql("""select count(name) as count from `tabAttendance` where employee = '%s' and time(in_time) > '09:30:00' and leave_application IS NULL and attendance_request IS NULL and attendance_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0
        attendance_perm = frappe.db.sql("""select count(*) as count from `tabAttendance Permission` where employee = '%s' and status in ('Approved','Open') and permission_date between '%s' and '%s' and session = "First Half" """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        on_duty = frappe.db.sql("""select count(*) as count from `tabAttendance Request` where employee = '%s' and docstatus=1 and from_date between '%s' and '%s' """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        leave = frappe.db.sql("""select count(*) as count from `tabLeave Application` where employee = '%s' and docstatus=1 and from_date between '%s' and '%s' and half_day = 0 """%(emp.name,from_date,to_date),as_dict=True)[0].count or 0 
        allowed_late = 3
        late = 0
        actual_late = late_list - (allowed_late + attendance_perm + on_duty + leave)
        
        if actual_late >= 0:
            at = actual_late
        else:
            at = 0
        if at >= 3:
            if at <= 5 :
                late = 0.5
            elif at <= 8 :
                late = 1
            elif at <= 11 :
                late = 1.5
            elif at <= 14 :
                late = 2
            elif at <= 17 :
                late = 2.5
            elif at <= 20 :
                late = 3
            elif at <= 23 :
                late = 3.5
            elif at <= 26 :
                late = 41
            elif at <= 29 :
                late = 4.5
        else:
            late = 0
        salary_st=frappe.db.get_value("Salary Structure Assignment",{"employee":emp.name,"docstatus":1},["name"])
        if salary_st:
            ad = frappe.db.sql("""select * from `tabSalary Structure Assignment` where employee = '%s' and docstatus = 1 ORDER BY from_date DESC LIMIT 1  """%(emp.name),as_dict=True)[0]
            if ad:
                days = date_diff(to_date,from_date) + 1
                if frappe.db.exists('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date}):
                    adsl = frappe.get_doc('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date})
                    adsl.emp_name = emp.name
                    adsl.deduction_days = late
                    adsl.actual_late = at
                    adsl.late_days = late_list 
                    adsl.on_duty= on_duty
                    adsl.permissions = attendance_perm
                    adsl.from_date = from_date
                    adsl.to_date = to_date
                    adsl.late_penalty = (late * (int(ad.base + ad.variable)/(days)))
                    adsl.save()
                else:
                    adsl = frappe.new_doc("Late Penalty")
                    adsl.emp_name = emp.name
                    adsl.deduction_days = late
                    adsl.actual_late = at
                    adsl.late_days = late_list 
                    adsl.on_duty= on_duty
                    adsl.permissions = attendance_perm
                    adsl.from_date = from_date
                    adsl.to_date = to_date
                    adsl.late_penalty = (late * (int(ad.base+ ad.variable)/(days)))
                    adsl.save()
    return "ok"

import datetime
@frappe.whitelist()
def additional_salary(from_date, to_date):
    # from_date="2024-11-01"
    # to_date="2024-11-30"
    frappe.errprint(from_date)
    frappe.errprint(to_date)
    employees = frappe.get_all("Employee",{"status":"Active",'date_of_joining':['<=',from_date]},["*"])
    for emp in employees:
        lp_list= frappe.db.sql("""select emp_name,deduction_days,from_date,late_penalty,to_date from `tabLate Penalty` where emp_name = '%s' and from_date = '%s' and to_date = '%s' """%(emp.name,from_date,to_date),as_dict=True)

        # lp= frappe.db.sql("""select emp_name,deduction_days,from_date,late_penalty,to_date from `tabLate Penalty` where emp_name = '%s' and from_date = '%s' and to_date = '%s' """%(emp.name,from_date,to_date),as_dict=True)[0]
        if lp_list:
            lp = lp_list[0]
            if lp.deduction_days>0:
                date_str1 = lp.from_date
                date_str2 = emp.date_of_joining
                date_obj1 = datetime.strptime(str(date_str1), '%Y-%m-%d').date()
                date_obj2 = datetime.strptime(str(date_str2),'%Y-%m-%d').date()
                if date_obj1 > date_obj2:
                    payroll_date = date_obj1
                else:
                    payroll_date = date_obj2
                if frappe.db.exists('Additional Salary',{'employee':emp.name,'payroll_date':payroll_date,"docstatus":1}):
                    ad = frappe.get_doc('Additional Salary',{'employee':emp.name,'payroll_date':payroll_date})
                    ad.employee = emp.name
                    ad.salary_component = "Late Penalty"
                    ad.company = emp.company
                    ad.amount = lp.late_penalty
                    ad.payroll_date = payroll_date
                    ad.save(ignore_permissions=True)
                    ad.submit()
                else:
                    ad = frappe.new_doc('Additional Salary')
                    ad.employee = emp.name
                    ad.salary_component = "Late Penalty"
                    ad.company = emp.company
                    ad.amount = lp.late_penalty
                    ad.payroll_date = payroll_date
                    ad.save(ignore_permissions=True)
                    ad.submit()
    return "ok"


from datetime import datetime
import calendar
@frappe.whitelist()
def create_update_leave_allocation():
    employees = frappe.get_all("Employee",{"status":"Active"},["*"],order_by='name ASC')
    current_date = datetime.now().date()
    # current_date=(get_first_day(current_date1))
    today_date = date.today()
    year = today_date.year
    year = int(year)
    next_year = year + 1
    from_date = f"01-04-{year}"
    to_date = f"31-03-{next_year}"
    from_date = datetime.strptime(from_date, "%d-%m-%Y").date() 
    to_date = datetime.strptime(to_date, "%d-%m-%Y").date()
    for emp in employees:
        doj = emp.date_of_joining
        diff = current_date - doj
        years = diff.days / 365.25  
        if(int(years)) > 0 :
            if frappe.db.exists("Leave Allocation",{'docstatus':1,'employee':emp.employee,'leave_type':"Casual Leave",'from_date': ('between', (from_date,to_date)),'to_date': ('between', (from_date,to_date))}):
                la = frappe.get_doc("Leave Allocation",{'docstatus':1,'employee':emp.employee,'leave_type':"Casual Leave",'from_date':('between', (from_date,to_date)),'to_date': ('between', (from_date,to_date))})
                la.new_leaves_allocated = la.new_leaves_allocated + 1.5
                la.to_date = to_date
                la.save(ignore_permissions=True)
                la.submit()   
            else:
                la = frappe.new_doc("Leave Allocation")
                la.employee = emp.name
                la.leave_type = "Casual Leave"
                la.new_leaves_allocated = 1.5
                la.from_date = current_date
                la.to_date = to_date
                la.save(ignore_permissions=True)
                la.submit()  

@frappe.whitelist()
def update_leave_ledger_entry():
    prev_date = datetime.now().date()
    start_time = datetime.combine(prev_date, datetime.min.time()) + timedelta(hours=0)
    end_time = datetime.combine(prev_date, datetime.min.time()) + timedelta(hours=20)
    llg=frappe.db.sql("""SELECT * FROM `tabLeave Ledger Entry` WHERE creation BETWEEN %s AND %s and transaction_type = 'Leave Allocation' and leave_type ='Casual Leave'""", (start_time, end_time), as_dict=True)
    i=0
    print(start_time)
    print(end_time)
    
    for l in llg:
        i+=1	
        llg_doc=frappe.get_doc("Leave Ledger Entry",l.name)
        print(llg_doc.name)
        frappe.db.set_value("Leave Ledger Entry",llg_doc.name,"from_date","prev_date")



# def transfer_website():
#     leads = frappe.get_all("Lead",['name','website'])
#     for l in leads:
#         frappe.db.set_value("Lead",l.name,"web",l.website)

# @frappe.whitelist()
# def transfer_contacts():
    # leads = frappe.get_all('Contact',['name','temp_mobile_no'])
    # for lead in leads:
#     contacts = frappe.db.sql("""select `tabContact`.name,`tabContact`.first_name,`tabContact`.middle_name,`tabContact`.last_name,`tabDynamic Link`.link_doctype, `tabDynamic Link`.link_name,`tabContact Email`.email_id
#     from `tabContact` Left Join `tabDynamic Link` on `tabContact`.name = `tabDynamic Link`.parent
#     Left Join `tabContact Email` on `tabContact`.name = `tabContact Email`.parent where `tabDynamic Link`.link_doctype = 'Lead'  """,as_dict=True)
#     for contact in contacts:
#         lead = frappe.get_doc('Lead',contact.link_name)
#         if not contact.last_name:
#             person_name = contact.first_name
#         else:
#             person_name = contact.first_name + contact.last_name
#         lead.append('lead_contacts',{
#             'person_name': person_name,
#             'mobile': lead.temp_mobile_no,
#             'email_id': contact.email_id
#         })
#         lead.save(ignore_permissions=True)
#         frappe.db.commit()
#         print(lead.name,person_name,lead.temp_mobile_no,contact.email_id)

# # @frappe.whitelist()
# def transfer_address():
#     addresses = frappe.db.sql("""select `tabAddress`.address_line1,`tabAddress`.address_line2,`tabAddress`.city,`tabAddress`.phone,`tabAddress`.state,`tabAddress`.country,`tabDynamic Link`.link_doctype, `tabDynamic Link`.link_name
#     from `tabAddress` Left Join `tabDynamic Link` on `tabAddress`.name = `tabDynamic Link`.parent where `tabDynamic Link`.link_doctype = 'Lead' """,as_dict=True)
#     for address in addresses:
#         lead = frappe.get_doc('Lead',address.link_name)
#         addr = ''
#         addr += address.address_line1 or '' + "\n"
#         addr += address.address_line2 or '' + "\n"
#         addr += address.city or '' + "\n"
#         addr += address.state or '' + "\n"
#         addr += address.country or '' + "\n"
#         addr += address.phone or '' + "\n"
#         lead.address = addr
#         lead.save(ignore_permissions=True)
#         frappe.db.commit()
        # print(addr)
    # print(lead.name,person_name,lead.temp_mobile_no,contact.email_id)


# @frappe.whitelist()
# def change_lead_owner():
#     leads = frappe.get_all("Lead",{'contact_by':'sr@groupteampro.com','territory':'Tamil Nadu'})
#     print(len(leads))
#     # for l in leads:
#     #     print(l.name)
#     #     # frappe.db.set_value("Lead",l.name,"contact_by","sales@groupteampro.com")
    #     frappe.db.set_value("Lead",l.name,"lead_owner","anil.p@groupteampro.com")

# def change_sams_owner(filename):
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     for pp in pps:
#         if pp[0]:
#             print(pp[0],pp[1])
#             frappe.db.set_value('SAMS',pp[0],'sa_owner',pp[1])


@frappe.whitelist()
def bulk_update_closure_status():
    count = 0
    closures = frappe.get_all('Closure',{'status':'PCC'},['visa','name'])
    for cl in closures:
        if not cl.visa:
            frappe.db.set_value('Closure',cl.name,'status','Visa')
            frappe.db.commit() 

@frappe.whitelist()
def rename_file(doc,method):
    hashcode = frappe.generate_hash()[:5]
    doc.file_name = hashcode + doc.file_name
