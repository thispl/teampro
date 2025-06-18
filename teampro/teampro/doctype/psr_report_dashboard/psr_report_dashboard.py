# Copyright (c) 2025, TeamPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import formatdate

class PSRReportDashboard(Document):
    pass

import frappe
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from io import BytesIO
from frappe.utils.file_manager import save_file
from teampro.teampro.doctype.psr_report_dashboard.psr_report_hour import make_xlsx_with_hour


@frappe.whitelist()
def download():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"Daily PSR Report {posting_date}"
    build_xlsx_response(filename)

def build_xlsx_response(filename_bcs):
    xlsx_file = make_xlsx(filename_bcs)
    frappe.response["filename"] = f"{filename_bcs}.xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"
    

def make_xlsx(data, sheet_name="Daily PSR Report", wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)

    # Styles
    fill_color_total = PatternFill(start_color="4C3B69", end_color="4C3B69", fill_type="solid")  # Total row
    fill_color_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue row
    fill_color_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White row

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    text_wrap = Alignment(wrap_text=True, horizontal="left", vertical="center")
    title_font = Font(bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)
    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    column_widths = {
        'A': 5, 'B': 30, 'C': 10, 'D': 7, 'E': 7, 'F': 7,
        'G': 7, 'H': 7, 'I': 7, 'J': 7
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Title
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "Daily PSR Report-Count (" + posting_date + ")"
    ws.merge_cells("A1:J1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment_center

    # Header
    header = ["S#", "Project Name", "Project Type", "#O", "#W", "#CD", "#PR", "#CR", "#IO", "#IR"]
    ws.append(header)
    for cell in ws[2]:
        cell.fill = fill_color_total
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border

    # Data rows
    data1 = get_data()
    total_row = ["", "Total", ""] + [0] * 7

    for idx, row in enumerate(data1):
        ws.append(row)
        is_even = idx % 2 == 0
        for cidx, cell in enumerate(ws[ws.max_row], start=1):
            if cidx in [2, 3]:
                cell.alignment = alignment_left
            else:
                cell.alignment = alignment_center
            cell.border = thin_border
            if is_even:
                cell.fill = fill_color_blue
                cell.font = black_font
            else:
                cell.fill = fill_color_white
                cell.font = black_font

        # Accumulate totals (D to J columns)
        for i in range(3, 10):
            try:
                total_row[i] += int(row[i]) if row[i] else 0
            except (ValueError, TypeError):
                total_row[i] += 0

    # Total row
    ws.append(total_row)
    for idx, cell in enumerate(ws[ws.max_row], start=1):
        if idx in [2, 3]:
            cell.alignment = alignment_left
        else:
            cell.alignment = alignment_center
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill_color_total
        cell.border = thin_border

    # Save to file
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def get_data():
    data = []
    s_no = 1
    projects = frappe.db.sql("""
        SELECT name, project_name, project_type
        FROM `tabProject`
        WHERE service = %s
          AND status NOT IN ('Completed', 'Cancelled', 'Hold') AND name NOT IN ('QPIC_ERP_10.04.22','SaudiFiber_05.04.2025','Pincode Global')
        ORDER BY 
            CASE 
                WHEN project_type = 'External' THEN 1
                WHEN project_type = 'AMC' THEN 2
                WHEN project_type = 'Enquiry' THEN 3
                WHEN project_type = 'Products' THEN 4
                WHEN project_type = 'Internal' THEN 5
                ELSE 6
            END
    """, ("IT-SW",), as_dict=True)
    for project in projects:
        cr_count = frappe.db.count("Task", {"project": project.name, "status": "Client Review"})
        pr_count = frappe.db.count("Task", {"project": project.name, "status": "Pending Review"})
        open_count = frappe.db.count("Task", {"project": project.name, "status": "Open"})
        code_review_count = frappe.db.count("Task", {"project": project.name, "status": "Code Review"})
        working_count = frappe.db.count("Task", {"project": project.name, "status": "Working"})
        issue_open_count=frappe.db.count("Issue",{"project":project.name,"status":"Open"})
        issue_replied_count=frappe.db.count("Issue",{"project":project.name,"status":"Replied"})
        data.append([
            s_no,project.project_name, project.project_type,open_count,working_count,code_review_count,pr_count,cr_count,
            issue_open_count,issue_replied_count
        ])
        s_no+=1

    return data


@frappe.whitelist()    
def send_daily_psr_report():
    frappe.sendmail(
        recipients=["dineshbabu.k@groupteampro.com","abdulla.pi@groupteampro.com","anil.p@groupteampro.com"],
        # recipients =['jothi.m@groupteampro.com'],
        subject=f"Daily Project Status Report : {formatdate(frappe.utils.today(), 'dd-mm-yyyy')}",
        message="""Dear Team,<br><br>
        Please find attached the Daily Project Status Report – Count Based and Hour Based, for your kind reference.<br><br>""",
        attachments=[
            {
                "fname": f"Daily PSR Report.xlsx",
                "fcontent": make_xlsx(data=None, sheet_name="Daily PSR Report", wb=None, column_widths=None).getvalue()
            },
            {
                "fname": f"Daily PSR Report(Hour).xlsx",
                "fcontent": make_xlsx_with_hour(data=None, sheet_name="Daily PSR Report - Hour", wb=None, column_widths=None).getvalue()
            }
        ]
    )


