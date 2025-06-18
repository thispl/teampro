import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from frappe.utils import today
import frappe

@frappe.whitelist()
def check_status_report():
    filename = f'Consolidated_Check_Report_{today()}.xlsx'
    xlsx_file = make_xlsx_case_status()

    # Sending file response
    frappe.response['filename'] = filename
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_case_status():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case Status Report"
    text_wrap_left = Alignment(vertical="center", horizontal="center")

    # Setting column widths
    for col in range(ord('A'), ord('Q') + 1):  
        ws.column_dimensions[chr(col)].width = 20

    # Adding headers
    headers = [
        "Sr.no", "Case ID", "Check ID", "Employee Name", "Customer", "Check Package", "Batch",
        "Check Status", "Check Report", "Client Employee Code", "Initiation Date",
        "Entry Allocated To", "Check Completion Date", "Check TAT Completion Date",
        "Insufficiency Closed", "Insufficiency Reported", "Actual Age"
    ]
    ws.append(headers)

    # Formatting headers
    header_fill = PatternFill(start_color="FF002060", end_color="FF002060", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = text_wrap_left

    # Fetching check details
    check_details = get_case_report_details()

    sr_no = 1
    for check in check_details:
        ws.append([
            sr_no,
            check.get("case_id"),
            check.get("check_id"),
            check.get("case_name"),
            check.get("customer"),
            check.get("check_package"),
            check.get("batch"),
            check.get("check_status"),
            check.get("check_report"),
            check.get("client_employee_code"),
            check.get("check_creation_date"),
            check.get("allocated_to"),
            check.get("check_completion_date"),
            check.get("end_date"),
            check.get("insufficiency_closed"),
            check.get("insufficiency_date"),
            check.get("actual_tat")
        ])
        
        sr_no += 1

        # Applying border to the last row
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Save to BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_case_report_details():
    # Fetch cases excluding specific statuses
    cases = frappe.get_all("Case", 
        filters={"case_status": ("not in", [
            "Final-QC", "Generate Report", "Case Report Completed",
            "Case Completed", "To be Billed", "SO Created", "Drop"
        ])}, 
        fields=["name", "case_status", "case_name", "customer", "insufficiency_closed","end_date"]
    )

    check_list = ["Education Checks", "Family", "Reference Check", "Court", 
                  "Social Media", "Criminal", "Employment", "Identity Aadhar", "Address Check"]

    report_data = []

    for case in cases:
        for check_type in check_list:
            checks = frappe.get_all(check_type, 
                filters={"case_id": case["name"]}, 
                fields=[
                    "name AS check_id", "check_status", "report_status",
                    "check_package", "batch", "client_employee_code", 
                    "check_creation_date", "allocated_to", "check_completion_date", 
                    "insufficiency_date", "actual_tat"
                ]
            )
            
            for check in checks:
                report_data.append({
                    "case_id": case["name"],
                    "case_name": case["case_name"],
                    "customer": case["customer"],
                    "check_id": check.get("check_id"),
                    "check_status": check.get("check_status"),
                    "check_report": check.get("report_status"),
                    "check_package": check.get("check_package"),
                    "batch": check.get("batch"),
                    "client_employee_code": check.get("client_employee_code"),
                    "check_creation_date": check.get("check_creation_date"),
                    "allocated_to": check.get("allocated_to"),
                    "check_completion_date": check.get("check_completion_date"),
                    "end_date": case.get("end_date"),
                    "insufficiency_closed": case.get("insufficiency_closed"),  # Fixed `.get()` usage
                    "insufficiency_date": check.get("insufficiency_date"),
                    "actual_tat": check.get("actual_tat")
                })

    return report_data
