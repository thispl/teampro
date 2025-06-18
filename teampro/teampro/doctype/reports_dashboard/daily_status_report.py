import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from frappe.utils import today, formatdate
import frappe
from collections import defaultdict

@frappe.whitelist()
def daily_status_report():
    filename = f'Daily_Status_Report_{today()}.xlsx'
    xlsx_file = generate_task_status_xlsx()

    frappe.response['filename'] = filename
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def generate_task_status_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    formatted_date = formatdate(today(), "dd-mm-yyyy")
    ws.title = "Daily Status Report"
    alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title Row (A1:R1)
    ws.merge_cells('A1:R1')
    ws['A1'].value = "Daily Status Report"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="614878", end_color="614878", fill_type="solid")
    ws['A1'].alignment = alignment
    ws['A1'].border = thin_border

    # Date (S1)
    ws['S1'].value = formatted_date
    ws['S1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['S1'].fill = PatternFill(start_color="614878", end_color="614878", fill_type="solid")
    ws['S1'].alignment = alignment
    ws['S1'].border = thin_border

    # Column Widths Adjustment
    column_widths = {
        'A': 6, 'B': 40, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10,
        'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 10, 'P': 10,
        'Q': 10, 'R': 10, 'S': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Row 2 Header (Light Blue)
    row_2_fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")
    ws.append(["", "", "Open", "", "Working", "", "Overdue", "", "Hold", "", 
               "Code Review", "", "Pending Review", "", "Client Review", "", "Total", "", "Reporting"])
    for cell in ws[2]:  
        cell.fill = row_2_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Row 3 Header (Dark Blue)
    row_3_fill = PatternFill(start_color="4f81bc", end_color="4f81bc", fill_type="solid")
    headers = [
        "Sr.NO", "Project", "Count", "EH", "Count", "EH", "Count", "EH", 
        "Count", "EH", "Count", "EH", "Count", "EH", "Count", "EH", 
        "Total Tasks", "EH", "Type"
    ]
    ws.append(headers)
    for cell in ws[3]:  
        cell.fill = row_3_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = alignment
        cell.border = thin_border

    # Fetch and append data
    project_task_details, project_types = get_project_task_report_details()
    sr_no = 1
    grand_total = {status: {"count": 0, "expected_time": 0} for status in 
                   ["Open", "Working", "Overdue", "Hold", "Code Review", "Pending Review", "Client Review"]}

    for project, task_data in project_task_details.items():
        row_data = [
            sr_no,
            project,
            task_data["Open"]["count"], task_data["Open"]["expected_time"],
            task_data["Working"]["count"], task_data["Working"]["expected_time"],
            task_data["Overdue"]["count"], task_data["Overdue"]["expected_time"],
            task_data["Hold"]["count"], task_data["Hold"]["expected_time"],
            task_data["Code Review"]["count"], task_data["Code Review"]["expected_time"],
            task_data["Pending Review"]["count"], task_data["Pending Review"]["expected_time"],
            task_data["Client Review"]["count"], task_data["Client Review"]["expected_time"],
            sum(task["count"] for task in task_data.values()),  # Total Task Count
            sum(task["expected_time"] for task in task_data.values()),  # Total Expected Time
            project_types.get(project, "N/A")  # Project Type
        ]
        ws.append(row_data)

        # Accumulate grand totals
        for status in grand_total.keys():
            grand_total[status]["count"] += task_data[status]["count"]
            grand_total[status]["expected_time"] += task_data[status]["expected_time"]

        # Apply border to the data row
        for cell in ws[sr_no + 3]:  
            cell.border = thin_border
            cell.alignment = alignment

        sr_no += 1

    # Add grand total row
    total_row = [
        "", "Grand Total",
        grand_total["Open"]["count"], grand_total["Open"]["expected_time"],
        grand_total["Working"]["count"], grand_total["Working"]["expected_time"],
        grand_total["Overdue"]["count"], grand_total["Overdue"]["expected_time"],
        grand_total["Hold"]["count"], grand_total["Hold"]["expected_time"],
        grand_total["Code Review"]["count"], grand_total["Code Review"]["expected_time"],
        grand_total["Pending Review"]["count"], grand_total["Pending Review"]["expected_time"],
        grand_total["Client Review"]["count"], grand_total["Client Review"]["expected_time"],
        sum(grand_total[status]["count"] for status in grand_total.keys()),  # Total Tasks
        sum(grand_total[status]["expected_time"] for status in grand_total.keys()),  # Total Expected Time
        ""
    ]
    ws.append(total_row)

    # Apply formatting to the Grand Total row
    for cell in ws[ws.max_row]:  
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Yellow background
        cell.alignment = alignment
        cell.border = thin_border

    # Save to BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_project_task_report_details():
    projects = frappe.get_all("Project", 
        filters={"status": ("not in", ["Hold", "Completed", "Cancelled", "Enquiry"])}, 
        fields=["name", "project_type"]
    )

    task_statuses = ["Open", "Working", "Overdue", "Hold", "Code Review", "Pending Review", "Client Review"]
    project_task_data = defaultdict(lambda: {status: {"count": 0, "expected_time": 0} for status in task_statuses})
    project_types = {}

    for project in projects:
        project_name = project["name"]
        project_types[project_name] = project.get("project_type", "N/A")  

        tasks = frappe.get_all("Task",
            filters={"project": project_name, "service": "IT-SW"},
            fields=["status", "expected_time"]
        )

        for task in tasks:
            if task["status"] in task_statuses:
                project_task_data[project_name][task["status"]]["count"] += 1
                project_task_data[project_name][task["status"]]["expected_time"] += task.get("expected_time", 0) or 0

    return project_task_data, project_types
