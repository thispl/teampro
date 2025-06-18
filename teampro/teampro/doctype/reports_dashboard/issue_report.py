import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from frappe.utils import today, formatdate
import frappe
from collections import defaultdict

@frappe.whitelist()
def issue_report():
    filename = f'Issue_Report_{today()}.xlsx'
    xlsx_file = generate_task_status_xlsx()

    frappe.response['filename'] = filename
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def generate_task_status_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    formatted_date = formatdate(today(), "dd-mm-yyyy")
    ws.title = "Issue Details"

    
    alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

   
    ws.merge_cells('A1:J1')
    ws['A1'].value = "Issue Details"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="614878", end_color="614878", fill_type="solid")
    ws['A1'].alignment = alignment
    ws['A1'].border = thin_border

   
    ws['K1'].value = formatted_date
    ws['K1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['K1'].fill = PatternFill(start_color="614878", end_color="614878", fill_type="solid")
    ws['K1'].alignment = alignment
    ws['K1'].border = thin_border

   
    column_widths = {'A': 6, 'B': 35, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

   
    headers = ["S.No", "Project", "Type", "Open", "Working", "Overdue", "Hold", "Code Review", "Pending Review", "Client Review", "Grand Total"]
    ws.append(headers)

    header_fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")
    for cell in ws[2]:  
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

   
    project_task_details, project_types = get_project_task_report_details()
    sr_no = 1
    total_columns = [0] * 8  

    for project, task_data in project_task_details.items():
        row_values = [
            sr_no,
            project,
            project_types.get(project, "N/A"),
            task_data["Open"]["count"], 
            task_data["Working"]["count"], 
            task_data["Overdue"]["count"], 
            task_data["Hold"]["count"],
            task_data["Code Review"]["count"], 
            task_data["Pending Review"]["count"], 
            task_data["Client Review"]["count"], 
            sum(task["count"] for task in task_data.values()), 
        ]
        
        
        for i in range(3, 11): 
            total_columns[i-3] += row_values[i]

        ws.append(row_values)

        
        for cell in ws[sr_no + 2]:  
            cell.border = thin_border
            cell.alignment = alignment

        sr_no += 1

   
    total_row = ["", "Grand Total", ""] + total_columns
    ws.append(total_row)
    
    last_row_index = ws.max_row
    for cell in ws[last_row_index]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange Background
        cell.border = thin_border
        cell.alignment = alignment

    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_project_task_report_details():
    projects = frappe.get_all(
        "Project", 
        filters={"status": ("not in", ["Hold", "Completed", "Cancelled", "Enquiry"])}, 
        fields=["name", "project_type"]
    )

    task_statuses = ["Open", "Working", "Overdue", "Hold", "Code Review", "Pending Review", "Client Review"]
    project_task_data = defaultdict(lambda: {status: {"count": 0} for status in task_statuses})
    project_types = {}

    for project in projects:
        project_name = project["name"]
        project_types[project_name] = project.get("project_type", "N/A")  

        
        issues = frappe.get_all(
            "Issue",
            filters={"project": project_name},
            fields=["name"]
        )

        
        issue_list = [issue["name"] for issue in issues if issue.get("name")]

       
        for issue in issues:
            status = issue["name"]
            if status in task_statuses:
                project_task_data[project_name][status]["count"] += 1

        
        if issue_list:
            tasks = frappe.get_all(
                "Task",
                filters={"project": project_name, 'service': 'IT-SW', 'issue': ["in", issue_list]},
                fields=["status"]
            )
        else:
            tasks = []

       
        for task in tasks:
            status = task["status"]
            if status in task_statuses:
                project_task_data[project_name][status]["count"] += 1

    return project_task_data, project_types
