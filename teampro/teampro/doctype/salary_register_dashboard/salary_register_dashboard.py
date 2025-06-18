# Copyright (c) 2025, TeamPRO and contributors
# For license information, please see license.txt

# import frappe
# from frappe.model.document import Document
import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from six import BytesIO
from datetime import datetime
from frappe.utils import getdate, nowdate, format_date

class SalaryRegisterDashboard(Document):
	pass

@frappe.whitelist()
def download():
    filename = 'Salary_Register_Reports'
    build_xlsx_response(filename)

def apply_common_styles(cell,header_font_data, align_left, border,align_right,row,sheet):
    row_fill = PatternFill(fgColor="f2f2f2", fill_type="solid")
    cell.fill = row_fill
    cell.alignment = align_left
    cell.font = header_font_data
    cell.alignment=align_right
    border_1 = Border(
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    cell.border=border
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    sheet.merge_cells(f"A{row}:E{row}")  # Correctly define the merge range
    cell = sheet[f"A{row}"]
    cell.alignment = align_left
    cell.font = header_font_data
    for col in ['F', 'G', 'H', 'I', 'J']:
        cell = sheet[f"{col}{row}"]
        cell.fill = row_fill
        cell.alignment = align_right
        cell.font = header_font_data
        # cell.border = border

def apply_common_styles_for_total(cell,header_font_data, align_center, border,align_right,row,sheet):
    row_fill = PatternFill(fgColor="f2f2f2", fill_type="solid")
    cell.fill = row_fill
    cell.alignment = align_center
    cell.font = header_font_data
    cell.alignment=align_right
    border_1 = Border(
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    cell.border=border
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    sheet.merge_cells(f"A{row}:E{row}")  # Correctly define the merge range
    cell = sheet[f"A{row}"]
    cell.alignment = align_center
    cell.font = header_font_data
    for col in ['F', 'G', 'H', 'I', 'J']:
        cell = sheet[f"{col}{row}"]
        cell.fill = row_fill
        cell.alignment = align_right
        cell.font = header_font_data

def apply_common_styles_left_employees(cell, left_emp_font,align_left, align_right,border,sheet,row):
    cell.alignment = align_left
    cell.font = left_emp_font
    cell.border = border
    for col in ['F', 'G', 'H', 'I', 'J']:
        # cell.fill = row_fill
        cell = sheet[f"{col}{row}"]
        cell.alignment = align_right
        cell.font = left_emp_font                      
        cell.border =border  
    # cell.border=border
def apply_common_styles_dept_font(cell,dept_font, align_left, border,align_center,row,sheet):
    cell.alignment = align_left
    cell.font = dept_font
    cell.alignment=align_center
    border_1 = Border(
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    cell.border=border
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    # for col in ['A','B']:
    # sheet.merge_cells(f"A{row}:B{row}")
    sheet.merge_cells(f"A{row}:E{row}")  # Correctly define the merge range
    cell = sheet[f"A{row}"]
    cell.alignment = align_left
    cell.font = dept_font
    for col in ['F', 'G', 'H', 'I', 'J']:
        cell = sheet[f"{col}{row}"]
        cell.alignment = align_center
        cell.font = dept_font
        # cell.border = border
def apply_header_styles(ws, header_font, align_center, border):
    header_fill = PatternFill(fgColor="002060", fill_type="solid")
    
    for cell in ws[1]:  # Apply to the header row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
def apply_common_styles_active_employees(cell, active_emp_font,align_left, border,align_right,sheet,row):
    
    cell.border=border
    cell.alignment = align_left
    cell.font = active_emp_font 
    for col in ['F', 'G', 'H', 'I', 'J','L']:
        # cell.fill = row_fill
        cell = sheet[f"{col}{row}"]
        cell.alignment = align_right
        cell.font = active_emp_font                      
        cell.border =border             
def set_column_widths(ws):
    """Set the column widths for the Excel sheet."""
    column_widths = [5, 20, 30, 20, 20, 10, 10, 15, 15, 15, 5,10,10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

def make_xlsx(sheet_name='Sheet1'):
    args = frappe.local.form_dict
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Create headers
    headerrow = ["S#", "E-Code", "Employee Name", "DOJ", "Designation", "Fixed", "PD", "Gross", "Deduction", "Net", "B","PR Score","Advance"]
    ws.append(headerrow)

    # Define common styles
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=False)
    align_right = Alignment(horizontal='right', vertical='top', wrap_text=False)
    header_font = Font(color="FFFFFF", bold=True, size=14)
    header_font_data = Font(color="002060", bold=True, size=12)
    left_emp_font = Font(color="ec6b08", bold=True, size=10)
    dept_font=Font(bold=True, size=12)
    active_emp_font=Font(bold=True, size=10)
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Apply header styles
    apply_header_styles(ws, header_font, align_center, border)

    # Set column widths
    set_column_widths(ws)

    from_date = args.get('from_date')
    to_date = args.get('to_date')

    if not from_date or not to_date:
        frappe.throw("From Date and To Date are required.")

    departments = frappe.get_all(
        "Department", 
        {"disabled": 0, 'is_group': 1, "name": ["!=", "All Departments"]}, 
        ["name"], 
        order_by="name asc"
    )

    index = 0
    total_fixed_salary_= 0
    total_gross_salary_ = 0
    total_deduction_ = 0
    total_netpay_ = 0
    total_advance=0
    for dept in departments:
        child_departments = frappe.get_all(
            "Department", 
            {"disabled": 0, "parent_department": dept.name}, 
            ["name"], 
            order_by="name asc"
        )
        
        department_names = [child_dept['name'] for child_dept in child_departments]

        total_fixed_salary = 0
        total_gross_salary = 0
        total_deduction = 0
        total_netpay = 0
        for dept_name in department_names:
            salary_slip_data = frappe.db.sql("""
                SELECT employee, salary_structure 
                FROM `tabSalary Slip`
                WHERE start_date = %s AND end_date = %s 
                AND docstatus != 2 AND department = %s 
                GROUP BY employee 
                ORDER BY employee
            """, (from_date, to_date, dept_name), as_dict=True)

            fixed_salary = 0
            for slip in salary_slip_data:
                fixed_base = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': slip['employee'], 'salary_structure': slip['salary_structure'], 'docstatus': 1},
                    'base', order_by='creation desc'
                ) or 0.0
                fixed_var = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': slip['employee'], 'salary_structure': slip['salary_structure'], 'docstatus': 1},
                    'variable', order_by='creation desc'
                ) or 0.0
                fixed =fixed_base+fixed_var
                fixed_salary += fixed

            total_fixed_salary += fixed_salary
        for child_dept in child_departments:
            salary_slip_list_data = frappe.db.sql("""
                SELECT department, employee,
                    SUM(gross_pay) AS dept_gross_pay, 
                    SUM(total_deduction) AS dept_total_deduction, 
                    SUM(net_pay) AS dept_net_pay, 
                    salary_structure 
                FROM `tabSalary Slip`
                WHERE start_date = %s AND end_date = %s 
                AND docstatus != 2 AND department = %s 
                GROUP BY department 
                ORDER BY department DESC
            """, (from_date, to_date, child_dept.name), as_dict=True)

            for s in salary_slip_list_data:
                total_gross_salary += s['dept_gross_pay']
                total_deduction += s['dept_total_deduction']
                total_netpay += s['dept_net_pay']
        # Append the department summary row
        header_data = [
            dept.name,
            "",
            "",
            "",
            "",
            format_currency(total_fixed_salary),
            "",
            format_currency(total_gross_salary),
            format_currency(total_deduction),
            format_currency(total_netpay)
        ]
        total_fixed_salary_+=total_fixed_salary
        total_gross_salary_+=total_gross_salary
        total_deduction_+=total_deduction
        total_netpay_+=total_netpay
        ws.append(header_data)
        for cell in ws[ws.max_row]:  
            row=ws.max_row
            sheet=ws          
            apply_common_styles(cell,header_font_data, align_left, border,align_right,row,sheet)
        # Process each child department
        
        for child_dept in child_departments:
            fixed_child_data=0
            salary_slip_data = frappe.db.sql("""
                    SELECT employee, salary_structure 
                    FROM `tabSalary Slip`
                    WHERE start_date = %s AND end_date = %s 
                    AND docstatus != 2 AND department = %s 
                    GROUP BY employee 
                    ORDER BY employee
                """, (from_date, to_date, child_dept.name), as_dict=True)
            for slip in salary_slip_data:
                fixed_child_base = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': slip['employee'], 'salary_structure': slip['salary_structure'], 'docstatus': 1},
                    'base', order_by='creation desc'
                ) or 0.0
                fixed_child_var = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': slip['employee'], 'salary_structure': slip['salary_structure'], 'docstatus': 1},
                    'variable', order_by='creation desc'
                ) or 0.0
                fixed_child = fixed_child_base + fixed_child_var
                fixed_child_data+=fixed_child
            salary_slip_list_data = frappe.db.sql("""
                SELECT department, employee,
                    SUM(gross_pay) AS dept_gross_pay, 
                    SUM(total_deduction) AS dept_total_deduction, 
                    SUM(net_pay) AS dept_net_pay, 
                    salary_structure 
                FROM `tabSalary Slip`
                WHERE start_date = %s AND end_date = %s 
                AND docstatus != 2 AND department = %s 
                GROUP BY department 
                ORDER BY department DESC
            """, (from_date, to_date, child_dept.name), as_dict=True)

            for s in salary_slip_list_data:
                
                child_row = [
                    child_dept.name,
                    "",
                    "",
                    "",
                    "",
                    format_currency(fixed_child_data),
                    "",
                    format_currency(s['dept_gross_pay']),
                    format_currency(s['dept_total_deduction']),
                    format_currency(s['dept_net_pay'])
                ]
                ws.append(child_row)

                # Apply styles to the child department row
                for cell in ws[ws.max_row]:
                    row=ws.max_row
                    sheet=ws
                    apply_common_styles_dept_font(cell,dept_font, align_left, border,align_right,row,sheet)

            # Process salary slips for each employee in the child department
            salary_slips = frappe.db.sql("""
                SELECT employee, employee_name, department, designation, payment_days, gross_pay, total_deduction, net_pay, salary_structure 
                FROM `tabSalary Slip`
                WHERE start_date = %s AND end_date = %s 
                AND docstatus != 2 AND department = %s 
                ORDER BY employee
            """, (from_date, to_date, child_dept.name), as_dict=True)

            for salary_slip in salary_slips:
                index += 1
                from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")  # Convert string to date object
                month_year = from_date_obj.strftime("%b %Y")  # Format to "Feb 2025"
                pr_score=frappe.db.get_value("Appraisal",{"employee":salary_slip['employee'],"custom_appraisal_cycle_month":month_year,"docstatus":1},["total_score"])
                # advance = frappe.db.sql("""
                #    SELECT SUM(c.debit_in_account_currency) AS debit
                #     FROM `tabJournal Entry` AS j
                #     INNER JOIN `tabJournal Entry Account` AS c ON j.name = c.parent
                #     WHERE c.party_type = "Employee" 
                #     AND c.account IN ("Staff Advance - THIS","staff advance - TFP")
                #     AND (c.reference_type IS NULL OR c.reference_type != "Expense Claim") 
                #     AND c.party =%s
                #     AND j.docstatus = 1;

                # """, (salary_slip['employee'],), as_list=True) 
               
                advance = frappe.db.sql("""
                    SELECT 
                    SUM(j.amount_in_account_currency) AS outstanding_advance
                    FROM `tabPayment Ledger Entry` j
                    WHERE 
                        j.party_type = 'Employee'
                        AND j.account IN ('Staff Advance - THIS', 'Staff Advance - TFP')
                        AND j.party = %s
                        AND j.delinked =0
                        AND j.account_type='Payable'
                """, (salary_slip['employee'],), as_list=True)

                # emp_advance = advance[0]['outstanding_advance'] if advance else 0.0
                # total_advance += emp_advance


                total_advance+=advance[0][0] if advance and advance[0][0] is not None else 0.0
                emp_advance = advance[0][0] if advance and advance[0][0] is not None else 0.0

                date_of_joining = frappe.db.get_value(
                    'Employee',
                    {'name': salary_slip['employee']},
                    'date_of_joining'
                )
                formatted_doj = datetime.strptime(str(date_of_joining), '%Y-%m-%d')
                doj = formatted_doj.strftime('%d-%m-%Y')
                fixed_base = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': salary_slip['employee'], 'salary_structure': salary_slip['salary_structure']},
                    'base'
                ) or 0.0
                fixed_var = frappe.db.get_value(
                    'Salary Structure Assignment',
                    {'employee': salary_slip['employee'], 'salary_structure': salary_slip['salary_structure']},
                    'variable'
                ) or 0.0
                fixed = fixed_base+fixed_var
                row = [
                    index,
                    salary_slip['employee'],
                    salary_slip['employee_name'],
                    doj,
                    salary_slip['designation'],
                    format_currency(fixed),
                    salary_slip['payment_days'],
                    format_currency(salary_slip['gross_pay']),
                    format_currency(salary_slip['total_deduction']),
                    format_currency(salary_slip['net_pay']),
                    "",
                    pr_score,
                    emp_advance
                ]
                ws.append(row)
                status = frappe.db.get_value(
                    'Employee',
                    {'name': salary_slip['employee']},
                    'status'
                )
                # Apply styles to the data row
                if status!='Active':
                    for cell in ws[ws.max_row]:
                        row=ws.max_row
                        sheet=ws
                        apply_common_styles_left_employees(cell, left_emp_font,align_left, align_right,border,sheet,row)
                else:
                    for cell in ws[ws.max_row]:
                        row=ws.max_row
                        sheet=ws
                        apply_common_styles_active_employees(cell, active_emp_font,align_left, border,align_right,sheet,row)

    
    
    total_data = [
        "Total",
        "",
        "",
        "",
        "",
        format_currency(total_fixed_salary_),
        "",
        format_currency(total_gross_salary_),
        format_currency(total_deduction_),
        format_currency(total_netpay_),
        "",
        "",
        format_currency(total_advance)

    ]
    ws.append(total_data)
    for cell in ws[ws.max_row]:  
        row=ws.max_row
        sheet=ws          
        apply_common_styles_for_total(cell,header_font_data, align_center, border,align_right,row,sheet)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def build_xlsx_response(filename):  
    xlsx_file = make_xlsx(sheet_name=filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def format_currency(value):
    if value is None:
        return "0"
    
    number_str = str(int(value))
    
    if len(number_str) > 3:
        last_three = number_str[-3:]  
        other_digits = number_str[:-3] 
        
        formatted_other = []
        while len(other_digits) > 2:
            formatted_other.append(other_digits[-2:])
            other_digits = other_digits[:-2]
        if other_digits:
            formatted_other.append(other_digits)        
        formatted_other.reverse()
        formatted_number = ','.join(formatted_other) + ',' + last_three
    else:
        formatted_number = number_str
    
    return formatted_number