# Copyright (c) 2024, TeamPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import frappe
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from io import BytesIO
from frappe.utils.file_manager import save_file


class FormattedReportsDownload(Document):
    pass

@frappe.whitelist()
def batch_status_report(batch_customer=None,batch=None):
    s_no = 1
    # filters = []
    # if batch_customer:
    #     filters.append(f"batch_customer = '{batch_customer}'")
    # if batch:
    #     filters.append(f"batch = '{batch}'")
    am = batch_customer
    dm = batch
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="18" style="text-align:center; font-weight:bold;">PR:04 – Batch Status Report (BSR)</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Initiation Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Batch Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Billing Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Completion Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">No of Cases</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Case</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Pending</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Comp</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Insuff</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">No of Checks</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Ch_Pending</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Ch_Comp</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#Ch_Insuff</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">EV</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ECD</td>' \
            '</tr>'

    query = """
    SELECT 
            s.name AS name, 
            s.customer AS customer, 
            s.expected_start_date AS expected_start_date, 
            s.batch_status AS batch_status, 
            s.billing_status AS billing_status, 
            s.expected_end_date AS expected_end_date, 
            s.no_of_cases AS no_of_cases, 
            s.`case` AS `case`,  -- Escaping the 'case' column
            s.pending AS pending, 
            s.comp AS comp, 
            s.insuff AS insuff,
            s.no_of_checks AS no_of_checks
        FROM 
            `tabBatch` s
    """
    
    # List to hold conditions
    conditions = []

    # Add conditions based on parameters
    if am:
        conditions.append("customer = %s")
        conditions.append(am)

    if dm:
        conditions.append("name = %s")
        conditions.append(dm)

    # Add conditions to the query if any
    if conditions:
        query += " WHERE " + " AND ".join(conditions[::2])

    # Execute the query with the parameters
    batch_data = frappe.db.sql(query, tuple(conditions[1::2]), as_dict=True)


    for i in batch_data:
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.customer}</td>' \
                f'<td style="text-align:center;">{i.expected_start_date}</td>' \
                f'<td style="text-align:center;">{i.batch_status}</td>' \
                f'<td style="text-align:center;">{i.billing_status}</td>' \
                f'<td style="text-align:center;">{i.expected_end_date}</td>' \
                f'<td style="text-align:center;">{i.no_of_cases}</td>' \
                f'<td style="text-align:center;">{i.case}</td>' \
                f'<td style="text-align:center;">{i.pending}</td>' \
                f'<td style="text-align:center;">{i.comp}</td>' \
                f'<td style="text-align:center;">{i.insuff}</td>' \
                f'<td style="text-align:center;">{i.no_of_checks}</td>' \
                f'<td style="text-align:center;"></td>' \
                f'<td style="text-align:center;"></td>' \
                f'<td style="text-align:center;"></td>' \
                f'<td style="text-align:center;"></td>' \
                f'<td style="text-align:center;"></td>' \
                '</tr>'
        s_no += 1
    
    data += '</table>'
    return data

import frappe
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

@frappe.whitelist()
def download_bcs_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_bcs = f"PR-04 – Batch Status Report (BSR) {posting_date}"
    build_xlsx_response_bcs(filename_bcs)
    
def build_xlsx_response_bcs(filename_bcs):
    xlsx_file = make_xlsx_bcs(filename_bcs)
    frappe.response["filename"] = f"{filename_bcs}.xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"

def make_xlsx_bcs(data, sheet_name="PR:04 – Batch Status Report (BSR)", wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    # Replace invalid characters in the sheet name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)

    # Define styles
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")  # Center alignment
    alignment_left = Alignment(horizontal="left", vertical="center")      # Left alignment
    text_wrap = Alignment(wrap_text=True, horizontal="left", vertical="center")  # Text wrap
    title_font = Font(bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)  # White font for headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # Set column widths
    column_widths = {
        'A': 5, 'B': 20, 'C': 30, 'D': 20, 'E': 20, 'F': 20,
        'G': 20, 'H': 20, 'I': 10, 'J': 10, 'K': 10, 'L': 10,
        'M': 10, 'N': 15, 'O': 10, 'P': 10, 'Q': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add title
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR:04 – Batch Status Report (BSR)(" + posting_date + ")"
    ws.merge_cells("A1:R1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment_center

    # Add header row
    header = [
        "S NO", "ID", "Customer", "Initiation Date", "Batch Status", "Billing Status",
        "Completion Date","#Case", "#Pending", "#Comp", "#Insuff",
        "#Checks", "#Ch_Pending", "#Ch_Comp", "#Ch_Insuff", "EV", "ECD"
    ]
    ws.append(header)

    # Style header row
    for cell in ws[2]:  # Assuming header is in row 2
        cell.fill = fill_color
        cell.font = header_font
        cell.alignment = alignment_center  # Center alignment for headers
        cell.border = thin_border

    # Add data rows
    data1 = get_data_of_bcs()
    for row in data1:
        ws.append(row)

        # Style the last added row
        for idx, cell in enumerate(ws[ws.max_row], start=1):
            if idx == 3:  # If it's the Customer column, wrap text
                cell.alignment = text_wrap
                cell.border = thin_border
            else:
                cell.alignment = alignment_left  # Left alignment for other columns
                cell.border = thin_border

    # Save to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def get_data_of_bcs():
    data = []
    s_no = 1
    # am = args.get("batch_customer")
    # dm = args.get("batch")
    
    # Initial part of the query (without WHERE clause)
    batch_data = frappe.db.sql("""
    SELECT 
        s.name AS name, 
        s.customer AS customer, 
        s.expected_start_date AS expected_start_date, 
        s.batch_status AS batch_status, 
        s.billing_status AS billing_status, 
        s.expected_end_date AS expected_end_date, 
        s.no_of_cases AS no_of_cases, 
        s.pending AS pending, 
        s.comp AS comp, 
        s.insuff AS insuff,
        s.no_of_checks AS no_of_checks,
        COUNT(CASE WHEN cs.verification_status = 'Pending' THEN 1 END) AS pending_check,
        COUNT(CASE WHEN cs.verification_status = 'Completed' THEN 1 END) AS completed,
        COUNT(CASE WHEN cs.verification_status = 'Insufficient' THEN 1 END) AS insuff_check
    FROM 
        `tabBatch` s
    LEFT JOIN `tabCheckwise Report` cs 
        ON s.name = cs.parent
    GROUP BY 
        s.name

    """,as_dict=True)

    
    for i in batch_data:
        formatted_date = frappe.utils.formatdate(i.expected_start_date, 'dd-mm-yyyy')
        end_date=frappe.utils.formatdate(i.expected_end_date, 'dd-mm-yyyy')
        data.append([ 
            s_no, i.name, i.customer,formatted_date, i.batch_status, i.billing_status,
            end_date, i.no_of_cases,float(i.pending), float(i.comp), float(i.insuff),
            i.no_of_checks,i.pending_check,i.completed,i.insuff_check, '',''
        ])
        s_no += 1

    return data

@frappe.whitelist()
def download_bcs_report_cust():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_bcs = f"PR-04 – Batch Status Report (BSR) {posting_date}"
    build_xlsx_response_bcs_cust(filename_bcs)
    
def build_xlsx_response_bcs_cust(filename_bcs):
    xlsx_file = make_xlsx_bcs_cust(filename_bcs)
    frappe.response["filename"] = f"{filename_bcs}.xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"

def make_xlsx_bcs_cust(data, sheet_name="PR:04 – Batch Status Report (BSR)", wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    # Replace invalid characters in the sheet name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)

    # Define styles
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")  # Center alignment
    alignment_left = Alignment(horizontal="left", vertical="center")      # Left alignment
    text_wrap = Alignment(wrap_text=True, horizontal="left", vertical="center")  # Text wrap
    title_font = Font(bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)  # White font for headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # Set column widths
    column_widths = {
        'A': 5, 'B': 20, 'C': 30, 'D': 20, 'E': 20, 'F': 20,
        'G': 20, 'H': 20, 'I': 10, 'J': 10, 'K': 10, 'L': 10,
        'M': 10, 'N': 15, 'O': 10, 'P': 10, 'Q': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add title
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR:04 – Batch Status Report (BSR)(" + posting_date + ")"
    ws.merge_cells("A1:R1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment_center

    # Add header row
    header = [
        "S NO", "ID", "Customer", "Initiation Date", "Batch Status", "Billing Status",
        "Completion Date","#Case", "#Pending", "#Comp", "#Insuff",
        "#Checks", "#Ch_Pending", "#Ch_Comp", "#Ch_Insuff", "EV", "ECD"
    ]
    ws.append(header)

    # Style header row
    for cell in ws[2]:  # Assuming header is in row 2
        cell.fill = fill_color
        cell.font = header_font
        cell.alignment = alignment_center  # Center alignment for headers
        cell.border = thin_border

    # Add data rows
    data1 = get_data_of_bcs_cust()
    for row in data1:
        ws.append(row)

        # Style the last added row
        for idx, cell in enumerate(ws[ws.max_row], start=1):
            if idx == 3:  # If it's the Customer column, wrap text
                cell.alignment = text_wrap
                cell.border = thin_border
            else:
                cell.alignment = alignment_left  # Left alignment for other columns
                cell.border = thin_border

    # Save to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_bcs_cust():
    data = []
    s_no = 1
    customer_name=frappe.db.get_single_value("Formatted Reports  Download","batch_customer")
    batch_data = frappe.db.sql("""
    SELECT 
        s.name AS name, 
        s.customer AS customer, 
        s.expected_start_date AS expected_start_date, 
        s.batch_status AS batch_status, 
        s.billing_status AS billing_status, 
        s.expected_end_date AS expected_end_date, 
        s.no_of_cases AS no_of_cases, 
        s.pending AS pending, 
        s.comp AS comp, 
        s.insuff AS insuff,
        s.no_of_checks AS no_of_checks,
        COUNT(CASE WHEN cs.verification_status = 'Pending' THEN 1 END) AS pending_check,
        COUNT(CASE WHEN cs.verification_status = 'Completed' THEN 1 END) AS completed,
        COUNT(CASE WHEN cs.verification_status = 'Insufficient' THEN 1 END) AS insuff_check
    FROM 
        `tabBatch` s
    LEFT JOIN `tabCheckwise Report` cs 
        ON s.name = cs.parent
    WHERE customer=%s
    GROUP BY 
        s.name

    """,customer_name,as_dict=True)

    for i in batch_data:
        formatted_date = frappe.utils.formatdate(i.expected_start_date, 'dd-mm-yyyy')
        end_date=frappe.utils.formatdate(i.expected_end_date, 'dd-mm-yyyy')
        data.append([ 
            s_no, i.name, i.customer,formatted_date, i.batch_status, i.billing_status,
            end_date, i.no_of_cases,i.pending, i.comp, i.insuff,
            i.no_of_checks,i.pending_check,i.completed,i.insuff_check, '',''
        ])
        s_no += 1

    return data

@frappe.whitelist()
def download_bcs_report_batch():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_bcs = f"PR-04 – Batch Status Report (BSR) {posting_date}"
    build_xlsx_response_bcs_batch(filename_bcs)
    
def build_xlsx_response_bcs_batch(filename_bcs):
    xlsx_file = make_xlsx_bcs_batch(filename_bcs)
    frappe.response["filename"] = f"{filename_bcs}.xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"

def make_xlsx_bcs_batch(data, sheet_name="PR:04 – Batch Status Report (BSR)", wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    # Replace invalid characters in the sheet name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)

    # Define styles
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")  # Center alignment
    alignment_left = Alignment(horizontal="left", vertical="center")      # Left alignment
    text_wrap = Alignment(wrap_text=True, horizontal="left", vertical="center")  # Text wrap
    title_font = Font(bold=True, size=14)
    header_font = Font(color="FFFFFF", bold=True)  # White font for headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    # Set column widths
    column_widths = {
        'A': 5, 'B': 20, 'C': 30, 'D': 20, 'E': 20, 'F': 20,
        'G': 20, 'H': 20, 'I': 10, 'J': 10, 'K': 10, 'L': 10,
        'M': 10, 'N': 15, 'O': 10, 'P': 10, 'Q': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add title
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR:04 – Batch Status Report (BSR)(" + posting_date + ")"
    ws.merge_cells("A1:R1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment_center

    # Add header row
    header = [
        "S NO", "ID", "Customer", "Initiation Date", "Batch Status", "Billing Status",
        "Completion Date","#Case", "#Pending", "#Comp", "#Insuff",
        "#Checks", "#Ch_Pending", "#Ch_Comp", "#Ch_Insuff", "EV", "ECD"
    ]
    ws.append(header)

    # Style header row
    for cell in ws[2]:  # Assuming header is in row 2
        cell.fill = fill_color
        cell.font = header_font
        cell.alignment = alignment_center  # Center alignment for headers
        cell.border = thin_border

    # Add data rows
    data1 = get_data_of_bcs_batch()
    for row in data1:
        ws.append(row)

        # Style the last added row
        for idx, cell in enumerate(ws[ws.max_row], start=1):
            if idx == 3:  # If it's the Customer column, wrap text
                cell.alignment = text_wrap
                cell.border = thin_border
            else:
                cell.alignment = alignment_left  # Left alignment for other columns
                cell.border = thin_border

    # Save to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_bcs_batch():
    data = []
    s_no = 1
    batch_name=frappe.db.get_single_value("Formatted Reports  Download","batch")
    batch_data = frappe.db.sql("""
    SELECT 
        s.name AS name, 
        s.customer AS customer, 
        s.expected_start_date AS expected_start_date, 
        s.batch_status AS batch_status, 
        s.billing_status AS billing_status, 
        s.expected_end_date AS expected_end_date, 
        s.no_of_cases AS no_of_cases, 
        s.pending AS pending, 
        s.comp AS comp, 
        s.insuff AS insuff,
        s.no_of_checks AS no_of_checks,
        COUNT(CASE WHEN cs.verification_status = 'Pending' THEN 1 END) AS pending_check,
        COUNT(CASE WHEN cs.verification_status = 'Completed' THEN 1 END) AS completed,
        COUNT(CASE WHEN cs.verification_status = 'Insufficient' THEN 1 END) AS insuff_check
    FROM 
        `tabBatch` s
    LEFT JOIN `tabCheckwise Report` cs 
        ON s.name = cs.parent
    WHERE s.name=%s
    GROUP BY 
        s.name

    """,batch_name,as_dict=True)

    for i in batch_data:
        formatted_date = frappe.utils.formatdate(i.expected_start_date, 'dd-mm-yyyy')
        end_date=frappe.utils.formatdate(i.expected_end_date, 'dd-mm-yyyy')
        data.append([ 
            s_no, i.name, i.customer,formatted_date, i.batch_status, i.billing_status,
            end_date, i.no_of_cases,i.pending, i.comp, i.insuff,
            i.no_of_checks,i.pending_check,i.completed,i.insuff_check, '',''
        ])
        s_no += 1

    return data

@frappe.whitelist()
def print_sales_invoice_outstanding_report(account_manager=None, delivery_manager=None, service=None, company=None):
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="12" style="text-align:center; font-weight:bold;">PR:06 – Collection Pending Report – SI Outstanding (CPR)</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Company</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Service</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">AM</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Grand Total</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Outstanding</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">DM</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer Group</td>' \
            '</tr>'
    
    filters = []
    if account_manager:
        filters.append(f"account_manager = '{account_manager}'")
    if delivery_manager:
        filters.append(f"delivery_manager = '{delivery_manager}'")
    if service:
        filters.append(f"services = '{service}'")
    if company:
        filters.append(f"company = '{company}'")
    
    filter_query = " AND ".join(filters)
    
    query = f"""
        SELECT 
            s.name AS name, 
            c.abbr AS company, 
            s.services AS services, 
            s.status AS status, 
            s.customer AS customer, 
            account_manager.short_code AS account_manager, 
            s.base_grand_total AS total, 
            s.outstanding_amount AS outstanding_amount, 
            s.posting_date AS posting_date, 
            delivery_manager.short_code AS delivery_manager, 
            s.customer_group AS customer_group
        FROM 
            `tabSales Invoice` s
        INNER JOIN 
            `tabCompany` c ON c.name = s.company
        LEFT JOIN 
            `tabEmployee` account_manager ON account_manager.user_id = s.account_manager
        LEFT JOIN 
            `tabEmployee` delivery_manager ON delivery_manager.user_id = s.delivery_manager
        WHERE 
            s.status NOT IN ('Return', 'Paid', 'Credit Note Issued', 'Cancelled')
            {f"AND {filter_query}" if filter_query else ""}
    """
    
    sales_invoices = frappe.db.sql(query, as_dict=True)
    
    total_of_total = total_outstanding = 0
    for i in sales_invoices:
        formatted_date = frappe.utils.formatdate(i.posting_date, 'dd-mm-yyyy')
        
        data += f'<tr>' 
        data += f'<td style="text-align:center;">{s_no}</td>' 
        data += f'<td style="text-align:center;">{i.name}</td>' 
        data += f'<td style="text-align:center;">{i.company}</td>' 
        data += f'<td style="text-align:center;">{i.services}</td>' 
        data += f'<td style="text-align:center;">{i.status}</td>' 
        data += f'<td style="text-align:left;">{i.customer}</td>' 
        data += f'<td style="text-align:center;">{i.account_manager}</td>' 
        data += f'<td style="text-align:center;">{i.total:,.2f}</td>' 
        data += f'<td style="text-align:center;">{i.outstanding_amount:,.2f}</td>' 
        data += f'<td style="text-align:center;">{formatted_date}</td>' 
        data += f'<td style="text-align:center;">{i.delivery_manager}</td>' 
        data += f'<td style="text-align:center;">{i.customer_group}</td>' 
        data += '</tr>'
        
        s_no += 1
        total_of_total += i.total
        total_outstanding += i.outstanding_amount
        total_of_total = round(total_of_total, 2)
    
    data += f'<tr style="background-color: #f79646;">' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:left;">Total</td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;">{total_of_total:,.2f}</td>' \
            f'<td style="text-align:center;">{total_outstanding:,.2f}</td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            '</tr>'
            
    data += '</table>'
    return data


from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import frappe

@frappe.whitelist()
def download_sales_invoice_outstanding_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR-06 – Collection Pending Report – SI Outstanding (CPR)" + posting_date
    build_xlsx_response_si(filename)

def build_xlsx_response_si(filename):
    xlsx_file = make_xlsx_si(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_si(sheet_name="PR-06 – Collection Pending Report – SI Outstanding (CPR)", wb=None):
    args = frappe.local.form_dict
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name

    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    alignment_header = Alignment(horizontal="center", vertical="center")
    text_wrap = Alignment(wrap_text=True, horizontal="left", vertical="center")
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
    inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    column_widths = [6, 15, 10, 9, 10, 45, 6,5, 13, 13, 14, 6, 17]
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L","M"]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[columns[i]].width = width
        
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR-06 – Collection Pending Report – SI Outstanding (CPR)(" + posting_date + ")"
    ws.merge_cells("A1:L1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment_header
    
    headers = ["S NO", "ID", "Company", "Services", "Status", "Customer Name",
               "AM","Age", "Grand Total", "Outstanding", "Date", "DM", "Customer Group"]
    ws.append(headers)

    for cell in ws[2]:
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment_header
        cell.border = thin_border

    total_invoice = 0
    total_outstanding = 0
    data = get_data_si(args)

    for row in data:
        ws.append(row)
        ws[f"H{ws.max_row}"].style = inr_format
        ws[f"I{ws.max_row}"].style = inr_format
        for cell in ws[ws.max_row]:
            cell.alignment = text_wrap
            cell.border = thin_border
            
        total_invoice += row[8]  
        total_outstanding += row[9] 

    ws.append(["", "", "", "", "", "Total", "","", total_invoice, total_outstanding, "", "", ""])

    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = bg_fill
        cell.border = thin_border
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):  # Adjust row range as needed
        for cell in row:
            col = cell.column_letter
            if col in ("I","J"):  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_si(args):
    am = args.get("account_manager")
    dm = args.get("delivery_manager")
    ser = args.get("service")
    comp = args.get("company")

    data = []
    s_no = 1

    filters = {"status": ["not in", ["Return", "Paid", "Credit Note Issued", "Cancelled"]]}

    if am:
        filters["account_manager"] = am
    if dm:
        filters["delivery_manager"] = dm
    if ser:
        filters["services"] = ser
    if comp:
        filters["company"] = comp

    sales_invoice = frappe.db.get_all(
        "Sales Invoice", filters=filters,
        fields=["name", "company", "services", "status", "customer", "account_manager", 
                "base_grand_total", "outstanding_amount", "posting_date", "delivery_manager", "customer_group"]
    )

    for invoice in sales_invoice:
        account = frappe.db.get_value("Employee", {"user_id": invoice["account_manager"]}, "short_code") or ""
        delivery = frappe.db.get_value("Employee", {"user_id": invoice["delivery_manager"]}, "short_code") or ""
        abbr = frappe.db.get_value("Company", {"name": invoice["company"]}, "abbr") or ""
        transaction_date = invoice.posting_date  # assuming this is a datetime/date object already
        todate = date.today()

        # Make sure both are date objects
        if isinstance(transaction_date, str):
            transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()

        age = str((todate - transaction_date).days)
        formatted_date = frappe.utils.formatdate(invoice["posting_date"], "dd-mm-yyyy")
        data.append([
            s_no, invoice["name"], abbr, invoice["services"], invoice["status"],
            invoice["customer"], account,age, invoice["base_grand_total"], 
            invoice["outstanding_amount"], formatted_date, delivery, 
            invoice["customer_group"]
        ])
        s_no += 1

    return data


@frappe.whitelist()
def opportunity_report(opportunity_owner=None, opp_am=None, opp_service=None):
    s_no = 1
    data = '<table border="1" style="width: 100%; white-space: nowrap;">'
    data += '<tr><td colspan="12" style="text-align:center; font-weight:bold; font-size: 15px;">PR-01 – Opportunity Status Report (OSR)</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Owner</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Service</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">From</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 20%;">Organization Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">Age</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Amount</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">PB%</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">ECD</td>' \
            '<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">Remarks</td>' \
            '</tr>'

    filters = {"status": ["in", ["Open", "Quotation", "Replied"]]}

    if opportunity_owner:
        filters["opportunity_owner"] = opportunity_owner

    if opp_am:
        filters["lead_owner"] = opp_am

    if opp_service:
        filters["service"] = opp_service

    opportunity_data = frappe.db.get_all("Opportunity", filters if (opportunity_owner or opp_am or opp_service) else {}, "*")

    grouped_data = {}
    for opportunity in opportunity_data:
        owner = opportunity.get("opportunity_owner")
        if owner not in grouped_data:
            grouped_data[owner] = []
        grouped_data[owner].append(opportunity)

    for owner, opportunities in grouped_data.items():
        # Calculate rowspan (number of opportunities for this owner)
        rowspan = len(opportunities)
        short_code = frappe.db.get_value("Employee", {"user_id": owner}, "short_code")

        # Add the first row with rowspan
        first_opportunity = opportunities[0]
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;" rowspan="{rowspan}">{short_code or owner}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("service")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("opportunity_from")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("status")}</td>' \
                f'<td style="text-align:left;">{first_opportunity.get("organization_name")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("transaction_date")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("custom_opportunity_age")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("opportunity_amount")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("probability")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("expected_closing")}</td>' \
                f'<td style="text-align:center;">{first_opportunity.get("remark")}</td>' \
                '</tr>'
        s_no += 1

        # Add remaining rows for the same owner
        for opportunity in opportunities[1:]:
            data += f'<tr>' \
                    f'<td style="text-align:center;">{s_no}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("service")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("opportunity_from")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("status")}</td>' \
                    f'<td style="text-align:left;">{opportunity.get("organization_name")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("transaction_date")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("custom_opportunity_age")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("opportunity_amount")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("probability")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("expected_closing")}</td>' \
                    f'<td style="text-align:center;">{opportunity.get("remark")}</td>' \
                    '</tr>'
            s_no += 1

    data += '</table>'
    return data


from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
import frappe

@frappe.whitelist()
def opportunity_excel_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_opp = "PR:01 – Opportunity Status Report (OSR)_" + posting_date
    build_xlsx_response_opp(filename_opp)

def build_xlsx_response_opp(filename_opp):
    xlsx_file = make_xlsx_opp(filename_opp)
    frappe.response['filename'] = filename_opp + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_opp(data, sheet_name="PR:01 – Opportunity Status Report (OSR)", wb=None, column_widths=None):
    default_column_widths = [7, 7, 10, 13, 10, 40, 13, 7, 10, 7, 13, 20]
    column_widths = column_widths or default_column_widths

    args = frappe.local.form_dict

    if wb is None:
        wb = Workbook()
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)

    # Define styles
    text_wrap_left = Alignment(wrap_text=True, vertical="center", horizontal="left")
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    bold_font = Font(bold=True)
    # bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    column_widths = [6, 8, 10, 20, 10, 35, 15, 6, 13, 13, 15, 45]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[columns[i]].width = width

    # Title
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR:01 – Opportunity Status Report (OSR) (" + posting_date + ")"
    ws.merge_cells("A1:L1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = center_alignment  # Center the title

    # Headers
    header = ["S NO", "Owner", "Service", "From", "Status", "Organization Name",
              "Date", "Age", "Amount", "PB%", "ECD", "Remark"]
    ws.append(header)

    for cell in ws[2]:
        cell.fill = fill_color
        cell.font = header_font
        cell.alignment = center_alignment  # Center the headers
        cell.border = thin_border

    # Data
    data1 = get_data_of_opp(args)
    current_row = 3
    start_row = current_row
    last_owner = data1[0][1] if data1 else None

    for row in data1:
        ws.append(row)
        owner = row[1]

        if owner != last_owner and start_row < current_row - 1:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
            ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
            start_row = current_row

        last_owner = owner
        for cell in ws[ws.max_row]:
            cell.alignment = text_wrap_left
            cell.border = thin_border

        current_row += 1

    if start_row < current_row - 1:
        ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
        ws.cell(row=start_row, column=2).alignment = Alignment(horizontal="left", vertical="center")

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.alignment = text_wrap_left
            cell.border = thin_border

    # Adjust row heights for added vertical spacing
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 50  # Adjust height for spacing

    # for cell in ws[ws.max_row]:
    #     cell.font = bold_font
        # cell.fill = bg_fill
        # cell.border = thin_border
        # cell.alignment = text_wrap_left

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def get_data_of_opp(args):

    own = args.get("opportunity_owner")
    am = args.get("opp_am")
    ser = args.get("opp_service")

    data = []
    s_no = 1
    filters = {"status": ["in", ["Open", "Quotation", "Replied"]]}

    if own:
        filters["opportunity_owner"] = own

    if am:
        filters["lead_owner"] = am

    if ser:
        filters["service"] = ser

    opportunity_data = frappe.db.get_all("Opportunity", filters if (own or am or ser) else {}, "*")
    grouped_data = {}
    
    for opportunity in opportunity_data:
        owner = opportunity.get("opportunity_owner")
        if owner not in grouped_data:
            grouped_data[owner] = []
        grouped_data[owner].append(opportunity)

    
    for owner, opportunities in grouped_data.items():
        for i in opportunities:
            employee_short_code = frappe.db.get_value("Employee", {"user_id": owner}, "short_code")
            formatted_transaction_date = frappe.utils.formatdate(i.transaction_date, 'dd-mm-yyyy')
            formatted_ecd_date = frappe.utils.formatdate(i.expected_closing, 'dd-mm-yyyy')
            data.append([
                s_no, 
                employee_short_code, 
                i.service, 
                i.opportunity_from, 
                i.status, 
                i.organization_name, 
                formatted_transaction_date,
                i.custom_opportunity_age, 
                i.opportunity_amount, 
                i.probability, 
                formatted_ecd_date,
                i.remark
            ])
            s_no += 1

    return data


# @frappe.whitelist()
# def download_PTSR():
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     filename = "MR:03 – Project Task Status Report – REC (PTSR - R)" + posting_date
#     build_xlsx_response_PTSR(filename)

# def build_xlsx_response_PTSR(filename):
#     xlsx_file = make_xlsx_PTSR(filename)
#     frappe.response['filename'] = filename + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'

# def make_xlsx_PTSR(sheet_name="MR:03 – Project Task Status Report – REC (PTSR - R)", wb=None, column_widths=None):
#     if wb is None:
#         wb = openpyxl.Workbook()
#     # ws = wb.create_sheet(sheet_name, 0)
#     valid_sheet_name = sheet_name.replace(":", "-")
#     ws = wb.create_sheet(valid_sheet_name, 0)
#     default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15,15,15,15,15,15, 15, 15, 25, 15, 7, 7, 7, 7, 7, 7]
#     column_widths = column_widths or default_column_widths
#     for i, width in enumerate(column_widths, start=1):
#         ws.column_dimensions[get_column_letter(i)].width = width
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     ftitle = "REC : Project – Task Status Report : - " + posting_date
#     ws.append([ftitle])
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
#     title_cell = ws.cell(row=1, column=1)
#     title_cell.alignment = Alignment(horizontal="center", vertical="center")
#     header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
#     header_font = Font(color="FFFFFF")
#     headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark",'SPOC Remark', 'Exp Value', 'Exp PSL','Completed Value','Cr. Exp. Value','Cr.Exp.PSL','Exp.Week', 'Sourcing Status', 'Territory', 'TASK', 'Task Priority', '#VAC', '#SP', '#FP', '#SL', '#PSL', '#LP']
#     black_border = Border(
#         left=Side(border_style="thin", color="000000"),
#         right=Side(border_style="thin", color="000000"),
#         top=Side(border_style="thin", color="000000"),
#         bottom=Side(border_style="thin", color="000000")
#     )
#     ws.append(headers)
#     ws.freeze_panes = "A3"
#     header_row = ws[ws.max_row]
#     for cell in header_row:
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = black_border
#     cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
#     row = 3
#     serial_number = 1
#     grand_totals = {'vac':0,'sp': 0,'fp': 0,'sl':0,'psl':0,'custom_lp':0}
#     for c in cust:
#         priority = {"High": 1, "Medium": 2, "Low": 3}
#         pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'],order_by= "priority ASC")
#         if not pname:
#             continue
#         task_totals = {'vac':0,'sp':0,'fp':0,'sl':0,'psl':0,'custom_lp':0}
#         project_data = []      
#         for p in pname:
#             pdata = []
#             print(p.project_name)        
#             taskid = frappe.get_all("Task", {"status": ("in",('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'],order_by= "priority ASC")              
#             # print(p['project_name'])
#             # for tn in taskid:
#                 # print(tn.name)
#             for t in taskid:
#                 pdata.append([p['project_name'] if p['project_name'] else "",p['priority'] if p['priority'] else "",p['remark'] if p['remark'] else "",p['account_manager_remark'] if p['account_manager_remark'] else "",p['custom_spoc_remark'] if p['custom_spoc_remark'] else "",p['expected_value'] if p['expected_value'] else "",p['expected_psl'] if p['expected_psl'] else "",p['sourcing_statu'] if p['sourcing_statu'] else "",p['territory'] if p['territory'] else "",t['subject'],t['priority'],t['vac'],t['sp'],t['fp'],t['sl'],t['psl'],t['custom_lp']])
#                 task_totals['vac'] +=t['vac']
#                 task_totals['sp'] +=t['sp']
#                 task_totals['fp']+= t['fp']
#                 task_totals['sl'] +=t['sl']
#                 task_totals['psl'] += t['psl']
#                 task_totals['custom_lp'] += t['custom_lp']
#             project_data.append({
#                 'project_name': p['project_name'],'priority': p['priority'],
#                 'remark': p['remark'],'account_manager_remark': p['account_manager_remark'],'custom_spoc_remark':p['custom_spoc_remark'],'sourcing_statu': p['sourcing_statu'],'territory': p['territory'],
#                 'expected_value': p['expected_value'],'expected_psl': p['expected_psl'],'tasks': pdata})
#         blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
#         row_data = [serial_number, c['name']] + [""] * 10 + [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]
#         ws.append(row_data)
#         row_to_fill = ws.max_row
#         for col, cell in enumerate(ws[row_to_fill], start=1):
#             cell.fill = blue_fill
#             if col > 11:
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#             else:
#                 cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
#             cell.border = black_border
#         ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
#         serial_number += 1
#         row += 1
#         current_row_start = row
#         for project in project_data:
#             project_row_start = row
#             for task_data in project['tasks']:
#                 ws.append([""] + task_data)
#                 for col in range(2, len(task_data) + 2):
#                     cell = ws.cell(row=row, column=col)
#                     if 2 <= col <= 11:
#                         cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
#                     else:
#                         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#                     cell.border = black_border
#                 row += 1
#             if project_row_start < row - 1:
#                 for col in range(2, 10):
#                     ws.merge_cells(start_row=project_row_start, start_column=col, end_row=row-1, end_column=col)
#                 ws.merge_cells(start_row=project_row_start, start_column=1, end_row=row-1, end_column=1)
#         grand_totals['vac'] += task_totals['vac']
#         grand_totals['sp'] += task_totals['sp']
#         grand_totals['fp'] += task_totals['fp']
#         grand_totals['sl'] += task_totals['sl']
#         grand_totals['psl'] += task_totals['psl']
#         grand_totals['custom_lp'] +=task_totals['custom_lp']
#     yellow_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
#     ws.append(['Total'] + [''] * 10 + [grand_totals['vac'], grand_totals['sp'], grand_totals['fp'], grand_totals['sl'], grand_totals['psl'],grand_totals['custom_lp']])
#     last_row = ws.max_row
#     for cell in ws[last_row]:
#         cell.fill = yellow_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#         cell.border = black_border
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)
#     return xlsx_file

@frappe.whitelist()
def download_PTSR():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "MR:03 – Project Task Status Report – REC (PTSR - R)" + posting_date
    build_xlsx_response_PTSR(filename)

def build_xlsx_response_PTSR(filename):
    xlsx_file = make_xlsx_PTSR(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def int_to_roman(num):
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4, 1
    ]
    syms = [
        "M", "CM", "D", "CD",
        "C", "XC", "L", "XL",
        "X", "IX", "V", "IV", "I"
    ]
    roman = ""
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman += syms[i]
            num -= val[i]
        i += 1
    return roman


def make_xlsx_PTSR(sheet_name="MR:03 – Project Task Status Report – REC (PTSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()
    # ws = wb.create_sheet(sheet_name, 0)
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15,15,15,15,15,15, 15, 15, 25, 15, 7, 7, 7, 7, 7, 7]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "REC : Project – Task Status Report : - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")
    header_font = Font(color="FFFFFF")
    white_font = Font(color="FFFFFF")
    headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark",'SPOC Remark', 'Exp Value', 'Exp PSL','Completed Value','Cr. Exp. Value','Cr.Exp.PSL','Exp.Week', 'Sourcing Status', 'Territory', 'TASK', 'Task Priority', '#VAC', '#SP', '#FP', '#SL', '#PSL', '#LP']
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    ws.append(headers)
    ws.freeze_panes = "A3"
    yellow_columns = {"Completed Value", "Cr. Exp. Value", "Cr.Exp.PSL", "Exp.Week"}
    header_row = ws[ws.max_row]
    for cell in header_row:
        if cell.value in yellow_columns:
            cell.fill = yellow_fill
            cell.font = black_font
        else:
            cell.fill = header_fill
            cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
    # for cell in header_row:
    #     cell.fill = header_fill
    #     cell.font = header_font
    #     cell.alignment = Alignment(horizontal="center", vertical="center")
    #     cell.border = black_border
    cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
    row = 3
    serial_number = 1
    s_no=1
    grand_totals = {'vac':0,'sp': 0,'fp': 0,'sl':0,'psl':0,'custom_lp':0,'exp_value':0.0,'exp_psl':0.0}
    proj_s_no = 1 
    for c in cust:
        priority = {"High": 1, "Medium": 2, "Low": 3}
        pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'],order_by= "priority ASC")
        if not pname:
            continue
        task_totals = {'exp_value':0.0,'vac':0,'sp':0,'fp':0,'sl':0,'psl':0,'custom_lp':0,'exp_psl':0.0}
        project_data = []      
        for p in pname:
            pdata = []
            print(p.project_name)        
            taskid = frappe.get_all("Task", {"status": ("in",('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'],order_by= "priority ASC")              
            # print(p['project_name'])
            # for tn in taskid:
                # print(tn.name)
            for t in taskid:
                pdata.append([p['project_name'] if p['project_name'] else "",p['priority'] if p['priority'] else "",p['remark'] if p['remark'] else "",p['account_manager_remark'] if p['account_manager_remark'] else "",p['custom_spoc_remark'] if p['custom_spoc_remark'] else "",p['expected_value'] if p['expected_value'] else "",p['expected_psl'] if p['expected_psl'] else "","","","","",p['sourcing_statu'] if p['sourcing_statu'] else "",p['territory'] if p['territory'] else "",t['subject'],t['priority'],t['vac'],t['sp'],t['fp'],t['sl'],t['psl'],t['custom_lp']])
                task_totals['vac'] +=t['vac']
                task_totals['sp'] +=t['sp']
                task_totals['fp']+= t['fp']
                task_totals['sl'] +=t['sl']
                task_totals['psl'] += t['psl']
                task_totals['custom_lp'] += t['custom_lp']
            task_totals['exp_value'] += float(p['expected_value']) if p['expected_value'] not in (None, '') else 0
            task_totals['exp_psl'] += float(p['expected_psl']) if p['expected_psl'] not in (None, '') else 0
            project_data.append({
                'project_name': p['project_name'],'priority': p['priority'],
                'remark': p['remark'],'account_manager_remark': p['account_manager_remark'],'custom_spoc_remark':p['custom_spoc_remark'],'sourcing_statu': p['sourcing_statu'],'territory': p['territory'],
                'expected_value': float(p['expected_value']) if p['expected_value'] not in (None, '') else 0,'expected_psl': p['expected_psl'],'completed':'','cr_exp':'','cr_psl':'','exp_week':'','tasks': pdata})
        s_no+=1
        blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
        # row_data = [serial_number, c['name']] + [""] * 14 + [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]
        row_data = [int_to_roman(serial_number), c['name']] + [""] * 4 +[task_totals['exp_value']]+[task_totals['exp_psl']]+[""] * 8+ [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]        
        ws.append(row_data)
        row_to_fill = ws.max_row
        for col, cell in enumerate(ws[row_to_fill], start=1):
            cell.fill = blue_fill
            if col > 11:
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
            cell.border = black_border
        ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
        serial_number += 1
        row += 1
        current_row_start = row
        
        for project in project_data:
            project_row_start = row
            for task_data in project['tasks']:
                ws.append([proj_s_no] + task_data)
                for col in range(1, len(task_data) + 1):
                    cell = ws.cell(row=row, column=col)
                    if 2 <= col <= 11:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = black_border
                row += 1
            if project_row_start < row - 1:
                for col in range(2, 13):
                    ws.merge_cells(start_row=project_row_start, start_column=col, end_row=row-1, end_column=col)
                ws.merge_cells(start_row=project_row_start, start_column=1, end_row=row-1, end_column=1)
            proj_s_no += 1
        grand_totals['vac'] += task_totals['vac']
        grand_totals['sp'] += task_totals['sp']
        grand_totals['fp'] += task_totals['fp']
        grand_totals['sl'] += task_totals['sl']
        grand_totals['psl'] += task_totals['psl']
        grand_totals['custom_lp'] +=task_totals['custom_lp']
        grand_totals['exp_value'] += float(task_totals['exp_value']) if task_totals['exp_value'] not in (None, '') else 0
        grand_totals['exp_psl'] += int(task_totals['exp_psl']) if task_totals['exp_psl'] not in (None, '') else 0

    yellow_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    ws.append(['Total'] + [''] * 5 +[grand_totals['exp_value']]+ [grand_totals['exp_psl']]+ [''] * 8 +[grand_totals['vac'], grand_totals['sp'], grand_totals['fp'], grand_totals['sl'], grand_totals['psl'],grand_totals['custom_lp']])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
    from openpyxl.cell.cell import MergedCell

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue  # Skip merged cells
            col_letter = get_column_letter(cell.column)  # Safe way to get column letter
            if col_letter in ("H","A"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter == "G":
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00'  # Example: INR formatting
    from openpyxl.utils.cell import column_index_from_string
    g_col_idx = column_index_from_string("G")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        cell = row[g_col_idx - 1]
        if isinstance(cell, MergedCell):
            continue

        # Try to convert string values to float
        try:
            if isinstance(cell.value, str) and cell.value.strip().isdigit():
                cell.value = float(cell.value.strip())
            elif isinstance(cell.value, str):
                # Handle comma-separated numbers like '1,05,000'
                cleaned = cell.value.replace(',', '').strip()
                if cleaned.isdigit():
                    cell.value = float(cleaned)

            # Now apply formatting if it's a number
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        except Exception as e:
            pass

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file



import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_PSR():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10, 5,15,15,15,15, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")
    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL",
        "Ex Value", "Ex PSL",'Completed Value','Cr. Exp. Value','Cr.Exp.PSL','Exp.Week', "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    ws.freeze_panes = "A3"
    yellow_columns = {"Completed Value", "Cr. Exp. Value", "Cr.Exp.PSL", "Exp.Week"}
    header_row = ws[ws.max_row]
    
    # for cell in header_row:
    #     cell.fill = header_fill
    #     cell.font = header_font
    #     cell.alignment = Alignment(horizontal="center", vertical="center")
    #     cell.border = black_border
    for cell in header_row:
        if cell.value in yellow_columns:
            cell.fill = yellow_fill
            cell.font = black_font
        else:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
    # Initialize variables for Customer and Project Data
    cust = frappe.get_all("Customer", fields=["name", "territory"])
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = 0
    
    proj_s_no = 1 
    for c in cust:
        pname = frappe.get_all("Project", {
            "status": ("in", ['Open', 'Enquiry']),
            "customer": c['name'],
            "service": ("in", ['REC-I', 'REC-D'])
        }, ['*'], order_by="priority ASC")

        if not pname:
            continue

        blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
        customer_with_territory = f"{c['name']} - {c['territory']}"
        row_data = [int_to_roman(serial_number), customer_with_territory] + [""] * 17
        ws.append(row_data)
        row_to_fill = ws.max_row
        customer_row_index = ws.max_row 

        ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
        ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
        ws.cell(row=customer_row_index, column=11).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=customer_row_index, column=11).number_format = '#,##0.00'
        ws.cell(row=customer_row_index, column=12).alignment = Alignment(horizontal="center", vertical="center")
        for col, cell in enumerate(ws[row_to_fill], start=1):
            cell.fill = blue_fill
            if col <= 10:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
        serial_number += 1
        row += 1
        cust_vac = cust_sp = cust_fp = cust_sl = cust_lp = cust_psl = cust_ev = cust_ex_psl = 0
        for p in pname:
            project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
            tasks = frappe.get_all("Task", {
                "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
                "project": p.name
            }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

            for task in tasks:
                project_vac += task.get('vac', 0) or 0
                project_sp += task.get('sp', 0) or 0
                project_fp += task.get('fp', 0) or 0
                project_sl += task.get('sl', 0) or 0
                project_lp += task.get('custom_lp', 0) or 0
                project_psl += task.get('psl', 0) or 0

            total_vac += project_vac
            total_sp += project_sp
            total_fp += project_fp
            total_sl += project_sl
            total_lp += project_lp
            total_psl += project_psl
            cust_vac += project_vac
            cust_sp += project_sp
            cust_fp += project_fp
            cust_sl += project_sl
            cust_lp += project_lp
            cust_psl += project_psl
            cust_ev += float(p.get('expected_value', 0) or 0)
            cust_ex_psl += float(p.get('expected_psl', 0) or 0)


            task_data = [
                p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,
                p.get('expected_value', 0), p.get('expected_psl', 0), "","","","",p.get('remark', ''), p.get('account_manager_remark', ''),
                p.get('custom_spoc_remark', '')
            ]
            ws.append([proj_s_no]+ task_data)
            

            try:
                ev_total += float(p.get('expected_value', 0) or 0)
                ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
            except ValueError:
                frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")
            from openpyxl.cell.cell import MergedCell
            # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
            for col in range(1, len(task_data) + 2):
                cell = ws.cell(row=row, column=col)
                

                if col in [18,19,20]:  # Left-align AM Remark, PM Remark, SPOC Remark
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col == 11:  # Expected Value column
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0.00'
                else:  # Center-align for all other columns
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = black_border
            row += 1
            proj_s_no += 1
        ws.cell(row=customer_row_index, column=5, value=cust_vac)
        ws.cell(row=customer_row_index, column=6, value=cust_sp)
        ws.cell(row=customer_row_index, column=7, value=cust_fp)
        ws.cell(row=customer_row_index, column=8, value=cust_sl)
        ws.cell(row=customer_row_index, column=9, value=cust_lp)
        ws.cell(row=customer_row_index, column=10, value=cust_psl)
        ws.cell(row=customer_row_index, column=11, value=cust_ev)
        ws.cell(row=customer_row_index, column=12, value=cust_ex_psl)

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,
        ev_total, ex_psl_total, "", "", "","","","",""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
    from openpyxl.utils.cell import column_index_from_string

    g_col_idx = column_index_from_string("K")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        cell = row[g_col_idx - 1]
        if isinstance(cell, MergedCell):
            continue

        # Try to convert string values to float
        try:
            if isinstance(cell.value, str) and cell.value.strip().isdigit():
                cell.value = float(cell.value.strip())
            elif isinstance(cell.value, str):
                # Handle comma-separated numbers like '1,05,000'
                cleaned = cell.value.replace(',', '').strip()
                if cleaned.isdigit():
                    cell.value = float(cleaned)

            # Now apply formatting if it's a number
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        except Exception as e:
            pass

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

# @frappe.whitelist()
# def download_PSR_customer():
# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
# 	xlsx_file = make_xlsx_PSR_cust(filename)
# 	frappe.response['filename'] = filename + '.xlsx'
# 	frappe.response['filecontent'] = xlsx_file
# 	frappe.response['type'] = 'binary'


# import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# from openpyxl.utils import get_column_letter
# import frappe
# import io
# from datetime import datetime

# def make_xlsx_PSR_cust(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
# 	if wb is None:
# 		wb = openpyxl.Workbook()
# 	# ws = wb.create_sheet(sheet_name, 0)
# 	valid_sheet_name = sheet_name.replace(":", "-")
# 	ws = wb.create_sheet(valid_sheet_name, 0)
# 	args = frappe.local.form_dict

# 	default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15, 15, 15]
# 	column_widths = column_widths or default_column_widths
# 	for i, width in enumerate(column_widths, start=1):
# 		ws.column_dimensions[get_column_letter(i)].width = width

# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	ftitle = "REC : Project – Status Report : - " + posting_date
# 	ws.append([ftitle])
# 	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
# 	title_cell = ws.cell(row=1, column=1)
# 	title_cell.alignment = Alignment(horizontal="center", vertical="center")
# 	title_cell.font = Font(bold=True)

# 	header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	header_font = Font(color="FFFFFF")
# 	black_border = Border(
# 		left=Side(border_style="thin", color="000000"),
# 		right=Side(border_style="thin", color="000000"),
# 		top=Side(border_style="thin", color="000000"),
# 		bottom=Side(border_style="thin", color="000000")
# 	)

# 	# Headers with borders and fill color
# 	headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark", 'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory']
# 	ws.append(headers)
# 	header_row = ws[ws.max_row]
# 	for cell in header_row:
# 		cell.fill = header_fill
# 		cell.font = header_font
# 		cell.alignment = Alignment(horizontal="center", vertical="center")
# 		cell.border = black_border
# 	row = 3
# 	serial_number = 1
# 	ev_total = 0
# 	cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')
# 	pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer":cust_name, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    
# 	# if not pname:
# 	# 	continue
# 	# Adding main customer row with blue fill
# 	blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
# 	row_data = [serial_number,cust_name] + [""] * 8
# 	ws.append(row_data)
# 	row_to_fill = ws.max_row

# 	# Apply fill and border only to outer edges of merged cells
# 	ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
# 	ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
    
# 	for col, cell in enumerate(ws[row_to_fill], start=1):
# 		cell.fill = blue_fill
# 		if col <= 10:
# 			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 	ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
# 	serial_number += 1
# 	row += 1
    
# 	# Add task data with border on all sides
# 	for p in pname:
# 		task_data = [p['project_name'], p['priority'], p['remark'], p['account_manager_remark'],
# 						p['custom_spoc_remark'], p['expected_value'], p['expected_psl'], p['sourcing_statu'], p['territory']]
# 		ws.append([""] + task_data)
        
# 		# Update totals
# 		try:
# 			ev_total += float(p.get('expected_value', 0) or 0)
# 		except ValueError:
# 			frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")


# 		for col in range(2, len(task_data) + 2):
# 			cell = ws.cell(row=row, column=col)
# 			cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
# 			cell.border = black_border
# 		row += 1
    
# 	# Total row formatting
# 	total_row = ["","Total", "", "", "", "", ev_total, "", "", ""]
# 	ws.append(total_row)
# 	last_row = ws[ws.max_row]
# 	total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	for cell in last_row:
# 		cell.fill = total_fill
# 		cell.font = Font(color="FFFFFF")
# 		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 		cell.border = black_border

# 	output = io.BytesIO()
# 	wb.save(output)
# 	output.seek(0)
# 	return output.read()

@frappe.whitelist()
def download_PSR_customer():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_cust(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR_cust(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

    # Initialize variables for Customer and Project Data
    cust = frappe.get_all("Customer", fields=["name", "territory"])
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = 0
    cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')
    cust_territory=frappe.db.get_value("Customer",{"name":cust_name},["territory"])
    pname = frappe.get_all("Project", {
        "status": ("in", ['Open', 'Enquiry']),
        "customer":cust_name ,
        "service": ("in", ['REC-I', 'REC-D'])
    }, ['*'], order_by="priority ASC")


    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{cust_name} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 10:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if col in [13, 14, 15]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:  # Center-align for all other columns
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# import openpyxl
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
# from openpyxl.utils import get_column_letter
# import frappe
# import io
# from datetime import datetime

# @frappe.whitelist()
# def download_PSR_proj():
# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
# 	xlsx_file = make_xlsx_PSR_proj(filename)
# 	frappe.response['filename'] = filename + '.xlsx'
# 	frappe.response['filecontent'] = xlsx_file
# 	frappe.response['type'] = 'binary'


# import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# from openpyxl.utils import get_column_letter
# import frappe
# import io
# from datetime import datetime

# def make_xlsx_PSR_proj(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
# 	if wb is None:
# 		wb = openpyxl.Workbook()
# 	# ws = wb.create_sheet(sheet_name, 0)
# 	valid_sheet_name = sheet_name.replace(":", "-")
# 	ws = wb.create_sheet(valid_sheet_name, 0)
# 	args = frappe.local.form_dict

# 	default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15, 15, 15]
# 	column_widths = column_widths or default_column_widths
# 	for i, width in enumerate(column_widths, start=1):
# 		ws.column_dimensions[get_column_letter(i)].width = width

# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	ftitle = "REC : Project – Status Report : - " + posting_date
# 	ws.append([ftitle])
# 	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
# 	title_cell = ws.cell(row=1, column=1)
# 	title_cell.alignment = Alignment(horizontal="center", vertical="center")
# 	title_cell.font = Font(bold=True)

# 	header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	header_font = Font(color="FFFFFF")
# 	black_border = Border(
# 		left=Side(border_style="thin", color="000000"),
# 		right=Side(border_style="thin", color="000000"),
# 		top=Side(border_style="thin", color="000000"),
# 		bottom=Side(border_style="thin", color="000000")
# 	)

# 	# Headers with borders and fill color
# 	headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark", 'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory']
# 	ws.append(headers)
# 	header_row = ws[ws.max_row]
# 	for cell in header_row:
# 		cell.fill = header_fill
# 		cell.font = header_font
# 		cell.alignment = Alignment(horizontal="center", vertical="center")
# 		cell.border = black_border
# 	row = 3
# 	serial_number = 1
# 	ev_total = 0
# 	pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
# 	pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
# 	project_customer=frappe.db.get_value("Project",{"name":pro_name},["customer"])
# 	# if not pname:
# 	# 	continue
# 	# Adding main customer row with blue fill
# 	blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
# 	row_data = [serial_number,project_customer] + [""] * 8
# 	ws.append(row_data)
# 	row_to_fill = ws.max_row

# 	# Apply fill and border only to outer edges of merged cells
# 	ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
# 	ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
    
# 	for col, cell in enumerate(ws[row_to_fill], start=1):
# 		cell.fill = blue_fill
# 		if col <= 10:
# 			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 	ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
# 	serial_number += 1
# 	row += 1
    
# 	# Add task data with border on all sides
# 	for p in pname:
# 		task_data = [p['project_name'], p['priority'], p['remark'], p['account_manager_remark'],
# 						p['custom_spoc_remark'], p['expected_value'], p['expected_psl'], p['sourcing_statu'], p['territory']]
# 		ws.append([""] + task_data)
        
# 		# Update totals
# 		try:
# 			ev_total += float(p.get('expected_value', 0) or 0)
# 		except ValueError:
# 			frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")


# 		for col in range(2, len(task_data) + 2):
# 			cell = ws.cell(row=row, column=col)
# 			cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
# 			cell.border = black_border
# 		row += 1
    
# 	# Total row formatting
# 	total_row = ["","Total", "", "", "", "", ev_total, "", "", ""]
# 	ws.append(total_row)
# 	last_row = ws[ws.max_row]
# 	total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	for cell in last_row:
# 		cell.fill = total_fill
# 		cell.font = Font(color="FFFFFF")
# 		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 		cell.border = black_border

# 	output = io.BytesIO()
# 	wb.save(output)
# 	output.seek(0)
# 	return output.read()

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_PSR_proj():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_proj(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR_proj(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

    # Initialize variables for Customer and Project Data
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = 0
    pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
    pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    project_customer=frappe.db.get_value("Project",{"name":pro_name},["customer"])
    cust_territory=frappe.db.get_value("Customer",{"name":project_customer},["territory"])

    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{project_customer} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 10:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if col in [13, 14, 15]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:  # Center-align for all other columns
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# import openpyxl
# from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
# from openpyxl.utils import get_column_letter
# import frappe
# import io
# from datetime import datetime

# @frappe.whitelist()
# def download_PSR_both():
# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
# 	xlsx_file = make_xlsx_PSR_both(filename)
# 	frappe.response['filename'] = filename + '.xlsx'
# 	frappe.response['filecontent'] = xlsx_file
# 	frappe.response['type'] = 'binary'


# import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# from openpyxl.utils import get_column_letter
# import frappe
# import io
# from datetime import datetime

# def make_xlsx_PSR_both(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
# 	if wb is None:
# 		wb = openpyxl.Workbook()
# 	# ws = wb.create_sheet(sheet_name, 0)
# 	valid_sheet_name = sheet_name.replace(":", "-")
# 	ws = wb.create_sheet(valid_sheet_name, 0)
# 	args = frappe.local.form_dict

# 	default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15, 15, 15]
# 	column_widths = column_widths or default_column_widths
# 	for i, width in enumerate(column_widths, start=1):
# 		ws.column_dimensions[get_column_letter(i)].width = width

# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	ftitle = "REC : Project – Status Report : - " + posting_date
# 	ws.append([ftitle])
# 	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
# 	title_cell = ws.cell(row=1, column=1)
# 	title_cell.alignment = Alignment(horizontal="center", vertical="center")
# 	title_cell.font = Font(bold=True)

# 	header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	header_font = Font(color="FFFFFF")
# 	black_border = Border(
# 		left=Side(border_style="thin", color="000000"),
# 		right=Side(border_style="thin", color="000000"),
# 		top=Side(border_style="thin", color="000000"),
# 		bottom=Side(border_style="thin", color="000000")
# 	)

# 	# Headers with borders and fill color
# 	headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark", 'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory']
# 	ws.append(headers)
# 	header_row = ws[ws.max_row]
# 	for cell in header_row:
# 		cell.fill = header_fill
# 		cell.font = header_font
# 		cell.alignment = Alignment(horizontal="center", vertical="center")
# 		cell.border = black_border
# 	row = 3
# 	serial_number = 1
# 	ev_total = 0
# 	pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
# 	project_customer=frappe.db.get_single_value("Formatted Reports  Download",'customer')
# 	pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name,"customer":project_customer, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
# 	# if not pname:
# 	# 	continue
# 	# Adding main customer row with blue fill
# 	blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
# 	row_data = [serial_number,project_customer] + [""] * 8
# 	ws.append(row_data)
# 	row_to_fill = ws.max_row

# 	# Apply fill and border only to outer edges of merged cells
# 	ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
# 	ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
    
# 	for col, cell in enumerate(ws[row_to_fill], start=1):
# 		cell.fill = blue_fill
# 		if col <= 10:
# 			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 	ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
# 	serial_number += 1
# 	row += 1
    
# 	# Add task data with border on all sides
# 	for p in pname:
# 		task_data = [p['project_name'], p['priority'], p['remark'], p['account_manager_remark'],
# 						p['custom_spoc_remark'], p['expected_value'], p['expected_psl'], p['sourcing_statu'], p['territory']]
# 		ws.append([""] + task_data)
        
# 		# Update totals
# 		try:
# 			ev_total += float(p.get('expected_value', 0) or 0)
# 		except ValueError:
# 			frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")


# 		for col in range(2, len(task_data) + 2):
# 			cell = ws.cell(row=row, column=col)
# 			cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
# 			cell.border = black_border
# 		row += 1
    
# 	# Total row formatting
# 	total_row = ["","Total", "", "", "", "", ev_total, "", "", ""]
# 	ws.append(total_row)
# 	last_row = ws[ws.max_row]
# 	total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
# 	for cell in last_row:
# 		cell.fill = total_fill
# 		cell.font = Font(color="FFFFFF")
# 		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 		cell.border = black_border

# 	output = io.BytesIO()
# 	wb.save(output)
# 	output.seek(0)
# 	return output.read()
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_PSR_both():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_both(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR_both(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

    # Initialize variables for Customer and Project Data
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = 0
    pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
    project_customer=frappe.db.get_single_value("Formatted Reports  Download",'customer')
    pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name,"customer":project_customer, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    cust_territory=frappe.db.get_value("Customer",{"name":project_customer},["territory"])

    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{project_customer} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 10:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if col in [13, 14, 15]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:  # Center-align for all other columns
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@frappe.whitelist()
def print_sales_order_outstanding_report(account_manager=None, delivery_manager=None, service=None, company=None):
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="17" style="text-align:center; font-weight:bold;">PR-05 – To Be Billed - SO Outstanding (TBB)</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Services</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Company</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">AM</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Net Total</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Grand Total</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">% Billed</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Amount Billed</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Pending Billing </td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Advance</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Pending Cashflow</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Reference Customer</td>' \
            '</tr>'

    filters = []
    if account_manager:
        filters.append(f"account_manager = '{account_manager}'")
    if delivery_manager:
        filters.append(f"delivery_manager = '{delivery_manager}'")
    if service:
        filters.append(f"services = '{service}'")
    if company:
        filters.append(f"company = '{company}'")
    
    filter_query = " AND ".join(filters)
    
    query = f"""
    SELECT 
        s.name AS name,
        s.customer AS customer,
        s.service AS services,
        c.abbr AS company, 
        account_manager.short_code AS account_manager,
        s.transaction_date AS posting_date,
        s.status AS status, 
        s.base_net_total AS base_net_total,
        s.base_grand_total AS total,
        s.per_billed AS per_billed,
        s.amount_billed_company_currency AS company_currency,
        s.advance_paid AS advance_paid,
        s.reference_customer_ AS reference_customer_
    FROM 
        `tabSales Order` s
    INNER JOIN 
        `tabCompany` c ON c.name = s.company
    LEFT JOIN 
        `tabEmployee` account_manager ON account_manager.user_id = s.account_manager
    WHERE 
        s.status NOT IN ('To Deliver', 'On Hold', 'Closed', 'Cancelled', 'Completed')
        {f"AND {filter_query}" if filter_query else ""}
"""
    sales_order = frappe.db.sql(query, as_dict=True)

    total_net_total = total_grand_total = total_amount_billed = total_pending_billing = total_advance = total_pending_collection = pending_collection = total_per_billed = 0
    for i in sales_order:
        formatted_date = frappe.utils.formatdate(i.posting_date, 'dd-mm-yyyy')
        # pending_billed = (i.base_net_total-i.company_currency)
        # pending_collection = (pending_billed-i.advance_paid)
        # to_be_billed = (i.total-(i.total*i.per_billed)+i.advance_paid)
        amount_billed=i["base_grand_total"] * i["per_billed"]
        pending_billing= i["base_net_total"]-amount_billed
        pending_cash_flow=pending_billing-i["advance_paid"]
            
        data += f'<tr>'
        data += f'<td style="text-align:center;">{s_no}</td>'
        data += f'<td style="text-align:center;">{i.name}</td>'
        data += f'<td style="text-align:center;">{i.customer}</td>'
        data += f'<td style="text-align:center;">{i.services}</td>'
        data += f'<td style="text-align:center;">{i.company}</td>'
        data += f'<td style="text-align:center;">{i.account_manager}</td>'
        data += f'<td style="text-align:center;">{formatted_date}</td>'
        data += f'<td style="text-align:center;">{i.status}</td>'
        data += f'<td style="text-align:center;">{i.base_net_total}</td>'
        data += f'<td style="text-align:center;">{i.total:,.2f}</td>'
        data += f'<td style="text-align:center;">{i.per_billed:,.2f}</td>'
        data += f'<td style="text-align:center;">{amount_billed:,.2f}</td>'
        data += f'<td style="text-align:center;">{pending_billing:,.2f}</td>'
        data += f'<td style="text-align:center;">{i.advance_paid:,.2f}</td>'
        data += f'<td style="text-align:center;">{pending_cash_flow:,.2f}</td>'
        data += f'<td style="text-align:center;">{i.reference_customer_}</td>'
        data += '</tr>'
        s_no += 1
        total_net_total += i.base_net_total
        total_grand_total += i.total
        total_amount_billed += amount_billed
        total_advance += i.advance_paid
        total_pending_billing += pending_billing
        total_pending_collection += pending_collection
        total_per_billed += i.per_billed

    data += f'<tr style="background-color: #f79646">' \
            f'<td style="text-align:center;"></td>' \
             f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:left;">Total</td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;"></td>' \
            f'<td style="text-align:center;">{total_net_total:,.2f}</td>' \
            f'<td style="text-align:center;">{total_grand_total:,.2f}</td>' \
            f'<td style="text-align:center;">{round(total_per_billed)}</td>' \
            f'<td style="text-align:center;">{total_amount_billed:,.2f}</td>' \
            f'<td style="text-align:center;">{total_pending_billing:,.2f}</td>' \
            f'<td style="text-align:center;">{total_advance:,.2f}</td>' \
            f'<td style="text-align:center;">{pending_collection:,.2f}</td>' \
            f'<td style="text-align:center;"></td>' \
            '</tr>'
    
    data += '</table>'
    return data

@frappe.whitelist()
def download_sales_order_outstanding_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR-05 – To Be Billed - SO Outstanding (TBB)" + posting_date
    build_xlsx_response_so(filename)

def build_xlsx_response_so(filename):
    xlsx_file = make_xlsx_so(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_so(sheet_name="PR-05 – To Be Billed - SO Outstanding (TBB)", wb=None):
    args = frappe.local.form_dict
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name

    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
    text_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")  # Centered with wrap
    bottom_alignment = Alignment(vertical="bottom")               # Bottom alignment if needed

    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
    inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    column_widths = [6, 25, 35, 9, 10, 9, 15, 20,5, 13, 13, 15, 15, 15, 18, 18, 35]
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P","Q"]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[columns[i]].width = width
        
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "PR-05 – To Be Billed - SO Outstanding (TBB)(" + posting_date + ")"
    ws.merge_cells("A1:Q1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment
    
    headers = ["S NO", "ID", "Customer Name", "Services", "Company",
               "AM", "Date", "Status","Age" ,"Net Total", "Grand Total", "% Billed", "Amount Billed", "Pending Billing",
               "Advance", "Pending Cash Flow", "Reference Customer"]
    ws.append(headers)

    for cell in ws[2]:
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border

    total_net_total = 0
    total_grand_total = 0
    total_per_billed = 0
    total_amount_billed = 0
    total_pending_billing = 0
    total_advance = 0
    pending_collection = 0
    to_be_billed = 0

    data = get_data_so(args)

    for row in data:
        ws.append(row)
        ws[f"I{ws.max_row}"].style = inr_format
        ws[f"J{ws.max_row}"].style = inr_format
        ws[f"K{ws.max_row}"].style = inr_format
        ws[f"L{ws.max_row}"].style = inr_format
        ws[f"M{ws.max_row}"].style = inr_format
        ws[f"N{ws.max_row}"].style = inr_format
        ws[f"O{ws.max_row}"].style = inr_format
        ws[f"P{ws.max_row}"].style = inr_format
        for cell in ws[ws.max_row]:
            cell.alignment = text_wrap
            cell.border = thin_border

        total_net_total += row[9]
        total_grand_total += row[10]
        total_per_billed += row[11]
        total_amount_billed += row[12]
        total_pending_billing += row[13]
        total_advance += row[14]
        pending_collection += row[15]

    ws.append(["", "", "Total", "", "","", "", "", "", total_net_total, total_grand_total, total_per_billed, total_amount_billed, total_pending_billing, total_advance, pending_collection, ""])

    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = bg_fill
        cell.border = thin_border
        cell.alignment = alignment
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):  # Adjust row range as needed
        for cell in row:
            col = cell.column_letter
            if cell.column_letter == "B":  # SO Number column
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
            elif col in ("J", "K", "L", "M", "N", "O", "P"):  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = text_wrap  # Wrap text with default alignment
            cell.border = thin_border
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

from frappe.utils import nowdate
from datetime import date, datetime

def get_data_so(args):

    am = args.get("account_manager")
    dm = args.get("delivery_manager")
    ser = args.get("service")
    comp = args.get("company")

    data = []
    s_no = 1

    filters = {"status": ["not in", ['To Deliver', 'On Hold', 'Closed', 'Cancelled', 'Completed']]}

    if am:
        filters["account_manager"] = am
    if dm:
        filters["delivery_manager"] = dm
    if ser:
        filters["service"] = ser
    if comp:
        filters["company"] = comp

    sales_order = frappe.db.get_all(
        "Sales Order", filters=filters,
        fields=["name", "company", "service", "status", "customer", "account_manager", 
                "base_grand_total", "transaction_date", "reference_customer_", "advance_paid", "amount_billed_company_currency",
                "per_billed", "base_net_total", "delivery_manager"]
    )

    for order in sales_order:
        account = frappe.db.get_value("Employee", {"user_id": order["account_manager"]}, "short_code") or ""
        delivery = frappe.db.get_value("Employee", {"user_id": order["delivery_manager"]}, "short_code") or ""
        abbr = frappe.db.get_value("Company", {"name": order["company"]}, "abbr") or ""

        formatted_date = frappe.utils.formatdate(order["transaction_date"], 'dd-mm-yyyy')
        # pending_billed = (order["base_net_total"]-order["amount_billed_company_currency"])
        # pending_collected = (pending_billed-order["advance_paid"])
        amount_billed=order["base_grand_total"] * order["per_billed"]/100
        # pending_billing= order["base_net_total"]-amount_billed
        pending_billing= order["base_grand_total"]-amount_billed
        pending_cash_flow=(pending_billing-order["advance_paid"])
        transaction_date = order.transaction_date  # assuming this is a datetime/date object already
        todate = date.today()

        # Make sure both are date objects
        if isinstance(transaction_date, str):
            transaction_date = datetime.strptime(transaction_date, "%Y-%m-%d").date()

        age = str((todate - transaction_date).days)
        data.append([
            s_no, order["name"], order["customer"], order["service"], abbr, account,
            formatted_date, order["status"],age, order["base_net_total"], order["base_grand_total"], order["per_billed"],
            amount_billed, pending_billing, order["advance_paid"], pending_cash_flow, 
            order["reference_customer_"]
        ])
        s_no += 1

    return data

@frappe.whitelist()
def print_psr_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += f'<tr><td colspan="15" style="text-align:center; font-weight:bold;">PR:02 – Project Status Report – REC (PSR - R) - {posting_date}</td></tr>'
    data += '<tr style="background-color: #002060; color: white !important;">' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width: 150px;">Customer / Project Name</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SS</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">VAC</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">FP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">LP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width:75px;">EV</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">EPSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">AM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">PM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">SPOC Remark</td>' \
            '</tr>'

    cust = frappe.get_all("Customer", fields=["name"])

    ev_total = vac_total = sp_total = fp_total = sl_total = lp_total = psl_total = epsl_total=0

    for c in cust:
        pname = frappe.get_all("Project", {
            "status": ("in", ['Open', 'Enquiry']),
            "customer": c['name'],
            "service": ("in", ['REC-I', 'REC-D'])
        }, ['*'], order_by="priority ASC")

        for p in pname:
            tasks = frappe.get_all("Task", {
                "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
                "project": p.name
            }, ['vac', 'sp', 'fp', 'sl', 'custom_lp', 'psl'])

            project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0

            for task in tasks:
                project_vac += task.get('vac', 0) or 0
                project_sp += task.get('sp', 0) or 0
                project_fp += task.get('fp', 0) or 0
                project_sl += task.get('sl', 0) or 0
                project_lp += task.get('custom_lp', 0) or 0
                project_psl += task.get('psl', 0) or 0

            vac_total += project_vac
            sp_total += project_sp
            fp_total += project_fp
            sl_total += project_sl
            lp_total += project_lp
            psl_total += project_psl

            try:
                ev_total += float(p.get('expected_value', 0) or 0)
                epsl_total += float(p.get('expected_psl', 0) or 0)
            except ValueError:
                frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

            data += f'<tr style="background-color: #98d7f5;">' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;">{s_no}</td>' \
                    f'<td colspan=14 style="text-align:left;border:1px solid black;">{c["name"]} - {p.get("territory", "")}</td>' \
                    f'</tr>'
            data += f'<tr>' \
                    f'<td style="text-align:center;border:1px solid black;font-size: 14px;"></td>' \
                    f'<td style="text-align:left;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("project_name", "")}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("priority", "")}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("sourcing_statu", "")}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_vac}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sp}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_fp}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sl}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_lp}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_psl}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_value", 0)}</td>' \
                    f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_psl", 0)}</td>' \
                    f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("remark", "")}</td>' \
                    f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("account_manager_remark", "")}</td>' \
                    f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("custom_spoc_remark", "")}</td>' \
                    f'</tr>'
            s_no += 1

    data += f'<tr style="background-color: #002060;; color: white !important;">' \
        f'<td colspan="4" style="text-align:center; font-weight: bold;color: white !important;font-size: 14px;">Total</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{vac_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{fp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{lp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{psl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{ev_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{epsl_total}</td>' \
        f'<td colspan="3"></td>' \
        f'</tr>'


    data += '</table>'
    return data


# @frappe.whitelist()
# def print_psr_report_for_cust(doc):
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     s_no = 1
#     data = '<table border="1" style="border-collapse: collapse; width: 100%; ">'
#     data += f'<tr><td colspan="12" style="text-align:center; font-weight:bold;">PR:02 – Project Status Report – REC (PSR - R)- {posting_date}</td></tr>'
#     data += '<tr style="background-color: #002060; color: white;">' \
#             '<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">AM Mark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">PM Mark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">EV</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">EPSL</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">SS</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;border-right: 2px solid black;">Territory</td>' \
#             '</tr>'
    
#     cust = frappe.get_all("Customer", fields=["name"])
#     cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')
#     ev_total = 0
#     # for c in cust:
#     pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": cust_name, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
#     for p in pname:
#         try:
#             ev_total += float(p.get('expected_value', 0) or 0)
#         except ValueError:
#             frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

#         data += f'<tr style="background-color: #98d7f5;">' \
#                 f'<td style="text-align:center;">{s_no}</td>' \
#                 f'<td colspan=3 style="text-align:left;">{cust_name}</td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
#                 f'</tr>'
#         data += f'<tr>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:left;">{p.project_name}</td>' \
#                 f'<td style="text-align:center;">{p.priority}</td>' \
#                 f'<td style="text-align:left;">{p.remark}</td>' \
#                 f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
#                 f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
#                 f'<td style="text-align:center;">{p.expected_value}</td>' \
#                 f'<td style="text-align:center;">{p.expected_psl}</td>' \
#                 f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
#                 f'<td style="text-align:center;border-right: 2px solid black;">{p.territory}</td>' \
#                 '</tr>'
#         s_no += 1
#     data += f'<tr style="background-color: #002060;">' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center; font-weight: bold; color: #ffffff;">Total</td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:left;"></td>' \
#             f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{ev_total}</td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
#             '</tr>'
            
#     data += '</table>'
#     return data
@frappe.whitelist()
def print_psr_report_for_cust(doc):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += f'<tr><td colspan="15" style="text-align:center; font-weight:bold;">PR:02 – Project Status Report – REC (PSR - R) - {posting_date}</td></tr>'
    data += '<tr style="background-color: #002060; color: white !important;">' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width: 150px;">Customer / Project Name</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SS</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">VAC</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">FP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">LP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width:75px;">EV</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">EPSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">AM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">PM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">SPOC Remark</td>' \
            '</tr>'

    cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')

    ev_total = vac_total = sp_total = fp_total = sl_total = lp_total = psl_total = epsl_total=0

    pname = frappe.get_all("Project", {
        "status": ("in", ['Open', 'Enquiry']),
        "customer":cust_name,
        "service": ("in", ['REC-I', 'REC-D'])
    }, ['*'], order_by="priority ASC")

    for p in pname:
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['vac', 'sp', 'fp', 'sl', 'custom_lp', 'psl'])

        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        vac_total += project_vac
        sp_total += project_sp
        fp_total += project_fp
        sl_total += project_sl
        lp_total += project_lp
        psl_total += project_psl

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            epsl_total += float(p.get('expected_psl', 0) or 0)
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        data += f'<tr style="background-color: #98d7f5;">' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;">{s_no}</td>' \
                f'<td colspan=14 style="text-align:left;border:1px solid black;">{cust_name} - {p.get("territory", "")}</td>' \
                f'</tr>'
        data += f'<tr>' \
                f'<td style="text-align:center;border:1px solid black;font-size: 14px;"></td>' \
                f'<td style="text-align:left;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("project_name", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("priority", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("sourcing_statu", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_vac}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_fp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sl}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_lp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_psl}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_value", 0)}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_psl", 0)}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("remark", "")}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("account_manager_remark", "")}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("custom_spoc_remark", "")}</td>' \
                f'</tr>'
        s_no += 1

    data += f'<tr style="background-color: #002060;; color: white !important;">' \
        f'<td colspan="4" style="text-align:center; font-weight: bold;color: white !important;font-size: 14px;">Total</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{vac_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{fp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{lp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{psl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{ev_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{epsl_total}</td>' \
        f'<td colspan="3"></td>' \
        f'</tr>'


    data += '</table>'
    return data


# @frappe.whitelist()
# def print_psr_report_for_proj(doc):
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     s_no = 1
#     data = '<table border="1" style="border-collapse: collapse; width: 100%; ">'
#     data += f'<tr><td colspan="12" style="text-align:center; font-weight:bold;">PR:02 – Project Status Report – REC (PSR - R)- {posting_date}</td></tr>'
#     data += '<tr style="background-color: #002060; color: white;">' \
#             '<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">AM Mark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">PM Mark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">EV</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">EPSL</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">SS</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;border-right: 2px solid black;">Territory</td>' \
#             '</tr>'
    
#     cust = frappe.get_all("Customer", fields=["name"])
#     proj_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
#     cust_name=frappe.db.get_value("Project",{"name":proj_name},["customer"])
#     ev_total = 0
#     # for c in cust:
#     pname = frappe.get_all("Project", {"name":proj_name,"status": ("in", ['Open', 'Enquiry']), "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
#     for p in pname:
#         try:
#             ev_total += float(p.get('expected_value', 0) or 0)
#         except ValueError:
#             frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

#         data += f'<tr style="background-color: #98d7f5;">' \
#                 f'<td style="text-align:center;">{s_no}</td>' \
#                 f'<td colspan=3 style="text-align:left;">{cust_name}</td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
#                 f'</tr>'
#         data += f'<tr>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:left;">{p.project_name}</td>' \
#                 f'<td style="text-align:center;">{p.priority}</td>' \
#                 f'<td style="text-align:left;">{p.remark}</td>' \
#                 f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
#                 f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
#                 f'<td style="text-align:center;">{p.expected_value}</td>' \
#                 f'<td style="text-align:center;">{p.expected_psl}</td>' \
#                 f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
#                 f'<td style="text-align:center;border-right: 2px solid black;">{p.territory}</td>' \
#                 '</tr>'
#         s_no += 1
#     data += f'<tr style="background-color: #002060;">' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center; font-weight: bold; color: #ffffff;">Total</td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:left;"></td>' \
#             f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{ev_total}</td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;"></td>' \
#             f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
#             '</tr>'
            
#     data += '</table>'
#     return data
@frappe.whitelist()
def print_psr_report_for_proj(doc):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += f'<tr><td colspan="15" style="text-align:center; font-weight:bold;">PR:02 – Project Status Report – REC (PSR - R) - {posting_date}</td></tr>'
    data += '<tr style="background-color: #002060; color: white !important;">' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width: 150px;">Customer / Project Name</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SS</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">VAC</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">FP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">SL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">LP</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">PSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;width:75px;">EV</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;color: white !important;font-size: 14px;">EPSL</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">AM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">PM Remark</td>' \
            '<td style="text-align:center; font-weight:bold;border:1px solid black;width: 250px;color: white !important;font-size: 14px;">SPOC Remark</td>' \
            '</tr>'

    # cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')

    ev_total = vac_total = sp_total = fp_total = sl_total = lp_total = psl_total = epsl_total=0

    proj_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
    cust_name=frappe.db.get_value("Project",{"name":proj_name},["customer"])
    ev_total = 0
    # for c in cust:
    pname = frappe.get_all("Project", {"name":proj_name,"status": ("in", ['Open', 'Enquiry']), "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    

    for p in pname:
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['vac', 'sp', 'fp', 'sl', 'custom_lp', 'psl'])

        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        vac_total += project_vac
        sp_total += project_sp
        fp_total += project_fp
        sl_total += project_sl
        lp_total += project_lp
        psl_total += project_psl

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            epsl_total += float(p.get('expected_psl', 0) or 0)
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        data += f'<tr style="background-color: #98d7f5;">' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;">{s_no}</td>' \
                f'<td colspan=14 style="text-align:left;border:1px solid black;">{cust_name} - {p.get("territory", "")}</td>' \
                f'</tr>'
        data += f'<tr>' \
                f'<td style="text-align:center;border:1px solid black;font-size: 14px;"></td>' \
                f'<td style="text-align:left;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("project_name", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("priority", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("sourcing_statu", "")}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_vac}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_fp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_sl}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_lp}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{project_psl}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_value", 0)}</td>' \
                f'<td style="text-align:center;vertical-align: middle !important;border:1px solid black;font-size: 14px;">{p.get("expected_psl", 0)}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("remark", "")}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("account_manager_remark", "")}</td>' \
                f'<td style="text-align:left;border:1px solid black;font-size: 14px;">{p.get("custom_spoc_remark", "")}</td>' \
                f'</tr>'
        s_no += 1

    data += f'<tr style="background-color: #002060;; color: white !important;">' \
        f'<td colspan="4" style="text-align:center; font-weight: bold;color: white !important;font-size: 14px;">Total</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{vac_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{fp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{sl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{lp_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{psl_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{ev_total}</td>' \
        f'<td style="text-align:center; font-weight: bold;border:1px solid black;color: white !important;font-size: 14px;">{epsl_total}</td>' \
        f'<td colspan="3"></td>' \
        f'</tr>'


    data += '</table>'
    return data

@frappe.whitelist()
# def ptsr_report(doc):
def ptsr_report():
    from datetime import datetime
    posting_date = datetime.now().strftime("%d-%m-%Y")
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += f'<tr><td colspan="17" style="text-align:center; font-weight:bold;">REC : Project - Task Status Report - {posting_date}</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">AM Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">PM Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Expected Value</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Expected PSL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Sourcing Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Task</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border-right-style: hidden;">Task Priority</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border-left-style: hidden;">#VAC</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#SP</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#FP</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#SL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#PSL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#LP</td>' \
            '</tr>'
    
    cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
    
    ev_total = 0
    grand_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}

    for c in cust:
        pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
        task_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}

        for p in pname:
            try:
                ev_total += float(p.get('expected_value', 0) or 0)
            except ValueError:
                frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")
            
            taskid = frappe.get_all("Task", {"status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'], order_by="priority ASC")
            
            # Project row
            data += f'<tr style="background-color: #98d7f5;">' \
                    f'<td style="text-align:center;">{s_no}</td>' \
                    f'<td colspan=4 style="text-align:left;">{c.name} - {p.territory}</td>' \
                    f'<td colspan=13></td>' \
                    '</tr>'
            
            # Each task row within a project
            for t in taskid:
                data += f'<tr>' \
                        f'<td style="text-align:center;"></td>' \
                        f'<td style="text-align:left;">{p.project_name}</td>' \
                        f'<td style="text-align:center;">{p.priority}</td>' \
                        f'<td style="text-align:left;">{p.remark}</td>' \
                        f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
                        f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
                        f'<td style="text-align:center;">{p.expected_value}</td>' \
                        f'<td style="text-align:center;">{p.expected_psl}</td>' \
                        f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
                        f'<td style="text-align:left;">{t.subject}</td>' \
                        f'<td style="text-align:center;border-right-style: hidden;">{t.priority}</td>' \
                        f'<td style="text-align:center;border-left-style: hidden;">{t.vac}</td>' \
                        f'<td style="text-align:center;">{t.sp}</td>' \
                        f'<td style="text-align:center;">{t.fp}</td>' \
                        f'<td style="text-align:center;">{t.sl}</td>' \
                        f'<td style="text-align:center;">{t.pl}</td>' \
                        f'<td style="text-align:center;">{t.custom_lp}</td>' \
                        '</tr>'
                
                # Update task totals
                task_totals['vac'] += t.get('vac', 0)
                task_totals['sp'] += t.get('sp', 0)
                task_totals['fp'] += t.get('fp', 0)
                task_totals['sl'] += t.get('sl', 0)
                task_totals['psl'] += t.get('psl', 0)
                task_totals['custom_lp'] += t.get('custom_lp', 0)
                s_no += 1

    # Grand total row
    data += f'<tr style="background-color: #002060;">' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">Total</td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{ev_total}</td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["vac"]}</td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{grand_totals["sp"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["fp"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["sl"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["psl"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["custom_lp"]}</td>' \
            '</tr>'
    
    data += '</table>'
    return data

@frappe.whitelist()
def print_closure_report(so_status=None, so_validate=None):
    s_no = 1
    data = '<div style="overflow-x: auto; page-break-before: auto; margin: 0;">'  # Wrapper for horizontal scrolling if necessary and avoid page break
    data += '<table border="1" style="border-collapse: collapse; width: 100%; font-size: 10px;">'
    data += '<tr><td colspan="46" style="text-align:center; font-weight:bold;">MR:05 – Closure Detailed Status Report (CDSR)</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Passport No</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Given Name / Surename(as per passport)</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Mobile / Whatsapp</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Position / Task Subject</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Territory</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate Owner</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate Collection Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Photo(as per visa specification)</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Signed Offer Letter</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Project</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Account Manager</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">SA Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">MOH</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">PCC</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">IAF / CV / Client Form</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate Outstanding</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">SA ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Stamped Visa</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Entry Visa</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Final Medical</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Client SI</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate SI</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ECR Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Pre-Medical</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">SO Confirmed Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Passport Number</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">PCC Original at</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">SA Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Visa Original at</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Final Medical Original at</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Visa Expiry Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Visa Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">SO Created</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Collection Priority</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">PP Original at</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Interview Location</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Nationality</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Candidate SC</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Client Payment Company Currency</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Billing Currency(Client / Associate)</td>' \
            '</tr>'
    
    conditions = []
    closure_query = """
        SELECT * FROM `tabClosure`
        WHERE status NOT IN ('Arrived', 'Dropped')
    """

    # Handle conditions based on checkbox values
    if so_validate is None:  # If 'so_validate' is not passed, ignore it
        if so_status:  # If 'so_status' is True, we add the filter for so_created
            conditions.append("so_created = 1")  # 1 for True (Checked)
        else:
            conditions.append("so_created = 0")  # 0 for False (Unchecked)

    # If conditions exist, add them to the query
    if conditions:
        closure_query += " AND " + " AND ".join(conditions)

    closure = frappe.db.sql(closure_query, as_dict=True)

    for i in closure:
        data += f'<tr>'
        data += f'<td style="text-align:center;">{s_no}</td>'
        data += f'<td style="text-align:center;">{i.name}</td>'
        data += f'<td style="text-align:center;">{i.passport_no}</td>'
        data += f'<td style="text-align:center;">{i.given_name}</td>'
        data += f'<td style="text-align:center;">{i.mobile}</td>'
        data += f'<td style="text-align:center;">{i.task_subject}</td>'
        data += f'<td style="text-align:center;">{i.customer}</td>'
        data += f'<td style="text-align:center;">{i.territory}</td>'
        data += f'<td style="text-align:center;">{i.candidate_owner}</td>'
        data += f'<td style="text-align:center;">{i.collection_status}</td>'
        data += f'<td style="text-align:center;">{i.photo}</td>'
        data += f'<td style="text-align:center;">{i.sol}</td>'
        data += f'<td style="text-align:center;">{i.project}</td>'
        data += f'<td style="text-align:center;">{i.account_manager}</td>'
        data += f'<td style="text-align:center;">{i.sa_name}</td>'
        data += f'<td style="text-align:center;">{i.moh}</td>'
        data += f'<td style="text-align:center;">{i.pcc}</td>'
        data += f'<td style="text-align:center;">{i.irf}</td>'
        data += f'<td style="text-align:center;">{i.outstanding_amount}</td>'
        data += f'<td style="text-align:center;">{i.sa_id}</td>'
        data += f'<td style="text-align:center;">{i.visa_stamping}</td>'
        data += f'<td style="text-align:center;">{i.visa}</td>'
        data += f'<td style="text-align:center;">{i.final_medical}</td>'
        data += f'<td style="text-align:center;">{i.client_payment_company_currency}</td>'
        data += f'<td style="text-align:center;">{i.candidate_payment_company_currenc}</td>'
        data += f'<td style="text-align:center;">{i.ecr_status}</td>'
        data += f'<td style="text-align:center;">{i.premedical}</td>'
        data += f'<td style="text-align:center;">{i.candidate}</td>'
        data += f'<td style="text-align:center;">{i.so_confirmed_date}</td>'
        data += f'<td style="text-align:center;">{i.passport_number}</td>'
        data += f'<td style="text-align:center;">{i.pcc_original}</td>'
        data += f'<td style="text-align:center;">{i.sams_name}</td>'
        data += f'<td style="text-align:center;">{i.visa_original_at}</td>'
        data += f'<td style="text-align:center;">{i.final_medical_original_at}</td>'
        data += f'<td style="text-align:center;">{i.visa_expiry_date}</td>'
        data += f'<td style="text-align:center;">{i.visa_status}</td>'
        data += f'<td style="text-align:center;">{i.status}</td>'
        data += f'<td style="text-align:center;">{i.so_created}</td>'
        data += f'<td style="text-align:center;">{i.collection_priority}</td>'
        data += f'<td style="text-align:center;">{i.pp_original_at}</td>'
        data += f'<td style="text-align:center;">{i.interview_location}</td>'
        data += f'<td style="text-align:center;">{i.posting_date}</td>'
        data += f'<td style="text-align:center;">{i.nationality}</td>'
        data += f'<td style="text-align:center;">{i.candidate_service_charge}</td>'
        data += f'<td style="text-align:center;">{i.client_payment_company_currency}</td>'
        data += f'<td style="text-align:center;">{i.billing_currency}</td>'
        data += '</tr>'
        s_no += 1

    data += '</table>'
    return data


# @frappe.whitelist()
# def download_closure_report():
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     filename_cl = "Closure Report " + posting_date
#     build_xlsx_response_cl(filename_cl)

# def build_xlsx_response_cl(filename_cl):
#     xlsx_file = make_xlsx_cl(filename_cl)
#     frappe.response['filename'] = filename_cl + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'

# def make_xlsx_cl(sheet_name="Closure Report", wb=None):
#     if wb is None:
#         wb = openpyxl.Workbook()

#     ws = wb.active
#     ws.title = sheet_name

#     fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
#     font = Font(bold=True, color="FFFFFF")
#     alignment = Alignment(horizontal="center")
#     text_wrap = Alignment(wrap_text=True)
#     title_font = Font(bold=True, size=14)
#     bold_font = Font(bold=True)
#     #bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
#     inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
#     thin_border = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )
    
#     column_widths = [6, 10, 13, 40, 20, 22, 25, 10, 37, 30, 30, 30, 15, 30, 20, 20, 20, 30, 25, 10, 30, 30, 30, 10, 15, 10, 20, 15, 20, 15, 15, 20, 15, 28, 15, 15, 15, 16, 20, 15, 20, 15, 15, 15, 32, 32]
#     columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U","V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT"]
#     for i, width in enumerate(column_widths):
#         ws.column_dimensions[columns[i]].width = width
        
#     posting_date = datetime.now().strftime("%d-%m-%Y")
#     title = "Closure Report (" + posting_date + ")"
#     ws.merge_cells("A1:AT1")
#     ws["A1"].value = title
#     ws["A1"].font = title_font
#     ws["A1"].alignment = alignment
    
#     headers = ["S NO", "ID", "Passport No", "Given Name/Surename(as per passport)",
#             "Mobile/Whatsapp", "Position/Task Subject", "Customer", "Territory", "Candidate Owner",
#             "Candidate Collection Status", "Photo(as per visa specification)", "Signed Offer Letter",
#             "Project", "Account Manager", "SA Name", "MOH", "PCC", "IAF / CV / Client Form", "Candidate Outstanding", "SA ID",
#             "Stamped Visa", "Entry Visa", "Final Medical", "Client SI", "Candidate SI", "ECR Status", "Pre-Medical", "Candidate",
#             "SO Confirmed Date", "Passport Number", "PCC Original at", "SA Name", "Visa Original at", "Final Medical Original at",
#             "Visa Expiry Date", "Visa Status", "Status", "SO Created", "Collection Priority", "PP Original at",
#             "Interview Location", "Date", "Nationality", "Candidate SC", "Client Payment Company Currency",
#             "Billing Currency(Client / Associate)"]
#     ws.append(headers)

#     for cell in ws[2]:
#         cell.fill = fill_color
#         cell.font = font
#         cell.alignment = alignment
#         cell.border = thin_border

#     data = get_data_closure()

#     for row in data:
#         ws.append(row)
#         ws[f"S{ws.max_row}"].style = inr_format
#         ws[f"X{ws.max_row}"].style = inr_format
#         ws[f"Y{ws.max_row}"].style = inr_format
#         ws[f"AR{ws.max_row}"].style = inr_format
#         ws[f"AS{ws.max_row}"].style = inr_format									
#         for cell in ws[ws.max_row]:
#             cell.alignment = text_wrap
#             cell.border = thin_border
            

#     #for cell in ws[ws.max_row]:
#     # 	cell.font = bold_font
#     # 	cell.fill = bg_fill
#     # 	cell.border = thin_border
        
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)
#     return xlsx_file

# def get_data_closure():
#     data = []
#     s_no = 1
#     closure = frappe.db.sql("""
#     SELECT * FROM `tabClosure`
#     WHERE 
#         status NOT IN ('Arrived', 'Dropped')
# """, as_dict=True)

#     for i in closure:
        
#         data.append([
#             s_no, i.name, i.passport_no, i.given_name, i.mobile, i.task_subject, i.customer, i.territory,
#             i.candidate_owner, i.collection_status, i.photo, i.sol, i.project, i.account_manager, i.sa_name,
#             i.moh, i.pcc, i.irf, i.outstanding_amount, i.sa_id, i.visa_stamping, i.visa, i.final_medical, i.client_payment_company_currency,
#             i.candidate_payment_company_currenc, i.ecr_status, i.premedical, i.candidate, i.so_confirmed_date, i.passport_number, i.pcc_original,
#             i.sams_name, i.visa_original_at, i.final_medical_original_at, i.visa_expiry_date, i.visa_status, i.status, i.so_created, i.collection_priority,
#             i.pp_original_at, i.interview_location, i.posting_date, i.nationality, i.candidate_service_charge,
#             i.client_payment_company_currency, i.billing_currency])
        
#         s_no += 1

#     return data
@frappe.whitelist()
def download_closure_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_cl = "MR-05 – Closure Detailed Status Report (CDSR)" + posting_date
    build_xlsx_response_cl(filename_cl)

def build_xlsx_response_cl(filename_cl):
    xlsx_file = make_xlsx_cl(filename_cl)
    frappe.response['filename'] = filename_cl + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_cl(sheet_name="MR-05 – Closure Detailed Status Report (CDSR)", wb=None):
    if wb is None:
        wb = openpyxl.Workbook()
    args = frappe.local.form_dict
    ws = wb.active
    ws.title = sheet_name

    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center")
    text_wrap = Alignment(wrap_text=True)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    #bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
    inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    column_widths = [6, 10, 13, 40, 20, 22, 25, 10, 37, 30, 30, 30, 15, 30, 20, 20, 20, 30, 25, 10, 30, 30, 30, 10, 15, 10, 20, 15, 20, 15, 15, 20, 15, 28, 15, 15, 15, 16, 20, 15, 20, 15, 15, 15, 32, 32]
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U","V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT"]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[columns[i]].width = width
        
    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "Closure Report (" + posting_date + ")"
    ws.merge_cells("A1:AT1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment
    
    headers = ["S NO", "ID", "Passport No", "Given Name/Surename(as per passport)",
            "Mobile/Whatsapp", "Position/Task Subject", "Customer", "Territory", "Candidate Owner",
            "Candidate Collection Status", "Photo(as per visa specification)", "Signed Offer Letter",
            "Project", "Account Manager", "SA Name", "MOH", "PCC", "IAF / CV / Client Form", "Candidate Outstanding", "SA ID",
            "Stamped Visa", "Entry Visa", "Final Medical", "Client SI", "Candidate SI", "ECR Status", "Pre-Medical", "Candidate",
            "SO Confirmed Date", "Passport Number", "PCC Original at", "SA Name", "Visa Original at", "Final Medical Original at",
            "Visa Expiry Date", "Visa Status", "Status", "SO Created", "Collection Priority", "PP Original at",
            "Interview Location", "Date", "Nationality", "Candidate SC", "Client Payment Company Currency",
            "Billing Currency(Client / Associate)"]
    ws.append(headers)

    for cell in ws[2]:
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border

    data = get_data_closure(args)

    for row in data:
        ws.append(row)
        ws[f"S{ws.max_row}"].style = inr_format
        ws[f"X{ws.max_row}"].style = inr_format
        ws[f"Y{ws.max_row}"].style = inr_format
        ws[f"AR{ws.max_row}"].style = inr_format
        ws[f"AS{ws.max_row}"].style = inr_format									
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
        
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_closure(args=None):
    data = []
    s_no = 1
    
    # Safely access the so_status and so_validate using .get() to avoid undefined errors
    am = args.get("so_status", False)  # Default to False if so_status is not found
    so_validate = args.get("so_validate", False)  # Default to False if so_validate is not found
    conditions = []
    
    # Base query to filter by status
    query = """
    SELECT *
    FROM 
        `tabClosure` 
    WHERE status NOT IN ('Arrived', 'Dropped')
    """
    
    # Add conditions based on so_validate
    if so_validate:
        # If so_validate is checked, fetch data regardless of so_created status (both True and False)
        pass  # No need to add a filter for so_created because we want both enabled and disabled data
    else:
        # If so_validate is not checked, only filter based on 'so_created'
        if am:
            conditions.append("so_created = %s")
            conditions.append(am)

    # Add the conditions to the query if any
    if conditions:
        query += " AND " + " AND ".join(conditions[::2])

    # Execute the query with the parameters
    closure = frappe.db.sql(query, tuple(conditions[1::2]), as_dict=True)

    for i in closure:
        data.append([
            s_no, i.name, i.passport_no, i.given_name, i.mobile, i.task_subject, i.customer, i.territory,
            i.candidate_owner, i.collection_status, i.photo, i.sol, i.project, i.account_manager, i.sa_name,
            i.moh, i.pcc, i.irf, i.outstanding_amount, i.sa_id, i.visa_stamping, i.visa, i.final_medical, i.client_payment_company_currency,
            i.candidate_payment_company_currenc, i.ecr_status, i.premedical, i.candidate, i.so_confirmed_date, i.passport_number, i.pcc_original,
            i.sams_name, i.visa_original_at, i.final_medical_original_at, i.visa_expiry_date, i.visa_status, i.status, i.so_created, i.collection_priority,
            i.pp_original_at, i.interview_location, i.posting_date, i.nationality, i.candidate_service_charge,
            i.client_payment_company_currency, i.billing_currency])

        s_no += 1

    return data



@frappe.whitelist()
def todo_report(allocated_to):
    s_no = 1
    posting_date = datetime.now().strftime("%d-%m-%Y")

    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="18" style="text-align:center; font-weight:bold;">TODO Report</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Subject</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Current Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '</tr>'
    todo_data = frappe.db.get_all("ToDo",{"allocated_to":allocated_to,"status":"Open"}, ["*"])
    for i in todo_data:
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.custom_subject}</td>' \
                f'<td style="text-align:center;">{i.current_status_}</td>' \
                f'<td style="text-align:center;">{i.created_on}</td>' \
                f'<td style="text-align:center;">{i.status}</td>' \
                '</tr>'
        s_no += 1

    data += '</table>'
    return data

from datetime import datetime
from openpyxl.styles import PatternFill
@frappe.whitelist()
def download_todo_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_todo = "ToDo Report"+ posting_date
    build_xlsx_response_todo(filename_todo)
    
def build_xlsx_response_todo(filename_todo):
    xlsx_file = make_xlsx_todo(filename_todo)
    frappe.response['filename'] = filename_todo + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
def make_xlsx_todo(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  # White color font
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20 
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    title_font = Font(bold=True, size=14)
    alignment = Alignment(horizontal="center")
    text_wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")  # Centered with wrap

    posting_date = datetime.now().strftime("%d-%m-%Y")
    title = "TODO Report (" + posting_date + ")"
    ws.merge_cells("A1:F1")
    ws["A1"].value = title
    ws["A1"].font = title_font
    ws["A1"].alignment = alignment
    header = ["S NO", "ID", "Subject", "Current Status", "Date", "Status"]
    ws.append(header) 
    for cell in ws[2]: 
        cell.font = header_font  # Apply white font to each header cell 
        cell.border = black_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_color
        cell.border = thin_border


    data1= get_data_of_todo(args)
    for row in data1:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = text_wrap
            cell.border = thin_border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_todo(args):
    data = []
    s_no=1
    todo_data = frappe.db.get_all("ToDo",{"allocated_to":args.allocated_to,"status":"Open"}, ["*"])
    for i in todo_data:
        formatted_creation_date = frappe.utils.formatdate(i.created_on, 'dd-mm-yyyy')

        data.append([s_no, i.name, i.custom_subject, i.current_status_,formatted_creation_date, i.status])
        s_no+=1
    return data


import frappe
from datetime import datetime

@frappe.whitelist()
def appointment_schedule_report(doc):
    s_no = 1
    posting_date = datetime.now().strftime("%d-%m-%Y")
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="8" style="text-align:center; font-weight:bold;">Appointment Schedule Report</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Territory</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Current Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Appointment On</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Appointment Created By</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Appointment To Be Taken By</td>' \
            '</tr>'

    app_data = frappe.db.get_all("Appointment", fields=["*"],order_by='scheduled_time DESC')
    for i in app_data:
        user_set = []
        territory=''
        appointment_doc = frappe.get_doc("Appointment", i.name)
        if i.custom_sales_follow_up_id:
            territory = frappe.db.get_value("Sales Follow Up", {"name":i.custom_sales_follow_up_id}, ["territory"])
        if appointment_doc.custom_appointment_to_be_created_by:
            for u in appointment_doc.custom_appointment_to_be_created_by:
                user_set.append(u.user)
        user_set_str=", ".join(user_set)
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.customer_name or ""}</td>' \
                f'<td style="text-align:center;">{territory or ""}</td>' \
                f'<td style="text-align:center;">{i.status}</td>' \
                f'<td style="text-align:center;">{i.scheduled_time or ""}</td>' \
                f'<td style="text-align:center;">{i.owner or ""}</td>' \
                f'<td style="text-align:center;">{user_set_str or ""}</td>' \
                '</tr>'

        s_no += 1
    data += '</table>'
    return data

from datetime import datetime
from openpyxl.styles import PatternFill
@frappe.whitelist()
def app_schedule_excel_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_as = "Appointment Schedule Report"+ posting_date
    build_xlsx_response_as(filename_as)
    
def build_xlsx_response_as(filename_as):
    xlsx_file = make_xlsx_as(filename_as)
    frappe.response['filename'] = filename_as + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_as(data, sheet_name=None, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  # White color font
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20 
    ws.column_dimensions['G'].width = 20 
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    header = ["Appointment Schedule Report"]
    ws.append(header)  
    ws.append(["S NO", "ID", "Name","Territory", "Current Status", "Appointment On","Appointment Created By","Appointment To Be Taken By"])
    for cell in ws[2]: 
        cell.fill = fill_color
        cell.font = header_font  # Apply white font to each header cell 
        cell.border = black_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    for cell in ws[1]: 
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data1= get_data_of_as_report()
    for row in data1:
        ws.append(row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_as_report():
    data = []
    s_no=1
    app_data = frappe.db.get_all("Appointment", fields=["*"])
    for i in app_data:
        user_set = []
        territory=''
        appointment_doc = frappe.get_doc("Appointment", i.name)
        if i.custom_sales_follow_up_id:
            territory = frappe.db.get_value("Sales Follow Up", {"name":i.custom_sales_follow_up_id}, ["territory"])
        if appointment_doc.custom_appointment_to_be_created_by:
            for u in appointment_doc.custom_appointment_to_be_created_by:
                user_set.append(u.user)
        user_set_str=", ".join(user_set)
        data.append([s_no, i.name, i.customer_name or "", territory or "", i.status,i.scheduled_time or "",i.owner or "",user_set_str or ""])
        s_no+=1
    return data

@frappe.whitelist()
def appointment_taken_report(doc):
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="18" style="text-align:center; font-weight:bold;">Appointment Taken Report</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Organization Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Territory</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Validation Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Service</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">NCB</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Remarks</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">First Name</td>' \
            '</tr>'

    app_data = frappe.db.get_all("Appointment", filters={"status": ("not in", ["Closed", "Completed"])}, fields=["*"])
    for i in app_data:
        if i.custom_sales_follow_up_id:
            sp = frappe.db.get_all("Sales Follow Up", {"name":i.custom_sales_follow_up_id}, ["*"])
        for j in sp:
            data += f'<tr>' \
                    f'<td style="text-align:center;">{s_no}</td>' \
                    f'<td style="text-align:center;">{i.name}</td>' \
                    f'<td style="text-align:center;">{i.customer_name or ""}</td>'  \
                    f'<td style="text-align:center;">{j.territory or ""}</td>' \
                    f'<td style="text-align:center;">{i.status}</td>' \
                    f'<td style="text-align:center;">{j.validation_status or ""}</td>' \
                    f'<td style="text-align:center;">{j.service or ""}</td>' \
                    f'<td style="text-align:center;">{j.next_contact_by or ""}</td>' \
                    f'<td style="text-align:center;">{j.remarks or ""}</td>' \
                    f'<td style="text-align:center;">{j.first_name or ""}</td>' \
                    '</tr>'

        s_no += 1
    data += '</table>'
    return data

from datetime import datetime
from openpyxl.styles import PatternFill
@frappe.whitelist()
def app_taken_excel_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_app = "Appointment Taken Report"+ posting_date
    build_xlsx_response_app(filename_app)
    
def build_xlsx_response_app(filename_app):
    xlsx_file = make_xlsx_app(filename_app)
    frappe.response['filename'] = filename_app + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_app(data, sheet_name=None, wb=None, column_widths=None):
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  # White color font
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20 
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20 
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    header = ["Appointment Taken Report"]
    ws.append(header)  
    ws.append(["S NO", "ID", "Organization Name","Territory", "Status", "Validation Status", "Service","NCB","Remarks","First Name"])
    for cell in ws[2]: 
        cell.fill = fill_color
        cell.font = header_font  # Apply white font to each header cell 
        cell.border = black_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    for cell in ws[1]: 
        cell.alignment = Alignment(horizontal="center", vertical="center")
    data1= get_data_of_app_report()
    for row in data1:
        ws.append(row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_app_report():
    data = []
    s_no=1
    app_data = frappe.db.get_all("Appointment", filters={"status": ("not in", ["Closed", "Completed"])}, fields=["*"])
    for i in app_data:
        if i.custom_sales_follow_up_id:
            sp = frappe.db.get_all("Sales Follow Up", {"name":i.custom_sales_follow_up_id}, ["*"])
        for j in sp:
            data.append([s_no, i.name, i.customer_name or "", j.territory or "", i.status,j.validation_status or "",j.service or "",j.next_contact_by or "",j.remarks or "",j.first_name or ""])
        s_no+=1
    return data

@frappe.whitelist(allow_guest=True)
def otp_verification(otpSent, otpValue, mobile):
    if otpSent == otpValue:
        user_data = frappe.db.sql("""select name, full_name from `tabUser` where mobile_no = '%s'""" %(mobile), as_dict=1)
        if user_data:
            user = user_data[0].name
            full_name = user_data[0].full_name
            candidate_data = frappe.db.sql("""select name from `tabCandidate` where mail_id = '%s'""" %(user), as_dict=1)
            if candidate_data:
                candidate = candidate_data[0].name
                from frappe.auth import LoginManager
                login_manager = LoginManager()
                login_manager.login_as(user)
                auth_token = frappe.generate_hash(length=32)
                frappe.cache().hset("auth_token", auth_token, user)
                result = {
                    "status": "success",
                    "message": candidate,
                    "auth_token": auth_token,
                    "full_name": full_name,
                    "user": user,
                }
            else:
                result = 'candidate not found'
        else:
            result = 'user not found'
    else:
        result = "invalid"
    return result


@frappe.whitelist()
def print_closure_count_report():
    from datetime import datetime
    posting_date = datetime.now().strftime("%d-%m-%Y")

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)

    data = '<table style="width: 100%;">'
    data += f'<tr><td colspan="18" style="text-align:center; font-weight:bold;">Closure - Count & Status Report - {posting_date}</td></tr>'
    data += (
        '<tr style="background-color: #dce6f1;">'
        '<td rowspan=2 colspan=2 style="text-align:center; font-weight:bold; color:black;">Row Labels</td>'
    )
    closure = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    for i in closure:
        if i.nationality:
            data += f'<td colspan=3 style="text-align:center; font-weight:bold; color:black;">{i.nationality}</td>' 
    data += (
        '<td style="text-align:center; font-weight:bold; color:black;">Total Count</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Cand.</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Client.</td>'
        '</tr>'
        '<tr style="background-color: #dce6f1;">'
    )
    
    for _ in closure:
        data += (
            '<td style="text-align:center; font-weight:bold; color:black;">Count</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
        )
    data += '<td style="text-align:center; font-weight:bold; color:black;">Count</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
    total_count = 0
    total_candidate_payment_company_currenc = 0
    total_client_payment_company_currency = 0
    nationality = [i.nationality for i in closure if i.nationality]
    data += '<tr>' \
             '<td colspan=2 style="text-align:left; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0 and nationality = '%s'""" %(nat), as_dict=True)

        for so_0 in so_created_0_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_0.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0.cand or 0)}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0.client or 0)}</td>'
     
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""", as_dict=True)
    if so_created_0_total:
        data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["count"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["cand"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["client"]) or 0}</td>'
    else:
        data += '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    data += '</tr>' 
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "0"})
        if s.status and has_customer:

            data += (
                '<tr>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                f'<td style="text-align:left; font-weight:bold; color:black;">{s.status}</td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '</tr>'
            )

            

            for j in closure:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 0
                    GROUP BY customer
                """, (s.status, j.nationality), as_dict=True)

                for i in col:
                    data += '<tr>'
                    data += '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                    data += f'<td style="text-align:left; color:black; padding-left: 10px;">{i.customer}</td>'

                    for nat in closure:
                        if nat.nationality:
                            total_count=total_candidate_payment_company_currenc=total_client_payment_company_currency=0
                            if nat.nationality == j.nationality:
                                data += f'<td style="text-align:right; color:black;">{i.count}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.candidate_payment_company_currenc)}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.client_payment_company_currency)}</td>'
                            else:
                                data += '<td style="text-align:right; color:black;"></td>' * 3
                    total_count += i.count		
                    total_candidate_payment_company_currenc += i.candidate_payment_company_currenc
                    total_client_payment_company_currency += i.client_payment_company_currency
                    data += f'<td style="text-align:right; color:black;">{total_count}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_candidate_payment_company_currenc)}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_client_payment_company_currency)}</td>'
                    data += '</tr>'
    nationality = [i.nationality for i in closure if i.nationality]
    data += '<tr>' \
             '<td colspan=2 style="text-align:left; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">1</td>'
    for nat in nationality:
        so_created_1_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1 and nationality = '%s'""" %(nat), as_dict=True)

        
        for so_1 in so_created_1_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.cand) or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.client) or 0}</td>'
     
    so_created_1_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""", as_dict=True)
    if so_created_1_total:
        data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_1_total[0]["count"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_1_total[0]["cand"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_1_total[0]["client"]) or 0}</td>'
    else:
        data += '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    data += '</tr>' 
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "1"})
        if s.status and has_customer:
            data += (
                '<tr>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                f'<td style="text-align:left; font-weight:bold; color:black;">{s.status}</td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '</tr>'
            )

            

            for j in closure:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 1
                    GROUP BY customer
                """, (s.status, j.nationality), as_dict=True)

                for i in col:
                    data += '<tr>'
                    data += '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                    data += f'<td style="text-align:left; color:black; padding-left: 10px;">{i.customer}</td>'

                    for nat in closure:
                        if nat.nationality:
                            total_count=total_candidate_payment_company_currenc=total_client_payment_company_currency=0
                            if nat.nationality == j.nationality:
                                data += f'<td style="text-align:right; color:black;">{i.count}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.candidate_payment_company_currenc)}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.client_payment_company_currency)}</td>'
                                        
                                total_count += i.count
                                total_candidate_payment_company_currenc += i.candidate_payment_company_currenc
                                total_client_payment_company_currency += i.client_payment_company_currency
                            else:
                                data += '<td style="text-align:right; color:black;"></td>' * 3
                    total_count += i.count
                    total_candidate_payment_company_currenc += i.candidate_payment_company_currenc
                    total_client_payment_company_currency += i.client_payment_company_currency
                    data += f'<td style="text-align:right; color:black;">{total_count}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_candidate_payment_company_currenc)}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_client_payment_company_currency)}</td>'
                    data += '</tr>'

    data += '<tr style="background-color: #dce6f1;">'
    data += '<td colspan=2 style="text-align:left; font-weight:bold; color:black;">Grand Total</td>'
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s'""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_0_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.cand) or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.client) or 0}</td>'
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived")""" , as_dict=True)
    data += f'<td style="text-align:right; font-weight:bold; color:black;">{total_so_created_nation[0]["count"]}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["cand"])}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["client"])}</td>' \
            
    data += '</tr>'

    data += '</table>'
    data += """<div style="display: flex; justify-content: center; margin-top: 100px;">
                    <ol style="list-style-type: disc;">
                        <h4>PivotTable Fields</h4>
                        <li>Customer</li>
                        <li>Candidate SI</li>
                        <li>Client SI</li>
                        <li>Status</li>
                        <li>SO Created</li>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">

                        <h4>Columns</h4>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Rows</h4>
                        <li>SO Created</li>
                        <li>Status</li>
                        <li>Customer</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Values</h4>
                        <li>Count</li>
                        <li>Cand.</li>
                        <li>Client.</li>
                    </ol>
                </div>"""
    print(data)
    return data

@frappe.whitelist()
def print_closure_count_report_so_true(doc):
    from datetime import datetime
    posting_date = datetime.now().strftime("%d-%m-%Y")

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)

    data = '<table style="width: 100%;">'
    data += f'<tr><td colspan="18" style="text-align:center; font-weight:bold;">Closure - Count & Status Report - {posting_date}</td></tr>'
    data += (
        '<tr style="background-color: #dce6f1;">'
        '<td rowspan=2 colspan=2 style="text-align:center; font-weight:bold; color:black;">Row Labels</td>'
    )
    closure = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    for i in closure:
        if i.nationality:
            data += f'<td colspan=3 style="text-align:center; font-weight:bold; color:black;">{i.nationality}</td>' 
    data += (
        '<td style="text-align:center; font-weight:bold; color:black;">Total Count</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Cand.</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Client.</td>'
        '</tr>'
        '<tr style="background-color: #dce6f1;">'
    )
    
    for _ in closure:
        data += (
            '<td style="text-align:center; font-weight:bold; color:black;">Count</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
        )
    data += '<td style="text-align:center; font-weight:bold; color:black;">Count</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
    total_count = 0
    total_candidate_payment_company_currenc = 0
    total_client_payment_company_currency = 0
    nationality = [i.nationality for i in closure if i.nationality]
    data += '<tr>' \
             '<td colspan=2 style="text-align:left; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1 and nationality = '%s'""" %(nat), as_dict=True)

        for so_1 in so_created_0_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.cand or 0)}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.client or 0)}</td>'
     
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""", as_dict=True)
    if so_created_0_total:
        data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["count"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["cand"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["client"]) or 0}</td>'
    else:
        data += '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    data += '</tr>' 
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created":"1"})
        if s.status and has_customer:
            data += (
                '<tr>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                f'<td style="text-align:left; font-weight:bold; color:black;">{s.status}</td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '</tr>'
            )

            

            for j in closure:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 1
                    GROUP BY customer
                """, (s.status, j.nationality), as_dict=True)

                for i in col:
                    data += '<tr>'
                    data += '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                    data += f'<td style="text-align:left; color:black; padding-left: 10px;">{i.customer}</td>'

                    for nat in closure:
                        if nat.nationality:
                            total_count=total_candidate_payment_company_currenc=total_client_payment_company_currency=0
                            if nat.nationality == j.nationality:
                                data += f'<td style="text-align:right; color:black;">{i.count}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.candidate_payment_company_currenc)}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.client_payment_company_currency)}</td>'
                            else:
                                data += '<td style="text-align:right; color:black;"></td>' * 3
                            
                    total_count += i.count
                    total_candidate_payment_company_currenc += i.candidate_payment_company_currenc
                    total_client_payment_company_currency += i.client_payment_company_currency
                    data += f'<td style="text-align:right; color:black;">{total_count}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_candidate_payment_company_currenc)}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_client_payment_company_currency)}</td>'
                    data += '</tr>'
    
    data += '<tr style="background-color: #dce6f1;">'
    data += '<td colspan=2 style="text-align:left; font-weight:bold; color:black;">Grand Total</td>'
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s' and so_created = 1""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_0_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.cand) or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.client) or 0}</td>'
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""" , as_dict=True)
    data += f'<td style="text-align:right; font-weight:bold; color:black;">{total_so_created_nation[0]["count"]}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["cand"])}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["client"])}</td>' \
            
    data += '</tr>'

    data += '</table>'
    data += """<div style="display: flex; justify-content: center; margin-top: 100px;">
                    <ol style="list-style-type: disc;">
                        <h4>PivotTable Fields</h4>
                        <li>Customer</li>
                        <li>Candidate SI</li>
                        <li>Client SI</li>
                        <li>Status</li>
                        <li>SO Created</li>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">

                        <h4>Columns</h4>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Rows</h4>
                        <li>SO Created</li>
                        <li>Status</li>
                        <li>Customer</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Values</h4>
                        <li>Count</li>
                        <li>Cand.</li>
                        <li>Client.</li>
                    </ol>
                </div>"""
    print(data)
    return data

@frappe.whitelist()
def print_closure_count_report_so(doc):
    from datetime import datetime
    posting_date = datetime.now().strftime("%d-%m-%Y")

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)

    data = '<table style="width: 100%;">'
    data += f'<tr><td colspan="18" style="text-align:center; font-weight:bold;">Closure - Count & Status Report - {posting_date}</td></tr>'
    data += (
        '<tr style="background-color: #dce6f1;">'
        '<td rowspan=2 colspan=2 style="text-align:center; font-weight:bold; color:black;">Row Labels</td>'
    )
    closure = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    for i in closure:
        if i.nationality:
            data += f'<td colspan=3 style="text-align:center; font-weight:bold; color:black;">{i.nationality}</td>' 
    data += (
        '<td style="text-align:center; font-weight:bold; color:black;">Total Count</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Cand.</td>'
        '<td style="text-align:center; font-weight:bold; color:black;">Total Client.</td>'
        '</tr>'
        '<tr style="background-color: #dce6f1;">'
    )
    
    for _ in closure:
        data += (
            '<td style="text-align:center; font-weight:bold; color:black;">Count</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>'
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
        )
    data += '<td style="text-align:center; font-weight:bold; color:black;">Count</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Cand.</td>' \
            '<td style="text-align:center; font-weight:bold; color:black;">Client.</td>'
    total_count = 0
    total_candidate_payment_company_currenc = 0
    total_client_payment_company_currency = 0
    nationality = [i.nationality for i in closure if i.nationality]
    data += '<tr>' \
             '<td colspan=2 style="text-align:left; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0 and nationality = '%s'""" %(nat), as_dict=True)

        for so_1 in so_created_0_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.cand or 0)}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_1.client or 0)}</td>'
     
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""", as_dict=True)
    if so_created_0_total:
        data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["count"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["cand"]) or 0}</td>' \
                f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_created_0_total[0]["client"]) or 0}</td>'
    else:
        data += '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>' \
                '<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">0</td>'
    data += '</tr>' 
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created":"0"})
        if s.status and has_customer:
            data += (
                '<tr>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                f'<td style="text-align:left; font-weight:bold; color:black;">{s.status}</td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                '</tr>'
            )

            

            for j in closure:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 0
                    GROUP BY customer
                """, (s.status, j.nationality), as_dict=True)

                for i in col:
                    data += '<tr>'
                    data += '<td style="text-align:right; font-weight:bold; color:black;"></td>'
                    data += f'<td style="text-align:left; color:black; padding-left: 10px;">{i.customer}</td>'

                    for nat in closure:
                        if nat.nationality:
                            total_count=total_candidate_payment_company_currenc=total_client_payment_company_currency=0
                            if nat.nationality == j.nationality:
                                data += f'<td style="text-align:right; color:black;">{i.count}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.candidate_payment_company_currenc)}</td>'
                                data += f'<td style="text-align:right; color:black;">{int(i.client_payment_company_currency)}</td>'
                            else:
                                data += '<td style="text-align:right; color:black;"></td>' * 3
                            
                    total_count += i.count
                    total_candidate_payment_company_currenc += i.candidate_payment_company_currenc
                    total_client_payment_company_currency += i.client_payment_company_currency
                    data += f'<td style="text-align:right; color:black;">{total_count}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_candidate_payment_company_currenc)}</td>' \
                            f'<td style="text-align:right; color:black;">{int(total_client_payment_company_currency)}</td>'
                    data += '</tr>'
    
    data += '<tr style="background-color: #dce6f1;">'
    data += '<td colspan=2 style="text-align:left; font-weight:bold; color:black;">Grand Total</td>'
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s' and so_created = 0""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            data += f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{so_0_1.count or 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.cand) if so_0_1.cand else 0}</td>' \
                    f'<td style="text-align:right; font-weight:bold; color:black;border: 1px solid #dce6f1; border-right-style: hidden;border-top-style: hidden;border-left-style: hidden;">{int(so_0_1.client) if so_0_1.client else 0}</td>'
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""" , as_dict=True)
    data += f'<td style="text-align:right; font-weight:bold; color:black;">{total_so_created_nation[0]["count"]}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["cand"])}</td>' \
            f'<td style="text-align:right; font-weight:bold; color:black;">{int(total_so_created_nation[0]["client"])}</td>' \
            
    data += '</tr>'

    data += '</table>'
    data += """<div style="display: flex; justify-content: center; margin-top: 100px;">
                    <ol style="list-style-type: disc;">
                        <h4>PivotTable Fields</h4>
                        <li>Customer</li>
                        <li>Candidate SI</li>
                        <li>Client SI</li>
                        <li>Status</li>
                        <li>SO Created</li>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">

                        <h4>Columns</h4>
                        <li>Nationality</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Rows</h4>
                        <li>SO Created</li>
                        <li>Status</li>
                        <li>Customer</li>
                    </ol>
                    <ol style="list-style-type: disc; margin-left: 10px;">
                        <h4>Values</h4>
                        <li>Count</li>
                        <li>Cand.</li>
                        <li>Client.</li>
                    </ol>
                </div>"""
    print(data)
    return data

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
import frappe
from frappe.utils import nowdate
from openpyxl.utils import get_column_letter

@frappe.whitelist()
def download_closure_status_report(posting_date=None):
    if not posting_date:
        posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"PR:03 – Closure Status Report (CSR) {posting_date}"
    build_xlsx_response_clsr(filename, posting_date)

def build_xlsx_response_clsr(filename, posting_date):
    xlsx_file = make_xlsx_clsr(filename, posting_date)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_clsr(sheet_name="PR:03 – Closure Status Report (CSR)", posting_date=None, wb=None):
    if wb is None:
        wb = Workbook()
    
    # ws = wb.active
    # ws.title = sheet_name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    fill_color = PatternFill(start_color="dce6f1", end_color="dce6f1", fill_type="solid")
    white_color = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    font = Font(bold=True, color="000000")
    alignment = Alignment(horizontal="center")
    alignment_right = Alignment(horizontal="right")
    text_wrap = Alignment(wrap_text=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    bottom_border = Border(bottom=Side(style="thin", color="dce6f1"))  
    nationalities = get_nationalities()

    total_columns = 2 + (len(nationalities) * 3) + 3  

    title_text = f"PR:03 – Closure Status Report (CSR)- {posting_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = alignment

    ws.row_dimensions[1].height = 25 
    ws.row_dimensions[2].height = 20 

    column_widths = [5, 50] 
    for _ in nationalities:
        column_widths.extend([12, 12, 12])  
    column_widths.extend([12, 12, 12]) 

    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    header_row_start = 2
    ws.merge_cells(f"A{header_row_start}:B{header_row_start + 1}")
    ws["A2"].value = "Row Labels"
    ws["A2"].alignment = alignment
    ws["A2"].font = font

    col_index = 3
    for nationality in nationalities:
        start_col = get_column_letter(col_index)
        end_col = get_column_letter(col_index + 2)
        ws.merge_cells(f"{start_col}{header_row_start}:{end_col}{header_row_start}")
        ws[f"{start_col}{header_row_start}"].value = nationality
        ws[f"{start_col}{header_row_start}"].alignment = alignment
        ws[f"{start_col}{header_row_start}"].font = font
        col_index += 3

    summary_start = get_column_letter(col_index)
    ws.merge_cells(f"{summary_start}{header_row_start}:{get_column_letter(col_index + 2)}{header_row_start}")
    ws[f"{summary_start}{header_row_start}"].value = "Total"
    ws[f"{summary_start}{header_row_start}"].alignment = alignment
    ws[f"{summary_start}{header_row_start}"].font = font

    header_row_second = header_row_start + 1
    col_index = 3
    for _ in nationalities:
        ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
        ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
        ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."
        col_index += 3

    ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
    ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
    ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."


    nationality = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    nationality = [nat.nationality for nat in nationality if nat.nationality]
    row = ["0", ""]
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0 and nationality = '%s'""" %(nat), as_dict=True)
        for so_1 in so_created_0_nation:
            row += [so_1.count or 0, so_1.cand or 0, so_1.client or 0]
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""", as_dict=True)
    if so_created_0_total:
        row += [so_created_0_total[0]["count"] or 0, so_created_0_total[0]["cand"] or 0, so_created_0_total[0]["client"] or 0]
    else:
        row += [0, 0, 0]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font
        cell.fill = white_color
        cell.border = bottom_border

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "0"})
        if s.status and has_customer:
            row = ["", s.status, "", "", "","", "", "", "", "", "", "", "", ""]
            
            ws.append(row)
            row_idx = ws.max_row  
            for col_idx, cell_value in enumerate(row, start=1): 
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font
                cell.fill = white_color
            for nat in nationality:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 0
                    GROUP BY customer
                """, (s.status, nat), as_dict=True)
                for i in col:
                    row = ["", i.customer]
                    for n in nationality:
                        total_count=total_cand=total_client=0
                        if n:
                            if nat == n:
                                row += [i.count, i.candidate_payment_company_currenc, i.client_payment_company_currency]
                            else:
                                row += ["", "", ""]
                    total_count += i.count
                    total_cand += i.candidate_payment_company_currenc
                    total_client += i.client_payment_company_currency
                    row += [total_count, total_cand, total_client]
                    ws.append(row)
                    row_idx = ws.max_row  
                    for col_idx, cell_value in enumerate(row, start=1): 
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = white_color
    
    row = ["1", ""]
    for nat in nationality:
        so_created_1_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1 and nationality = '%s'""" %(nat), as_dict=True)
        for so_1 in so_created_1_nation:
            row += [so_1.count or 0, so_1.cand or 0, so_1.client or 0]
    so_created_1_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""", as_dict=True)
    if so_created_0_total:
        row += [so_created_1_total[0]["count"] or 0, so_created_1_total[0]["cand"] or 0, so_created_1_total[0]["client"] or 0]
    else:
        row += [0, 0, 0]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font
        cell.fill = white_color
        cell.border = bottom_border
    
    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)
    
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "1"})
        if s.status and has_customer:
            row = ["", s.status, "", "", "","", "", "", "", "", "", "", "", ""]
            
            ws.append(row)
            row_idx = ws.max_row  
            for col_idx, cell_value in enumerate(row, start=1): 
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font
                cell.fill = white_color
            for nat in nationality:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 1
                    GROUP BY customer
                """, (s.status, nat), as_dict=True)
                for i in col:
                    row = ["", i.customer]
                    for n in nationality:
                        total_count=total_cand=total_client=0
                        if n:
                            if nat == n:
                                row += [i.count, i.candidate_payment_company_currenc, i.client_payment_company_currency]
                            else:
                                row += ["", "", ""]
                    total_count += i.count
                    total_cand += i.candidate_payment_company_currenc
                    total_client += i.client_payment_company_currency
                    row += [total_count, total_cand, total_client]
                    ws.append(row)
                    row_idx = ws.max_row  
                    for col_idx, cell_value in enumerate(row, start=1): 
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = white_color
    row = ["", "Grand Total"]
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s'""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            row += [so_0_1.count or 0, so_0_1.cand or 0, so_0_1.client or 0]
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived")""" , as_dict=True)
    row += [total_so_created_nation[0]["count"], total_so_created_nation[0]["cand"], total_so_created_nation[0]["client"]]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment_right
  
    for row in ws.iter_rows(min_row=header_row_start, max_row=header_row_second, max_col=total_columns):
        for cell in row:
            cell.fill = fill_color
            cell.font = font
            cell.alignment = alignment
    
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO
import frappe
from frappe.utils import nowdate
from openpyxl.utils import get_column_letter

@frappe.whitelist()
def download_closure_status_report_so_false(posting_date=None):
    if not posting_date:
        posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"PR:03 – Closure Status Report (CSR) {posting_date}"
    build_xlsx_response_clsr_so_false(filename, posting_date)

def build_xlsx_response_clsr_so_false(filename, posting_date):
    xlsx_file = make_xlsx_clsr_so_false(filename, posting_date)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_clsr_so_false(sheet_name="PR:03 – Closure Status Report (CSR)", posting_date=None, wb=None):
    if wb is None:
        wb = Workbook()
    
    # ws = wb.active
    # ws.title = sheet_name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    fill_color = PatternFill(start_color="dce6f1", end_color="dce6f1", fill_type="solid")
    white_color = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    font = Font(bold=True, color="000000")
    alignment = Alignment(horizontal="center")
    alignment_right = Alignment(horizontal="right")
    text_wrap = Alignment(wrap_text=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    bottom_border = Border(bottom=Side(style="thin", color="dce6f1"))  
    nationalities = get_nationalities()

    total_columns = 2 + (len(nationalities) * 3) + 3  

    title_text = f"PR:03 – Closure Status Report (CSR)- {posting_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = alignment

    ws.row_dimensions[1].height = 25 
    ws.row_dimensions[2].height = 20 

    column_widths = [5, 50] 
    for _ in nationalities:
        column_widths.extend([12, 12, 12])  
    column_widths.extend([12, 12, 12]) 

    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    header_row_start = 2
    ws.merge_cells(f"A{header_row_start}:B{header_row_start + 1}")
    ws["A2"].value = "Row Labels"
    ws["A2"].alignment = alignment
    ws["A2"].font = font

    col_index = 3
    for nationality in nationalities:
        start_col = get_column_letter(col_index)
        end_col = get_column_letter(col_index + 2)
        ws.merge_cells(f"{start_col}{header_row_start}:{end_col}{header_row_start}")
        ws[f"{start_col}{header_row_start}"].value = nationality
        ws[f"{start_col}{header_row_start}"].alignment = alignment
        ws[f"{start_col}{header_row_start}"].font = font
        col_index += 3

    summary_start = get_column_letter(col_index)
    ws.merge_cells(f"{summary_start}{header_row_start}:{get_column_letter(col_index + 2)}{header_row_start}")
    ws[f"{summary_start}{header_row_start}"].value = "Total"
    ws[f"{summary_start}{header_row_start}"].alignment = alignment
    ws[f"{summary_start}{header_row_start}"].font = font

    header_row_second = header_row_start + 1
    col_index = 3
    for _ in nationalities:
        ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
        ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
        ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."
        col_index += 3

    ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
    ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
    ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."


    nationality = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    nationality = [nat.nationality for nat in nationality if nat.nationality]
    row = ["0", ""]
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0 and nationality = '%s'""" %(nat), as_dict=True)
        for so_1 in so_created_0_nation:
            row += [so_1.count or 0, so_1.cand or 0, so_1.client or 0]
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""", as_dict=True)
    if so_created_0_total:
        row += [so_created_0_total[0]["count"] or 0, so_created_0_total[0]["cand"] or 0, so_created_0_total[0]["client"] or 0]
    else:
        row += [0, 0, 0]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font
        cell.fill = white_color
        cell.border = bottom_border

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "0"})
        if s.status and has_customer:
            row = ["", s.status, "", "", "","", "", "", "", "", "", "", "", ""]
            
            ws.append(row)
            row_idx = ws.max_row  
            for col_idx, cell_value in enumerate(row, start=1): 
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font
                cell.fill = white_color
            for nat in nationality:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 0
                    GROUP BY customer
                """, (s.status, nat), as_dict=True)
                for i in col:
                    row = ["", i.customer]
                    for n in nationality:
                        total_count=total_cand=total_client=0
                        if n:
                            if nat == n:
                                row += [i.count, i.candidate_payment_company_currenc, i.client_payment_company_currency]
                            else:
                                row += ["", "", ""]
                    total_count += i.count
                    total_cand += i.candidate_payment_company_currenc
                    total_client += i.client_payment_company_currency
                    row += [total_count, total_cand, total_client]
                    ws.append(row)
                    row_idx = ws.max_row  
                    for col_idx, cell_value in enumerate(row, start=1): 
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = white_color
    
        row = ["", "Grand Total"]
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s' and so_created = 0""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            row += [so_0_1.count or 0, so_0_1.cand or 0, so_0_1.client or 0]
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 0""" , as_dict=True)
    row += [total_so_created_nation[0]["count"], total_so_created_nation[0]["cand"], total_so_created_nation[0]["client"]]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment_right
  
    for row in ws.iter_rows(min_row=header_row_start, max_row=header_row_second, max_col=total_columns):
        for cell in row:
            cell.fill = fill_color
            cell.font = font
            cell.alignment = alignment
    
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

@frappe.whitelist()
def download_closure_status_report_so_true(posting_date=None):
    if not posting_date:
        posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = f"PR:03 – Closure Status Report (CSR) {posting_date}"
    build_xlsx_response_clsr_so_true(filename, posting_date)

def build_xlsx_response_clsr_so_true(filename, posting_date):
    xlsx_file = make_xlsx_clsr_so_true(filename, posting_date)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx_clsr_so_true(sheet_name="PR:03 – Closure Status Report (CSR)", posting_date=None, wb=None):
    if wb is None:
        wb = Workbook()
    
    # ws = wb.active
    # ws.title = sheet_name
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    fill_color = PatternFill(start_color="dce6f1", end_color="dce6f1", fill_type="solid")
    white_color = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    font = Font(bold=True, color="000000")
    alignment = Alignment(horizontal="center")
    alignment_right = Alignment(horizontal="right")
    text_wrap = Alignment(wrap_text=True)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    bottom_border = Border(bottom=Side(style="thin", color="dce6f1"))  
    nationalities = get_nationalities()

    total_columns = 2 + (len(nationalities) * 3) + 3  

    title_text = f"PR:03 – Closure Status Report (CSR)- {posting_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = alignment

    ws.row_dimensions[1].height = 25 
    ws.row_dimensions[2].height = 20 

    column_widths = [5, 50] 
    for _ in nationalities:
        column_widths.extend([12, 12, 12])  
    column_widths.extend([12, 12, 12]) 

    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    header_row_start = 2
    ws.merge_cells(f"A{header_row_start}:B{header_row_start + 1}")
    ws["A2"].value = "Row Labels"
    ws["A2"].alignment = alignment
    ws["A2"].font = font

    col_index = 3
    for nationality in nationalities:
        start_col = get_column_letter(col_index)
        end_col = get_column_letter(col_index + 2)
        ws.merge_cells(f"{start_col}{header_row_start}:{end_col}{header_row_start}")
        ws[f"{start_col}{header_row_start}"].value = nationality
        ws[f"{start_col}{header_row_start}"].alignment = alignment
        ws[f"{start_col}{header_row_start}"].font = font
        col_index += 3

    summary_start = get_column_letter(col_index)
    ws.merge_cells(f"{summary_start}{header_row_start}:{get_column_letter(col_index + 2)}{header_row_start}")
    ws[f"{summary_start}{header_row_start}"].value = "Total"
    ws[f"{summary_start}{header_row_start}"].alignment = alignment
    ws[f"{summary_start}{header_row_start}"].font = font

    header_row_second = header_row_start + 1
    col_index = 3
    for _ in nationalities:
        ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
        ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
        ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."
        col_index += 3

    ws[f"{get_column_letter(col_index)}{header_row_second}"].value = "Count"
    ws[f"{get_column_letter(col_index + 1)}{header_row_second}"].value = "Cand."
    ws[f"{get_column_letter(col_index + 2)}{header_row_second}"].value = "Client."


    nationality = frappe.db.sql("""
            SELECT DISTINCT nationality 
            FROM `tabClosure` where status not in ("Dropped", "Arrived")
        """, as_dict=True)
    nationality = [nat.nationality for nat in nationality if nat.nationality]
    row = ["0", ""]
    for nat in nationality:
        so_created_0_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1 and nationality = '%s'""" %(nat), as_dict=True)
        for so_1 in so_created_0_nation:
            row += [so_1.count or 0, so_1.cand or 0, so_1.client or 0]
    so_created_0_total = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""", as_dict=True)
    if so_created_0_total:
        row += [so_created_0_total[0]["count"] or 0, so_created_0_total[0]["cand"] or 0, so_created_0_total[0]["client"] or 0]
    else:
        row += [0, 0, 0]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = font
        cell.fill = white_color
        cell.border = bottom_border

    status_list = frappe.db.sql("""
        SELECT DISTINCT status 
        FROM `tabClosure` 
        WHERE status NOT IN ("Dropped", "Arrived")
    """, as_dict=True)
    for s in status_list:
        has_customer = frappe.db.exists("Closure", {"status": s.status, "customer": ["!=", ""], "so_created": "1"})
        if s.status and has_customer:
            row = ["", s.status, "", "", "","", "", "", "", "", "", "", "", ""]
            
            ws.append(row)
            row_idx = ws.max_row  
            for col_idx, cell_value in enumerate(row, start=1): 
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font
                cell.fill = white_color
            for nat in nationality:
                col = frappe.db.sql("""
                    SELECT customer, 
                        COUNT(customer) AS count, 
                        SUM(candidate_payment_company_currenc) AS candidate_payment_company_currenc, 
                        SUM(client_payment_company_currency) AS client_payment_company_currency 
                    FROM `tabClosure` 
                    WHERE status = %s 
                    AND nationality = %s
                    AND so_created = 1
                    GROUP BY customer
                """, (s.status, nat), as_dict=True)
                for i in col:
                    row = ["", i.customer]
                    for n in nationality:
                        total_count=total_cand=total_client=0
                        if n:
                            if nat == n:
                                row += [i.count, i.candidate_payment_company_currenc, i.client_payment_company_currency]
                            else:
                                row += ["", "", ""]
                    total_count += i.count
                    total_cand += i.candidate_payment_company_currenc
                    total_client += i.client_payment_company_currency
                    row += [total_count, total_cand, total_client]
                    ws.append(row)
                    row_idx = ws.max_row  
                    for col_idx, cell_value in enumerate(row, start=1): 
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = white_color
    
        row = ["", "Grand Total"]
    for nat in nationality:
        so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and nationality = '%s' and so_created = 1""" %(nat), as_dict=True)
        for so_0_1 in so_created_nation:
            row += [so_0_1.count or 0, so_0_1.cand or 0, so_0_1.client or 0]
    total_so_created_nation = frappe.db.sql("""select count(customer) as count, sum(candidate_payment_company_currenc) as cand, sum(client_payment_company_currency) as client from `tabClosure` where status not in ("Dropped", "Arrived") and so_created = 1""" , as_dict=True)
    row += [total_so_created_nation[0]["count"], total_so_created_nation[0]["cand"], total_so_created_nation[0]["client"]]
    ws.append(row)
    row_idx = ws.max_row  
    for col_idx, cell_value in enumerate(row, start=1): 
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill_color
        cell.font = font
        cell.alignment = alignment_right
  
    for row in ws.iter_rows(min_row=header_row_start, max_row=header_row_second, max_col=total_columns):
        for cell in row:
            cell.fill = fill_color
            cell.font = font
            cell.alignment = alignment
    
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_nationalities():
    nationalities = frappe.db.sql("""select  distinct nationality from `tabClosure` where status not in ("Dropped", "Arrived")""", as_dict=True)
    if nationalities:
        return [d['nationality'] for d in nationalities]
    return []


