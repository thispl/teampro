import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime
from io import BytesIO


    
@frappe.whitelist()
def download_PSR_new():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def make_xlsx_PSR(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 11,10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    # yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")
    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL","PSL in Value",
        "Ex Value", "Ex PSL","AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    ws.freeze_panes = "A3"
    # yellow_columns = {"Completed Value", "Cr. Exp. Value", "Cr.Exp.PSL", "Exp.Week"}
    header_row = ws[ws.max_row]

    for cell in header_row:
        # if cell.value in yellow_columns:
        #     cell.fill = yellow_fill
        #     cell.font = black_font
        # else:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
    cust = frappe.get_all("Customer", fields=["name", "territory"])
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = total_psl_value = 0
    
    proj_s_no = 1 
    for c in cust:
        pname = frappe.get_all("Project", {
            "status": ("in", ['Open', 'Enquiry']),
            "customer": c['name'],
            "service": ("in", ['REC-I', 'REC-D'])
        }, ['*'], order_by="priority ASC")

        if not pname:
            continue

        blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
        customer_with_territory = f"{c['name']} - {c['territory']}"
        row_data = [int_to_roman(serial_number), customer_with_territory] + [""] * 17
        ws.append(row_data)
        row_to_fill = ws.max_row
        customer_row_index = ws.max_row 

        ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
        ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
        ws.cell(row=customer_row_index, column=11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=customer_row_index, column=12).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=customer_row_index, column=12).number_format = '#,##0'
        ws.cell(row=customer_row_index, column=13).alignment = Alignment(horizontal="center", vertical="center")
        for col, cell in enumerate(ws[row_to_fill], start=1):
            if col <= 16:
                cell.fill = blue_fill
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if 2 < col <= 11:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col == 12:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
        serial_number += 1
        row += 1
        cust_vac = cust_sp = cust_fp = cust_sl = cust_lp = cust_psl = cust_ev = cust_ex_psl = cust_psl_value = 0
        for p in pname:
            project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = project_psl_value = 0
            project_psl_value =p.get('custom_psl_value') or 0
            tasks = frappe.get_all("Task", {
                "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
                "project": p.name
            }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

            for task in tasks:
                project_vac += task.get('vac', 0) or 0
                project_sp += task.get('sp', 0) or 0
                project_fp += task.get('fp', 0) or 0
                project_sl += task.get('sl', 0) or 0
                project_lp += task.get('custom_lp', 0) or 0
                project_psl += task.get('psl', 0) or 0

            total_vac += project_vac
            total_sp += project_sp
            total_fp += project_fp
            total_sl += project_sl
            total_lp += project_lp
            total_psl += project_psl
            total_psl_value +=project_psl_value
            cust_vac += project_vac
            cust_sp += project_sp
            cust_fp += project_fp
            cust_sl += project_sl
            cust_lp += project_lp
            cust_psl += project_psl
            cust_psl_value += project_psl_value
            cust_ev += float(p.get('expected_value', 0) or 0)
            cust_ex_psl += float(p.get('expected_psl', 0) or 0)


            task_data = [
                p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,p.get('custom_psl_value') or 0,
                p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
                p.get('custom_spoc_remark', '')
            ]
            ws.append([proj_s_no]+ task_data)
            try:
                ev_total += float(p.get('expected_value', 0) or 0)
                ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
            except ValueError:
                frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")
            from openpyxl.cell.cell import MergedCell
            # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
            for col in range(1, len(task_data) + 2):
                cell = ws.cell(row=row, column=col)
                if col in [14,15,16]:  # Left-align AM Remark, PM Remark, SPOC Remark
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col == 12:  # Expected Value column
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0'
                elif col == 2:  # Expected Value column
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else: 
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = black_border
            row += 1
            proj_s_no += 1
        ws.cell(row=customer_row_index, column=5, value=cust_vac)
        ws.cell(row=customer_row_index, column=6, value=cust_sp)
        ws.cell(row=customer_row_index, column=7, value=cust_fp)
        ws.cell(row=customer_row_index, column=8, value=cust_sl)
        ws.cell(row=customer_row_index, column=9, value=cust_lp)
        ws.cell(row=customer_row_index, column=10, value=cust_psl)
        ws.cell(row=customer_row_index, column=11, value=cust_psl_value)
        ws.cell(row=customer_row_index, column=12, value=cust_ev)
        ws.cell(row=customer_row_index, column=13, value=cust_ex_psl)

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,total_psl_value,
        ev_total, ex_psl_total, "","",""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for col_idx, cell in enumerate(last_row, start=1):
        if col_idx <= 16:
            cell.fill = total_fill
            cell.border = black_border
        # cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    from openpyxl.utils.cell import column_index_from_string

    g_col_idx = column_index_from_string("L")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        cell = row[g_col_idx - 1]
        if isinstance(cell, MergedCell):
            continue

        # Try to convert string values to float
        try:
            if isinstance(cell.value, str) and cell.value.strip().isdigit():
                cell.value = float(cell.value.strip())
            elif isinstance(cell.value, str):
                # Handle comma-separated numbers like '1,05,000'
                cleaned = cell.value.replace(',', '').strip()
                if cleaned.isdigit():
                    cell.value = float(cleaned)

            # Now apply formatting if it's a number
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        except Exception as e:
            pass

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@frappe.whitelist()
def int_to_roman(num):
    val = [
        1000, 900, 500, 400,
        100, 90, 50, 40,
        10, 9, 5, 4, 1
    ]
    syms = [
        "M", "CM", "D", "CD",
        "C", "XC", "L", "XL",
        "X", "IX", "V", "IV", "I"
    ]
    roman = ""
    i = 0
    while num > 0:
        for _ in range(num // val[i]):
            roman += syms[i]
            num -= val[i]
        i += 1
    return roman



@frappe.whitelist()
def download_PTSR_new():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "MR:03 – Project Task Status Report – REC (PTSR - R)" + posting_date
    build_xlsx_response_PTSR(filename)


@frappe.whitelist()
def build_xlsx_response_PTSR(filename):
    xlsx_file = make_xlsx_PTSR(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def make_xlsx_PTSR(sheet_name="MR:03 – Project Task Status Report – REC (PTSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()
    # ws = wb.create_sheet(sheet_name, 0)
    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15,15, 15, 25, 15, 15, 7, 7, 7, 7, 7, 7]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "REC : Project – Task Status Report : - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    black_font = Font(color="000000")
    header_font = Font(color="FFFFFF")
    white_font = Font(color="FFFFFF")
    headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark",'SPOC Remark', 'Exp Value', 'Exp PSL', 'Sourcing Status', 'Territory', 'TASK', 'Task Priority', '#VAC', '#SP', '#FP', '#SL', '#PSL', '#LP']
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    ws.append(headers)
    ws.freeze_panes = "A3"
    # yellow_columns = {"Completed Value", "Cr. Exp. Value", "Cr.Exp.PSL", "Exp.Week"}
    header_row = ws[ws.max_row]
    for cell in header_row:
        # if cell.value in yellow_columns:
        #     cell.fill = yellow_fill
        #     cell.font = black_font
        # else:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
    # for cell in header_row:
    #     cell.fill = header_fill
    #     cell.font = header_font
    #     cell.alignment = Alignment(horizontal="center", vertical="center")
    #     cell.border = black_border
    cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
    row = 3
    serial_number = 1
    s_no=1
    grand_totals = {'vac':0,'sp': 0,'fp': 0,'sl':0,'psl':0,'custom_lp':0,'exp_value':0.0,'exp_psl':0.0}
    proj_s_no = 1 
    for c in cust:
        priority = {"High": 1, "Medium": 2, "Low": 3}
        pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'],order_by= "priority ASC")
        if not pname:
            continue
        task_totals = {'exp_value':0.0,'vac':0,'sp':0,'fp':0,'sl':0,'psl':0,'custom_lp':0,'exp_psl':0.0}
        project_data = []      
        for p in pname:
            pdata = []
            print(p.project_name)        
            taskid = frappe.get_all("Task", {"status": ("in",('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'],order_by= "priority ASC")              
            # print(p['project_name'])
            # for tn in taskid:
                # print(tn.name)
            for t in taskid:
                pdata.append([p['project_name'] if p['project_name'] else "",p['priority'] if p['priority'] else "",p['remark'] if p['remark'] else "",p['account_manager_remark'] if p['account_manager_remark'] else "",p['custom_spoc_remark'] if p['custom_spoc_remark'] else "",p['expected_value'] if p['expected_value'] else "",p['expected_psl'] if p['expected_psl'] else "",p['sourcing_statu'] if p['sourcing_statu'] else "",p['territory'] if p['territory'] else "",t['subject'],t['priority'],t['vac'],t['sp'],t['fp'],t['sl'],t['psl'],t['custom_lp']])
                task_totals['vac'] +=t['vac']
                task_totals['sp'] +=t['sp']
                task_totals['fp']+= t['fp']
                task_totals['sl'] +=t['sl']
                task_totals['psl'] += t['psl']
                task_totals['custom_lp'] += t['custom_lp']
            task_totals['exp_value'] += float(p['expected_value']) if p['expected_value'] not in (None, '') else 0
            task_totals['exp_psl'] += float(p['expected_psl']) if p['expected_psl'] not in (None, '') else 0
            project_data.append({
                'project_name': p['project_name'],'priority': p['priority'],
                'remark': p['remark'],'account_manager_remark': p['account_manager_remark'],'custom_spoc_remark':p['custom_spoc_remark'],'sourcing_statu': p['sourcing_statu'],'territory': p['territory'],
                'expected_value': float(p['expected_value']) if p['expected_value'] not in (None, '') else 0,'expected_psl': p['expected_psl'],'tasks': pdata})
        s_no+=1
        blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
        # row_data = [serial_number, c['name']] + [""] * 14 + [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]
        row_data = [int_to_roman(serial_number), c['name']] + [""] * 4 +[task_totals['exp_value']]+[task_totals['exp_psl']]+[""] * 4+ [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]        
        ws.append(row_data)
        row_to_fill = ws.max_row
        for col, cell in enumerate(ws[row_to_fill], start=1):
            if col <= ws.max_column +1:
                cell.fill = blue_fill
            if col > 11:
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
            cell.border = black_border
        ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
        serial_number += 1
        row += 1
        current_row_start = row
        
        for project in project_data:
            project_row_start = row
            for task_data in project['tasks']:
                ws.append([proj_s_no] + task_data)
                for col in range(1, len(task_data) + 1):
                    cell = ws.cell(row=row, column=col)
                    if 2 <= col <= 11:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = black_border
                row += 1
            if project_row_start < row - 1:
                for col in range(2, 13):
                    ws.merge_cells(start_row=project_row_start, start_column=col, end_row=row-1, end_column=col)
                ws.merge_cells(start_row=project_row_start, start_column=1, end_row=row-1, end_column=1)
            proj_s_no += 1
        grand_totals['vac'] += task_totals['vac']
        grand_totals['sp'] += task_totals['sp']
        grand_totals['fp'] += task_totals['fp']
        grand_totals['sl'] += task_totals['sl']
        grand_totals['psl'] += task_totals['psl']
        grand_totals['custom_lp'] +=task_totals['custom_lp']
        grand_totals['exp_value'] += float(task_totals['exp_value']) if task_totals['exp_value'] not in (None, '') else 0
        grand_totals['exp_psl'] += int(task_totals['exp_psl']) if task_totals['exp_psl'] not in (None, '') else 0

    yellow_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    ws.append(['Total'] + [''] * 5 +[grand_totals['exp_value']]+ [grand_totals['exp_psl']]+ [''] * 4 +[grand_totals['vac'], grand_totals['sp'], grand_totals['fp'], grand_totals['sl'], grand_totals['psl'],grand_totals['custom_lp']])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
    from openpyxl.cell.cell import MergedCell

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = black_border
            if isinstance(cell, MergedCell):
                continue  # Skip merged cells
            col_letter = get_column_letter(cell.column)  # Safe way to get column letter
            if col_letter in ("H","A"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_letter in ("B","D","E","F","K"):
                cell.alignment = Alignment(horizontal="left", vertical="center",wrap_text=True)
            elif col_letter == "G":
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'  # Example: INR formatting
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
    from openpyxl.utils.cell import column_index_from_string
    g_col_idx = column_index_from_string("G")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        cell = row[g_col_idx - 1]
        if isinstance(cell, MergedCell):
            continue

        # Try to convert string values to float
        try:
            if isinstance(cell.value, str) and cell.value.strip().isdigit():
                cell.value = float(cell.value.strip())
            elif isinstance(cell.value, str):
                # Handle comma-separated numbers like '1,05,000'
                cleaned = cell.value.replace(',', '').strip()
                if cleaned.isdigit():
                    cell.value = float(cleaned)

            # Now apply formatting if it's a number
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        except Exception as e:
            pass

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


@frappe.whitelist()
def download_PSR_proj():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_proj(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR_proj(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10,10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL","PSL in Value",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

    # Initialize variables for Customer and Project Data
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = total_psl_value = 0
    pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
    pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    project_customer=frappe.db.get_value("Project",{"name":pro_name},["customer"])
    cust_territory=frappe.db.get_value("Customer",{"name":project_customer},["territory"])

    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{project_customer} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 10:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        project_psl_value = p.get('custom_psl_value') or 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl
        total_psl_value += project_psl_value

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,float(p.get('custom_psl_value',0) or 0),
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if cell.column in [16, 14, 15,2]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cell.column == 12:  # Expected Value column
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:  # Center-align for all other columns
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,total_psl_value,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.border = black_border
        if cell.column ==2:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif cell.column == 12:  # Expected Value column
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@frappe.whitelist()
def download_PSR_both():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_both(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_PSR_both(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10,10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL","PSL in Value",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border

    # Initialize variables for Customer and Project Data
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = total_psl_value = 0
    pro_name=frappe.db.get_single_value("Formatted Reports  Download",'project')
    project_customer=frappe.db.get_single_value("Formatted Reports  Download",'customer')
    pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "name":pro_name,"customer":project_customer, "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
    cust_territory=frappe.db.get_value("Customer",{"name":project_customer},["territory"])

    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{project_customer} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 11:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col == 2:  
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col == 12: 
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl
        total_psl_value += p.get('custom_psl_value') or 0

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,p.get('custom_psl_value') or 0,
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if col in [16, 14, 15]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif cell.column == 2:  
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif cell.column == 12: 
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else: 
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,total_psl_value,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
        if cell.column == 2:  
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif cell.column == 12: 
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@frappe.whitelist()
def download_PSR_customer():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PR:02 - Project Status Report - REC (PSR - R)_" + posting_date
    xlsx_file = make_xlsx_PSR_cust(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def make_xlsx_PSR_cust(sheet_name="PR:02 - Project Status Report - REC (PSR - R)", wb=None, column_widths=None):
    if wb is None:
        wb = openpyxl.Workbook()

    valid_sheet_name = sheet_name.replace(":", "-")
    ws = wb.create_sheet(valid_sheet_name, 0)
    args = frappe.local.form_dict

    default_column_widths = [8, 25, 10, 10, 5, 5, 5, 5, 5, 5, 10,10, 5, 43, 43, 43]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "PR:02 - Project Status Report - REC (PSR - R): - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(bold=True)

    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    headers = [
        "SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "SR Status", "VAC", "SP", "FP", "SL", "LP", "PSL","PSL in Value",
        "Ex Value", "Ex PSL", "AM Remark", "PM Remark", "SPOC Remark"
    ]
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
        if cell.column == 2:  
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Initialize variables for Customer and Project Data
    cust = frappe.get_all("Customer", fields=["name", "territory"])
    row = 3
    serial_number = 1
    ev_total = 0
    ex_psl_total = 0  # Initialize Ex PSL total
    total_vac = total_sp = total_fp = total_sl = total_lp = total_psl = total_psl_value = 0
    cust_name=frappe.db.get_single_value("Formatted Reports  Download",'customer')
    cust_territory=frappe.db.get_value("Customer",{"name":cust_name},["territory"])
    pname = frappe.get_all("Project", {
        "status": ("in", ['Open', 'Enquiry']),
        "customer":cust_name ,
        "service": ("in", ['REC-I', 'REC-D'])
    }, ['*'], order_by="priority ASC")


    blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
    customer_with_territory = f"{cust_name} - {cust_territory}"
    row_data = [serial_number, customer_with_territory] + [""] * 7
    ws.append(row_data)
    row_to_fill = ws.max_row

    ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
    ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)

    for col, cell in enumerate(ws[row_to_fill], start=1):
        cell.fill = blue_fill
        if col <= 11:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col ==2:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=4)
    serial_number += 1
    row += 1

    for p in pname:
        project_vac = project_sp = project_fp = project_sl = project_lp = project_psl = 0
        tasks = frappe.get_all("Task", {
            "status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')),
            "project": p.name
        }, ['subject', 'priority', 'vac', 'sp', 'fp', 'sl', 'psl', 'custom_lp'])

        for task in tasks:
            project_vac += task.get('vac', 0) or 0
            project_sp += task.get('sp', 0) or 0
            project_fp += task.get('fp', 0) or 0
            project_sl += task.get('sl', 0) or 0
            project_lp += task.get('custom_lp', 0) or 0
            project_psl += task.get('psl', 0) or 0

        total_vac += project_vac
        total_sp += project_sp
        total_fp += project_fp
        total_sl += project_sl
        total_lp += project_lp
        total_psl += project_psl
        total_psl_value += float(p.get('custom_psl_value',0) or 0)

        task_data = [
            p['project_name'], p['priority'], p.get('sourcing_statu', ''), project_vac, project_sp, project_fp, project_sl, project_lp, project_psl,float(p.get('custom_psl_value', 0) or 0),
            p.get('expected_value', 0), p.get('expected_psl', 0), p.get('remark', ''), p.get('account_manager_remark', ''),
            p.get('custom_spoc_remark', '')
        ]
        ws.append([""] + task_data)

        try:
            ev_total += float(p.get('expected_value', 0) or 0)
            ex_psl_total += float(p.get('expected_psl', 0) or 0)  # Summing Ex PSL total
        except ValueError:
            frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

        # Alignment Logic: Columns 13, 14, 15 left-aligned, others center-aligned
        for col in range(2, len(task_data) + 2):
            cell = ws.cell(row=row, column=col)
            if col in [16, 14, 15]:  # Left-align AM Remark, PM Remark, SPOC Remark
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col ==2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 12:  # Expected Value column
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:  # Center-align for all other columns
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = black_border
        row += 1

    # Total Row with Ex PSL Total
    total_row = [
        "", "Total", "", "", total_vac, total_sp, total_fp, total_sl, total_lp, total_psl,total_psl_value,
        ev_total, ex_psl_total, "", "", ""
    ]

    ws.append(total_row)
    last_row = ws[ws.max_row]
    total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    for cell in last_row:
        cell.fill = total_fill
        cell.font = Font(color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
        if cell.column == 12:  
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
        elif cell.column ==2:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Save to BytesIO and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

