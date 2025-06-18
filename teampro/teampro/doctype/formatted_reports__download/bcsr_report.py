import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_bcsr():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "MR:04 – Batch Check Status Report (BCSR)_" + posting_date
    xlsx_file = make_xlsx_bcsR(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

def make_xlsx_bcsR(sheet_name="MR:04 – Batch Check Status Report (BCSR)", wb=None, column_widths=None):
    from openpyxl.utils import get_column_letter

    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name.replace(":", "-")

    # Header row 1 and 2
    header_row_1 = [
        "CUSTOMER/BATCH NAME","Batch TAT","AM Remark","PM Remark","SPOC Remark","Expected Value","SO Created","Cr.Exp.Value","Exp.Week","Batch Status","Checks","Check TAT", "#Cases", "TAT", "", "", "", "Case Status", "", "", "", "", "", "", "", "", "", ""
    ]

    header_row_2 = [
        "", "","","","","","","","","","","","", "0-5", "6-10", "11-15", ">15",
        "Draft", "Entry Completed", "Entry QC", "Entry Insuff",
        "Execution", "Execution insuff", "Final QC", "Generate Report", "Case Completed", "To Be Billed"
    ]
        # Define colors and fonts
    yellow_columns = {"Cr.Exp.Value", "Exp.Week"}
    blue_columns = {"TAT","Case Status"}
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="007FBF", end_color="007FBF", fill_type="solid")
    header_fill = PatternFill(start_color="0F1568", end_color="0F1568", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000", bold=True)

    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Set headers and apply styling
    for col_num, (header1, header2) in enumerate(zip(header_row_1, header_row_2), 1):
        col_letter = get_column_letter(col_num)

        # Set header text
        ws[f"{col_letter}1"] = header1
        ws[f"{col_letter}2"] = header2

        for row in [1, 2]:
            cell = ws[f"{col_letter}{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = black_border
            if row == 1 and header1 in blue_columns:
                cell.fill = blue_fill
                cell.font = white_font
            elif header1 in yellow_columns or header2 in yellow_columns:
                cell.fill = yellow_fill
                cell.font = black_font
            else:
                cell.fill = header_fill
                cell.font = white_font

    # Merge cells for grouped headers (already correct)
    ws.merge_cells('A1:A2')  # CUSTOMER/BATCH NAME
    ws.merge_cells('B1:B2')  # Batch TAT
    ws.merge_cells('C1:C2')  # AM Remark
    ws.merge_cells('D1:D2')  # PM Remark
    ws.merge_cells('E1:E2')  # SPOC Remark
    ws.merge_cells('F1:F2')  # Expected Value
    ws.merge_cells('G1:G2')  # SO Created
    ws.merge_cells('H1:H2')  # Cr.Exp.Value
    ws.merge_cells('I1:I2')  # Exp.Week
    ws.merge_cells('J1:J2')  # Batch Status
    ws.merge_cells('K1:K2')  # Checks
    ws.merge_cells('L1:L2')  # Check TAT
    ws.merge_cells('M1:M2')  # #Cases
    ws.merge_cells('N1:Q1')  # TAT sub-columns
    ws.merge_cells('R1:AA1')  # Case Status sub-columns
    data = get_data()
    # for row in data:
    #     ws.append(row)
    start_row = 3  # Because headers are on rows 1 and 2
    for data_row_idx, row in enumerate(data, start=start_row):
        ws.append(row)
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.border = black_border
            if col_idx not in [1, 6, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    from openpyxl.utils import get_column_letter

    # Calculate totals for numeric columns
    total_row = ["Grand Total"]  # First cell
    numeric_columns = list(range(6, len(data[0]) + 1))  # Columns from 'Expected Value' onwards
    # Transpose data to calculate column-wise sums
    columns_data = list(zip(*data))  # Transpose rows to columns

    for col_idx in range(1, len(data[0])):  # Skip column A (index 0)
        try:
            total = sum(float(x) for x in columns_data[col_idx] if isinstance(x, (int, float)))
            total_row.append(total)
        except:
            total_row.append("")  # Leave blank for non-numeric columns

    # Append total row to sheet
    ws.append(total_row)

    # Style the total row (borders, center alignment except "Expected Value")
    total_row_idx = ws.max_row
    for col_idx in range(1, len(total_row) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.border = black_border
        if col_idx != 6:  # Column 6 is "Expected Value"
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

@frappe.whitelist()
def get_data():
    data = []
    batch=frappe.db.get_all("Batch",{"batch_status":("not in",["Completed"])},["*"])
    for i in batch:
        check_value=frappe.db.get_value("Check Package",i.check_package,"total_sp")
        check_tat=frappe.db.get_value("Check Package",i.check_package,"package_tat")
        cases = frappe.get_all("Case", filters={"batch": i.name}, fields=["actual_tat","case_status"])
        zero_to_five = len([case for case in cases if 0 <= case["actual_tat"] <= 5])
        six_to_ten = len([case for case in cases if 6 <= case["actual_tat"] <= 10])
        eleven_to_fifteen = len([case for case in cases if 11 <= case["actual_tat"] <= 15])
        more_than_fifteen = len([case for case in cases if case["actual_tat"] > 15])
        draft=len([case for case in cases if case["case_status"] == "Draft"])
        entry_qc=len([case for case in cases if case["case_status"] == "Entry-QC"])
        entry_comp=len([case for case in cases if case["case_status"] == "Entry Completed"])
        entry_insuff=len([case for case in cases if case["case_status"] == "Entry-Insuff"])
        execution=len([case for case in cases if case["case_status"] == "Execution"])
        execution_insuff=len([case for case in cases if case["case_status"] == "Execution-Insuff"])
        final_qc=len([case for case in cases if case["case_status"] == "Final-QC"])
        generate_rep=len([case for case in cases if case["case_status"] == "Generate Report"])
        case_comp=len([case for case in cases if case["case_status"] == "Case Completed"])
        to_be_billed=len([case for case in cases if case["case_status"] == "To be Billed"])
        if check_value:
            exp_value = (float(check_value)) * float(i.no_of_cases)
        else:
            exp_value=0
        data.append([
            f"{i.customer}/{i.name}",i.package_tat,"","","",exp_value,"","","",i.batch_status,
            i.no_of_checks,check_tat if check_tat else '',i.no_of_cases,zero_to_five if zero_to_five else ''
            ,six_to_ten if six_to_ten else '',eleven_to_fifteen if eleven_to_fifteen else '',more_than_fifteen if more_than_fifteen else '',
            draft if draft else '',entry_comp if entry_comp else '',entry_qc if entry_qc else '',
            entry_insuff if entry_insuff else '',execution if execution else '',execution_insuff if execution_insuff else '',
            final_qc if final_qc else '',generate_rep if generate_rep else '',case_comp if case_comp else '',to_be_billed if to_be_billed else ''
        ])

    return data


