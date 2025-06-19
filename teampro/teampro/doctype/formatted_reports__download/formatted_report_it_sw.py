from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import frappe
from datetime import datetime

@frappe.whitelist()
def download(customer=None, project=None):
    filename = "PR:02 – Project Status Report –(PSR - R)"
    build_xlsx_response(filename, customer, project)

def build_xlsx_response(filename, customer, project):
    xlsx_file = make_xlsx(customer, project)
    frappe.response['filename'] = f"{filename}.xlsx"
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(customer=None, project=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "IT-SW Project Status Report"

    headers = [
        "Project Name","Project Type", "Account Manager Remark", "Project Manager Remark", "SPOC Remark",
        "# Task", "# Open", "# Working","# CRD","# PR", "# CR","#Issue Open","#Issue Replied", "SO Value", "Pending Billing"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    sub_header_font = Font(bold=True, color="000000")
    
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark Blue
    sub_header_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")  # Sky Blue
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = [30,20, 40, 40, 40, 10, 10, 10, 10, 10,10,10,10, 15, 20]

    
    current_date = datetime.today().strftime("%d-%m-%Y")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"IT - SW Project Status Report (As on {current_date})")
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = alignment
    title_cell.border = thin_border

    
    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width  

    
    
    
    
    project_types =frappe.db.get_all('Project Type',{'sequence_number':['!=',0]},['name'],order_by='sequence_number')
    for project_type in project_types:
        filters = {"docstatus": ("!=", "2"), "service": "IT-SW"}
        filters_1 = {"status": ("not in", ["Hold", "Completed", "Cancelled"]), "service": "IT-SW",'project_type':project_type.name}
        if customer and not project:
            filters["customer"] = customer  
        elif project and not customer:
            filters_1["name"] = project 
        elif customer and project:
            filters["customer"] = customer
            filters["name"] = project  
        report = frappe.db.get_all(
            "Project",
            filters=filters_1,
            fields=["name", "project_name", "account_manager_remark", "remark", "custom_spoc_remark","status"],
            order_by='name'
        )

        row = 3  
        for i in report:
            task_count = frappe.db.count(
                "Task",
                filters={"project": i.name, "status": "Client Review", "service": "IT-SW"}
            )
            overall = frappe.db.count("Task", filters={"project":i.name, "status":("!=", "Cancelled"), "service":"IT-SW"})
            task_open=frappe.db.count("Task",filters={"project":i.name,"status":"Open","service":"IT-SW"})
            task_working=frappe.db.count("Task",filters={"project":i.name,"status":"Working","service":"IT-SW"})
            task_pr=frappe.db.count("Task",filters={"project":i.name,"status":"Pending Review","service":"IT-SW"})
            task_review=frappe.db.count("Task",filters={"project":i.name,"status":"Code Review","service":"IT-SW"})
            issue_open=frappe.db.count("Issue",filters={"project":i.name,"status":"Open"})
            issue_replied=frappe.db.count("Issue",filters={"project":i.name,"status":"Replied"})
            so_value = frappe.db.sql("""
                SELECT SUM(base_grand_total) 
                FROM `tabSales Order`
                WHERE project = %s AND docstatus != 2
            """, (i.name,))[0][0] or 0.0  

            sales_order = frappe.db.sql("""
                SELECT 
                    s.base_grand_total AS base_grand_total,
                    s.per_billed AS per_billed,
                    s.advance_paid AS advance_paid
                FROM 
                    `tabSales Order` s
                WHERE 
                    s.status NOT IN ('To Deliver', 'On Hold', 'Closed', 'Cancelled', 'Completed') 
                    AND s.project=%s
            """, (i.name,), as_dict=True)

            if sales_order:
                base_grand_total = sales_order[0].get("base_grand_total", 0.0)  
                per_billed = sales_order[0].get("per_billed", 0.0) or 0  
                advance_paid = sales_order[0].get("advance_paid", 0.0) or 0  

                amount_billed = (base_grand_total * per_billed) / 100  
                pending_billing = base_grand_total - (amount_billed + advance_paid)  
            else:
                base_grand_total = 0.0
                pending_billing = 0.0

            ws.append([
                i.project_name,project_type.name, i.account_manager_remark, i.remark, i.custom_spoc_remark,
                overall, task_open, task_working,task_review,task_pr, task_count, issue_open,issue_replied,
                f"₹ {so_value:,.0f}", f"₹ {pending_billing:,.0f}"
            ])

        
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                
                cell.border = thin_border
                if col in (1, 2, 3, 4,5):
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif col in (14, 15):
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                else:
                    cell.alignment = alignment

            row += 1

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

