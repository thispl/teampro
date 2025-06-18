import frappe
import pandas as pd
from frappe.utils import get_site_path, now_datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

@frappe.whitelist()
def download_acc_manager_individual():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    # Fetch Filters
    user_list = args.get("acc_manager")
    fiscal_year = args.get("fiscal_year")

    if isinstance(user_list, str):
        user_list = [user_list]  
    if not user_list:
        return 0  

    # Get Employee Name using user_id
    employee = frappe.db.get_value("Employee", {"user_id": ["in", user_list]}, ["name", "employee_name"], as_dict=True)
    if not employee:
        return {"error": "Employee not found"}

    employee_name = employee["employee_name"]
    employee_id = employee["name"]

    # Get Target Manager record for the employee
    target_manager = frappe.db.get_value("Target Manager", {"employee": employee_id}, "name")
    target_based = frappe.db.get_value("Target Manager", {"employee": employee_id},["target_based_unit"])
    if not target_manager:
        return {"error": "No target data found for this employee"}
    # Fetch Monthly Data from Target Child
    monthly_data = frappe.db.sql("""
        SELECT 
            month, 
            achieved AS so_achieved,
            sr AS sr,
            ct_yta AS ct_yta,
            ct AS ct
        FROM `tabTarget Child`
        WHERE parent = %(target_manager)s
        GROUP BY month
        ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
    """, {"target_manager": target_manager}, as_dict=True)

    # Ensure all months are covered
    all_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    month_data_dict = {data["month"]: data for data in monthly_data}
    
    # Create data structure for DataFrame
    data_rows = []

    for month in all_months:
        data_rows.append({
            "Month": month,
            "SO Target": month_data_dict.get(month, {}).get("ct", 0) if target_based=="Sales Order" else '',
            "SO Achieved": month_data_dict.get(month, {}).get("so_achieved", 0) if target_based=="Sales Order" else '',
            "SO Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if target_based=="Sales Order" else '',
            "SO Yet to Achieve":month_data_dict.get(month, {}).get("ct_yta", 0) if target_based=="Sales Order" else '',
            "SI Target": month_data_dict.get(month, {}).get("ct", 0) if target_based=="Sales Invoice" else '',
            "SI Achieved": month_data_dict.get(month, {}).get("so_achieved", 0) if target_based=="Sales Invoice" else '',
            "SI Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if target_based=="Sales Invoice" else '',
            "SI Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if target_based=="Sales Invoice" else ''
        })

    # Convert to DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel (Start from row 5 to leave space for headers)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Account Manager", index=False, startrow=4)
        writer._save()

    # Load and format the Excel file
    workbook = load_workbook(file_path)
    ws = workbook.active

    # Insert Employee Name, Fiscal Year, and Target Based On
    ws["A1"] = "Employee Name:"
    ws["B1"] = employee_name
    ws["A2"] = "Fiscal year:"
    ws["B2"] = fiscal_year
    ws["A3"] = "Target Based On:"
    ws["B3"] = target_based

    # Apply Styling
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Format Header Row (Row 5)
    for cell in ws[1]:  
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    # Format Header Row (Row 5)
    for cell in ws[5]:  
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Apply Borders to Data Cells
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Adjust Column Widths
    column_widths = {
        "A": 10, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15, "I": 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save & Close Workbook
    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

import frappe
import pandas as pd
from frappe.utils import get_site_path
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def get_employee_targets(fiscal_year):
    filters = {"custom_fiscal_year": fiscal_year}
    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])
    
    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue
       

        monthly_data = frappe.db.sql(
            """
            SELECT month, achieved AS achieved,
            sr AS sr,
            ct_yta AS ct_yta,
            ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s
            GROUP BY month
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
            """,
            {"target_manager": tm.name}, as_dict=True)
        
        all_months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        month_data_dict = {data["month"]: data for data in monthly_data}
        
        data_rows = []
        for month in all_months:
            data_rows.append({
                "Month": month,
                "SO Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
                "SI Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None,
            })
        
        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })
    
    return employee_data

# @frappe.whitelist()
# def download_acc_manager():
#     args = frappe.local.form_dict
#     fiscal_year = args.get("fiscal_year")
#     employee_targets = get_employee_targets(fiscal_year)
#     if not employee_targets:
#         return {"error": "No target data found"}
    
#     file_name = "Target_Status_Report.xlsx"
#     file_path = get_site_path("private", "files", secure_filename(file_name))
    
#     with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
#         row_offset = 0
#         for emp in employee_targets:
#             ws = writer.book.create_sheet(title=emp["employee_name"])
#             ws["A1"] = "Employee Name:"
#             ws["B1"] = emp["employee_name"]
#             ws["A2"] = "Fiscal year:"
#             ws["B2"] = fiscal_year
#             ws["A3"] = "Target Based On:"
#             ws["B3"] = emp["target_based"]
            
#             df = pd.DataFrame(emp["data"])
#             df.to_excel(writer, sheet_name=emp["employee_name"], index=False, startrow=4)
#             row_offset += len(emp["data"]) + 6  # Adjust for next employee
    
#     workbook = load_workbook(file_path)
#     for sheet in workbook.sheetnames:
#         ws = workbook[sheet]
        
#         header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
#         header_font = Font(color="FFFFFF", bold=True)
#         thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
#         for cell in ws[5]:
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.border = thin_border
        
#         for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#             for cell in row:
#                 cell.border = thin_border
        
#     workbook.save(file_path)
#     workbook.close()
    
#     with open(file_path, "rb") as f:
#         file_content = f.read()
    
#     return {
#         "filename": file_name,
#         "content": file_content
#     }
@frappe.whitelist()
def download_acc_manager():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    employee_targets = get_employee_targets(fiscal_year)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1  # Start from row 1
        
        for emp in employee_targets:
            # Add Employee Header
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            row_offset += 3  # Move down to write table headers
            
            # Column Headers for each employee
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  # Move to data rows
            
            # Append Employee Data
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  # Move to next row
            
            # Add an empty row after each employee
            ws.append([])
            row_offset += 1

    # Apply styles
    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Format all headers
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }

# For All data based on quarter and fiscal year

def get_employee_targets_quarter(fiscal_year, quarter=None):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }

    filters = {"custom_fiscal_year": fiscal_year}
    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])
    
    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue
        
        # Fetch only relevant months if quarter is provided
        query = """
            SELECT month, achieved AS achieved,sr AS sr,ct_yta AS ct_yta,ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s
        """
        if quarter:
            query += " AND month IN %(months)s"
            quarter_months = quarters.get(quarter, [])
        else:
            quarter_months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']

        query += " GROUP BY month ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')"
        
        monthly_data = frappe.db.sql(query, {"target_manager": tm.name, "months": quarter_months}, as_dict=True)
        
        month_data_dict = {data["month"]: data for data in monthly_data}
        
        data_rows = []
        for month in quarter_months:
            data_rows.append({
                "Month": month,
                "SO Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
                "SI Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
            })
        
        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })
    
    return employee_data

@frappe.whitelist()
def download_acc_manager_quarter():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    quarter = args.get("quarter")  # Fetch the selected quarter

    employee_targets = get_employee_targets_quarter(fiscal_year, quarter)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            # Employee header
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            ws.append(["Quarter:", quarter if quarter else "Full Year"])
            row_offset += 4  # Adjust for table headers
            
            # Headers
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  # Move to data rows
            
            # Employee data
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  # Move to next row
            
            # Empty row separator
            ws.append([])
            row_offset += 1

    # Apply formatting
    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }

def get_employee_targets_acc(fiscal_year, quarter=None, account_manager=None):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }

    # Base filters for Target Manager
    filters = {"custom_fiscal_year": fiscal_year}
    if account_manager:
        filters["user_id"] = account_manager  # Match Account Manager by user_id

    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])
    
    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue

        # Get the relevant months for the selected quarter
        if quarter and quarter in quarters:
            quarter_months = quarters[quarter]
            month_filter = f"AND month IN {tuple(quarter_months)}"
        else:
            quarter_months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
            month_filter = ""
        # Fetch only the required months
        query = f"""
            SELECT month, achieved AS achieved,
            sr AS sr,
            ct_yta AS ct_yta,
            ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s {month_filter}
            GROUP BY month 
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
        """

        monthly_data = frappe.db.sql(query, {"target_manager": tm.name}, as_dict=True)
        
        month_data_dict = {data["month"]: data for data in monthly_data}

        data_rows = []
        for month in quarter_months:
            data_rows.append({
                "Month": month,
                "SO Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
                "SI Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
            })

        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })
    
    return employee_data

import frappe
import pandas as pd
from frappe.utils import get_site_path
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

@frappe.whitelist()
def download_acc():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    quarter = args.get("quarter")  # Fetch the selected quarter
    account_manager = args.get("acc_manager")  # Fetch the selected Account Manager

    employee_targets = get_employee_targets_acc(fiscal_year, quarter, account_manager)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            # Employee header
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            ws.append(["Quarter:", quarter if quarter else "Full Year"])
            row_offset += 4  # Adjust for table headers
            
            # Headers
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  # Move to data rows
            
            # Employee data
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  # Move to next row
            
            # Empty row separator
            ws.append([])
            row_offset += 1

    # Apply formatting
    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:", "Quarter:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }

# Based on acc,fiscal,quarter,month
def get_employee_targets_acc_month(fiscal_year, quarter=None, month=None, account_manager=None):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }

    # Base filters for Target Manager
    filters = {"custom_fiscal_year": fiscal_year}
    if account_manager:
        filters["user_id"] = account_manager  # Ensure we only fetch data for this account manager

    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])

    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue

        # Determine months to include
        if month:  
            selected_months = [month]  # Use only the selected month
        elif quarter and quarter in quarters:
            selected_months = quarters[quarter]  # Use the quarter months
        else:
            selected_months = sum(quarters.values(), [])  # Use all months

        # Ensure month filter is correctly formatted for SQL
        selected_months_tuple = tuple(selected_months)
        if len(selected_months_tuple) == 1:
            month_filter = f"AND month = '{selected_months_tuple[0]}'"
        else:
            month_filter = f"AND month IN {selected_months_tuple}"
        # Fetch only the required months
        query = f"""
            SELECT month,achieved AS achieved,
            sr AS sr,
            ct_yta AS ct_yta,
            ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s {month_filter}
            GROUP BY month 
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
        """

        monthly_data = frappe.db.sql(query, {"target_manager": tm.name}, as_dict=True)
        month_data_dict = {data["month"]: data for data in monthly_data}

        data_rows = []
        for month in selected_months:
            data_rows.append({
                "Month": month,
                "SO Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
                "SI Target": month_data_dict.get(month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Achieved": month_data_dict.get(month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Strike Rate": month_data_dict.get(month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Yet to Achieve": month_data_dict.get(month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
            })

        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })

    return employee_data

@frappe.whitelist()
def download_acc_month():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    quarter = args.get("quarter")  # Fetch the selected quarter
    month = args.get("month")  # Fetch the selected month
    account_manager = args.get("acc_manager")  # Fetch the selected Account Manager

    employee_targets = get_employee_targets_acc_month(fiscal_year, quarter,month, account_manager)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            # Employee header
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            # ws.append(["Account Manager:", account_manager if account_manager else "All"])
            ws.append(["Target Based On:", emp["target_based"]])
            ws.append(["Quarter:", quarter if quarter else "Full Year"])
            row_offset += 4  # Adjust for table headers
            
            # Headers
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  # Move to data rows
            
            # Employee data
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  # Move to next row
            
            # Empty row separator
            ws.append([])
            row_offset += 1

    # Apply formatting
    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }
# Without acc,with fiscalyear,quarter,month
import frappe
import pandas as pd
from frappe.utils import get_site_path
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

def get_employee_targets_quarter_month(fiscal_year, quarter=None, month=None):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }

    filters = {"custom_fiscal_year": fiscal_year}
    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])
    
    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue
        
        if month:  
            selected_months = [month]  
        elif quarter and quarter in quarters:
            selected_months = quarters[quarter]  
        else:
            selected_months = sum(quarters.values(), [])  

        if len(selected_months) == 1:
            month_filter = f"AND month = '{selected_months[0]}'"
        else:
            month_filter = f"AND month IN {tuple(selected_months)}"
        query = f"""
            SELECT month, achieved AS achieved,
            sr AS sr,
            ct_yta AS ct_yta,
            ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s {month_filter}
            GROUP BY month 
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
        """

        monthly_data = frappe.db.sql(query, {"target_manager": tm.name}, as_dict=True)
        month_data_dict = {data["month"]: data for data in monthly_data}

        data_rows = []
        for m in selected_months:
            data_rows.append({
                "Month": m,
                "SO Target": month_data_dict.get(m, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Achieved": month_data_dict.get(m, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Strike Rate": month_data_dict.get(m, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
                "SO Yet to Achieve": month_data_dict.get(m, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
                "SI Target": month_data_dict.get(m, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Achieved": month_data_dict.get(m, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Strike Rate": month_data_dict.get(m, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
                "SI Yet to Achieve": month_data_dict.get(m, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
            })

        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })

    return employee_data

@frappe.whitelist()
def download_quarter_month_report():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    quarter = args.get("quarter")  
    month = args.get("month")  

    employee_targets = get_employee_targets_quarter_month(fiscal_year, quarter, month)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            ws.append(["Quarter:", quarter if quarter else "Full Year"])
            row_offset += 4
            
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  
            
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  
            
            ws.append([])
            row_offset += 1

    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }

import frappe
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from werkzeug.utils import secure_filename

def get_employee_targets_by_date(fiscal_year, account_manager=None, date=None):
    if not date:
        return []

    # Extract Month and Year from Date
    extracted_month = frappe.utils.getdate(date).strftime("%b")  # Convert to 'Apr', 'May', etc.
    extracted_year = frappe.utils.getdate(date).year  # Extract year

    filters = {"custom_fiscal_year": fiscal_year}
    if account_manager:
        filters["account_manager"] = account_manager  # Filter by account manager

    target_managers = frappe.get_all("Target Manager", filters=filters, fields=["name", "employee", "target_based_unit"])

    employee_data = []
    for tm in target_managers:
        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue

        # SQL Filtering by Extracted Month and Year
        query = f"""
            SELECT month, achieved AS achieved, sr AS sr, ct_yta AS ct_yta, ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s 
            AND month = %(month)s 
            GROUP BY month
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
        """

        monthly_data = frappe.db.sql(query, {"target_manager": tm.name, "month": extracted_month, "year": extracted_year}, as_dict=True)
        month_data_dict = {data["month"]: data for data in monthly_data}

        data_rows = [{
            "Month": extracted_month,
            "SO Target": month_data_dict.get(extracted_month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Achieved": month_data_dict.get(extracted_month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Strike Rate": month_data_dict.get(extracted_month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Yet to Achieve": month_data_dict.get(extracted_month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
            "SI Target": month_data_dict.get(extracted_month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Achieved": month_data_dict.get(extracted_month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Strike Rate": month_data_dict.get(extracted_month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Yet to Achieve": month_data_dict.get(extracted_month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
        }]

        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })

    return employee_data

@frappe.whitelist()
def download_employee_targets_report():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    account_manager = args.get("account_manager")  # Get account manager filter
    date = args.get("date")  # Get date filter

    employee_targets = get_employee_targets_by_date(fiscal_year, account_manager, date)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Employee_Target_Report.xlsx"
    file_path = frappe.get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            row_offset += 3
            
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  
            
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  
            
            ws.append([])
            row_offset += 1

    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }

import frappe
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from werkzeug.utils import secure_filename

def get_employee_targets_by_date_acc(fiscal_year, acc_manager=None, date=None):
    if not date:
        return []

    # Extract Month and Year from Date
    extracted_month = frappe.utils.getdate(date).strftime("%b")  # Convert to 'Apr', 'May', etc.
    extracted_year = frappe.utils.getdate(date).year  # Extract year

    # Apply filters properly
    filters = {"custom_fiscal_year": fiscal_year}
    if acc_manager:
        filters["user_id"] = acc_manager  # Ensure correct filtering

    # Fetch only the target managers matching the fiscal year and account manager
    target_managers = frappe.get_all(
        "Target Manager",
        filters=filters,
        fields=["name", "employee", "target_based_unit", "user_id"]
    )

    employee_data = []
    for tm in target_managers:
        # Ensure we only process records matching the specified account manager
        if acc_manager and tm.user_id != acc_manager:
            continue

        employee = frappe.get_value("Employee", tm.employee, ["name", "employee_name"], as_dict=True)
        if not employee:
            continue

        # SQL Filtering by Extracted Month and Year
        query = f"""
            SELECT month, achieved AS achieved, sr AS sr, ct_yta AS ct_yta, ct AS ct
            FROM `tabTarget Child`
            WHERE parent = %(target_manager)s 
            AND month = %(month)s 
            GROUP BY month
            ORDER BY FIELD(month, 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar')
        """

        monthly_data = frappe.db.sql(query, {"target_manager": tm.name, "month": extracted_month}, as_dict=True)
        month_data_dict = {data["month"]: data for data in monthly_data}

        data_rows = [{
            "Month": extracted_month,
            "SO Target": month_data_dict.get(extracted_month, {}).get("ct", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Achieved": month_data_dict.get(extracted_month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Strike Rate": month_data_dict.get(extracted_month, {}).get("sr", 0) if tm.target_based_unit == "Sales Order" else None,
            "SO Yet to Achieve": month_data_dict.get(extracted_month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Order" else None,
            "SI Target": month_data_dict.get(extracted_month, {}).get("ct", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Achieved": month_data_dict.get(extracted_month, {}).get("achieved", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Strike Rate": month_data_dict.get(extracted_month, {}).get("sr", 0) if tm.target_based_unit == "Sales Invoice" else None,
            "SI Yet to Achieve": month_data_dict.get(extracted_month, {}).get("ct_yta", 0) if tm.target_based_unit == "Sales Invoice" else None
        }]

        employee_data.append({
            "employee_name": employee["employee_name"],
            "target_based": tm.target_based_unit,
            "data": data_rows
        })

    return employee_data

@frappe.whitelist()
def download_employee_targets_report_acc():
    args = frappe.local.form_dict
    fiscal_year = args.get("fiscal_year")
    acc_manager = args.get("acc_manager")  # Get account manager filter
    date = args.get("date")  # Get date filter

    employee_targets = get_employee_targets_by_date_acc(fiscal_year, acc_manager, date)
    if not employee_targets:
        return {"error": "No target data found"}
    
    file_name = "Employee_Target_Report.xlsx"
    file_path = frappe.get_site_path("private", "files", secure_filename(file_name))
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        ws = writer.book.create_sheet(title="Target Report")

        row_offset = 1
        
        for emp in employee_targets:
            ws.append(["Employee Name:", emp["employee_name"]])
            ws.append(["Fiscal Year:", fiscal_year])
            ws.append(["Target Based On:", emp["target_based"]])
            row_offset += 3
            
            headers = ["Month","SO Target", "SO Achieved", 
                       "SO Strike Rate", "SO Yet to Achieve",
                       "SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"]
            ws.append(headers)
            row_offset += 1  
            
            for data in emp["data"]:
                row = [
                    data["Month"],
                    data["SO Target"],
                    data["SO Achieved"],
                    data["SO Strike Rate"],
                    data["SO Yet to Achieve"],
                    data["SI Target"],
                    data["SI Achieved"],
                    data["SI Strike Rate"],
                    data["SI Yet to Achieve"],
                ]
                ws.append(row)
                row_offset += 1  
            
            ws.append([])
            row_offset += 1

    workbook = load_workbook(file_path)
    ws = workbook["Target Report"]
    
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    for row in ws.iter_rows():
        if row[0].value in ["Month", "Employee Name:"]:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    workbook.save(file_path)
    workbook.close()
    
    with open(file_path, "rb") as f:
        file_content = f.read()
    
    return {
        "filename": file_name,
        "content": file_content
    }
