    
import frappe
import pandas as pd
from frappe.utils import get_site_path, now_datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

@frappe.whitelist()
def download_excel():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))

    # DataFrame initialization
    df1 = pd.DataFrame(columns=["Service", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"])
    df2 = pd.DataFrame(columns=["Account Manager","Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"])
    df3 = pd.DataFrame(columns=["Month", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve"])

    # Fetch the achieved value
    user_list = args.get("acc_manager")
    if isinstance(user_list, str):
        user_list = [user_list]  
    if not user_list:
        return 0  
    achieved_values = calculate_achieve(args)
    achieved_values_si = calculate_achieve_si(args)
    for manager, data in achieved_values.items():
        df2 = pd.concat([df2, pd.DataFrame([{
            "Account Manager": data["Short Code"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":achieved_values_si.get(manager, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
           
        }])], ignore_index=True)
    # Fetch service values
    service_values = calculate_service(args)
    service_values_si=calculate_service_si(args)
    # Populate df1 with service data
    for service, data in service_values.items():
        df1 = pd.concat([df1, pd.DataFrame([{
            "Service": data["Service"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":service_values_si.get(service, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
           
        }])], ignore_index=True)

    monthly_values = calculate_monthly_data(args)
    monthly_values_si=calculate_monthly_data_si(args)
    # Clear and populate df3 with the filtered month data
    df3 = pd.DataFrame(columns=["Month", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    for month, data in monthly_values.items():
        df3 = pd.concat([df3, pd.DataFrame([{
            "Month": data["Month"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":monthly_values_si.get(month, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
        }])], ignore_index=True)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Service Data", index=False)
        df2.to_excel(writer, sheet_name="Account Manager Data", index=False)
        df3.to_excel(writer, sheet_name="Monthly Data", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 10, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border  
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

@frappe.whitelist()
def calculate_achieve(args):
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    # month = map_months.get(args.get("month"), "01")  # Default to January if missing
    current_date = now_datetime()
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    
    # If no month is selected, use the current month
    month_number = map_months.get(selected_month, str(current_date.month).zfill(2))
    year = now_datetime().year

    user_list = args.get("acc_manager")
    if not user_list:
        user_list = []  # Ensure it is always a list
    elif isinstance(user_list, str):
        user_list = [user_list]
    if user_list:  # Filter by selected managers
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Order` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE so.account_manager IN ({}) 
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            GROUP BY so.account_manager, emp.short_code
        """.format(", ".join(["%s"] * len(user_list)))
        params = tuple(user_list) + (month_number, year)
    else:  # Fetch all managers if no filter is applied
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Order` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            GROUP BY so.account_manager, emp.short_code
        """
        params = (month_number, year)

    results = frappe.db.sql(query, params, as_dict=True)
    achieved_data = {}
    for res in results:
        account_manager = res["account_manager"]
        achieved_data[account_manager] = {
            "Achieved": res["total"] or 0,
            "Short Code": res["short_code"] or "N/A"
        }

    return achieved_data

@frappe.whitelist()
def calculate_achieve_si(args):
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    # month = map_months.get(args.get("month"), "01")  # Default to January if missing
    current_date = now_datetime()
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    
    # If no month is selected, use the current month
    month_number = map_months.get(selected_month, str(current_date.month).zfill(2))
    year = now_datetime().year

    user_list = args.get("acc_manager")
    if not user_list:
        user_list = []  # Ensure it is always a list
    elif isinstance(user_list, str):
        user_list = [user_list]
    if user_list:  # Filter by selected managers
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Invoice` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE so.account_manager IN ({}) 
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            GROUP BY so.account_manager, emp.short_code
        """.format(", ".join(["%s"] * len(user_list)))
        params = tuple(user_list) + (month_number, year)
    else:  # Fetch all managers if no filter is applied
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Invoice` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            GROUP BY so.account_manager, emp.short_code
        """
        params = (month_number, year)

    results = frappe.db.sql(query, params, as_dict=True)
    achieved_data = {}
    for res in results:
        account_manager = res["account_manager"]
        achieved_data[account_manager] = {
            "Achieved": res["total"] or 0,
            "Short Code": res["short_code"] or "N/A"
        }

    return achieved_data

@frappe.whitelist()
def calculate_service(args):
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    # month = map_months.get(args.get("month"), "01")  # Default to January if missing
    current_date = now_datetime()
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    
    # If no month is selected, use the current month
    month_number = map_months.get(selected_month, str(current_date.month).zfill(2))
    year = now_datetime().year
    target_service = args.get("target_service")

    service_data = {}  # ✅ Ensure proper initialization

    if target_service:  # If a service is selected, filter by it
        service_query = """
            SELECT 
                so.service, 
                SUM(so.base_total) AS total
            FROM `tabSales Order` so
            WHERE so.service = %s 
            AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.service
        """
        service_params = (target_service, month_number, year)
    else:  # If no service is selected, fetch all services
        service_query = """
            SELECT 
                so.service, 
                SUM(so.base_total) AS total
            FROM `tabSales Order` so
            WHERE so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.service
        """
        service_params = (month_number, year)

    service_results = frappe.db.sql(service_query, service_params, as_dict=True)

    for res in service_results:
        service = res["service"]
        service_data[service] = {
            "Service": res["service"] or "N/A",
            "Achieved": res["total"] or 0
        }

    return service_data

@frappe.whitelist()
def calculate_service_si(args):
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    # month = map_months.get(args.get("month"), "01")  # Default to January if missing
    current_date = now_datetime()
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    
    # If no month is selected, use the current month
    month_number = map_months.get(selected_month, str(current_date.month).zfill(2))
    year = now_datetime().year
    target_service = args.get("target_service")

    service_data = {}  # ✅ Ensure proper initialization

    if target_service:  # If a service is selected, filter by it
        service_query = """
            SELECT 
                so.services, 
                SUM(so.base_total) AS total
            FROM `tabSales Invoice` so
            WHERE so.services = %s 
            AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.services
        """
        service_params = (target_service, month_number, year)
    else:  # If no service is selected, fetch all services
        service_query = """
            SELECT 
                so.services, 
                SUM(so.base_total) AS total
            FROM `tabSales Invoice` so
            WHERE so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.services
        """
        service_params = (month_number, year)

    service_results = frappe.db.sql(service_query, service_params, as_dict=True)

    for res in service_results:
        service = res["services"]
        service_data[service] = {
            "Service": res["services"] or "N/A",
            "Achieved": res["total"] or 0
        }

    return service_data

@frappe.whitelist()
def calculate_monthly_data(args):
    """Fetch total achieved values grouped by month, with an optional filter for a specific month."""
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    year = now_datetime().year

    if selected_month:  # If a specific month is selected
        month_number = map_months.get(selected_month, "01")
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Order`
            WHERE status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND YEAR(creation) = %s 
            AND MONTH(creation) = %s
            GROUP BY month_name
        """
        params = (year, month_number)
    else:  # If no month is selected, fetch data for all months
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Order`
            WHERE status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND YEAR(creation) = %s 
            GROUP BY month_name
            ORDER BY STR_TO_DATE(month_name, '%%b')
        """
        params = (year,)

    results = frappe.db.sql(query, params, as_dict=True)

    monthly_data = {}
    for res in results:
        month_name = res["month_name"]
        monthly_data[month_name] = {
            "Month": month_name,
            "Achieved": res["total"] or 0
        }

    return monthly_data

@frappe.whitelist()
def calculate_monthly_data_si(args):
    """Fetch total achieved values grouped by month, with an optional filter for a specific month."""
    map_months = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 
        'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
    selected_month = args.get("month")  # User-selected month (e.g., "Jan")
    year = now_datetime().year

    if selected_month:  # If a specific month is selected
        month_number = map_months.get(selected_month, "01")
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Invoice`
            WHERE status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND YEAR(creation) = %s 
            AND MONTH(creation) = %s
            GROUP BY month_name
        """
        params = (year, month_number)
    else:  # If no month is selected, fetch data for all months
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Invoice`
            WHERE status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND YEAR(creation) = %s 
            GROUP BY month_name
            ORDER BY STR_TO_DATE(month_name, '%%b')
        """
        params = (year,)

    results = frappe.db.sql(query, params, as_dict=True)

    monthly_data = {}
    for res in results:
        month_name = res["month_name"]
        monthly_data[month_name] = {
            "Month": month_name,
            "Achieved": res["total"] or 0
        }

    return monthly_data

import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from frappe.utils import now_datetime

@frappe.whitelist()
def download_acc_manager():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    df2 = pd.DataFrame(columns=["Account Manager", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    # Fetch the achieved value
    user_list = args.get("acc_manager")
    if isinstance(user_list, str):
        user_list = [user_list]  
    if not user_list:
        return 0  
    achieved_values = calculate_achieve(args)
    achieved_values_si = calculate_achieve_si(args)
    for manager, data in achieved_values.items():
        df2 = pd.concat([df2, pd.DataFrame([{
            "Account Manager": data["Short Code"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":achieved_values_si.get(manager, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df2.to_excel(writer, sheet_name="Account Manager", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 10, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

@frappe.whitelist()
def download_service():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    df2 = pd.DataFrame(columns=["Service", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    # Fetch the achieved value
    user_list = args.get("acc_manager")
    if isinstance(user_list, str):
        user_list = [user_list]  
    if not user_list:
        return 0  
    service_values = calculate_service(args)
    service_values_si=calculate_service_si(args)
    # Populate df1 with service data
    for service, data in service_values.items():
        df2 = pd.concat([df2, pd.DataFrame([{
            "Service": data["Service"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":service_values_si.get(service, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df2.to_excel(writer, sheet_name="Service", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 10, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

@frappe.whitelist()
def download_month():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    # Fetch the achieved value
    user_list = args.get("acc_manager")
    if isinstance(user_list, str):
        user_list = [user_list]  
    if not user_list:
        return 0  
    monthly_values = calculate_monthly_data(args)
    monthly_values_si=calculate_monthly_data_si(args)
    # Clear and populate df3 with the filtered month data
    df3 = pd.DataFrame(columns=["Month", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    for month, data in monthly_values.items():
        df3 = pd.concat([df3, pd.DataFrame([{
            "Month": data["Month"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":monthly_values_si.get(month, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df3.to_excel(writer, sheet_name="Service", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 10, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

@frappe.whitelist()
def download_excel_filter():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    df1 = pd.DataFrame(columns=["Service", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])    
    df2 = pd.DataFrame(columns=["Account Manager", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    df3 = pd.DataFrame(columns=["Month", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    achieved_values = calculate_achieve_filter_so_acc(args)
    achieved_values_si = calculate_achieve_filter_si_acc(args)
    for manager, data in achieved_values.items():
        df2 = pd.concat([df2, pd.DataFrame([{
            "Account Manager": data["Short Code"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":achieved_values_si.get(manager, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    service_values = calculate_service_filter_so(args)
    service_values_si=calculate_service_filter_si(args)
    # Populate df1 with service data
    for service, data in service_values.items():
        df1 = pd.concat([df1, pd.DataFrame([{
            "Service": data["Service"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":service_values_si.get(service, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    monthly_values = calculate_monthly_data_filter_so(args)
    monthly_values_si=calculate_monthly_data_filter_si(args)
    # Clear and populate df3 with the filtered month data
    for month, data in monthly_values.items():
        df3 = pd.concat([df3, pd.DataFrame([{
            "Month": data["Month"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":monthly_values_si.get(month, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Service", index=False)
        df2.to_excel(writer, sheet_name="Account Manager", index=False)
        df3.to_excel(writer, sheet_name="Month", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 15, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }
import frappe
from datetime import datetime

@frappe.whitelist()
def calculate_achieve_filter_so_acc(args):
    frappe.log_error(title="Method", message="Inside Method")
    
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    
    # user_list = args.get("acc_manager")
    
    # if isinstance(user_list, str):
    #     user_list = [user_list]
    
    achieved_data = {}

    for month_num, year in month_years:
        # if user_list:  
        #     # If specific account managers are selected, filter by them
        #     query = """
        #         SELECT 
        #             so.account_manager, 
        #             SUM(so.base_total) AS total,
        #             emp.short_code
        #         FROM `tabSales Order` so
        #         LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
        #         WHERE so.account_manager IN ({}) 
        #         AND MONTH(so.creation) = %s 
        #         AND YEAR(so.creation) = %s 
        #         AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
        #         GROUP BY so.account_manager, emp.short_code
        #     """.format(", ".join(["%s"] * len(user_list)))
        #     params = tuple(user_list) + (month_num, year)
        # else:  
        #     # If no account manager is selected, fetch all account managers
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Order` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            GROUP BY so.account_manager, emp.short_code
        """
        params = (month_num, year)
    
        results = frappe.db.sql(query, params, as_dict=True)
        for res in results:
            account_manager = res["account_manager"]
            if account_manager not in achieved_data:
                achieved_data[account_manager] = {
                    "Achieved": 0,
                    "Short Code": res["short_code"] or "N/A"
                }
            achieved_data[account_manager]["Achieved"] += res["total"] or 0

    frappe.log_error(title="Filter Achieved Data", message=achieved_data)
    
    return achieved_data

@frappe.whitelist()
def calculate_achieve_filter_si_acc(args):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    
    # user_list = args.get("acc_manager")
    
    # if isinstance(user_list, str):
    #     user_list = [user_list]
    
    achieved_data = {}

    for month_num, year in month_years:
        # if user_list:  
        #     query = """
        #     SELECT 
        #         so.account_manager, 
        #         SUM(so.base_total) AS total,
        #         emp.short_code
        #     FROM `tabSales Invoice` so
        #     LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
        #     WHERE so.account_manager IN ({}) 
        #     AND MONTH(so.creation) = %s 
        #     AND YEAR(so.creation) = %s 
        #     AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
        #     GROUP BY so.account_manager, emp.short_code
        # """.format(", ".join(["%s"] * len(user_list)))
        #     params = tuple(user_list) + (month_num, year)
        # else:  # Fetch all managers if no filter is applied
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Invoice` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            GROUP BY so.account_manager, emp.short_code
        """
        params = (month_num, year)
        
        results = frappe.db.sql(query, params, as_dict=True)
        for res in results:
            account_manager = res["account_manager"]
            if account_manager not in achieved_data:
                achieved_data[account_manager] = {
                    "Achieved": 0,
                    "Short Code": res["short_code"] or "N/A"
                }
            achieved_data[account_manager]["Achieved"] += res["total"] or 0

    frappe.log_error(title="Filter Achieved Data", message=achieved_data)
    
    return achieved_data

@frappe.whitelist()
def calculate_service_filter_so(args):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    
    service_data = {}

    for month_num, year in month_years:
        service_query = """
            SELECT 
                so.service, 
                SUM(so.base_total) AS total
            FROM `tabSales Order` so
            WHERE so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.service
        """
        service_params = (month_num, year)

        service_results = frappe.db.sql(service_query, service_params, as_dict=True)

        for res in service_results:
            service = res["service"]
            if service not in service_data:
                service_data[service] = {
                    "Service": res["service"] or "N/A",
                    "Achieved": 0
                }
            service_data[service]["Achieved"] += res["total"] or 0
    frappe.log_error(title="service",message=service_data)
    return service_data

@frappe.whitelist()
def calculate_service_filter_si(args):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    
    service_data = {}

    for month_num, year in month_years:
        service_query = """
            SELECT 
                so.services, 
                SUM(so.base_total) AS total
            FROM `tabSales Invoice` so
            WHERE so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND MONTH(so.creation) = %s 
            AND YEAR(so.creation) = %s 
            GROUP BY so.services
        """
        service_params = (month_num, year)

        service_results = frappe.db.sql(service_query, service_params, as_dict=True)

        for res in service_results:
            service = res["services"]
            if service not in service_data:
                service_data[service] = {
                    "Service": res["services"] or "N/A",
                    "Achieved":0
                }
            service_data[service]["Achieved"] += res["total"] or 0
    frappe.log_error(title="service",message=service_data)

    return service_data

@frappe.whitelist()
def calculate_monthly_data_filter_so(args):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    monthly_data = {}
    for month_num, year in month_years:
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Order`
            WHERE status NOT IN ('Cancelled', 'Closed', 'On Hold')
            AND YEAR(creation) = %s 
            AND MONTH(creation) = %s 
            GROUP BY month_name
            ORDER BY STR_TO_DATE(month_name, '%%b')
        """
        params = (year,month_num)

        results = frappe.db.sql(query, params, as_dict=True)

        for res in results:
            month_name = res["month_name"]
            if month_name not in monthly_data:
                monthly_data[month_name] = {
                    "Month": month_name,
                    "Achieved":0
                }
            monthly_data[month_name]["Achieved"] += res["total"] or 0

    return monthly_data

@frappe.whitelist()
def calculate_monthly_data_filter_si(args):
    quarters = {
        'Q1': ['Apr', 'May', 'Jun'],
        'Q2': ['Jul', 'Aug', 'Sep'],
        'Q3': ['Oct', 'Nov', 'Dec'],
        'Q4': ['Jan', 'Feb', 'Mar']
    }
    
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3,
        'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9,
        'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    
    fiscal_year = args.get("fiscal_year")  
    quarter = args.get("quarter")  
    
    if not fiscal_year or not quarter:
        frappe.throw("Fiscal year and quarter are required.")
    
    start_year, end_year = map(int, fiscal_year.split('-'))
    selected_months = quarters.get(quarter)
    
    if not selected_months:
        frappe.throw("Invalid quarter. Please select from Q1, Q2, Q3, or Q4.")
    
    month_years = [(month_map[month], end_year if month in ['Jan', 'Feb', 'Mar'] else start_year) for month in selected_months]
    frappe.log_error(title="Method", message=month_years)
    monthly_data = {}
    for month_num, year in month_years:
        query = """
            SELECT 
                DATE_FORMAT(creation, '%%b') AS month_name, 
                SUM(base_total) AS total
            FROM `tabSales Invoice`
            WHERE status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
            AND YEAR(creation) = %s 
            AND MONTH(creation) = %s 
            GROUP BY month_name
            ORDER BY STR_TO_DATE(month_name, '%%b')
        """
        params = (year,month_num)

        results = frappe.db.sql(query, params, as_dict=True)

        for res in results:
            month_name = res["month_name"]
            if month_name not in monthly_data:
                monthly_data[month_name] = {
                    "Month": month_name,
                    "Achieved":0
                }
            monthly_data[month_name]["Achieved"] += res["total"] or 0

    return monthly_data

@frappe.whitelist()
def download_date_filter():
    args = frappe.local.form_dict
    file_name = "Target_Status_Report.xlsx"
    file_path = get_site_path("private", "files", secure_filename(file_name))
    df1 = pd.DataFrame(columns=["Service", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])    
    df2 = pd.DataFrame(columns=["Account Manager", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    df3 = pd.DataFrame(columns=["Month", "Sales Order","SO Target", "SO Achieved", "SO Strike Rate", "SO Yet to Achieve", "Sales Invoice","SI Target", "SI Achieved", "SI Strike Rate", "SI Yet to Achieve", "Collection", "Target", "Achieved", "Strike Rate", "Yet to Achieve"])
    achieved_values = calculate_achieve_date_so_acc(args)
    achieved_values_si = calculate_achieve_date_si_acc(args)
    for manager, data in achieved_values.items():
        df2 = pd.concat([df2, pd.DataFrame([{
            "Account Manager": data["Short Code"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":achieved_values_si.get(manager, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    service_values = calculate_service_date_so(args)
    service_values_si=calculate_service_date_si(args)
    # Populate df1 with service data
    for service, data in service_values.items():
        df1 = pd.concat([df1, pd.DataFrame([{
            "Service": data["Service"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":service_values_si.get(service, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    monthly_values = calculate_monthly_date_so(args)
    monthly_values_si=calculate_monthly_date_si(args)
    # Clear and populate df3 with the filtered month data
    for month, data in monthly_values.items():
        df3 = pd.concat([df3, pd.DataFrame([{
            "Month": data["Month"],
            "Sales Order": None,  
            "SO Target": None,
            "SO Achieved": data["Achieved"],
            "SO Strike Rate": None,
            "SO Yet to Achieve": None,
            "Sales Invoice": None,
            "SI Target": None,
            "SI Achieved":monthly_values_si.get(month, {}).get("Achieved", 0),
            "SI Strike Rate": None,
            "SI Yet to Achieve": None,
            "Collection": None,
            "Target": None,
            "Achieved":None,
            "Strike Rate": None,
            "Yet to Achieve": None
        }])], ignore_index=True)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Service", index=False)
        df2.to_excel(writer, sheet_name="Account Manager", index=False)
        df3.to_excel(writer, sheet_name="Month", index=False)
        writer._save()
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Load and format the Excel file
    workbook = load_workbook(file_path)
    header_fill = PatternFill(start_color="031273", end_color="031273", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    column_widths = {
        "A": 15, "B": 15, "C": 15, "D": 15,
        "E": 15, "F": 15, "G": 15, "H": 15,"I":15,"J":15,"K":15,"L":15,"M":15,"N":15,"O":15,"P":15
    }

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border  # Apply border to header

        # Apply border to all rows
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border  
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    workbook.save(file_path)
    workbook.close()

    with open(file_path, "rb") as f:
        file_content = f.read()

    return {
        "filename": file_name,
        "content": file_content
    }

@frappe.whitelist()
def calculate_achieve_date_so_acc(args):
    date_filter = args.get("date")  # Get the date filter if provided

    if date_filter:
        # If date is provided, filter only for that date
        query = """
            SELECT 
                so.account_manager, 
                SUM(so.base_total) AS total,
                emp.short_code
            FROM `tabSales Order` so
            LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
            WHERE DATE(so.creation) = %s
            AND so.status NOT IN ('Cancelled', 'Closed', 'On Hold')
            GROUP BY so.account_manager, emp.short_code
        """
        params = (date_filter,)
    achieved_data = {}
    results = frappe.db.sql(query, params, as_dict=True)
    for res in results:
        account_manager = res["account_manager"]
        if account_manager not in achieved_data:
            achieved_data[account_manager] = {
                "Achieved":res["total"] or 0,
                "Short Code": res["short_code"] or "N/A"
            }

    frappe.log_error(title="Filter Achieved Data", message=achieved_data)
    
    return achieved_data

@frappe.whitelist()
def calculate_achieve_date_si_acc(args):
    date_filter = args.get("date") 
    achieved_data = {}
    query = """
        SELECT 
            so.account_manager, 
            SUM(so.base_total) AS total,
            emp.short_code
        FROM `tabSales Invoice` so
        LEFT JOIN `tabEmployee` emp ON so.account_manager = emp.user_id
        WHERE DATE(so.creation) = %s 
        AND so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
        GROUP BY so.account_manager, emp.short_code
    """
    params = (date_filter,)
    
    results = frappe.db.sql(query, params, as_dict=True)
    for res in results:
        account_manager = res["account_manager"]
        if account_manager not in achieved_data:
            achieved_data[account_manager] = {
                "Achieved": res["total"],
                "Short Code": res["short_code"] or "N/A"
            }

    frappe.log_error(title="Filter Achieved Data", message=achieved_data)
    
    return achieved_data

@frappe.whitelist()
def calculate_service_date_so(args):
    date_filter = args.get("date") 
    service_data = {}

    service_query = """
        SELECT 
            so.service, 
            SUM(so.base_total) AS total
        FROM `tabSales Order` so
        WHERE DATE(so.creation) = %s 
        GROUP BY so.service
    """
    service_params = (date_filter,)

    service_results = frappe.db.sql(service_query, service_params, as_dict=True)

    for res in service_results:
        service = res["service"]
        if service not in service_data:
            service_data[service] = {
                "Service": res["service"] or "N/A",
                "Achieved":res["total"] or 0
            }
    frappe.log_error(title="service",message=service_data)
    return service_data

@frappe.whitelist()
def calculate_service_date_si(args):
    date_filter = args.get("date") 
    service_data = {}
    service_query = """
        SELECT 
            so.services, 
            SUM(so.base_total) AS total
        FROM `tabSales Invoice` so
        WHERE so.status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
        AND DATE(so.creation) = %s 
        GROUP BY so.services
    """
    service_params = (date_filter, )

    service_results = frappe.db.sql(service_query, service_params, as_dict=True)

    for res in service_results:
        service = res["services"]
        if service not in service_data:
            service_data[service] = {
                "Service": res["services"] or "N/A",
                "Achieved":res["total"] or 0
            }
    frappe.log_error(title="service",message=service_data)

    return service_data

@frappe.whitelist()
def calculate_monthly_date_so(args):
    date_filter = args.get("date") 
    monthly_data = {}
    query = """
        SELECT 
            DATE_FORMAT(creation, '%%b') AS month_name, 
            SUM(base_total) AS total
        FROM `tabSales Order`
        WHERE status NOT IN ('Cancelled', 'Closed', 'On Hold')
        AND DATE(creation) = %s 
        GROUP BY month_name
        ORDER BY STR_TO_DATE(month_name, '%%b')
    """
    params = (date_filter,)

    results = frappe.db.sql(query, params, as_dict=True)

    for res in results:
        month_name = res["month_name"]
        if month_name not in monthly_data:
            monthly_data[month_name] = {
                "Month": month_name,
                "Achieved":res["total"] or 0
            }
    return monthly_data

@frappe.whitelist()
def calculate_monthly_date_si(args):
    date_filter = args.get("date") 
    monthly_data = {}
    query = """
        SELECT 
            DATE_FORMAT(creation, '%%b') AS month_name, 
            SUM(base_total) AS total
        FROM `tabSales Invoice`
        WHERE status NOT IN ('Cancelled', 'Credit Note Issued', 'Return')
        AND DATE(creation) = %s 
        GROUP BY month_name
        ORDER BY STR_TO_DATE(month_name, '%%b')
    """
    params = (date_filter,)

    results = frappe.db.sql(query, params, as_dict=True)

    for res in results:
        month_name = res["month_name"]
        if month_name not in monthly_data:
            monthly_data[month_name] = {
                "Month": month_name,
                "Achieved":res["total"] or 0
            }
    return monthly_data