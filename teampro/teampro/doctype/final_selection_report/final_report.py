from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import frappe
from collections import defaultdict

@frappe.whitelist()
def download(project):
    filename = 'Interview Final Selection'
    build_xlsx_response(filename, project)

def build_xlsx_response(filename, project):
    xlsx_file = make_xlsx(project)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(project):
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Report"

    # Define headers
    headers = [
        "S.No", "Count", "CDID", "Given Name / Surname", "Nationality", "DOB", "Age", 
        "Passport No", "PP Validity", "Position", "Qualification", 
        "Specialization", "Local Exp.", "Gulf Exp.", "Total Exp.", 
        "Interview Date", "Interview Location", "TT Result", 
        "Grade", "Basic Salary", "IAF (Interview Application Form)", 
        "Offer Letter", "E Wakkala Received", "Biometric", 
        "Medical", "Visa Stamped", "Ready for Onboard", "Travel Date", "Remarks"
    ]

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    position_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Fetch candidate data
    candidates = frappe.db.get_all(
        'Candidate',
        filters={'pending_for':'Proposed PSL','project': project},
        fields=[
            'name', 'given_name', 'nationality', 'date_of_birth', 'age', 
            'passport_number', 'position', 'highest_degree', 'specialization', 
            'india_experience', 'overseas_experience', 'total_experience', 
            'interviewed_date', 'interview_location', 'grade', 
            'basic', 'irf', 'offer_letter','passport_expiry_date'
        ]
    )

    # Organize candidates by position
    position_wise_candidates = defaultdict(list)
    for candidate in candidates:
        position_wise_candidates[candidate['position']].append(candidate)

    # Write data to the worksheet, position-wise
    global_count = 1  # Initialize a global count for S.No
    row_start = 2  # Start writing data from the second row
    for position, candidates in position_wise_candidates.items():
        # Add position header with full merge
        position_row = ws.cell(row=row_start, column=1, value=position)
        position_row.font = Font(bold=True, size=12)
        position_row.alignment = left_alignment  # Left align the position header
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=len(headers))
        position_row.fill = position_fill
        row_start += 1  # Move to the next row for the data

        # Write candidates for the position
        for index, candidate in enumerate(candidates, start=1):  # Start index at 1 for position-wise count
            ws.append([
                global_count,  # S.No (global sequential count)
                index,  # Count (position-wise count)
                candidate.get('name'),  # CDID
                candidate.get('given_name'),  # Given Name / Surname
                candidate.get('nationality'),  # Nationality
                candidate.get('date_of_birth'),  # DOB
                candidate.get('age'),  # Age
                candidate.get('passport_number'),  # Passport No
                candidate.get('passport_expiry_date'),  # PP Validity (placeholder)
                candidate.get('position'),  # Position
                candidate.get('highest_degree'),  # Qualification
                candidate.get('specialization'),  # Specialization
                candidate.get('india_experience'),  # Local Exp.
                candidate.get('overseas_experience'),  # Gulf Exp.
                candidate.get('total_experience'),  # Total Exp.
                candidate.get('interviewed_date'),  # Interview Date
                candidate.get('interview_location'),  # Interview Location
                '',  # TT Result (placeholder)
                candidate.get('grade'),  # Grade
                candidate.get('basic'),  # Basic Salary
                '',  # IAF
                candidate.get('offer_letter'),  # Offer Letter
                '',  # E Wakkala Received (placeholder)
                '',  # Biometric (placeholder)
                '',  # Medical (placeholder)
                '',  # Visa Stamped (placeholder)
                '',  # Ready for Onboard (placeholder)
                '',  # Travel Date (placeholder)
                '',  # Remarks
            ])
            global_count += 1  # Increment the global count for S.No
        row_start += len(candidates) + 1  # Move down for the next position section

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add border for the entire table
    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=row_start-1):
        for cell in row:
            cell.border = thin_border

    # Save the workbook to a BytesIO object for file handling
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)  # Move the cursor to the beginning of the file

    return xlsx_file
