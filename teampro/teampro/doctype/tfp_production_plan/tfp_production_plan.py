# Copyright (c) 2025, TeamPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document


class TFPProductionPlan(Document):
    pass



@frappe.whitelist()
def tfp_production_plan_report():
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="26" style="text-align:center; font-weight:bold;">TFP - Packing Plan</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Sr</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">SO ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">PRT</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Customer Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Packing</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Delivery</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Item Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">QTY</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">UOM</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">St.QTY</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">UOM</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Cover type</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">MRP</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Mnfg. On</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;"># Covers</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">2P Type</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Per 2P</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;"># 2P</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Name Print</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Material Request</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Purchase Order</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Purchase Receipt</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Purchase Invoice</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Delivery Note</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Sales Invoice</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">Stock Issued</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border:1px solid black;">DN - Closing</td>' \
            '</tr>'

    so_list = frappe.db.get_all("Sales Order", {"service": "TFP", "status": "To Deliver and Bill"}, ["name", "customer", "custom_packing_on", "delivery_date"])
    total_qty = 0
    total_stock_qty = 0
    total_covers=0
    per_2p=0
    for so in so_list:
        items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, ["item_name", "qty", "uom"])
        rowspan = len(items)
        first_row = True
        for item in items:
            total_qty += item.qty or 0
            total_stock_qty += item.stock_qty or 0
            total_covers+=item.custom_covers or 0
            per_2p+=item.custom_2nd_packing or 0
            data += "<tr>"
            if first_row:
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}">{s_no}</td>'
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}">{so.name}</td>'
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}"></td>'  # PRT
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}">{so.customer}</td>'
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}">{so.custom_packing_on if so.custom_packing_on else ""}</td>'
                data += f'<td style="text-align:center;border:1px solid black;" rowspan="{rowspan}">{so.delivery_date}</td>'
                first_row = False
            data += f'<td style="text-align:center;border:1px solid black;">{item.item_name or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.qty or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.uom or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.stock_qty or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.stock_uom or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_cover_type or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.mrp or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_mfg_on or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_covers or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_packing_type or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_per_2p or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_2nd_packing or ""}</td>'
            data += f'<td style="text-align:center;border:1px solid black;">{item.custom_name_print or ""}</td>'
            data += '<td style="text-align:center;border:1px solid black;"></td>' * 8  
            data += "</tr>"
        data += '<tr style="background-color:#D3D3D3; font-weight:bold;border:1px solid black;">'
        data += '<td style="border:1px solid black;"></td>' * 7
        data += f'<td style="text-align:center;border:1px solid black;">{total_qty}</td>'
        data += '<td style="border:1px solid black;"></td>'  # UOM
        data += f'<td style="text-align:center;border:1px solid black;">{total_stock_qty}</td>'
        data += '<td style="border:1px solid black;"></td>' * 4  # Fill remaining columns
        data += f'<td style="text-align:center;border:1px solid black;">{total_covers}</td>'
        data += '<td style="border:1px solid black;"></td>' * 2  
        data += f'<td style="text-align:center;border:1px solid black;">{per_2p}</td>'
        data += '<td style="border:1px solid black;"></td>' * 9
        data += '</tr>'
        s_no += 1

    data += '</table>'
    return data



import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_tfp_production_plan_excel():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "TFP Packing Plan" + ("as on"+posting_date)
    xlsx_file = make_xlsx_tfp(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file
    frappe.response['type'] = 'binary'

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime
def make_xlsx_tfp(sheet_name="TFP Packing Plan", wb=None, column_widths=None):
    

    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = sheet_name.replace(":", "-")
    ws.title = "TFP Packing Plan"

    # Style Definitions
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    fill_header = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

    # Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=26)
    cell = ws.cell(row=1, column=1, value="TFP - Packing Plan")
    cell.font = Font(bold=True)
    cell.alignment = center_align

    headers = [
        "Sr", "SO ID", "PRT", "Customer Name", "Packing", "Delivery", "Item Name", "QTY", "UOM",
        "St.QTY", "UOM", "Cover type", "MRP", "Mnfg. On", "# Covers", "2P Type", "Per 2P", "# 2P",
        "Name Print", "Material Request", "Purchase Order", "Purchase Receipt",
        "Purchase Invoice", "Delivery Note", "Sales Invoice", "Stock Issued", "DN - Closing"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = fill_header
        cell.border = border

    # Data Rows
    row_idx = 3
    s_no = 1

    grand_total_qty = 0
    grand_total_stock_qty = 0
    grand_total_covers = 0
    grand_total_per_2p = 0

    so_list = frappe.db.get_all("Sales Order", {"service": "TFP", "status": "To Deliver and Bill"}, ["name", "customer", "custom_packing_on", "delivery_date"])
    for so in so_list:
        items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, ["item_name", "qty", "uom", "stock_qty", "stock_uom",
            "custom_cover_type", "mrp", "custom_mfg_on", "custom_covers", "custom_packing_type",
            "custom_per_2p", "custom_2nd_packing", "custom_name_print"
        ])
        rowspan = len(items)
        start_row = row_idx
        total_qty = total_stock_qty = total_covers = per_2p = 0

        for item in items:
            ws.cell(row=row_idx, column=7, value=item.item_name or "")
            ws.cell(row=row_idx, column=8, value=item.qty or "")
            ws.cell(row=row_idx, column=9, value=item.uom or "")
            ws.cell(row=row_idx, column=10, value=item.stock_qty or "")
            ws.cell(row=row_idx, column=11, value=item.stock_uom or "")
            ws.cell(row=row_idx, column=12, value=item.custom_cover_type or "")
            ws.cell(row=row_idx, column=13, value=item.mrp or "")
            formatted_mfg_on = item.custom_mfg_on.strftime("%d-%m-%Y") if item.custom_mfg_on else ""
            ws.cell(row=row_idx, column=14, value=formatted_mfg_on)
            ws.cell(row=row_idx, column=15, value=item.custom_covers or "")
            ws.cell(row=row_idx, column=16, value=item.custom_packing_type or "")
            ws.cell(row=row_idx, column=17, value=item.custom_per_2p or "")
            ws.cell(row=row_idx, column=18, value=item.custom_2nd_packing or "")
            ws.cell(row=row_idx, column=19, value=item.custom_name_print or "")

            total_qty += item.qty or 0
            total_stock_qty += item.stock_qty or 0
            total_covers += item.custom_covers or 0
            per_2p += item.custom_2nd_packing or 0

            # Accumulate grand totals
            grand_total_qty += item.qty or 0
            grand_total_stock_qty += item.stock_qty or 0
            grand_total_covers += item.custom_covers or 0
            grand_total_per_2p += item.custom_2nd_packing or 0

            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = border
            for col in range(7, 20):
                if col in (9, 11, 12, 16):  # UOM and description columns
                    ws.cell(row=row_idx, column=col).alignment = left_align
                elif col in (10, 13):
                    ws.cell(row=row_idx, column=col).alignment = right_align
                else:
                    ws.cell(row=row_idx, column=col).alignment = center_align
                ws.cell(row=row_idx, column=col).border = border
            for col in range(20, 28):
                ws.cell(row=row_idx, column=col).border = border
            row_idx += 1

        for i in range(start_row, row_idx):
            ws.cell(row=i, column=1, value=s_no)
            s_no += 1

        for col, val in zip(range(2, 7), [so.name, "", so.customer, so.custom_packing_on.strftime("%d-%m-%Y") if so.custom_packing_on else "", so.delivery_date.strftime("%d-%m-%Y") if so.delivery_date else ""]):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=row_idx - 1, end_column=col)
            cell = ws.cell(row=start_row, column=col, value=val)
            cell.alignment = left_align
            cell.border = border

        # Totals Row
        for col in range(1, 28):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = border

        ws.cell(row=row_idx, column=8, value=total_qty).alignment = center_align
        ws.cell(row=row_idx, column=10, value=total_stock_qty).alignment = right_align
        ws.cell(row=row_idx, column=15, value=total_covers).alignment = center_align
        ws.cell(row=row_idx, column=18, value=per_2p).alignment = center_align

        row_idx += 1
        s_no += 1

    # Grand Total Row
    for col in range(1, 28):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.border = border
        cell.alignment = center_align
        cell.font = Font(bold=True)

    ws.cell(row=row_idx, column=1, value="Grand Total").alignment = left_align
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)

    ws.cell(row=row_idx, column=8, value=grand_total_qty).alignment = center_align
    ws.cell(row=row_idx, column=10, value=grand_total_stock_qty).alignment = right_align
    ws.cell(row=row_idx, column=15, value=grand_total_covers).alignment = center_align
    ws.cell(row=row_idx, column=18, value=grand_total_per_2p).alignment = center_align

    row_idx += 1

    # Final formatting
    # for i in range(1, 27):
    #     ws.column_dimensions[get_column_letter(i)].width = 14
    column_widths = {
        1: 5,  # Sr
        2: 20, 3: 8, 4: 18, 5: 17, 6: 17,
        7: 18, 8: 8, 9: 6, 10: 10, 11: 6,
        12: 12, 13: 8, 14: 17, 15: 8, 16: 10,
        17: 8, 18: 8, 19: 10,
        20: 12, 21: 12, 22: 12, 23: 12, 24: 12, 25: 12, 26: 12, 27: 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.border is None or cell.border == Border():
                cell.border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# import frappe
# import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# from frappe.utils import formatdate, flt
# from io import BytesIO

# @frappe.whitelist()
# def download_tfp_plan_excel():
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "TFP Packing Plan"

#     headers = [
#         "Sr", "SO ID", "PRT", "Customer Name", "Packing", "Delivery", "Item Name", "QTY", "UOM",
#         "St.QTY", "UOM", "MRP", "Packing Details", "WRD Details",
#         "Name Print", "Material Request", "Purchase Order", "Purchase Receipt",
#         "Purchase Invoice", "Delivery Note", "Sales Invoice", "Stock Issued", "DN - Closing"
#     ]

#     # Styles
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill("solid", fgColor="002060")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
#     right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
#     thin_border = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )
#     total_fill = PatternFill("solid", fgColor="D9D9D9")  # Light ash gray

#     def style_cell(cell, align=center_align, font=None, fill=None):
#         cell.alignment = align
#         cell.border = thin_border
#         if font:
#             cell.font = font
#         if fill:
#             cell.fill = fill

#     def style_merged_range(ws, start_row, end_row, col, value, align=left_align, font=None, fill=None):
#         for r in range(start_row, end_row + 1):
#             cell = ws.cell(row=r, column=col, value=value if r == start_row else None)
#             style_cell(cell, align=align, font=font, fill=fill)
#         ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

#     # Add header row
#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num, value=header)
#         style_cell(cell, align=center_align, font=header_font, fill=header_fill)

#     row_idx = 2
#     s_no = 1
#     grand_total_qty = grand_total_stock_qty = 0

#     so_list = frappe.db.get_all("Sales Order", {
#         "service": "TFP",
#         "status": "To Deliver and Bill"
#     }, ["name", "customer", "custom_packing_on", "delivery_date"])

#     for so in so_list:
#         items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, [
#             "item_name", "qty", "uom", "stock_qty", "stock_uom", "custom_cover_type", "mrp",
#             "custom_mfg_on", "custom_covers", "custom_packing_type", "custom_per_2p",
#             "custom_2nd_packing", "custom_name_print", "custom_tertiary_packingbox",
#             "custom_bag", "custom_box", "custom_wrd_uom", "custom_wrd_rate"
#         ])
#         rowspan = len(items)
#         total_qty = total_stock_qty = 0
#         primary = ''
#         secondary = ''
#         tertiary = ''
#         packing_details = ''
#         for idx, item in enumerate(items):
#             row = row_idx

#             if idx == 0:
#                 merge_values = [
#                     s_no, so.name, '', so.customer,
#                     formatdate(so.custom_packing_on), formatdate(so.delivery_date)
#                 ]
#                 for i, val in enumerate(merge_values, start=1):
#                     align = left_align if i == 4 else center_align
#                     style_merged_range(ws, row, row + rowspan - 1, i, val, align=align)

#             # Item-specific details
#             ws.cell(row=row, column=7, value=item.item_name or "")
#             ws.cell(row=row, column=8, value=item.qty or 0)
#             ws.cell(row=row, column=9, value=item.uom or "")
#             ws.cell(row=row, column=10, value=item.stock_qty or 0)
#             ws.cell(row=row, column=11, value=item.stock_uom or "")
#             ws.cell(row=row, column=12, value=item.mrp or 0)
#             style_cell(ws.cell(row=row, column=12), align=right_align)

#             if item.custom_cover_type:
#                 primary = frappe.db.get_value("Item", item.custom_cover_type, "item_name") or ''
#             if item.custom_packing_type:
#                 secondary = frappe.db.get_value("Item", item.custom_packing_type, "item_name") or ''
#             if item.custom_tertiary_packingbox:
#                 tertiary = frappe.db.get_value("Item", item.custom_tertiary_packingbox, "item_name") or ''
#             packing_details = f"(C): {primary or 'None'}: {item.custom_covers or 0}\n(B): {secondary or 'None'}: {item.custom_bag or 0}\n(BX): {tertiary or 'None'}: {item.custom_box or 0}"
#             ws.cell(row=row, column=13, value=packing_details)

#             wrd_details = f"(W): {item.custom_wrd_uom or 'None'}\n(R): {item.custom_wrd_rate or '0'}\n(D): {formatdate(item.custom_mfg_on) if item.custom_mfg_on else ''}"
#             ws.cell(row=row, column=14, value=wrd_details)

#             ws.cell(row=row, column=15, value=item.custom_name_print or "")

#             for i in range(16, len(headers) + 1):
#                 ws.cell(row=row, column=i, value="")

#             # Apply styles
#             for col_idx in range(7, len(headers) + 1):
#                 align = left_align if col_idx in [7, 13, 14, 15] else center_align
#                 style_cell(ws.cell(row=row, column=col_idx), align=align)

#             total_qty += flt(item.qty)
#             total_stock_qty += flt(item.stock_qty)

#             row_idx += 1

#         # Subtotal row
#         ws.cell(row=row_idx, column=7, value="Total")
#         ws.cell(row=row_idx, column=8, value=total_qty)
#         ws.cell(row=row_idx, column=10, value=total_stock_qty)
#         for col_idx in range(1, len(headers) + 1):
#             style_cell(ws.cell(row=row_idx, column=col_idx), align=center_align, fill=total_fill)

#         row_idx += 1
#         s_no += 1
#         grand_total_qty += total_qty
#         grand_total_stock_qty += total_stock_qty

#     # Grand Total row
#     ws.cell(row=row_idx, column=1, value="Grand Total")
#     ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
#     for i in range(1, 8):
#         style_cell(ws.cell(row=row_idx, column=i), align=center_align, font=header_font, fill=header_fill)
#     ws.cell(row=row_idx, column=8, value=grand_total_qty)
#     ws.cell(row=row_idx, column=10, value=grand_total_stock_qty)
#     for col_idx in range(8, len(headers) + 1):
#         style_cell(ws.cell(row=row_idx, column=col_idx), align=center_align, font=header_font, fill=header_fill)

#     # Adjust column widths
#     for col in ws.columns:
#         max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
#         ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

#     # Save and return file
#     file_content = BytesIO()
#     wb.save(file_content)
#     frappe.response['filename'] = "TFP_Packing_Plan.xlsx"
#     frappe.response['filecontent'] = file_content.getvalue()
#     frappe.response['type'] = 'binary'


# import frappe
# import openpyxl
# from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
# from frappe.utils import formatdate, flt
# from io import BytesIO

# @frappe.whitelist()
# def download_tfp_plan_excel():
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "TFP Packing Plan"

#     headers = [
#         "Sr", "SO ID", "PRT", "Customer Name", "Packing", "Delivery", "Item Name", "QTY", "UOM",
#         "St.QTY", "Cr. Stock", "Stock Status", "UOM", "MRP", "Packing Details", "WRD Details",
#         "Name Print", "Material Request", "Purchase Order", "Purchase Receipt",
#         "Purchase Invoice", "Delivery Note", "Sales Invoice", "Stock Issued", "DN - Closing"
#     ]

#     # Styles
#     header_font = Font(bold=True, color="FFFFFF")
#     stock_header_fill = PatternFill("solid", fgColor="FF0000")  # Dark red
#     header_fill = PatternFill("solid", fgColor="002060")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
#     right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
#     thin_border = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin")
#     )
#     total_fill = PatternFill("solid", fgColor="D9D9D9")  # Light ash gray
#     in_stock_fill = PatternFill("solid", fgColor="92D050")  # Green
#     out_stock_fill = PatternFill("solid", fgColor="FF0000")  # Red
#     white_font = Font(color="FFFFFF", bold=True)

#     def style_cell(cell, align=center_align, font=None, fill=None):
#         cell.alignment = align
#         cell.border = thin_border
#         if font:
#             cell.font = font
#         if fill:
#             cell.fill = fill

#     def style_merged_range(ws, start_row, end_row, col, value, align=left_align, font=None, fill=None):
#         for r in range(start_row, end_row + 1):
#             cell = ws.cell(row=r, column=col, value=value if r == start_row else None)
#             style_cell(cell, align=align, font=font, fill=fill)
#         ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

#     # Header row
#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num, value=header)
#         if header in ["Cr. Stock", "Stock Status"]:
#             style_cell(cell, align=center_align, font=white_font, fill=stock_header_fill)
#         else:
#             style_cell(cell, align=center_align, font=header_font, fill=header_fill)

#     row_idx = 2
#     s_no = 1
#     grand_total_qty = grand_total_stock_qty = 0

#     so_list = frappe.db.get_all("Sales Order", {
#         "service": "TFP",
#         "status": "To Deliver and Bill"
#     }, ["name", "customer", "custom_packing_on", "delivery_date"])

#     for so in so_list:
#         items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, [
#             "item_code", "item_name", "qty", "uom", "stock_qty", "stock_uom", "custom_cover_type", "mrp",
#             "custom_mfg_on", "custom_covers", "custom_packing_type", "custom_per_2p",
#             "custom_2nd_packing", "custom_name_print", "custom_tertiary_packingbox",
#             "custom_bag", "custom_box", "custom_wrd_uom", "custom_wrd_rate"
#         ])
#         rowspan = len(items)
#         total_qty = total_stock_qty = 0
#         primary = secondary = tertiary = packing_details = ""

#         for idx, item in enumerate(items):
#             row = row_idx

#             if idx == 0:
#                 merge_values = [
#                     s_no, so.name, '', so.customer,
#                     formatdate(so.custom_packing_on), formatdate(so.delivery_date)
#                 ]
#                 for i, val in enumerate(merge_values, start=1):
#                     align = left_align if i == 4 else center_align
#                     style_merged_range(ws, row, row + rowspan - 1, i, val, align=align)

#             # Item-specific
#             ws.cell(row=row, column=7, value=item.item_name or "")
#             ws.cell(row=row, column=8, value=item.qty or 0)
#             ws.cell(row=row, column=9, value=item.uom or "")
#             ws.cell(row=row, column=10, value=item.stock_qty or 0)

#             # Cr. Stock
#             cr_stock = frappe.db.get_value("Bin", {
#                 "item_code": item.item_code,
#                 "warehouse": "Stores - TFP"
#             }, "actual_qty") or 0
#             ws.cell(row=row, column=11, value=cr_stock)

#             # Stock Status
#             status = "In Stock" if flt(item.stock_qty) <= flt(cr_stock) else "Out Of Stock"
#             status_cell = ws.cell(row=row, column=12, value=status)
#             status_font = Font(color="00B050", bold=True) if status == "In Stock" else Font(color="FF0000", bold=True)
#             style_cell(status_cell, align=center_align, font=status_font)


#             # UOM
#             ws.cell(row=row, column=13, value=item.stock_uom or "")

#             # MRP
#             ws.cell(row=row, column=14, value=item.mrp or 0)
#             style_cell(ws.cell(row=row, column=14), align=right_align)

#             # Packing details
#             if item.custom_cover_type:
#                 primary = frappe.db.get_value("Item", item.custom_cover_type, "item_name") or ''
#             if item.custom_packing_type:
#                 secondary = frappe.db.get_value("Item", item.custom_packing_type, "item_name") or ''
#             if item.custom_tertiary_packingbox:
#                 tertiary = frappe.db.get_value("Item", item.custom_tertiary_packingbox, "item_name") or ''
#             packing_details = f"(C): {primary or 'None'}: {item.custom_covers or 0}\n(B): {secondary or 'None'}: {item.custom_bag or 0}\n(BX): {tertiary or 'None'}: {item.custom_box or 0}"
#             ws.cell(row=row, column=15, value=packing_details)

#             # WRD details
#             wrd_details = f"(W): {item.custom_wrd_uom or 'None'}\n(R): {item.custom_wrd_rate or '0'}\n(D): {formatdate(item.custom_mfg_on) if item.custom_mfg_on else ''}"
#             ws.cell(row=row, column=16, value=wrd_details)

#             ws.cell(row=row, column=17, value=item.custom_name_print or "")

#             # Empty placeholders
#             for i in range(18, len(headers) + 1):
#                 ws.cell(row=row, column=i, value="")

#             # Styling for entire row
#             for col_idx in range(7, len(headers) + 1):
#                 align = left_align if col_idx in [7, 15, 16, 17] else center_align
#                 style_cell(ws.cell(row=row, column=col_idx), align=align)

#             total_qty += flt(item.qty)
#             total_stock_qty += flt(item.stock_qty)
#             row_idx += 1
#         all_in_stock = all(
#             flt(it.stock_qty) <= (frappe.db.get_value("Bin", {"item_code": it.item_code, "warehouse": "Stores - TFP"}, "actual_qty") or 0)
#             for it in items
#         )
#         status_label = "CREATE DN" if all_in_stock else "CREATE MR"
#         # Subtotal
#         ws.cell(row=row_idx, column=7, value="Total")
#         ws.cell(row=row_idx, column=8, value=total_qty)
#         ws.cell(row=row_idx, column=10, value=total_stock_qty)
#         merged_status_cell = ws.cell(row=row_idx, column=11, value=status_label)
#         ws.merge_cells(start_row=row_idx, start_column=11, end_row=row_idx, end_column=12)
#         style_cell(merged_status_cell, align=center_align, font=Font(bold=True), fill=total_fill)
#         for col_idx in list(range(1, 11)) + list(range(13, len(headers) + 1)):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             style_cell(cell, align=center_align, fill=total_fill)

#         row_idx += 1
#         s_no += 1
#         grand_total_qty += total_qty
#         grand_total_stock_qty += total_stock_qty

#     # Grand total row
#     ws.cell(row=row_idx, column=1, value="Grand Total")
#     ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
#     for i in range(1, 8):
#         style_cell(ws.cell(row=row_idx, column=i), align=center_align, font=header_font, fill=header_fill)
#     ws.cell(row=row_idx, column=8, value=grand_total_qty)
#     ws.cell(row=row_idx, column=10, value=grand_total_stock_qty)
#     for col_idx in range(8, len(headers) + 1):
#         style_cell(ws.cell(row=row_idx, column=col_idx), align=center_align, font=header_font, fill=header_fill)

#     # Adjust column widths
#     for col in ws.columns:
#         max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
#         ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

#     # Return file
#     file_content = BytesIO()
#     wb.save(file_content)
#     frappe.response['filename'] = "TFP_Packing_Plan.xlsx"
#     frappe.response['filecontent'] = file_content.getvalue()
#     frappe.response['type'] = 'binary'

import frappe
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from frappe.utils import formatdate, flt
from io import BytesIO

@frappe.whitelist()
def download_tfp_plan_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TFP Packing Plan"

    headers = [
        "Sr", "SO ID", "PRT", "Customer Name", "Packing", "Delivery", "Item Name", "QTY", "UOM",
        "St.QTY", "Cr. Stock", "Stock Status", "UOM", "MRP", "Packing Details", "WRD Details",
        "Name Print"
    ]

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    stock_header_fill = PatternFill("solid", fgColor="FF0000")  # Dark red
    header_fill = PatternFill("solid", fgColor="002060")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    total_fill = PatternFill("solid", fgColor="D9D9D9")  # Light ash gray
    in_stock_fill = PatternFill("solid", fgColor="92D050")  # Green
    out_stock_fill = PatternFill("solid", fgColor="FF0000")  # Red
    white_font = Font(color="FFFFFF", bold=True)

    def style_cell(cell, align=center_align, font=None, fill=None):
        cell.alignment = align
        cell.border = thin_border
        if font:
            cell.font = font
        if fill:
            cell.fill = fill

    def style_merged_range(ws, start_row, end_row, col, value, align=left_align, font=None, fill=None):
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col, value=value if r == start_row else None)
            style_cell(cell, align=align, font=font, fill=fill)
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    # Header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        if header in ["Cr. Stock", "Stock Status"]:
            style_cell(cell, align=center_align, font=white_font, fill=stock_header_fill)
        else:
            style_cell(cell, align=center_align, font=header_font, fill=header_fill)

    row_idx = 2
    s_no = 1
    grand_total_qty = grand_total_stock_qty = 0

    so_list = frappe.db.get_all("Sales Order", {
        "service": "TFP",
        "status": "To Deliver and Bill"
    }, ["name", "customer", "custom_packing_on", "delivery_date"])

    for so in so_list:
        items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, [
            "item_code", "item_name", "qty", "uom", "stock_qty", "stock_uom", "custom_cover_type", "mrp",
            "custom_mfg_on", "custom_covers", "custom_packing_type", "custom_per_2p",
            "custom_2nd_packing", "custom_name_print", "custom_tertiary_packingbox",
            "custom_bag", "custom_box", "custom_wrd_uom", "custom_wrd_rate"
        ])
        rowspan = len(items)
        total_qty = total_stock_qty = 0
        primary = secondary = tertiary = packing_details = ""

        for idx, item in enumerate(items):
            row = row_idx

            if idx == 0:
                merge_values = [
                    s_no, so.name, '', so.customer,
                    formatdate(so.custom_packing_on), formatdate(so.delivery_date)
                ]
                for i, val in enumerate(merge_values, start=1):
                    align = left_align if i == 4 else center_align
                    style_merged_range(ws, row, row + rowspan - 1, i, val, align=align)

            # Item-specific
            ws.cell(row=row, column=7, value=item.item_name or "")
            ws.cell(row=row, column=8, value=item.qty or 0)
            ws.cell(row=row, column=9, value=item.uom or "")
            ws.cell(row=row, column=10, value=item.stock_qty or 0)

            # Cr. Stock
            cr_stock = frappe.db.get_value("Bin", {
                "item_code": item.item_code,
                "warehouse": "Stores - TFP"
            }, "actual_qty") or 0
            ws.cell(row=row, column=11, value=cr_stock)

            # Stock Status
            status = "In Stock" if flt(item.stock_qty) <= flt(cr_stock) else "Out Of Stock"
            status_cell = ws.cell(row=row, column=12, value=status)
            status_font = Font(color="00B050", bold=True) if status == "In Stock" else Font(color="FF0000", bold=True)
            style_cell(status_cell, align=center_align, font=status_font)


            # UOM
            ws.cell(row=row, column=13, value=item.stock_uom or "")

            # MRP
            ws.cell(row=row, column=14, value=item.mrp or 0)
            style_cell(ws.cell(row=row, column=14), align=right_align)

            # Packing details
            if item.custom_cover_type:
                primary = frappe.db.get_value("Item", item.custom_cover_type, "item_name") or ''
            if item.custom_packing_type:
                secondary = frappe.db.get_value("Item", item.custom_packing_type, "item_name") or ''
            if item.custom_tertiary_packingbox:
                tertiary = frappe.db.get_value("Item", item.custom_tertiary_packingbox, "item_name") or ''
            packing_details = f"(C): {primary or 'None'}: {item.custom_covers or 0}\n(B): {secondary or 'None'}: {item.custom_bag or 0}\n(BX): {tertiary or 'None'}: {item.custom_box or 0}"
            ws.cell(row=row, column=15, value=packing_details)

            # WRD details
            wrd_details = f"(W): {item.custom_wrd_uom or 'None'}\n(R): {item.custom_wrd_rate or '0'}\n(D): {formatdate(item.custom_mfg_on) if item.custom_mfg_on else ''}"
            ws.cell(row=row, column=16, value=wrd_details)

            ws.cell(row=row, column=17, value=item.custom_name_print or "")

            # Empty placeholders
            for i in range(18, len(headers) + 1):
                ws.cell(row=row, column=i, value="")

            # Styling for entire row
            for col_idx in range(7, len(headers) + 1):
                align = left_align if col_idx in [7, 15, 16, 17] else center_align
                style_cell(ws.cell(row=row, column=col_idx), align=align)

            total_qty += flt(item.qty)
            total_stock_qty += flt(item.stock_qty)
            row_idx += 1
        all_in_stock = all(
            flt(it.stock_qty) <= (frappe.db.get_value("Bin", {"item_code": it.item_code, "warehouse": "Stores - TFP"}, "actual_qty") or 0)
            for it in items
        )
        status_label = "CREATE DN" if all_in_stock else "CREATE MR"
        # Subtotal
        ws.cell(row=row_idx, column=7, value="Total")
        ws.cell(row=row_idx, column=8, value=total_qty)
        ws.cell(row=row_idx, column=10, value=total_stock_qty)
        merged_status_cell = ws.cell(row=row_idx, column=11, value=status_label)
        ws.merge_cells(start_row=row_idx, start_column=11, end_row=row_idx, end_column=12)
        style_cell(merged_status_cell, align=center_align, font=Font(bold=True), fill=total_fill)
        for col_idx in list(range(1, 11)) + list(range(13, len(headers) + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            style_cell(cell, align=center_align, fill=total_fill)

        row_idx += 1
        s_no += 1
        grand_total_qty += total_qty
        grand_total_stock_qty += total_stock_qty

    # Grand total row
    ws.cell(row=row_idx, column=1, value="Grand Total")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    for i in range(1, 8):
        style_cell(ws.cell(row=row_idx, column=i), align=center_align, font=header_font, fill=header_fill)
    ws.cell(row=row_idx, column=8, value=grand_total_qty)
    ws.cell(row=row_idx, column=10, value=grand_total_stock_qty)
    for col_idx in range(8, len(headers) + 1):
        style_cell(ws.cell(row=row_idx, column=col_idx), align=center_align, font=header_font, fill=header_fill)

    # Adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Return file
    file_content = BytesIO()
    wb.save(file_content)
    frappe.response['filename'] = "TFP_Packing_Plan.xlsx"
    frappe.response['filecontent'] = file_content.getvalue()
    frappe.response['type'] = 'binary'

