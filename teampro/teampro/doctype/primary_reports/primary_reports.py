# Copyright (c) 2024, TeamPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import frappe
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from io import BytesIO
from frappe.utils.file_manager import save_file


class PrimaryReports(Document):
	pass

@frappe.whitelist()
def batch_status_report(doc):
	s_no = 1
	data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
	data += '<tr><td colspan="18" style="text-align:center; font-weight:bold;">BCS - Batch Status Report</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Initiation Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Batch Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Billing Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Completion Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">No of Cases</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Case</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Pending</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Comp</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Insuff</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">No of Checks</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Ch_Pending</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Ch_Comp</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#Ch_Insuff</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">EV</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ECD</td>' \
			'</tr>'

	batch_data = frappe.db.get_all("Batch", {"batch_status": ("in", ["Open", "Open with Insuff"])}, ["*"])

	for i in batch_data:
		data += f'<tr>' \
				f'<td style="text-align:center;">{s_no}</td>' \
				f'<td style="text-align:center;">{i.name}</td>' \
				f'<td style="text-align:center;">{i.customer}</td>' \
				f'<td style="text-align:center;">{i.expected_start_date}</td>' \
				f'<td style="text-align:center;">{i.batch_status}</td>' \
				f'<td style="text-align:center;">{i.billing_status}</td>' \
				f'<td style="text-align:center;">{i.expected_end_date}</td>' \
				f'<td style="text-align:center;">{i.no_of_cases}</td>' \
				f'<td style="text-align:center;">{i.case}</td>' \
				f'<td style="text-align:center;">{i.pending}</td>' \
				f'<td style="text-align:center;">{i.comp}</td>' \
				f'<td style="text-align:center;">{i.insuff}</td>' \
				f'<td style="text-align:center;">{i.no_of_checks}</td>' \
				f'<td style="text-align:center;"></td>' \
				f'<td style="text-align:center;"></td>' \
				f'<td style="text-align:center;"></td>' \
				f'<td style="text-align:center;"></td>' \
				f'<td style="text-align:center;"></td>' \
				'</tr>'
		s_no += 1
	
	data += '</table>'
	return data

from datetime import datetime
from openpyxl.styles import PatternFill
@frappe.whitelist()
def download_bcs_report():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	filename_bcs = "BCS - Batch Status Report"+ posting_date
	build_xlsx_response_bcs(filename_bcs)
	
def build_xlsx_response_bcs(filename_bcs):
	xlsx_file = make_xlsx_bcs(filename_bcs)
	frappe.response['filename'] = filename_bcs + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
def make_xlsx_bcs(data, sheet_name=None, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)  # White color font
	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20 
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 20
	ws.column_dimensions['I'].width = 20
	ws.column_dimensions['J'].width = 20 
	ws.column_dimensions['K'].width = 20 
	ws.column_dimensions['L'].width = 20 
	ws.column_dimensions['M'].width = 20 
	ws.column_dimensions['N'].width = 20 
	ws.column_dimensions['O'].width = 20 
	ws.column_dimensions['P'].width = 20 
	ws.column_dimensions['Q'].width = 20 
	ws.column_dimensions['R'].width = 20 
	header = ["S NO", "ID", "Customer", "Initiation Date", "Batch Status", "Billing Status",
			   "Completion Date", "No of Cases", "#Case", "#Pending", "#Comp", "#Insuff", 
			   "No of Checks", "#Ch_Pending", "#Ch_Comp", "#Ch_Insuff", "EV", "ECD"]
	ws.append(header)  
	for cell in ws[1]: 
		cell.fill = fill_color
		cell.font = header_font  # Apply white font to each header cell 

	data1= get_data_of_bcs()
	for row in data1:
		ws.append(row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def get_data_of_bcs():
	data = []
	s_no=1
	batch_data = frappe.db.get_all("Batch", {"batch_status": ("in", ["Open", "Open with Insuff"])}, ["*"])

	for i in batch_data:
		data.append([s_no, i.name, i.customer, i.expected_start_date, i.batch_status, i.billing_status,
			i.expected_end_date, i.no_of_cases, i.case, i.pending, i.comp, i.insuff,
			i.no_of_checks, '', '', '', '', ''])
		s_no+=1
	return data
	
	
@frappe.whitelist()
def print_sales_invoice_outstanding_report():
	s_no = 1
	data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
	data += '<tr><td colspan="12" style="text-align:center; font-weight:bold;">Sales Invoice Outstanding Report</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Company</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Service</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">AM</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Total</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Outstanding</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">DM</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer Group</td>' \
			'</tr>'
	
	sales_invoices = frappe.db.sql("""
	SELECT 
		s.name AS name, 
		c.abbr AS company, 
		s.services AS services, 
		s.status AS status, 
		s.customer AS customer, 
		account_manager.short_code AS account_manager, 
		s.total AS total, 
		s.outstanding_amount AS outstanding_amount, 
		s.posting_date AS posting_date, 
		delivery_manager.short_code AS delivery_manager, 
		s.customer_group AS customer_group
	FROM 
		`tabSales Invoice` s
	INNER JOIN 
		`tabCompany` c ON c.name = s.company
	LEFT JOIN 
		`tabEmployee` account_manager ON account_manager.user_id = s.account_manager
	LEFT JOIN 
		`tabEmployee` delivery_manager ON delivery_manager.user_id = s.delivery_manager
	WHERE 
		s.status NOT IN ('Return', 'Paid', 'Credit Note Issued', 'Cancelled')
""", as_dict=True)

	
	total_of_total = total_outstanding = 0
	for i in sales_invoices:
		formatted_date = frappe.utils.formatdate(i.posting_date, 'dd-mm-yyyy')
		
		data += f'<tr>' 
		data += f'<td style="text-align:center;">{s_no}</td>' 
		data += f'<td style="text-align:center;">{i.name}</td>' 
		data += f'<td style="text-align:center;">{i.company}</td>' 
		data += f'<td style="text-align:center;">{i.services}</td>' 
		data += f'<td style="text-align:center;">{i.status}</td>' 
		data += f'<td style="text-align:left;">{i.customer}</td>' 
		data += f'<td style="text-align:center;">{i.account_manager}</td>' 
		data += f'<td style="text-align:center;">{i.total:,.2f}</td>' 
		data += f'<td style="text-align:center;">{i.outstanding_amount:,.2f}</td>' 
		data += f'<td style="text-align:center;">{formatted_date}</td>' 
		data += f'<td style="text-align:center;">{i.delivery_manager}</td>' 
		data += f'<td style="text-align:center;">{i.customer_group}</td>' 
		data += '</tr>'
		s_no += 1
		total_of_total += i.total
		total_outstanding += i.outstanding_amount
		total_of_total = round(total_of_total, 2)
	data += f'<tr style="background-color: #f79646;">' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:left;">Total</td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;">{total_of_total:,.2f}</td>' \
			f'<td style="text-align:center;">{total_outstanding:,.2f}</td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			'</tr>'
			
	data += '</table>'
	return data

from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import frappe

@frappe.whitelist()
def download_sales_invoice_outstanding_report():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	filename = "Sales Invoice Outstanding Report " + posting_date
	build_xlsx_response_si(filename)

def build_xlsx_response_si(filename):
	xlsx_file = make_xlsx_si(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx_si(sheet_name="Sales Invoice Outstanding Report", wb=None):
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.active
	ws.title = sheet_name

	fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
	font = Font(bold=True, color="FFFFFF")
	alignment = Alignment(horizontal="center")
	text_wrap = Alignment(wrap_text=True)
	title_font = Font(bold=True, size=14)
	bold_font = Font(bold=True)
	bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
	inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin")
	)
	
	column_widths = [6, 15, 10, 9, 10, 45, 6, 13, 13, 14, 6, 17]
	columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
	for i, width in enumerate(column_widths):
		ws.column_dimensions[columns[i]].width = width
		
	posting_date = datetime.now().strftime("%d-%m-%Y")
	title = "Sales Invoice Outstanding Report (" + posting_date + ")"
	ws.merge_cells("A1:L1")
	ws["A1"].value = title
	ws["A1"].font = title_font
	ws["A1"].alignment = alignment
	
	headers = ["S NO", "ID", "Company", "Services", "Status", "Customer Name",
			   "AM", "Total", "Outstanding", "Date", "DM", "Customer Group"]
	ws.append(headers)

	for cell in ws[2]:
		cell.fill = fill_color
		cell.font = font
		cell.alignment = alignment
		cell.border = thin_border

	total_invoice = 0
	total_outstanding = 0
	data = get_data_si()

	for row in data:
		ws.append(row)
		ws[f"H{ws.max_row}"].style = inr_format
		ws[f"I{ws.max_row}"].style = inr_format
		for cell in ws[ws.max_row]:
			cell.alignment = text_wrap
			cell.border = thin_border
			
		total_invoice += row[7]  
		total_outstanding += row[8] 

	ws.append(["", "", "", "", "", "Total", "", total_invoice, total_outstanding, "", "", ""])

	for cell in ws[ws.max_row]:
		cell.font = bold_font
		cell.fill = bg_fill
		cell.border = thin_border
		
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def get_data_si():
	data = []
	s_no = 1
	sales_invoices = frappe.db.sql("""
	SELECT 
		s.name AS name, 
		c.abbr AS company, 
		s.services AS services, 
		s.status AS status, 
		s.customer AS customer, 
		account_manager.short_code AS account_manager, 
		s.total AS total, 
		s.outstanding_amount AS outstanding_amount, 
		s.posting_date AS posting_date, 
		delivery_manager.short_code AS delivery_manager, 
		s.customer_group AS customer_group
	FROM 
		`tabSales Invoice` s
	INNER JOIN 
		`tabCompany` c ON c.name = s.company
	LEFT JOIN 
		`tabEmployee` account_manager ON account_manager.user_id = s.account_manager
	LEFT JOIN 
		`tabEmployee` delivery_manager ON delivery_manager.user_id = s.delivery_manager
	WHERE 
		s.status NOT IN ('Return', 'Paid', 'Credit Note Issued', 'Cancelled')
""", as_dict=True)

	for invoice in sales_invoices:
		formatted_date = frappe.utils.formatdate(invoice.posting_date, 'dd-mm-yyyy')
		data.append([
			s_no, invoice.name, invoice.company, invoice.services, invoice.status,
			invoice.customer, invoice.account_manager, invoice.total, 
			invoice.outstanding_amount, formatted_date, invoice.delivery_manager, 
			invoice.customer_group
		])
		
		s_no += 1

	return data




# @frappe.whitelist()
# def opportunity_report(doc):
#     s_no = 1
#     data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     data += '<tr><td colspan="11" style="text-align:center; font-weight:bold;">BCS - Batch Status Report</td></tr>'
#     data += '<tr style="background-color: #002060; color: white;">' \
#             '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Owner</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Service</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">From</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Organization Name</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Age</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">Amount</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">PB%</td>' \
#             '<td style="text-align:center; font-weight:bold; color:white;">ECD</td>' \
#             '</tr>'

#     opportunity_data = frappe.db.get_all("Opportunity", {"status": ("in", ["Open","Quotation"])}, ["*"],group_by='opportunity_owner asc')
#     grouped_data = {}
#     for i in opportunity_data:
#         data += f'<tr>' \
#                 f'<td style="text-align:center;">{s_no}</td>' \
#                 f'<td style="text-align:center;">{i.opportunity_owner}</td>' \
#                 f'<td style="text-align:center;">{i.service}</td>' \
#                 f'<td style="text-align:center;">{i.opportunity_from}</td>' \
#                 f'<td style="text-align:center;">{i.status}</td>' \
#                 f'<td style="text-align:center;">{i.organization_name}</td>' \
#                 f'<td style="text-align:center;">{i.transaction_date}</td>' \
#                 f'<td style="text-align:center;">{i.custom_opportunity_age}</td>' \
#                 f'<td style="text-align:center;">{i.opportunity_amount}</td>' \
#                 f'<td style="text-align:center;"></td>' \
#                 f'<td style="text-align:center;">{i.expected_closing}</td>' \
#                 '</tr>'
#         s_no += 1
	
#     data += '</table>'
#     return data
@frappe.whitelist()
def opportunity_report(doc):
	s_no = 1
	data = '<table border="1" style=" width: 100%; whitespace: nowrap;">'
	data += '<tr><td colspan="11" style="text-align:center; font-weight:bold; font-size: 15px;">Opportunity Report</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">S.NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Owner</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Service</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">From</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 20%;">Organization Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">Age</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 10%;">Amount</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">PB%</td>' \
			'<td style="text-align:center; font-weight:bold; color:white; min-width: 5%;">ECD</td>' \
			'</tr>'

	# Fetch all opportunities with status 'Open' or 'Quotation'
	opportunity_data = frappe.db.get_all("Opportunity", 
		{"status": ("in", ["Open", "Quotation", "Replied"])}, 
		["*"])

	# Group the data by the owner (opportunity_owner)
	grouped_data = {}
	for opportunity in opportunity_data:
		owner = opportunity.get("opportunity_owner")
		if owner not in grouped_data:
			grouped_data[owner] = []
		grouped_data[owner].append(opportunity)

	# Loop through each owner and display their opportunities
	for owner, opportunities in grouped_data.items():
		# Calculate rowspan (number of opportunities for this owner)
		rowspan = len(opportunities)
		short_code=frappe.db.get_value("Employee",{"user_id":owner},["short_code"], order_by = 'short_code asc')
		# Add the owner row once with rowspan
		data += f'<tr>' \
				f'<td style="text-align:center;">{s_no}</td>' \
				f'<td style="text-align:center;" rowspan="{rowspan}">{short_code}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("service")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("opportunity_from")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("status")}</td>' \
				f'<td style="text-align:left;">{opportunities[0].get("organization_name")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("transaction_date")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("custom_opportunity_age")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("opportunity_amount")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("probability")}</td>' \
				f'<td style="text-align:center;">{opportunities[0].get("expected_closing")}</td>' \
				'</tr>'
		s_no += 1
		# Add the remaining opportunity rows for the same owner
		for opportunity in opportunities[1:]:
			data += f'<tr>' \
					f'<td style="text-align:center;">{s_no}</td>' \
					f'<td style="text-align:center;">{opportunity.get("service")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("opportunity_from")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("status")}</td>' \
					f'<td style="text-align:left;">{opportunity.get("organization_name")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("transaction_date")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("custom_opportunity_age")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("opportunity_amount")}</td>' \
					f'<td style="text-align:center;">{opportunities[0].get("probability")}</td>' \
					f'<td style="text-align:center;">{opportunity.get("expected_closing")}</td>' \
					'</tr>'
			s_no += 1

	data += '</table>'
	return data
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
import frappe

@frappe.whitelist()
def opportunity_excel_report():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	filename_opp = "Opportunity_Report_" + posting_date
	build_xlsx_response_opp(filename_opp)

def build_xlsx_response_opp(filename_opp):
	xlsx_file = make_xlsx_opp(filename_opp)
	frappe.response['filename'] = filename_opp + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx_opp(data, sheet_name="Opportunities", wb=None, column_widths=None):
	
	default_column_widths = [7, 7, 10, 13, 10, 40, 13, 7, 10, 7, 13]
	column_widths = column_widths or default_column_widths
	
	if wb is None:
		wb = Workbook()
	ws = wb.active
	ws.title = sheet_name
	
	fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	align_center = Alignment(horizontal="center", vertical="center")
	title_font = Font(bold=True, size=14)
	alignment = Alignment(horizontal="center")
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin")
	)
	
	columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
	for i, col in enumerate(columns):
		ws.column_dimensions[col].width = column_widths[i] if i < len(column_widths) else 20
	posting_date = datetime.now().strftime("%d-%m-%Y")
	title = "Oppurtunity Report (" + posting_date + ")"
	ws.merge_cells("A1:K1")
	ws["A1"].value = title
	ws["A1"].font = title_font
	ws["A1"].alignment = alignment
	
	header = ["S NO", "Owner", "Service", "From", "Status", "Organization Name",
			   "Date", "Age", "Amount", "PB%", "ECD"]
	ws.append(header)
	
	for cell in ws[2]:
		cell.fill = fill_color
		cell.font = header_font
		cell.alignment = align_center
		cell.border = thin_border

	data1 = get_data_of_opp()
	current_row = 3 
	start_row = current_row
	last_owner = data1[0][1] if data1 else None 

	for row in data1:
		ws.append(row)
		owner = row[1] 

		if owner != last_owner and start_row < current_row - 1:
			ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
			ws.cell(row=start_row, column=2).alignment = align_center  
			start_row = current_row  
		
		last_owner = owner
		current_row += 1

	if start_row < current_row - 1:
		ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row - 1, end_column=2)
		ws.cell(row=start_row, column=2).alignment = align_center
		
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(header)):
		for cell in row:
			cell.border = thin_border

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def get_data_of_opp():
	data = []
	s_no = 1
	opportunity_data = frappe.db.get_all("Opportunity", filters={"status": ["in", ["Open", "Quotation"]]}, fields=["*"])
	grouped_data = {}
	
	for opportunity in opportunity_data:
		owner = opportunity.get("opportunity_owner")
		if owner not in grouped_data:
			grouped_data[owner] = []
		grouped_data[owner].append(opportunity)
	
	for owner, opportunities in grouped_data.items():
		for i in opportunities:
			employee_short_code = frappe.db.get_value("Employee", {"user_id": owner}, "short_code")
			
			data.append([
				s_no, 
				employee_short_code, 
				i.service, 
				i.transaction_date, 
				i.status, 
				i.organization_name, 
				i.transaction_date, 
				i.custom_opportunity_age, 
				i.opportunity_amount, 
				i.probability, 
				i.expected_closing
			])
			s_no += 1

	return data



@frappe.whitelist()
def download_PTSR():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	filename = "Project - Task Status Report " + posting_date
	build_xlsx_response_PTSR(filename)

def build_xlsx_response_PTSR(filename):
	xlsx_file = make_xlsx_PTSR(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx_PTSR(sheet_name="PTSR", wb=None, column_widths=None):
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15, 15, 15, 25, 15, 7, 7, 7, 7, 7, 7]
	column_widths = column_widths or default_column_widths
	for i, width in enumerate(column_widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = width
	posting_date = datetime.now().strftime("%d-%m-%Y")
	ftitle = "REC : Project – Task Status Report : - " + posting_date
	ws.append([ftitle])
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
	title_cell = ws.cell(row=1, column=1)
	title_cell.alignment = Alignment(horizontal="center", vertical="center")
	header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
	header_font = Font(color="FFFFFF")
	headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark",'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory', 'TASK', 'Task Priority', '#VAC', '#SP', '#FP', '#SL', '#PSL', '#LP']
	black_border = Border(
		left=Side(border_style="thin", color="000000"),
		right=Side(border_style="thin", color="000000"),
		top=Side(border_style="thin", color="000000"),
		bottom=Side(border_style="thin", color="000000")
	)
	ws.append(headers)
	header_row = ws[ws.max_row]
	for cell in header_row:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = black_border
	cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
	row = 3
	serial_number = 1
	grand_totals = {'vac':0,'sp': 0,'fp': 0,'sl':0,'psl':0,'custom_lp':0}
	for c in cust:
		priority = {"High": 1, "Medium": 2, "Low": 3}
		pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'],order_by= "priority ASC")
		if not pname:
			continue
		task_totals = {'vac':0,'sp':0,'fp':0,'sl':0,'psl':0,'custom_lp':0}
		project_data = []      
		for p in pname:
			pdata = []
			print(p.project_name)        
			taskid = frappe.get_all("Task", {"status": ("in",('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'],order_by= "priority ASC")              
			# print(p['project_name'])
			# for tn in taskid:
				# print(tn.name)
			for t in taskid:
				pdata.append([p['project_name'] if p['project_name'] else "",p['priority'] if p['priority'] else "",p['remark'] if p['remark'] else "",p['account_manager_remark'] if p['account_manager_remark'] else "",p['custom_spoc_remark'] if p['custom_spoc_remark'] else "",p['expected_value'] if p['expected_value'] else "",p['expected_psl'] if p['expected_psl'] else "",p['sourcing_statu'] if p['sourcing_statu'] else "",p['territory'] if p['territory'] else "",t['subject'],t['priority'],t['vac'],t['sp'],t['fp'],t['sl'],t['psl'],t['custom_lp']])
				task_totals['vac'] +=t['vac']
				task_totals['sp'] +=t['sp']
				task_totals['fp']+= t['fp']
				task_totals['sl'] +=t['sl']
				task_totals['psl'] += t['psl']
				task_totals['custom_lp'] += t['custom_lp']
			project_data.append({
				'project_name': p['project_name'],'priority': p['priority'],
				'remark': p['remark'],'account_manager_remark': p['account_manager_remark'],'custom_spoc_remark':p['custom_spoc_remark'],'sourcing_statu': p['sourcing_statu'],'territory': p['territory'],
				'expected_value': p['expected_value'],'expected_psl': p['expected_psl'],'tasks': pdata})
		blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
		row_data = [serial_number, c['name']] + [""] * 10 + [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]
		ws.append(row_data)
		row_to_fill = ws.max_row
		for col, cell in enumerate(ws[row_to_fill], start=1):
			cell.fill = blue_fill
			if col > 11:
				cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
			else:
				cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
			cell.border = black_border
		ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
		serial_number += 1
		row += 1
		current_row_start = row
		for project in project_data:
			project_row_start = row
			for task_data in project['tasks']:
				ws.append([""] + task_data)
				for col in range(2, len(task_data) + 2):
					cell = ws.cell(row=row, column=col)
					if 2 <= col <= 11:
						cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
					else:
						cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
					cell.border = black_border
				row += 1
			if project_row_start < row - 1:
				for col in range(2, 10):
					ws.merge_cells(start_row=project_row_start, start_column=col, end_row=row-1, end_column=col)
				ws.merge_cells(start_row=project_row_start, start_column=1, end_row=row-1, end_column=1)
		grand_totals['vac'] += task_totals['vac']
		grand_totals['sp'] += task_totals['sp']
		grand_totals['fp'] += task_totals['fp']
		grand_totals['sl'] += task_totals['sl']
		grand_totals['psl'] += task_totals['psl']
		grand_totals['custom_lp'] +=task_totals['custom_lp']
	yellow_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
	ws.append(['Total'] + [''] * 10 + [grand_totals['vac'], grand_totals['sp'], grand_totals['fp'], grand_totals['sl'], grand_totals['psl'],grand_totals['custom_lp']])
	last_row = ws.max_row
	for cell in ws[last_row]:
		cell.fill = yellow_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = black_border
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file


import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import frappe
import io
from datetime import datetime

@frappe.whitelist()
def download_PSR():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	filename = "Project - Status Report " + posting_date
	xlsx_file = make_xlsx_PSR(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file
	frappe.response['type'] = 'binary'

def make_xlsx_PSR(sheet_name="PSR", wb=None, column_widths=None):
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)

	default_column_widths = [8, 25, 10, 43, 60, 60, 15, 15, 15, 15]
	column_widths = column_widths or default_column_widths
	for i, width in enumerate(column_widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = width

	posting_date = datetime.now().strftime("%d-%m-%Y")
	ftitle = "REC : Project – Task Status Report : - " + posting_date
	ws.append([ftitle])
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
	title_cell = ws.cell(row=1, column=1)
	title_cell.alignment = Alignment(horizontal="center", vertical="center")
	title_cell.font = Font(bold=True)

	header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
	header_font = Font(color="FFFFFF")
	black_border = Border(
		left=Side(border_style="thin", color="000000"),
		right=Side(border_style="thin", color="000000"),
		top=Side(border_style="thin", color="000000"),
		bottom=Side(border_style="thin", color="000000")
	)

	# Headers with borders and fill color
	headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark", 'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory']
	ws.append(headers)
	header_row = ws[ws.max_row]
	for cell in header_row:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.border = black_border

	# Initialize variables for Customer and Project Data
	cust = frappe.get_all("Customer", fields=["name"])  # Example for fetching customers
	row = 3
	serial_number = 1
	ev_total = 0
	
	for c in cust:
		pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
		
		if not pname:
			continue

		# Adding main customer row with blue fill
		blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
		row_data = [serial_number, c['name']] + [""] * 8
		ws.append(row_data)
		row_to_fill = ws.max_row

		# Apply fill and border only to outer edges of merged cells
		ws.cell(row=row_to_fill, column=1).border = Border(left=black_border.left, top=black_border.top, bottom=black_border.bottom)
		ws.cell(row=row_to_fill, column=10).border = Border(right=black_border.right, top=black_border.top, bottom=black_border.bottom)
		
		for col, cell in enumerate(ws[row_to_fill], start=1):
			cell.fill = blue_fill
			if col <= 10:
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

		ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
		serial_number += 1
		row += 1
		
		# Add task data with border on all sides
		for p in pname:
			task_data = [p['project_name'], p['priority'], p['remark'], p['account_manager_remark'],
						 p['custom_spoc_remark'], p['expected_value'], p['expected_psl'], p['sourcing_statu'], p['territory']]
			ws.append([""] + task_data)
			
			# Update totals
			try:
				ev_total += float(p.get('expected_value', 0) or 0)
			except ValueError:
				frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")


			for col in range(2, len(task_data) + 2):
				cell = ws.cell(row=row, column=col)
				cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
				cell.border = black_border
			row += 1
	
	# Total row formatting
	total_row = ["","Total", "", "", "", "", ev_total, "", "", ""]
	ws.append(total_row)
	last_row = ws[ws.max_row]
	total_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
	for cell in last_row:
		cell.fill = total_fill
		cell.font = Font(color="FFFFFF")
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = black_border

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	return output.read()



@frappe.whitelist()
def print_sales_order_outstanding_report():
	s_no = 1

	data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
	data += '<tr><td colspan="17" style="text-align:center; font-weight:bold;">Sales Order Outstanding Report</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Services</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Company</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">AM</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Net Total</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Grand Total</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">% Billed</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Amount Billed</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Pending Billing </td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Advance</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Pending Collection</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">To Be Billed</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Reference Customer</td>' \
			'</tr>'
	
	sales_order = frappe.db.sql("""
	SELECT 
		s.name AS name,
		s.customer AS customer,
		s.service AS services,
		c.abbr AS company, 
		account_manager.short_code AS account_manager,
		s.transaction_date AS posting_date,
		s.status AS status, 
		s.net_total AS net_total,
		s.grand_total AS grand_total,
		s.per_billed AS per_billed,
		s.amount_billed_company_currency AS company_currency,
		s.advance_paid AS advance_paid,
		s.reference_customer_ AS reference_customer_
	FROM 
		`tabSales Order` s
	INNER JOIN 
		`tabCompany` c ON c.name = s.company
	LEFT JOIN 
		`tabEmployee` account_manager ON account_manager.user_id = s.account_manager
	WHERE 
		s.status NOT IN ('To Deliver', 'On Hold', 'Closed', 'Cancelled', 'Completed')
""", as_dict=True)
	
	total_net_total = total_grand_total = total_amount_billed = total_pending_billing = total_advance = total_pending_collection = pending_collection = total_per_billed = 0
	for i in sales_order:
		formatted_date = frappe.utils.formatdate(i.posting_date, 'dd-mm-yyyy')
		pending_billed = (i.net_total-i.company_currency)
		pending_collection = (pending_billed-i.advance_paid)
		to_be_billed = (i.grand_total-(i.grand_total*i.per_billed)+i.advance_paid)
			
		data += f'<tr>'
		data += f'<td style="text-align:center;">{s_no}</td>'
		data += f'<td style="text-align:center;">{i.name}</td>'
		data += f'<td style="text-align:center;">{i.customer}</td>'
		data += f'<td style="text-align:center;">{i.services}</td>'
		data += f'<td style="text-align:center;">{i.company}</td>'
		data += f'<td style="text-align:center;">{i.account_manager}</td>'
		data += f'<td style="text-align:center;">{formatted_date}</td>'
		data += f'<td style="text-align:center;">{i.status}</td>'
		data += f'<td style="text-align:center;">{i.net_total}</td>'
		data += f'<td style="text-align:center;">{i.grand_total:,.2f}</td>'
		data += f'<td style="text-align:center;">{i.per_billed}</td>'
		data += f'<td style="text-align:center;">{i.company_currency:,.2f}</td>'
		data += f'<td style="text-align:center;">{pending_billed:,.2f}</td>'
		data += f'<td style="text-align:center;">{i.advance_paid:,.2f}</td>'
		data += f'<td style="text-align:center;">{pending_collection:,.2f}</td>'
		data += f'<td style="text-align:center;">{to_be_billed:,.2f}</td>'
		data += f'<td style="text-align:center;">{i.reference_customer_}</td>'
		data += '</tr>'


		s_no += 1
		total_net_total += i.net_total
		total_grand_total += i.grand_total
		total_amount_billed += i.company_currency
		total_advance += i.advance_paid
		total_pending_billing += pending_billed
		total_pending_collection += pending_collection
		total_per_billed += i.per_billed

	data += f'<tr style="background-color: #f79646">' \
			f'<td style="text-align:center;"></td>' \
			 f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:left;">Total</td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;">{total_net_total:,.2f}</td>' \
			f'<td style="text-align:center;">{total_grand_total:,.2f}</td>' \
			f'<td style="text-align:center;">{total_per_billed}</td>' \
			f'<td style="text-align:center;">{total_amount_billed:,.2f}</td>' \
			f'<td style="text-align:center;">{total_pending_billing:,.2f}</td>' \
			f'<td style="text-align:center;">{total_advance:,.2f}</td>' \
			f'<td style="text-align:center;">{pending_collection:,.2f}</td>' \
			f'<td style="text-align:center;">{to_be_billed:,.2f}</td>' \
			f'<td style="text-align:center;"></td>' \
			'</tr>'
	
	data += '</table>'
	return data

# @frappe.whitelist()
# def download_sales_order_outstanding_report():
# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	filename = "Sales Order Outstanding Report " + posting_date
# 	build_xlsx_response_so(filename)

# def build_xlsx_response_so(filename):
# 	xlsx_file = make_xlsx_so(filename)
# 	frappe.response['filename'] = filename + '.xlsx'
# 	frappe.response['filecontent'] = xlsx_file.getvalue()
# 	frappe.response['type'] = 'binary'

# def make_xlsx_so(sheet_name="Sales Order Outstanding Report", wb=None):
# 	if wb is None:
# 		wb = openpyxl.Workbook()

# 	ws = wb.active
# 	ws.title = sheet_name

# 	fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
# 	font = Font(bold=True, color="FFFFFF")
# 	alignment = Alignment(horizontal="center")
# 	text_wrap = Alignment(wrap_text=True)
# 	title_font = Font(bold=True, size=14)
# 	bold_font = Font(bold=True)
# 	bg_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")  # Background color for last row
# 	inr_format = NamedStyle(name="inr_format", number_format="#,##0.00")
# 	thin_border = Border(
# 		left=Side(style="thin"),
# 		right=Side(style="thin"),
# 		top=Side(style="thin"),
# 		bottom=Side(style="thin")
# 	)
	
# 	column_widths = [6, 25, 35, 9, 10, 9, 15, 20, 13, 13, 15, 15, 15, 18, 18, 18, 35]
# 	columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
# 	for i, width in enumerate(column_widths):
# 		ws.column_dimensions[columns[i]].width = width
		
# 	posting_date = datetime.now().strftime("%d-%m-%Y")
# 	title = "Sales Order Outstanding Report (" + posting_date + ")"
# 	ws.merge_cells("A1:Q1")
# 	ws["A1"].value = title
# 	ws["A1"].font = title_font
# 	ws["A1"].alignment = alignment
	
# 	headers = ["S NO", "ID", "Customer Name", "Services", "Company",
# 			   "AM", "Date", "Status", "Net Total", "Grand Total", "% Billed", "Amount Billed", "Pending Billing",
# 			   "Advance", "Pending Collection", "To Be Billed", "Reference Customer"]
# 	ws.append(headers)

# 	for cell in ws[2]:
# 		cell.fill = fill_color
# 		cell.font = font
# 		cell.alignment = alignment
# 		cell.border = thin_border

# 	total_net_total = 0
# 	total_grand_total = 0
# 	total_per_billed = 0
# 	total_amount_billed = 0
# 	total_pending_billing = 0
# 	total_advance = 0
# 	pending_collection = 0
# 	to_be_billed = 0

# 	data = get_data_so()

# 	for row in data:
# 		ws.append(row)
# 		ws[f"H{ws.max_row}"].style = inr_format
# 		ws[f"I{ws.max_row}"].style = inr_format
# 		for cell in ws[ws.max_row]:
# 			cell.alignment = text_wrap
# 			cell.border = thin_border
			
# 		total_net_total += row[8]
# 		total_grand_total += row[9]
# 		total_per_billed += row[10]
# 		total_amount_billed += row[11]
# 		total_pending_billing += row[12]
# 		total_advance += row[13]
# 		pending_collection += row[14]
# 		to_be_billed += row[15]

# 	ws.append(["", "", "Total", "", "", "", "", "", total_net_total, total_grand_total, total_per_billed, total_amount_billed, total_pending_billing, total_advance, pending_collection, to_be_billed, ""])

# 	for cell in ws[ws.max_row]:
# 		cell.font = bold_font
# 		cell.fill = bg_fill
# 		cell.border = thin_border
		
# 	xlsx_file = BytesIO()
# 	wb.save(xlsx_file)
# 	xlsx_file.seek(0)
# 	return xlsx_file

# def get_data_so():
# 	data = []
# 	s_no = 1
# 	sales_order = frappe.db.sql("""
# 	SELECT 
# 		s.name AS name,
# 		s.customer AS customer,
# 		s.service AS services,
# 		c.abbr AS company, 
# 		account_manager.short_code AS account_manager,
# 		s.transaction_date AS posting_date,
# 		s.status AS status, 
# 		s.net_total AS net_total,
# 		s.grand_total AS grand_total,
# 		s.per_billed AS per_billed,
# 		s.amount_billed_company_currency AS company_currency,
# 		s.advance_paid AS advance_paid,
# 		s.reference_customer_ AS reference_customer_
# 	FROM 
# 		`tabSales Order` s
# 	INNER JOIN 
# 		`tabCompany` c ON c.name = s.company
# 	LEFT JOIN 
# 		`tabEmployee` account_manager ON account_manager.user_id = s.account_manager
# 	WHERE 
# 		s.status NOT IN ('To Deliver', 'On Hold', 'Closed', 'Cancelled', 'Completed')
# """, as_dict=True)

# 	for invoice in sales_order:
# 		formatted_date = frappe.utils.formatdate(invoice.posting_date, 'dd-mm-yyyy')
# 		pending_billed = (invoice.net_total-invoice.company_currency)
# 		pending_collected = (pending_billed-invoice.advance_paid)
# 		to_be_billed = (invoice.grand_total-(invoice.grand_total*invoice.per_billed)+invoice.advance_paid)
		
# 		data.append([
# 			s_no, invoice.name, invoice.customer, invoice.services, invoice.company, invoice.account_manager,
# 			formatted_date, invoice.status, invoice.net_total, invoice.grand_total, invoice.per_billed,
# 			invoice.company_currency, pending_billed, invoice.advance_paid, pending_collected, to_be_billed, 
# 			invoice.reference_customer_
# 		])
		
# 		s_no += 1

# 	return data



@frappe.whitelist()
def print_psr_report():
	posting_date = datetime.now().strftime("%d-%m-%Y")
	s_no = 1
	data = '<table border="1" style="border-collapse: collapse; width: 100%; ">'
	data += f'<tr><td colspan="12" style="text-align:center; font-weight:bold;">REC : Project - Status Report - {posting_date}</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">AM Mark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">PM Mark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">EV</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">EPSL</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SS</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;border-right: 2px solid black;">Territory</td>' \
			'</tr>'
	
	cust = frappe.get_all("Customer", fields=["name"])

	
	ev_total = 0
	for c in cust:
		pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
		for p in pname:
			try:
				ev_total += float(p.get('expected_value', 0) or 0)
			except ValueError:
				frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")

			data += f'<tr style="background-color: #98d7f5;">' \
					f'<td style="text-align:center;">{s_no}</td>' \
					f'<td colspan=3 style="text-align:left;">{c.name}</td>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
					f'</tr>'
			data += f'<tr>' \
					f'<td style="text-align:center;"></td>' \
					f'<td style="text-align:left;">{p.project_name}</td>' \
					f'<td style="text-align:center;">{p.priority}</td>' \
					f'<td style="text-align:left;">{p.remark}</td>' \
					f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
					f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
					f'<td style="text-align:center;">{p.expected_value}</td>' \
					f'<td style="text-align:center;">{p.expected_psl}</td>' \
					f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
					f'<td style="text-align:center;border-right: 2px solid black;">{p.territory}</td>' \
					'</tr>'
			s_no += 1
	data += f'<tr style="background-color: #002060;">' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">Total</td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:left;"></td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{ev_total}</td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;"></td>' \
			f'<td style="text-align:center;border-right: 2px solid black;"></td>' \
			'</tr>'
			
	data += '</table>'
	return data

@frappe.whitelist()
def ptsr_report(doc):
	from datetime import datetime
	posting_date = datetime.now().strftime("%d-%m-%Y")
	frappe.log_error(message="ptsr", title="error")
	s_no = 1
	data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
	data += f'<tr><td colspan="18" style="text-align:center; font-weight:bold;">REC : Project - Task Status Report - {posting_date}</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">AM Remark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">PM Remark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Expected Value</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Expected PSL</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Sourcing Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Territory</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Task</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;border-right-style: hidden;">Task Priority</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;border-left-style: hidden;">#VAC</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#SP</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#FP</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#SL</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#PSL</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">#LP</td>' \
			'</tr>'
	
	cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
	
	ev_total = 0
	grand_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}

	for c in cust:
		pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
		task_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}
		
		for p in pname:
			try:
				ev_total += float(p.get('expected_value', 0) or 0)
			except ValueError:
				frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")
			
			taskid = frappe.get_all("Task", {"status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'], order_by="priority ASC")
			
			# Project row
			data += f'<tr style="background-color: #98d7f5;">' \
					f'<td style="text-align:center;">{s_no}</td>' \
					f'<td colspan=2 style="text-align:left;">{c.name}</td>' \
					f'<td colspan=15></td>' \
					'</tr>'

			# Each task row within a project
			for t in taskid:
				data += f'<tr>' \
						f'<td style="text-align:center;"></td>' \
						f'<td style="text-align:left;">{p.project_name}</td>' \
						f'<td style="text-align:center;">{p.priority}</td>' \
						f'<td style="text-align:left;">{p.remark}</td>' \
						f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
						f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
						f'<td style="text-align:center;">{p.expected_value}</td>' \
						f'<td style="text-align:center;">{p.expected_psl}</td>' \
						f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
						f'<td style="text-align:center;">{p.territory}</td>' \
						f'<td style="text-align:left;">{t.subject}</td>' \
						f'<td style="text-align:center;border-right-style: hidden;">{t.priority}</td>' \
						f'<td style="text-align:center;border-left-style: hidden;">{t.vac}</td>' \
						f'<td style="text-align:center;">{t.sp}</td>' \
						f'<td style="text-align:center;">{t.fp}</td>' \
						f'<td style="text-align:center;">{t.sl}</td>' \
						f'<td style="text-align:center;">{t.pl}</td>' \
						f'<td style="text-align:center;">{t.custom_lp}</td>' \
						'</tr>'
				
				# Update task totals
				task_totals['vac'] += t.get('vac', 0)
				task_totals['sp'] += t.get('sp', 0)
				task_totals['fp'] += t.get('fp', 0)
				task_totals['sl'] += t.get('sl', 0)
				task_totals['psl'] += t.get('psl', 0)
				task_totals['custom_lp'] += t.get('custom_lp', 0)
				s_no += 1

	# Grand total row
	data += f'<tr style="background-color: #002060;">' \
			f'<td></td>' \
			f'<td style="text-align:center; font-weight: bold;color: #ffffff;">Total</td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{ev_total}</td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td></td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["vac"]}</td>' \
			f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{grand_totals["sp"]}</td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["fp"]}</td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["sl"]}</td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["psl"]}</td>' \
			f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["custom_lp"]}</td>' \
			'</tr>'
	
	data += '</table>'
	return data

@frappe.whitelist()
def print_closure_report():
	s_no = 1
	data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
	data += '<tr><td colspan="17" style="text-align:center; font-weight:bold;">Sales Order Outstanding Report</td></tr>'
	data += '<tr style="background-color: #002060; color: white;">' \
			'<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Passport No</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Given Name/Surename(as per passport)</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Mobile/Whatsapp</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Position/Task Subject</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Customer</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Territory</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate Owner</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate Collection Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Photo(as per visa specification)</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Signed Offer Letter</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Project</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Account Manager</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SA Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">MOH</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">PCC</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">IAF / CV / Client Form</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate Outstanding</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SA ID</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Stamped Visa</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Entry Visa</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Final Medical</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Client SI</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate SI</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">ECR Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Pre-Medical</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SO Confirmed Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Passport Number</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">PCC Original at</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SA Name</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Visa Original at</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Final Medical Original at</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Visa Expiry Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Visa Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">SO Created</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Collection Priority</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">PP Original at</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Interview Location</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Nationality</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Candidate SC</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Client Payment Company Currency</td>' \
			'<td style="text-align:center; font-weight:bold; color:white;">Billing Currency(Client / Associate)</td>' \

	closure = frappe.db.sql("""
	SELECT * FROM 'tabClosure'
	""", as_dict=True)

	for i in closure:

		data += f'<tr>'
		data += f'<td style="text-align:center;">{s_no}</td>'
		data += f'<td style="text-align:center;">{i.name}</td>'
		data += f'<td style="text-align:center;">{i.passport_no}</td>'
		data += f'<td style="text-align:center;">{i.given_name}</td>'
		data += f'<td style="text-align:center;">{i.mobile}</td>'
		data += f'<td style="text-align:center;">{i.task_subject}</td>'
		data += f'<td style="text-align:center;">{i.customer}</td>'
		data += f'<td style="text-align:center;">{i.territory}</td>'
		data += f'<td style="text-align:center;">{i.candidate_owner}</td>'
		data += f'<td style="text-align:center;">{i.collection_status}</td>'
		data += f'<td style="text-align:center;">{i.photo}</td>'
		data += f'<td style="text-align:center;">{i.sol}</td>'
		data += f'<td style="text-align:center;">{i.project}</td>'
		data += f'<td style="text-align:center;">{i.account_manager}</td>'
		data += f'<td style="text-align:center;">{i.sa_name}</td>'
		data += f'<td style="text-align:center;">{i.mog}</td>'
		data += f'<td style="text-align:center;">{i.pcc}</td>'
		data += f'<td style="text-align:center;">{i.irf}</td>'
		data += f'<td style="text-align:center;">{i.outstanding_amount}</td>'
		data += f'<td style="text-align:center;">{i.sa_id}</td>'
		data += f'<td style="text-align:center;">{i.visa_stamped}</td>'
		data += f'<td style="text-align:center;">{i.visa}</td>'
		data += f'<td style="text-align:center;">{i.final_medical}</td>'
		data += f'<td style="text-align:center;">{i.client_si}</td>'
		data += f'<td style="text-align:center;">{i.candidate_si}</td>'
		data += f'<td style="text-align:center;">{i.ecr_status}</td>'
		data += f'<td style="text-align:center;">{i.premedical}</td>'
		data += f'<td style="text-align:center;">{i.candidate}</td>'
		data += f'<td style="text-align:center;">{i.so_confirmed_date}</td>'
		data += f'<td style="text-align:center;">{i.passport_number}</td>'
		data += f'<td style="text-align:center;">{i.pcc_original}</td>'
		data += f'<td style="text-align:center;">{i.sa_name}</td>'
		data += f'<td style="text-align:center;">{i.visa_original_at}</td>'
		data += f'<td style="text-align:center;">{i.final_medical_original_at}</td>'
		data += f'<td style="text-align:center;">{i.visa_expiry_date}</td>'
		data += f'<td style="text-align:center;">{i.visa_status}</td>'
		data += f'<td style="text-align:center;">{i.status}</td>'
		data += f'<td style="text-align:center;">{i.so_created}</td>'
		data += f'<td style="text-align:center;">{i.collection_priority}</td>'
		data += f'<td style="text-align:center;">{i.pp_original_at}</td>'
		data += f'<td style="text-align:center;">{i.interview_location}</td>'
		data += f'<td style="text-align:center;">{i.posting_date}</td>'
		data += f'<td style="text-align:center;">{i.nationality}</td>'
		data += f'<td style="text-align:center;">{i.candidate_service_charge}</td>'
		data += f'<td style="text-align:center;">{i.client_payment_company_currency}</td>'
		data += f'<td style="text-align:center;">{i.billing_currency}</td>'
		data += '</tr>'
		
		s_no += 1

	data += '</table>'
	return data

	
