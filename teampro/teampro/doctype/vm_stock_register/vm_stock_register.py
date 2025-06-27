# Copyright (c) 2025, TeamPRO and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document


class VMStockRegister(Document):
    pass

import frappe
import requests
import openpyxl
import os
from io import BytesIO
from frappe import _
from frappe.utils import get_url


def download_external_file_url(file_url):
    if not file_url.startswith("http"):
        file_url = get_url() + file_url
    response = requests.get(file_url, stream=True)
    if response.status_code != 200:
        frappe.throw(_("Failed to download file from external storage."))

    return BytesIO(response.content)


@frappe.whitelist()
def vm_stock_import_excel(docname):
    doc = frappe.get_doc("VM Stock Register", docname)

    file_url = frappe.db.get_value(
        "File", 
        {"attached_to_doctype": "VM Stock Register", "attached_to_name": docname}, 
        "file_url"
    )

    if not file_url:
        frappe.throw(_("No file attached for this document."))

    try:
        in_memory_file = download_external_file_url(file_url)
        wb = openpyxl.load_workbook(in_memory_file)
        ws = wb.active
    except Exception as e:
        frappe.throw(f"Failed to read Excel file: {str(e)}")


    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_id, product_name, qty = row[0], row[1], row[2]
        for child in doc.slot_a:
            if child.slot_id == slot_id:
                child.cr_stock = product_name
                child.cr_stock_qty = qty
        for b in doc.slot_b:
            if b.slot_id == slot_id:
                b.cr_stock = product_name
                b.cr_stock_qty = qty
        for c in doc.slot_c:
            if c.slot_id == slot_id:
                c.cr_stock = product_name
                c.cr_stock_qty = qty
        for d in doc.slot_d:
            if d.slot_id == slot_id:
                d.cr_stock = product_name
                d.cr_stock_qty = qty
        for e in doc.slot_e:
            if e.slot_id == slot_id:
                e.cr_stock = product_name
                e.cr_stock_qty = qty
        for f in doc.slot_f:
            if f.slot_id == slot_id:
                f.cr_stock = product_name
                f.cr_stock_qty = qty
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return "Excel data imported and saved successfully."
   

@frappe.whitelist()
def create_re_filling_stock_entries(doc, method):
    doc = frappe.get_doc("VM Stock Register", doc.name)

    stock_entry = frappe.new_doc("Stock Entry")
    stock_entry.stock_entry_type = "Material Transfer"
    stock_entry.posting_date = frappe.utils.nowdate()
    stock_entry.company = "TEAMPRO Food Products"
    stock_entry.from_warehouse = "Stores - TFP"
    stock_entry.to_warehouse = "VM1_Precision - TFP"
    stock_entry.custom_vm_stock_register=doc.name
    has_items = False
    slot_tables = ["slot_a", "slot_b", "slot_c", "slot_d", "slot_e","slot_f"]

    for table in slot_tables:
        rows = getattr(doc, table) or []
        for row in rows:
            if row.item_code and row.new_stock_qty and row.new_stockuom and row.new_stock_qty>0:
                stock_entry.append("items", {
                    "item_code": row.item_code,
                    "qty": row.new_stock_qty,
                    "uom": row.new_stockuom,
                    "s_warehouse": stock_entry.from_warehouse,
                    "t_warehouse": stock_entry.to_warehouse
                })
                has_items = True

    if has_items:
        stock_entry.insert(ignore_permissions=True)
        stock_entry.save()
        return f"Stock Entry {stock_entry.name} created with {len(stock_entry.items)} items."
    else:
        return "No valid re-filling items found in slot tables. Stock Entry not created."

from frappe.utils import flt
@frappe.whitelist()
def create_packing_stock_entries(doc, method):
    doc = frappe.get_doc("VM Stock Register", doc.name)

    stock_entry = frappe.new_doc("Stock Entry")
    stock_entry.stock_entry_type = "Material Issue"
    stock_entry.posting_date = frappe.utils.nowdate()
    stock_entry.company = "TEAMPRO Food Products"
    stock_entry.from_warehouse = "Stores - TFP"
    stock_entry.custom_vm_stock_register = doc.name

    has_items = False

    slot_tables = ["slot_a", "slot_b", "slot_c", "slot_d", "slot_e", "slot_f"]

    for table in slot_tables:
        rows = getattr(doc, table) or []
        for row in rows:
            if flt(row.custom_covers) > 0 and row.custom_primary_packing_cover:
                stock_entry.append("items", {
                    "item_code": row.custom_primary_packing_cover,
                    "qty": flt(row.custom_covers),
                    "uom": "Nos",
                    "s_warehouse": stock_entry.from_warehouse,
                    "allow_zero_valuation_rate": 1
                })
                has_items = True

            if flt(row.custom_bag) > 0 and row.custom_secondary_packing_bag:
                stock_entry.append("items", {
                    "item_code": row.custom_secondary_packing_bag,
                    "qty": flt(row.custom_bag),
                    "uom": "Nos",
                    "s_warehouse": stock_entry.from_warehouse,
                    "allow_zero_valuation_rate": 1
                })
                has_items = True

            if flt(row.custom_box) > 0 and row.custom_tertiary_packingbox:
                stock_entry.append("items", {
                    "item_code": row.custom_tertiary_packingbox,
                    "qty": flt(row.custom_box),
                    "uom": "Nos",
                    "s_warehouse": stock_entry.from_warehouse,
                    "allow_zero_valuation_rate": 1
                })
                has_items = True

    if has_items:
        stock_entry.insert(ignore_permissions=True)
        stock_entry.save()
        # stock_entry.submit()  # Uncomment if auto-submission needed
        return f"Stock Entry {stock_entry.name} created with {len(stock_entry.items)} items."
    else:
        return "No valid items found in slot tables. Stock Entry not created."



@frappe.whitelist()
def get_uom_conversion(item_code, uom):
# def get_uom_conversion():
#     item_code ='LL-SV-00024'
#     uom ='80gm'
    conversion = frappe.db.get_value('UOM Conversion Detail', {
        'parent': item_code,
        'uom': uom
    }, 'conversion_factor')
    return conversion

@frappe.whitelist()
def get_parent_item_group(parent_item_group):
    return frappe.db.get_list('Item Group', {
           'parent_item_group': parent_item_group,
           'name':['not in',['Sub Assemblies','Raw Material']]
        }, pluck ='name')

@frappe.whitelist()
def get_available_balance(item_code,warehouse):
# def get_available_balance():
#     item_code = 'LL-SV-00024'
#     warehouse = ['Stores - TFP']
    bin_record = frappe.get_list('Bin',
        filters={
            'item_code': item_code,
            'warehouse': warehouse
        },
        fields=['actual_qty'],
        order_by='creation desc',
        limit=1
    )

    return bin_record[0].actual_qty if bin_record else 0

@frappe.whitelist()
def update_vm_status(doc,method):
    if doc.custom_vm_stock_register:
        frappe.db.set_value("VM Stock Register",doc.custom_vm_stock_register,"status","Schedule")

import frappe
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from frappe.utils.response import build_response
from io import BytesIO

@frappe.whitelist()
def download_stock_details(docname=None):
    if not docname:
        frappe.throw("VM Stock Register document name is required")

    doc = frappe.get_doc("VM Stock Register", docname)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Register"

    slot_fields = {
        'A': 'slot_a',
        'B': 'slot_b',
        'C': 'slot_c',
        'D': 'slot_d',
        'E': 'slot_e',
        'F': 'slot_f',  # Optional
    }

    # Styles
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light green
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # Light yellow

    # 🎯 Add heading
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    heading_cell = ws.cell(row=1, column=1, value="VM Stock Details")
    heading_cell.alignment = center
    heading_cell.font = Font(size=14, bold=True)

    row_index = 3  # Start table content from row 3
    section_totals = []

    for section, fieldname in slot_fields.items():
        rows = getattr(doc, fieldname)
        if not rows:
            continue

        slot_ids = [f"{section} x {i}" for i in range(1, 11)]
        data_map = {r.slot_id.strip(): r for r in rows if r.slot_id}

        # Merge and style section letter (A, B...)
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + 3, end_column=1)
        cell = ws.cell(row=row_index, column=1, value=section)
        cell.alignment = center
        cell.border = border
        cell.fill = section_fill

        # Row 1: Slot headers
        for idx, sid in enumerate(slot_ids):
            cell = ws.cell(row=row_index, column=2 + idx, value=sid)
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill
        cell = ws.cell(row=row_index, column=12, value="Total")
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill
        row_index += 1

        # Row 2: Item names
        for idx, sid in enumerate(slot_ids):
            item = data_map.get(sid).item_name if data_map.get(sid) else ""
            cell = ws.cell(row=row_index, column=2 + idx, value=item)
            cell.alignment = center
            cell.border = border
        ws.cell(row=row_index, column=12).border = border
        row_index += 1

        # Row 3: Quantities
        total = 0
        for idx, sid in enumerate(slot_ids):
            qty = float(data_map.get(sid).new_stock_qty or 0) if data_map.get(sid) else 0
            total += qty
            cell = ws.cell(row=row_index, column=2 + idx, value=qty)
            cell.alignment = center
            cell.border = border
        cell = ws.cell(row=row_index, column=12, value=total)
        cell.alignment = center
        cell.border = border
        section_totals.append(total)
        row_index += 1

        # Row 4: UOMs
        for idx, sid in enumerate(slot_ids):
            uom = data_map.get(sid).new_stockuom if data_map.get(sid) else ""
            cell = ws.cell(row=row_index, column=2 + idx, value=uom or "")
            cell.alignment = center
            cell.border = border
        ws.cell(row=row_index, column=12).border = border
        row_index += 1

    # 🟨 Grand Total row
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=11)
    for col in range(1, 12):
        cell = ws.cell(row=row_index, column=col)
        cell.border = border
        cell.fill = total_fill
        cell.alignment = center
    ws.cell(row=row_index, column=1, value="Grand Total")

    grand_total_cell = ws.cell(row=row_index, column=12, value=sum(section_totals))
    grand_total_cell.alignment = center
    grand_total_cell.fill = total_fill
    grand_total_cell.border = border

    # Column width
    for col in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # Download response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = f"Stock Register - {docname}.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"
    frappe.response["doctype"] = None
    return build_response("download")
