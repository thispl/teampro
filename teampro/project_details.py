from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles


@frappe.whitelist()
def download():
    filename = 'Project/Task Details'
    test = build_xlsx_response(filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["Project details"," "," "," "])
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='009dd1', fill_type = "solid")
    data1= get_data(args)
    for row in data1:
        ws.append(row)
   
    ws.append(["Task details"," "," "," "])
    ws.append(["Position","#Vac","#SP","#FP","#SL","#PSL"])
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='009dd1', fill_type = "solid")
    for header in ws.iter_rows(min_row=8, max_row=8, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
    data2= get_task(args)
    for row in data2:
        ws.append(row)

    ws.append(["Task plan"," "," "," "])
    ws.append(["Position","Date"," ","Count"," ",''])
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=len((get_task(args))) + 9, max_row=len((get_task(args))) + 9, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='009dd1', fill_type = "solid")
    for header in ws.iter_rows(min_row=len((get_task(args))) + 10, max_row=len((get_task(args))) + 10, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
    data3= task_plan(args)
    for row in data3:
        ws.append(row)

    ws.append(["Status of Commitment"," "," "," "])
    ws.append(["Details","Commitment"," ","Acheivement"," "," "])
    ws.append(["Position","Date","Count","Date","Count","Remarks"])
    align_center = Alignment(horizontal='center',vertical='center')
    for header in ws.iter_rows(min_row=len((get_task(args))) + len((task_plan(args))) + 11 , max_row=len((get_task(args))) + len((task_plan(args))) + 11, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='009dd1', fill_type = "solid")
    for header in ws.iter_rows(min_row=len((get_task(args))) + len((task_plan(args))) + 12 , max_row=len((get_task(args))) + len((task_plan(args))) + 12, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='FFFF00', fill_type = "solid")
    for header in ws.iter_rows(min_row=len((get_task(args))) + len((task_plan(args))) + 13 , max_row=len((get_task(args))) + len((task_plan(args))) + 13, min_col=1, max_col=6):
        for cell in header:
            cell.font = Font(bold=True)
    data4= task_plan(args)
    for row in data4:
        ws.append(row)
        
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 )
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3 )
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6 )
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3 )
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=6 )
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=3 )
    ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=6 )
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3 )
    ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=6 )
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3 )
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=6 )
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6 )
    ws.merge_cells(start_row=len((get_task(args))) + 9, start_column=1, end_row=len((get_task(args))) + 9, end_column=6 )
    ws.merge_cells(start_row=len((get_task(args))) + 10, start_column=2, end_row=len((get_task(args))) + 10, end_column=3 )
    ws.merge_cells(start_row=len((get_task(args))) + 10, start_column=4, end_row=len((get_task(args))) + 10, end_column=6 )
    ws.merge_cells(start_row=len((get_task(args))) + len((task_plan(args))) + 11, start_column=1, end_row=len((get_task(args))) + len((task_plan(args))) + 11, end_column=6 )
    ws.merge_cells(start_row=len((get_task(args))) + len((task_plan(args))) + 12, start_column=2, end_row=len((get_task(args))) + len((task_plan(args))) + 12, end_column=3 )
    ws.merge_cells(start_row=len((get_task(args))) + len((task_plan(args))) + 12, start_column=4, end_row=len((get_task(args))) + len((task_plan(args))) + 12, end_column=6 )


    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
@frappe.whitelist()
def get_data(args):
    data = []
    # row = []
    pro = frappe.get_doc("Project",args.name)
    row1 = ["Project Name",pro.project_name,"","Project ID",pro.project_id,""]
    row2 = ["Customer Name",pro.customer,"","Country",pro.territory,""]
    row3 = ["Account Type"," ","","Account Manager",pro.account_manager,""]
    row4 = ["Mode of Int",pro.mode_of_interview,"","Project Manager",pro.project_manager,""]
    row5 = ["Position",pro.task,"","Vacancy",pro.tvac,""] 
    data.append(row1)
    data.append(row2)
    data.append(row3)
    data.append(row4)
    data.append(row5)
    return data

@frappe.whitelist()
def get_task(args):
    data = []
    tasks = frappe.get_all("Task",{"project":args.name},['*'])
    for i in tasks:
        row = [i.subject,i.vac,i.sp,i.fp,i.sl,i.psl]
        data.append(row)
    return data

@frappe.whitelist()
def task_plan(args):
    data = []
    tasks = frappe.get_all("Task",{"project":args.name},['*'])
    for i in tasks:
        if i.sp > 0:
            row = [i.subject]
            data.append(row)
    return data
