import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from io import BytesIO
@frappe.whitelist()
def download():
    filename = 'Case Report'
    test = build_xlsx_response(filename)

def get_batch_data(batch=None, from_date=None, to_date=None, case=None, report=None, empname=None, empid=None):
    filters = []

    if batch:
        filters.append(['batch', '=', batch])
    if case:
        filters['case_status'] = case
    if report:
        filters['case_report'] = report
    if empname:
        filters['case_name'] = empname
    if empid:
        filters['client_employee_code'] = empid
   
    if from_date and to_date:
        filters.append(['date_of_initiating', 'between', [from_date, to_date]])
    elif from_date:
        filters.append(['date_of_initiating', '>=', from_date])
    elif to_date:
        filters.append(['date_of_initiating', '<=', to_date])

    cases = frappe.get_list('Case',
                            filters=filters,
                            fields=['name', 'case_name', 'batch',
                                    'date_of_initiating', 'end_date',
                                    'case_report', 'case_status'])
    
    filtered_cases = []
    for case in cases:
        filtered_cases.append(case)
    return filtered_cases


def make_xlsx(filename, batch, from_date, to_date):
    args = frappe.local.form_dict
    cases = get_batch_data(args.batch, args.from_date, args.to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename 
    headers = ["SI NO", "Case ID", "Name", "Batch", "Date of Initiation", "Completion Date", "Case Status", "Report Status"]
    ws.append(headers)
    align_center = Alignment(horizontal='center', vertical='center')
    for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8):
        for cell in header:
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.fill = PatternFill(fgColor='009dd1', fill_type="solid")
    count = 1
    for case in cases:
        ws.append([
            count,
            case['name'],
            case['case_name'],
            case['batch'],
            case['date_of_initiating'],
            case['end_date'],
            case['case_status'],
            case['case_report']
        ])
        count += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            cell.alignment = align_center  
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    
    ws.column_dimensions['A'].width = 5

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename, batch=None, from_date=None, to_date=None):
    xlsx_file = make_xlsx(filename, batch=batch, from_date=from_date, to_date=to_date)
    frappe.local.response.filename = filename + '.xlsx'
    frappe.local.response.filecontent = xlsx_file.getvalue()
    frappe.local.response.type = "binary"
