import frappe
from frappe.utils.csvutils import read_csv_content
from frappe.utils import get_first_day, get_last_day, format_datetime, get_url_to_form
from frappe.utils import cint
from frappe.utils.data import date_diff, now_datetime, nowdate, today, add_days
import datetime
from frappe import _
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate, nowdate
from frappe import throw, msgprint
import frappe
from datetime import date
from frappe import throw, _
from frappe.utils import getdate, today
today = date.today()
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from datetime import date
from six import BytesIO, string_types
from frappe.utils import time_diff
from frappe.utils.csvutils import read_csv_content
from erpnext.buying.doctype.purchase_order.purchase_order import update_status
from frappe.utils.file_manager import get_file
from jobpro.jobpro.doctype.closure.closure import create_sale_order
from urllib.parse import urlencode



# @frappe.whitelist()
# def skip_dn_so(doc, method):
# 	skip_delivery_note = 0

# 	for i in doc.items:
# 		maintain_stock = frappe.get_value("Item", i.item_code, "is_stock_item")
# 		if maintain_stock:
# 			skip_delivery_note = 0
# 		else:
# 			skip_delivery_note = 1

# 	doc.skip_delivery_note = skip_delivery_note
# @frappe.whitelist()
# def skip_dn_so(doc, method):
# 	skip_delivery_note = 0

# 	for i in doc.items:
# 		maintain_stock = frappe.get_value("Item", i.item_code, "is_stock_item")
# 		if not maintain_stock:
# 			skip_delivery_note = 1
# 			break
# 	else:
# 		skip_delivery_note = 0

# 	doc.skip_delivery_note = skip_delivery_note


@frappe.whitelist()
def create_project_completion_task(doc, method):
    if doc.service == 'IT-SW':
        task = frappe.db.exists('Task', {'project': doc.name, "subject": (
            "like", 'Project Completion Certificate')})
        if not task:
            task_id = frappe.new_doc('Task')
            task_id.update({
                "subject": "Project Completion Certificate",
                "customer": doc.customer,
                "project": doc.name,
                "service": 'IT-SW',
                "description":"Project Completion Certificate",
                "type":"Others",
                "priority":"Low",
                "custom_dev_team":"Others",
                "account_manager": doc.account_manager,
                "project_manager": doc.project_manager,

            })
            task_id.save(ignore_permissions=True)

@frappe.whitelist()
def fetch_start_time(doc,method):
    min_time = frappe.db.sql("""select min(from_time) as min_time from `tabTimesheet Detail` where parent='%s'""" % doc.name,as_dict=1)[0]
    if not doc.start_time:
        frappe.db.set_value("Timesheet",doc.name,"start_time",min_time['min_time'])

@frappe.whitelist()
def intimate_task_pr(task):
    taskid = frappe.get_doc('Task', task)
    message = frappe.render_template(
        "teampro/templates/intimate_task_pr.html",
        {"doc": taskid},
    )
    frappe.sendmail(
        recipients=[taskid.account_manager],
        cc=[taskid.spoc,taskid.custom_allocated_to],
        message=message,
        subject=_("Task Pending Review - Intimation"),
    )


@frappe.whitelist()
def intimate_task_completion(task):
    taskid = frappe.get_doc('Task', task)
    message = frappe.render_template(
        "teampro/templates/intimate_task_completion.html",
        {"doc": taskid},
    )
    frappe.sendmail(
        recipients=[taskid.account_manager,
                    taskid.project_manager, taskid.custom_allocated_to],
        cc=[taskid.spoc],
        message=message,
        subject=_("Task Pending Review - Intimation"),
    )


@frappe.whitelist()
def get_task_name(task):
    task_name = frappe.db.sql(
        """select name from `tabTask` where status = 'Open' and name = '%s'  """ % (task), as_dict=1)[0]
    return task_name['name']



def remove_private():
    cand = frappe.db.sql(
        """select name,irf from `tabCandidate` where irf != '' """, as_dict=True)
    print(len(cand))
    for can in cand:
        if can.irf:
            print("----------------------")
            print(can.irf)
            frappe.db.set_value("Candidate", can.name, "irf",
                                (can.irf).replace("/private", ""))
            print("------------------------")




def bulk_update_from_csv(filename):
    # below is the method to get file from Frappe File manager
    from frappe.utils.file_manager import get_file
    # Method to fetch file using get_doc and stored as _file
    _file = frappe.get_doc("File", {"file_name": filename})
    # Path in the system
    filepath = get_file(filename)
    # CSV Content stored as pps

    pps = read_csv_content(filepath[1])
    count = 0
    for pp in pps[:5]:
        mobile = pp[0]
        project = pp[1]
        position = pp[2]
        given_name = pp[3]
        customer = pp[4]
        print(mobile, project, position, given_name, customer)
        # ld = frappe.db.exists("Lead",{'name':pp[0]})
        # if ld:
        #     # items = frappe.get_all("Lead",{'name':pp[0]})
        #     # for item in items:
        #     i = frappe.get_doc('Lead',pp[0])
        #     if not i.contact_date:
        #         i.contact_date = pp[1]
        #         # i.append("supplier_items",{
        #         #     'supplier' : pp[1]
        #         # })
        #         i.save(ignore_permissions=True)
        #         frappe.db.commit()


def update_nxd():
    leads = frappe.get_list("Lead")
    for lead in leads:
        l = frappe.get_doc("Lead", lead.name)
        if not l.contact_date:
            l.contact_date = "2020-08-31 0:00:00"
            l.save(ignore_permissions=True)
            frappe.db.commit()


def update_lead():
    leads = frappe.get_list("Lead", {"organization_lead": "0"})
    for lead in leads:
        doc = frappe.get_doc("Lead", lead.name)
        doc.organization_lead = 1
        doc.save(ignore_permissions=True)
        frappe.db.commit()


@frappe.whitelist()
def add_project_id(project):
    projects = frappe.db.sql(
        """select project_id from `tabProject` where project_id is not null order by creation""", as_dict=True)
    project_id = projects[-1].project_id
    return 'PRO' + str(int(project_id.strip('PRO'))+1)


@frappe.whitelist()
def update_task_fields(project):
    tasks = frappe.get_all('Task', {'project': project}, [
                           'name', 'project_manager', 'account_manager', 'service'])
    proj = frappe.get_doc('Project', {'name': project})
    for t in tasks:
        if t.project_manager != proj.project_manager:
            frappe.db.set_value(
                'Task', t.name, 'project_manager', proj.project_manager)
        if t.account_manager != proj.account_manager:
            frappe.db.set_value(
                'Task', t.name, 'account_manager', proj.account_manager)
        if t.service != proj.service:
            frappe.db.set_value('Task', t.name, 'service', proj.service)


@frappe.whitelist()
def opportunity_send_mail(self, method):
    if self.service == 'TGT':
        link = get_url_to_form("Opportunity", self.name)
        subject = 'Reg.Opportunity- %s' % self.name
        content = """Dear Mam<br>Kindly find the new Opportunity.
        Click on <a href='%s'>View</a> to open the opportunity.<br>Thanks & Regards,<br>ERP
        """ % link
        frappe.sendmail(recipients=['saraswathi.p@groupteampro.com'],
                        subject=subject,
                        message=content)


@frappe.whitelist()
def checkin_alerts():
    # yesterday = add_days(today(),-2)
    yesterday = add_days(datetime.datetime.today(), -17).date()
    current_month = yesterday.month
    late_checkins = frappe.db.sql(
        "select employee from `tabEmployee Checkin` where date(time) = '%s' and time(time) > '09:35:00' " % (yesterday), as_dict=True)
    if late_checkins:
        for emp in late_checkins:
            apid = frappe.db.exists('Attendance Permission', {
                                    'employee': emp.employee, 'month': current_month})
            print(emp)
            if apid:
                pc = frappe.get_value(
                    'Attendance Permission', apid, 'permission_count')
                if pc < 3:
                    # adding permission
                    ap = frappe.get_doc('Attendance Permission', apid)
                    ap.update({
                        'permission_count': pc + 1
                    })
                    ap.save(ignore_permissions=True)
                    frappe.db.commit()

                    lp = frappe.db.exists('Leave Application', {
                                          'employee': emp.employee, 'from_date': yesterday, 'to_date': yesterday})
                    if lap:
                        pass
                else:
                    lw = frappe.new_doc('Leave Application')
                    lw.update({
                        'employee': emp.employee,
                        'leave_type': "Leave Without Pay",
                        'from_date': yesterday,
                        'to_date': yesterday,
                        'description': 'Auto marked as LWP due to Exceed Monthly Permission Limit'
                    })
                    lw.save(ignore_permissions=True)
                    frappe.db.commit()

            else:
                ap = frappe.new_doc('Attendance Permission')
                ap.update({
                    'employee': emp.employee,
                    'month': current_month,
                    'permission_count': 1
                })
                ap.save(ignore_permissions=True)
                frappe.db.commit()


@frappe.whitelist()
def lwp_alert():
    yesterday = add_days(today(), -19)
    employees = frappe.get_all('Employee', {'status': 'Active'}, [
                               'name', 'employee_name', 'employee'])
    for emp in employees:
        lc = frappe.db.sql("select employee from `tabEmployee Checkin` where date(time) = '%s' and employee = '%s'" % (
            yesterday, emp.name), as_dict=True)
        lap = frappe.db.exists('Leave Application', {
                               'employee': emp.employee, 'from_date': yesterday, 'to_date': yesterday})
        if lap:
            pass
        if len(lc) < 2:
            lp = frappe.new_doc('Leave Application')
            lp.update({
                'employee': emp.employee,
                'leave_type': "Leave Without Pay",
                'from_date': yesterday,
                'to_date': yesterday,
                'description': 'Auto marked as LWP due to not responding Miss Punch Alert'
            })
            lp.save()
            frappe.db.commit()


@frappe.whitelist()
def update_task_count(doc, method):
    if doc.task:
        submit_spoc = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for': 'Submit(SPOC)'}) or 0
        submit_client = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for': 'Submitted(Client)'}) or 0
        psl = frappe.db.count('Candidate', {'task': doc.task, 'pending_for': (
            'in', ('Client Offered', 'Proposed PSL'))}) or 0
        shortlisted = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for':'Shortlisted'}) or 0
        # linedup = frappe.db.count(
        # 	'Candidate', {'task': doc.task, 'pending_for': 'Linedup'}) or 0
        linedup = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for':('in', ('Linedup','Linedup Confirmed'))}) or 0
        interviewed = frappe.db.count(
            'Candidate', {'task': doc.task, 'pending_for': 'Interviewed'}) or 0
        result_pending =frappe.db.count('Candidate',{'task':doc.task,'pending_for':'Result Pending'}) or 0
        # submit_interviewed=(submitted + interviewed)
        frappe.db.set_value('Task', doc.task, 'psl', psl)
        frappe.db.set_value('Task', doc.task, 'fp',(submit_spoc + interviewed + submit_client))
        frappe.db.set_value('Task',doc.task,'custom_rp',result_pending)
        frappe.db.set_value('Task', doc.task, 'sl', shortlisted)
        frappe.db.set_value('Task', doc.task, 'custom_lp',linedup)

        task_status = frappe.db.get_value('Task', doc.task, 'status')

        if task_status in ('Completed', 'Cancelled'):
            # if pps == 0:
            frappe.db.set_value('Task', doc.task, 'sp', 0)

        else:
            vac = frappe.db.get_value('Task', doc.task, 'vac')
            prop = frappe.db.get_value('Task', doc.task, 'prop')
            pps = (vac - psl) * prop - (submit_spoc + submit_client+
                                        interviewed + shortlisted +linedup)
            frappe.db.set_value('Task', doc.task, 'sp', pps)

def update_candidate_tcount():
    tasks = frappe.get_all('Task',{'service':('in',('REC-I', 'REC-D'))})
    for task in tasks:
        # task = task.name
#         print(task)
        shortlisted = frappe.db.count('Candidate',{'task':task.name,'pending_for':'Shortlisted'}) or 0
        linedup = frappe.db.count('Candidate',{'task':task.name,'pending_for':'Linedup'}) or 0
        # interviewed = frappe.db.count('Candidate',{'task':task,'pending_for':'Interviewed'}) or 0
        print(task.name)
        print(shortlisted)
        print(linedup)
        frappe.db.sql(""" update `tabTask` set sl='%s' 
                    where name = '%s' """%(shortlisted,task.name),as_dict=True)
        # frappe.db.update('Task',task,'custom_lp',linedup)
        # frappe.db.update('Task',task,'fp',submitted + interviewed)
        # frappe.db.update('Task',task,'sl',shortlisted)

        # print([psl,submitted,interviewed,shortlisted,linedup])

        # task_status = frappe.db.get_value('Task',task,'status')
        # if task_status in ('Completed','Cancelled'):
        #     frappe.db.update('Task',task,'sp',0)
        # else:
        #     vac = frappe.db.get_value('Task',task,'vac')
        #     prop = frappe.db.get_value('Task',task,'prop')
        #     pps = (vac - psl) * prop - (submitted + interviewed + shortlisted + linedup)
        #     frappe.db.update('Task',task,'sp',pps)


def test_hook():
    frappe.log_error(title='hi', message='ok')


# def update_submission_date(doc, method):
# 	frappe.db.set_value("Sales Invoice", doc.name, "posting_date", nowdate())
# 	frappe.db.commit()


@frappe.whitelist()
def attendance():
    att = frappe.db.sql("""update `tabEmployee Checkin` set skip_auto_attendance = 0 where date(time) between '2023-09-01' and '2023-09-30'   """)
    print(att)

@frappe.whitelist()
def get_delivery_note(doc, method):
    # if not frappe.db.exist("Sales Invoice",doc.name)
    # nd= frappe.new_doc("Sales Invoice")
    si = frappe.get_doc("Sales Invoice", {"name": doc.sales_invoice_no})
    si.delivery_note_no = doc.name
    si.save(ignore_permissions=True)


@frappe.whitelist()
def set_customer_group():
    sales = frappe.get_all('Sales Invoice', ["*"])
    for i in sales:
        customer_group = frappe.get_value(
            'Customer', {'name': i.customer}, ['customer_group'])
        frappe.db.set_value('Sales Invoice', i.name,
                            'customer_group', customer_group)


@frappe.whitelist()
def date(time):
    print('Hi')
    log_date = datetime.strptime(str(time), '%Y-%m-%d %H:%M:%S')
    x = log_date.strftime('%Y-%m-%d')

    return x


@frappe.whitelist()
def mark_bulk_drop():
    closure = frappe.db.get_all('Closure', {
                                'customer': 'Arabian Castles for General Contracting Co., LLC', 'nationality': 'Nepali'}, ['*'])
    print(closure)
    for ca in closure:
        frappe.db.set_value('Closure', ca.name, 'status', 'Dropped')


@frappe.whitelist()
def organization_name():
    opportunity = frappe.db.get_all('Opportunity', ["*"])
    for i in opportunity:
        print(i.title)
        frappe.db.set_value("Opportunity", i.name,
                            "organization_name", i.title)


@frappe.whitelist()
def address_html():
    lead = frappe.db.get_all('Lead', ["*"])
    for i in lead:

        print(i.address)
        frappe.db.set_value("Lead", i.name, "address", i.address)


@frappe.whitelist()
def late():
    late = frappe.db.sql("""delete from `tabLate Penalty` """)
    print(late)


# @frappe.whitelist()
# def update_attendance(doc, method):
# 	if doc.half_day != 1 :
# 		attendance_req = frappe.get_all("Attendance", {'attendance_date': (
# 			'between', (doc.from_date, doc.to_date)), 'employee_name': doc.employee_name}, ['name'])
# 		for att in attendance_req:
# 			if att.name:
# 				at = frappe.get_doc("Attendance", att.name)
# 				at.delete()


@frappe.whitelist()
def get_child(parent):
    if parent:
        so = frappe.get_doc("Candidate", parent)
        if so.table_28:
            return so.table_28


@frappe.whitelist()
def update_location_route():
    cust = frappe.db.get_all('Customer',{'customer_group':'Retail Shops'},['*'])
    for cus in cust:
        if cus.location:
            sales_order = frappe.db.get_all('Sales Order', {'customer': cus.name,'company':'TEAMPRO Food Products'}, ['*'])
            for sales in sales_order:
                frappe.db.set_value('Sales Order',sales.name,'location',cus.location)

@frappe.whitelist()
def update_user():
    user_id = frappe.get_value('Employee',{'employee':'TI00005'},['user_id'])
    print (user_id)
    hod = frappe.get_value('User',{'email':user_id},['name'])
    role = "HOD"
    hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
    if hod:
        print("HI")
    else:
        hod = frappe.get_doc('Has Role',{'parent':user_id},['name'])
        hod.role = "HOD"
        print("HII")

@frappe.whitelist()
def get_item_codes(company_name):
    item_codes = []
    item_defaults = frappe.get_all('Item Default',
        filters={'company': ('in', frappe.get_all('Item Default Company',
            filters={'company': company_name},
            pluck='parent'
        ))},
        pluck='item_code'
    )
    item_codes.extend(item_defaults)
    return item_codes
# update route number 
@frappe.whitelist()
def update_route_no():
    route = frappe.db.get_all('Customer',{'customer_group':'Retail Shops'},['*'])
    for ro in route :
        if ro.route_no:
            sales_invoice = frappe.db.get_all('Sales Invoice', {'customer': ro.name}, ['*'])
            for sales in sales_invoice:
                print(ro.name)
                frappe.db.set_value('Sales Invoice', sales.name,'route_no',ro.route_no)

# @frappe.whitelist()
# def get_against_so(doc,method):
#     if doc.is_return == 1 and doc.update_stock == 0:
#         sales_order = frappe.db.sql("""select `tabSales Invoice Item`.sales_order from `tabSales Invoice` left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent where `tabSales Invoice`.name = '%s' """%(doc.name),as_dict=True)[0]
#         dn = frappe.db.sql("""select parent from `tabDelivery Note Item` where `tabDelivery Note Item`.against_sales_order = '%s' """%(sales_order['sales_order']),as_dict=True)[0]
#         copy_dn = frappe.get_doc("Delivery Note",dn["parent"])
#         dnr = frappe.copy_doc(copy_dn)
#         dnr.is_return = 1
#         # dnr.return_against = dn["parent"]
#         dnr.set("items",[])
#         for j in doc.items:
#             dnr.append("items",{
#                 "item_code":j.item_code,
#                 "item_name":j.item_name,
#                 "description":j.description,
#                 "qty":j.qty,
#                 "cost_center":j.cost_center,
#                 "uom":j.uom,
#                 "stock_uom":j.uom,
#                 "rate":j.rate,
#                 "conversion_factor":1,
#                 "base_rate":j.base_rate,
#                 "amount":j.amount,
#                 "base_amount":j.base_amount,
#                 "expense_account": "Cost of Goods Sold - TFP"
#             })
#         dn_list = frappe.db.sql("""select parent from `tabDelivery Note Item` where `tabDelivery Note Item`.against_sales_order = '%s' """%(sales_order['sales_order']),as_dict=True)
#         for k in dn_list:
#             do = frappe.get_doc("Delivery Note",k.parent)
#             if do.docstatus == 1:
#                 dnr.append("return_against_dn",{
#                     "delivery_note":k.parent,
#                 })
#         dnr.save(ignore_permissions = True)
#         dnr.submit()

@frappe.whitelist()
def address(lead):
    if frappe.db.exists('Address',{'address_title':lead}):
        ad = frappe.get_doc('Address',{'address_title':lead})
        ad.address_title = lead
        ad.address_type = frappe.db.get_value('Lead',{'name':lead},['address_type'])
        ad.address_line1 = frappe.db.get_value('Lead',{'name':lead},['address_line_1'])
        ad.address_line2 = frappe.db.get_value('Lead',{'name':lead},['address_line_2'])
        ad.city = frappe.db.get_value('Lead',{'name':lead},['city_town'])
        ad.state = frappe.db.get_value('Lead',{'name':lead},['state__province'])
        ad.country = frappe.db.get_value('Lead',{'name':lead},['country__'])
        ad.pincode = frappe.db.get_value('Lead',{'name':lead},['postal_code'])
        ad.save(ignore_permissions=True)
        frappe.db.commit()
    else:
        ad = frappe.new_doc('Address')
        ad.address_title = lead
        ad.address_type = frappe.db.get_value('Lead',{'name':lead},['address_type'])
        ad.address_line1 = frappe.db.get_value('Lead',{'name':lead},['address_line_1'])
        ad.address_line2 = frappe.db.get_value('Lead',{'name':lead},['address_line_2'])
        ad.city = frappe.db.get_value('Lead',{'name':lead},['city_town'])
        ad.state = frappe.db.get_value('Lead',{'name':lead},['state__province'])
        ad.country = frappe.db.get_value('Lead',{'name':lead},['country__'])
        ad.pincode = frappe.db.get_value('Lead',{'name':lead},['postal_code'])
        ad.save(ignore_permissions=True)
        frappe.db.commit()


@frappe.whitelist()
def contact(lead):
    co = frappe.db.sql("""select * from `tabLead Contacts` where `tabLead Contacts`.parent = '%s' """%(lead),as_dict=True)
    for c in co:
        if frappe.db.exists('Contact',{'first_name':c.person_name,'mobile_no':c.mobile}):
            cn = frappe.get_doc('Contact',{'first_name':c.person_name,'mobile_no':c.mobile})
            cn.first_name = c.person_name
            cn.company_name = frappe.db.get_value('Lead',{'name':lead},['company_name'])
            cn.lead = lead
            for i in cn.email_ids:
                i.append("email_ids",{
                    "email_id":c.email_id,
                    "is_primary":c.is_primaryemail
                })
            for i in cn.phone_nos:
                i.append("phone_nos",{
                    "phone":c.mobile,
                    "is_primary_phone":c.is_primary,
                    "has_whatsapp":c.has_whatsapp
                })
            cn.save(ignore_permissions=True)
            frappe.db.commit()
        else:
            cn = frappe.new_doc('Contact')
            cn.first_name = c.person_name
            cn.company_name = frappe.db.get_value('Lead',{'name':lead},['company_name'])
            cn.lead = lead
            for i in cn.email_ids:
                i.append("email_ids",{
                    "email_id":c.email_id,
                    "is_primary":c.is_primaryemail
                })
            for i in cn.phone_nos:
                i.append("phone_nos",{
                    "phone":c.mobile,
                    "is_primary_phone":c.is_primary,
                    "has_whatsapp":c.has_whatsapp
                })
            cn.save(ignore_permissions=True)
            frappe.db.commit()
    
    
@frappe.whitelist()
def calc_cut_off_prize(doc,method):
    for f in doc.items:
        tfp_item = frappe.db.sql("""select tfp from `tabItem` where name = '%s' """%(f.item_code),as_dict=1)[0]
        tfp = tfp_item['tfp']
        if tfp == 1:
            price_list = frappe.db.sql("""select price_list_rate from `tabItem Price` where price_list = 'Cut Off Price' and item_code = '%s' """%(f.item_code),as_dict=1)
            for p in price_list:
                if f.uom == 'Gram':
                    item_price = (p.price_list_rate / 1000)
                    item_rate = round((item_price),2)
                    if f.rate < item_rate:
                        frappe.throw(_(' %s Rate is lesser than Cut Off Price')%(f.item_name))
                elif f.uom == 'Kg':
                    if f.rate < p.price_list_rate:
                        frappe.throw(_(' %s Rate is lesser than Cut Off Price')%(f.item_name))	

@frappe.whitelist()
def calc_cost_prize(doc,method):
    if doc.workflow_state!="Approved":
        for f in doc.items:
            tfp_item = frappe.db.sql("""select tfp from `tabItem` where name = '%s' """%(f.item_code),as_dict=1)[0]
            tfp = tfp_item['tfp']
            if tfp == 1:
                price_list = frappe.db.sql("""select price_list_rate from `tabItem Price` where price_list = 'Cost Price TFP' and item_code = '%s' """%(f.item_code),as_dict=1)
                for p in price_list:
                    if f.uom == 'Gram':
                        item_price = (p.price_list_rate / 1000)
                        item_rate = round((item_price),2)
                        if f.rate > item_rate:
                            frappe.msgprint(_(' %s Rate is Greater than Cost Price')%(f.item_name))
                    elif f.uom == 'Kg':
                        if f.rate > p.price_list_rate:
                            frappe.msgprint(_(' %s Rate is Greater than Cost Price')%(f.item_name))	

    
@frappe.whitelist()
def file_list(candidate):
    url_ls =[]
    file_list = frappe.get_all("File",{"attached_to_name": candidate},['file_url'])
    for a in file_list:
        url_ls.append(a.file_url)
    return url_ls


import requests
import io
from frappe import _
import PyPDF2
from frappe.utils.file_manager import get_file
from urllib.parse import urljoin

@frappe.whitelist()
def merge_and_open_pdf(candidate):
    url_list = []
    file_list = frappe.get_all("File", {"attached_to_name": candidate}, ["file_url"])

    merger = PyPDF2.PdfFileMerger()
    for file in file_list:
        base_url = frappe.utils.get_url()
        file_urls = frappe.utils.get_url(file.file_url)
        file_url = urljoin(base_url, file_urls)
        if file_url:
            try:
                response = requests.get(file_url)
                if response.headers.get("content-type") != "application/pdf":
                    # Skip non-PDF files
                    continue

                temp_file = io.BytesIO(response.content)
                merger.append(temp_file)
            except Exception as e:
                # Log the specific error message
                frappe.log_error(f"Error merging PDF: {str(e)}", _("Merge and Open PDF Error"))
                continue
        else:
            # Failed to retrieve the file, log the error
            frappe.log_error(f"Failed to retrieve file: {file_url}", _("Merge and Open PDF Error"))

    merged_pdf = io.BytesIO()
    merger.write(merged_pdf)
    merger.close()

    if len(merged_pdf.getvalue()) == 0:
        # Merged PDF is empty or invalid
        frappe.log_error("Merged PDF is empty or invalid", _("Merge and Open PDF Error"))
        return None

    return {
        "content": merged_pdf.getvalue(),
        "file_name": "merged_pdf.pdf",
        "file_type": "application/pdf"
    }



def is_valid_pdf(file):
    try:
        pdf_reader = PyPDF2.PdfFileReader(file)
        return pdf_reader.numPages > 0
    except PyPDF2.utils.PdfReadError:
        return False


import json
@frappe.whitelist()
def get_supporting_docs(selected_docs):
    selected_docs = json.loads(selected_docs)
    file_list = []
    for s in selected_docs:
        file_name = frappe.get_value("File", {"file_url": s},"name")
        file_list.append(file_name)
    return file_list


from datetime import datetime
from frappe import throw, _

def validate_date(doc, method):
    current_date = datetime.now().date()
    date_str = doc.posting_date
    if isinstance(date_str, str):
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        date_obj =date_str
    if date_obj < current_date:
        throw(_("Submitting the Purchase Invoice on back date will have impact on the GST .Please click on the 'Edit Posting Date' and change the Date to Today."))


def validate_date_salesinvoice(doc, method):
    current_date = datetime.now().date()
    date_str = doc.posting_date
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    if date_obj < current_date:
        throw(_("Submitting the Sales Invoice on a past date will have an impact on the GST. Please click on 'Edit Posting Date' and change the date to today."))


# def validate_date_salesinvoice(doc, method):
# 	today = date.today()
# 	date_str = doc.due_date
# 	date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
# 	if date_obj < today:
# 		throw(_("Submitting the Sales Invoice on back date will have impact on the GST .Please click on the 'Edit Posting Date' and change the Date to Today."))

@frappe.whitelist()
def update_pi():
    doc = frappe.db.sql(""" update `tabPurchase Invoice` set workflow_state = "Cancelled" where name = "ACC-PINV-2023-00076"  """)
    print(doc)


@frappe.whitelist()
def calac_invoice_discount(doc,method):
    discount = 0
    if doc.tfp_items:
        for i in doc.tfp_items:
            discount += (i.discount_amount * i.packet) * i.qty__packet
        frappe.db.set_value("Sales Invoice",doc.name,"tfp_discount",discount)

@frappe.whitelist()
def child_table_calc(doc,method):
    if doc.tfp_items:
        for i in doc.tfp_items:
            i.qty = i.packet * i.qty__packet
            i.rate_packet = round(i.qty__packet * (i.price_list_rate))
            i.special_rate = round(i.rate_packet - (i.rate_packet * (i.discount_percentage/100)))
            i.amount = round(i.special_rate) * i.packet
            for j in doc.items:
                if i.item_code == j.item_code:
                    j.amount = i.amount
                    j.delivery_date = doc.delivery_date


# @frappe.whitelist()
# def return_detailed_ts(timesheet):
    # timesheet = frappe.db.sql(""" select `tabTimesheet Detail`.task,`tabTimesheet Detail`.subject,`tabTimesheet Detail`.project,sum(`tabTimesheet Detail`.hours) as hours, GROUP_CONCAT(`tabTimesheet Detail`.description separator ', ') as description from `tabTimesheet`
    # left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent where `tabTimesheet`.name = '%s' group by `tabTimesheet Detail`.task"""%(timesheet),as_dict = 1)
#     return timesheet
    
# @frappe.whitelist()
# def return_detailed_ts(timesheet):
#     task = frappe.db.sql(""" select `tabTimesheet Detail`.task,`tabTimesheet Detail`.subject,`tabTimesheet Detail`.project,sum(`tabTimesheet Detail`.hours) as hours, GROUP_CONCAT(`tabTimesheet Detail`.description separator ', ') as description from `tabTimesheet`
#     left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent where `tabTimesheet Detail`.task is not null and `tabTimesheet`.name = '%s' group by `tabTimesheet Detail`.task"""%(timesheet),as_dict = 1)
#     # if time
#     meeting = frappe.db.sql(""" select `tabTimesheet Detail`.custom_meeting,`tabTimesheet Detail`.custom_subject_meeting, sum(`tabTimesheet Detail`.hours) as hours from `tabTimesheet`
#     left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent where `tabTimesheet Detail`.custom_meeting is not null and `tabTimesheet Detail`.task`tabTimesheet`.name = '%s' group by `tabTimesheet Detail`.custom_meeting"""%(timesheet),as_dict = 1)
#     issue = frappe.db.sql(""" select `tabTimesheet Detail`.custom_issue,`tabTimesheet Detail`.custom_subject_issue,`tabTimesheet Detail`.project,sum(`tabTimesheet Detail`.hours) as hours from `tabTimesheet`
#     left join `tabTimesheet Detail` on `tabTimesheet`.name = `tabTimesheet Detail`.parent where tabTimesheet Detail`.custom_issue is not null and `tabTimesheet`.name = '%s' group by `tabTimesheet Detail`.custom_issue"""%(timesheet),as_dict = 1)
#     return task,meeting,issue
@frappe.whitelist()
def return_detailed_ts(timesheet):
    # Parameterized query for tasks
    tdoc=frappe.get_doc('Timesheet',timesheet)
    alloc=frappe.db.get_value('Employee',{'name':tdoc.employee},['user_id'])
    task_query ="""
        SELECT `tabTimesheet Detail`.task,
               `tabTimesheet Detail`.subject,
               `tabTimesheet Detail`.project,
               SUM(`tabTimesheet Detail`.hours) AS hours,
               `tabTimesheet Detail`.task_status,
               GROUP_CONCAT(`tabTimesheet Detail`.description SEPARATOR ', ') AS description
        FROM `tabTimesheet`
        LEFT JOIN `tabTimesheet Detail` 
        ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
        WHERE `tabTimesheet Detail`.task IS NOT NULL 
          AND `tabTimesheet`.name = %s
        GROUP BY `tabTimesheet Detail`.task
    """
    task = frappe.db.sql(task_query, (timesheet,), as_dict=True)
    task_list = [t["task"] for t in task]  

    emp = frappe.db.get_value('Timesheet', {'name': timesheet}, 'employee')
    start_date = frappe.db.get_value('Timesheet', {'name': timesheet}, 'start_date')
    alloc = frappe.db.get_value('Employee', {'employee': emp}, 'user_id')

    additional_task_query = """
        SELECT 
            task.name,
            task.subject,
            task.project,
            task.status
        FROM `tabTask` task
        WHERE task.custom_allocated_to = %s
        AND task.custom_production_date = %s
        AND task.name NOT IN (
            SELECT `tabTimesheet Detail`.task
            FROM `tabTimesheet`
            LEFT JOIN `tabTimesheet Detail`
            ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
            WHERE `tabTimesheet Detail`.task IS NOT NULL
                AND `tabTimesheet`.name = %s
        )
    """
    additional_tasks = frappe.db.sql(additional_task_query, (alloc, start_date, timesheet), as_dict=True)

    for t in additional_tasks:
        tstatus=frappe.db.get_all('Task',{'name':t['name']},['status'])
        task.append({
            'task': t['name'],
            'subject': t.get('subject', ''),
            'project': t.get('project', ''),
            'hours': 0,
            'task_status': t.get('status', ''),
            'description': ''
        })
    # Parameterized query for meetings
    meeting_query = """
        SELECT `tabTimesheet Detail`.custom_meeting,
               `tabTimesheet Detail`.custom_subject_meeting,
               `tabTimesheet Detail`.project,
               SUM(`tabTimesheet Detail`.hours) AS hours
        FROM `tabTimesheet`
        LEFT JOIN `tabTimesheet Detail`
        ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
        WHERE `tabTimesheet Detail`.custom_meeting IS NOT NULL 
          AND `tabTimesheet`.name = %s
        GROUP BY `tabTimesheet Detail`.custom_meeting
    """
    meeting = frappe.db.sql(meeting_query, (timesheet,), as_dict=True)
    
    # Parameterized query for issues
    issue_query = """
        SELECT `tabTimesheet Detail`.custom_issue,
               `tabTimesheet Detail`.custom_subject_issue,
               `tabTimesheet Detail`.project,
               SUM(`tabTimesheet Detail`.hours) AS hours
        FROM `tabTimesheet`
        LEFT JOIN `tabTimesheet Detail`
        ON `tabTimesheet`.name = `tabTimesheet Detail`.parent
        WHERE `tabTimesheet Detail`.custom_issue IS NOT NULL 
          AND `tabTimesheet`.name = %s
        GROUP BY `tabTimesheet Detail`.custom_issue
    """
    issue = frappe.db.sql(issue_query, (timesheet,), as_dict=True)
    # task_list = [t["task"] for t in task]
    # emp=frappe.db.get_value('Timesheet',{'name':timesheet},['employee'])
    # sd=frappe.db.get_value('Timesheet',{'name':timesheet},['start_date'])
    # alloc=frappe.db.get_value('Employee',{'employee':emp},['user_id'])
    # tasks=frappe.db.get_all('Task',{'custom_allocated_to':alloc,'custom_production_date':sd},['name'])
    # for t in tasks:
    #     if t.name not in task_list:
            
    

    return task, meeting, issue





#For downloading Timesheet status as excel
@frappe.whitelist()
def make_time_sheet():
    args = frappe.local.form_dict
    filename = args.name
    xlsx_file = build_xlsx_response(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def make_xlsx(filename, sheet_name=None, column_widths=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or 'Sheet1'
    column_widths = column_widths or []
    
    # Fetching document data
    doc = frappe.get_doc("Timesheet", filename)
    if not doc:
        frappe.throw("Timesheet not found.")

    # Fetching the Purchase Order short code
    cb = frappe.db.get_value("Purchase Order", doc.name, 'short_code')
    
    # Adding headers
    ws.append(["Task", "Subject", "Project Name", "CB", "Status", "TU", "Description"])
    
    # Adding data rows
    for i in doc.timesheet_summary:
        ws.append([i.task, i.subject, i.project, cb, i.status, round(i.tu, 2), i.description])

    # Save to BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)  # Move to the beginning of the file
    return xlsx_file

def build_xlsx_response(filename):
    return make_xlsx(filename)

import bleach
from frappe.utils import strip
@frappe.whitelist()
def make_minutes_for_mom_points():
    args = frappe.local.form_dict
    filename = args.name
    test = build_xlsx_response_mom(filename)

def make_xlsx_mom(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    doc = frappe.get_doc("Meeting",args.name)
    if doc:
        ws.append(["Description","Action","Task"])
        for i in doc.minutes:
            description_without_html = bleach.clean(i.description, tags=[], strip=True)
            ws.append([description_without_html, i.custom_action, i.custom_id])
            # ws.append([i.description,i.custom_action,i.custom_id])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response_mom(filename):
    xlsx_file = make_xlsx_mom(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 	

@frappe.whitelist()
def get_all_quot(doc,method):
    if doc.tfp_items:
        for i in doc.tfp_items:
            i.qty = i.packet * i.qty__packet
            i.rate_packet = round(i.qty__packet * (i.price_list_rate))
            i.special_rate = round(i.rate_packet - (i.rate_packet * (i.discount_percentage/100)))
            i.amount = round(i.special_rate) * i.packet
            for j in doc.items:
                if i.item_code == j.item_code:
                    j.amount = i.amount

            
@frappe.whitelist()
def get_all_so(name):
    so = frappe.get_doc("Sales Order",name)
    return so.tfp_items

@frappe.whitelist()
def get_all_quote(name):
    quote = frappe.get_doc("Quotation",name)
    return quote.tfp_items

@frappe.whitelist()
def update_pi():
    # edu=frappe.db.get_all("Education Checks",{'epi_affiliated_university_name':['!=','']},['epi_affiliated_university_name'])
    # for e in edu:
    #     if not frappe.db.exists('University',{'name':e.name}):
            # univ=frappe.new_doc('University')
            # univ.university=e.name
            # univ.save(ignore_permissions=True)
   uni=frappe.db.get_all('University',{'name':['!=','']},['name'])
   for u in uni:
       doc=frappe.delete_doc('University',u.name)
       doc.delete()
@frappe.whitelist()
def update_custodian(doc,method):
    if doc.status == "Left":
        asset = frappe.get_all("Asset",{"custodian":doc.name},["name"])
        for i in asset:
            cust = frappe.get_doc("Asset",i.name)
            cust.custodian = ''
            cust.custodian_name = ''
            cust.department = ''
            cust.save(ignore_permissions = True)

@frappe.whitelist()
def create_food_count():
    from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
    holiday_list_name = 'TEAMPRO 2023'
    start_date = getdate(today())
    if not is_holiday(holiday_list_name, start_date):
        emp = ["TI00149","TC00042"]
        for i in emp:
            if not frappe.db.exists("Food Count",{'employee':i,'date':nowdate()}):
                doc = frappe.new_doc("Food Count")
                doc.employee = i
                doc.department="IT"
                doc.food_type="Veg"
                doc.date = nowdate()
                doc.save(ignore_permissions=True)
    
@frappe.whitelist()
def delete_document(name,checks_list):
    checks_list = json.loads(checks_list)
    nam = frappe.get_doc("Case",name)
    nam.delete()
    for i in checks_list:
        doc = frappe.get_doc(i["checks"],i["check_id"])
        doc.delete()


@frappe.whitelist()
def update_batch_status(doc,method):
    if doc.get("batch"):
        batch_doc = frappe.get_doc("Batch", doc.get("batch"))
        batch_doc.batch_status = "Proposed SO"
        batch_doc.save()

@frappe.whitelist()
def sales_order_batch(doc, method):
    sales_order = frappe.get_doc("Sales Order", doc.sales_order)
    if hasattr(sales_order, "batch") and sales_order.batch:
        doc.batch = sales_order.batch
        batch_doc = frappe.get_doc("Batch", doc.batch)
        batch_doc.billing_status = "Billed"
        batch_doc.save()

@frappe.whitelist()
def get_po_qty(item,company):
    new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order` left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' """ % (item,company), as_dict=True)[0]
    if not new_po['qty']:
        new_po['qty'] = 0
    if not new_po['d_qty']:
        new_po['d_qty'] = 0
    ppoc_total = new_po['qty'] - new_po['d_qty']
    return ppoc_total


@frappe.whitelist()
def get_salesorder_qty(item,company):
    new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty *`tabSales Order Item`.conversion_factor) as qty,sum(`tabSales Order Item`.delivered_qty *`tabSales Order Item`.conversion_factor) as d_qty from `tabSales Order` left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' and status != 'Closed' """ % (item,company), as_dict=True)[0]
    if not new_so['qty']:
        new_so['qty'] = 0
    if not new_so['d_qty']:
        new_so['d_qty'] = 0
    frappe.errprint([new_so['qty'], new_so['d_qty']])
    del_total = new_so['qty'] - new_so['d_qty']
    return del_total


@frappe.whitelist()
def get_values(name):
    ch=[]
    # list=["Generate Report","Generate Report with Insuff","Report Completed"]
    case = frappe.db.get_all("Case",{"batch":name,"case_status":["in",["Generate Report","Generate Report with Insuff","Completed"]]},["name"])
    for c in case:
        ch.append({
            "case_id":c.name,

        })
        # row = len(ch)
        # frappe.db.set_value("Batch",name,"rows",row)
    return ch

@frappe.whitelist()
def set_values(name):
    row = frappe.db.count("Case",{"batch":name,"case_status":["in",["Generate Report","Generate Report with Insuff","Completed"]]})
    billed = frappe.db.count("Case",{"batch": name, "billing_status": ["in", ["Billed", "Partially Billed"]]})
    pending_bill = row- billed
    return row,billed,pending_bill

    
@frappe.whitelist()
def create_so(doctype,batch,case_id):
    batch=frappe.get_doc("Batch",batch)
    case=frappe.get_doc("Case",case_id)
    item = frappe.new_doc("Item")
    item.item_code = case_id
    item.item_name= case.case_name
    item.item_group = "BCS Cases"
    item.item_group_code= "BCS"
    item.stock_uom = "Nos"
    item.qty = "1"
    item.gst_hsn_code = '998521'
    item.is_stock_item = "0"
    item.include_item_in_manufacturing = "0"
    dict_list = []
    dict_list.append(frappe._dict({"item_tax_template":"GST 18% - THIS","tax_category":"Tamil Nadu","valid_from": today()}))
    dict_list.append(frappe._dict({"item_tax_template":"I - GST @ 18% - THIS","tax_category":"Inter State","valid_from": today()}))
    for i in dict_list:
        item.append("taxes", {
            "item_tax_template":i.item_tax_template,
            "tax_category":i.tax_category,
            "valid_from": i.valid_from
            })
    item.append("item_defaults", {
                "company": "TeamPRO HR & IT Services Pvt. Ltd.",
                "buying_cost_center":"Main - THIS",
                "selling_cost_center":"Main - THIS",
                "income_account":"Sales - THIS",
                "expense_account":"Cost of Goods Sold - THIS"
            })
    item.insert()
    item.save(ignore_permissions=True)

    so = frappe.new_doc("Sales Order")
    so.company = "TEAMPRO HR & IT Services Pvt. Ltd."
    so.customer = batch.customer
    so.service = "BCS"
    so.order_type = "Sales"
    so.delivery_date = today()  
    so.transaction_date = today() 
    so.po_no = batch.customers_purchase_order
    # so.delivery_manager = batch.delivery_manager
    so.posa_notes:case.case_report
    so.tc_name="Account Details - THIS"
    rate = frappe.db.get_value("Check Package", {"name":batch.check_package},["total_sp"])
    
    so.append('items', {
        'item_code': case_id,
        'item_name':case.case_name,
        'description':batch.name,
        'qty':1,
        'posa_notes':case.case_report,
        'rate':rate,
        })
    case_status=case.case_status
    billing_status = case.billing_status
    if case_status =="Generate Report" or case_status=="Completed":
        frappe.set_value("Case",case_id,"billing_status","Billed")
    if case_status=="Generate Report with insuff":
        frappe.set_value("Case",case_id,"billing_status","Partially Billed")
    
    so.insert()
    so.save(ignore_permissions=True)

def on_task_save(doc, method):
    if doc.service == 'IT-SW':
        if doc.status == "Pending Review":
            issue = frappe.get_doc("Issue", doc.issue)
            task = frappe.db.count("Task",{'issue':doc.issue})
            if task == 1:
                if issue and issue.status != "Resolved":
                    issue.status = "Resolved"
                    issue.save()

        if doc.status == "Completed":
            issue = frappe.get_doc("Issue", doc.issue)
            if issue and issue.status != "Closed":
                issue.status = "Closed"
                issue.save()

        

@frappe.whitelist()
def issue_status(doc,method):
    if doc.service == 'IT-SW':
        if doc.issue is not None:
            if doc.status in ["Open","Working"]:
                issue = frappe.get_doc("Issue", doc.issue)
                if issue and issue.status != "Replied":
                    issue.status = "Replied"
                    issue.task = doc.name
                    issue.assigned_to = doc.completed_by
                    issue.project = doc.project
                    issue.save()
            if doc.status == "Pending Review":
                issue = frappe.get_doc("Issue", doc.issue)
                if issue and issue.status != "Resolved":
                    issue.status ="Resolved"
                    issue.assigned_to = doc.completed_by
                    issue.project = doc.project
                    issue.save()
            if doc.status == "Completed":
                issue = frappe.get_doc("Issue", doc.issue)
                if issue and issue.status != "Closed":
                    issue.status ="Closed"
                    issue.assigned_to = doc.completed_by
                    issue.project = doc.project
                    issue.save()

@frappe.whitelist()
def update_cb(doc,method):
    if doc.service == 'IT-SW' and doc.completed_by != "":
        cb = frappe.get_value("Employee",{'user_id':doc.custom_allocated_to},['short_code'])
        frappe.db.set_value("Task",doc.name,"cb",cb)

@frappe.whitelist()
def return_val():
    total_expected_time= frappe.db.sql(""" select sum(expected_time) as et from `tabTask` where project = '%s' """%("GOTRADEPRO"),as_dict=1)
    

@frappe.whitelist()
def update_actual_tat(date1,date2,date3,date4,pac_tat):
    list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
    actual_tat=0
    tat = 0
    tat_monitor = ''
    date = 0
    dat = 0
    variation = 0
    for i in list:
        doc=frappe.db.get_list(i,["name","workflow_state"])
        for j in doc:
            if(date2 and j.workflow_state=="Report Completed"):
                date=(date_diff(date2,date1))+1
                dat=(sum([int(date3),int(date4)]))
                actual_tat=date - dat
                variation = int(actual_tat)-int(pac_tat)

                if variation < 0:
                    tat=0
                    tat_monitor = "In TAT"
                else:
                    tat=variation
                    tat_monitor = "Out TAT"
        return actual_tat,tat,tat_monitor

@frappe.whitelist()
def case_drop_status(name,remark):
    frappe.db.set_value("Case",name,"case_status","Drop")
    frappe.db.set_value("Case",name,"reason_of_drop",remark)
    frappe.db.set_value("Case",name,"dropped",1)
    frappe.db.set_value("Case",name,"case_report","Drop")  

@frappe.whitelist()
def drop_status(name,date,remark):
    frappe.db.set_value("Social Media",name,"workflow_state","Drop")
    frappe.db.set_value("Social Media",name,"check_status","Drop")
    frappe.db.set_value("Social Media",name,"drop_date",date)
    frappe.db.set_value("Social Media",name,"remarks3",remark)
    frappe.db.set_value("Social Media",name,"dropped",1)
    frappe.db.set_value("Social Media",name,"report_status","Drop")

@frappe.whitelist()
def na_status(name,date,remark):
    frappe.db.set_value("Social Media",name,"workflow_state","Report Completed")
    frappe.db.set_value("Social Media",name,"check_completion_date",today())
    frappe.db.set_value("Social Media",name,"mark_na_on",date)
    frappe.db.set_value("Social Media",name,"remarks2",remark)
    frappe.db.set_value("Social Media",name,"na",1)
    frappe.db.set_value("Social Media",name,"report_status","Not Applicable")

@frappe.whitelist()
def check_status(name,date,remark):
    frappe.db.set_value("Education Checks",name,"workflow_state","Drop")
    frappe.db.set_value("Education Checks",name,"check_status","Drop")
    frappe.db.set_value("Education Checks",name,"drop_date",date)
    frappe.db.set_value("Education Checks",name,"remarks3",remark)
    frappe.db.set_value("Education Checks",name,"drop",1)
    frappe.db.set_value("Education Checks",name,"report_status","Drop")

@frappe.whitelist()
def app_state(name,date,remark):
    frappe.db.set_value("Education Checks",name,"workflow_state","Report Completed")
    frappe.db.set_value("Education Checks",name,"check_completion_date",today())
    frappe.db.set_value("Education Checks",name,"mark_na_on",date)
    frappe.db.set_value("Education Checks",name,"remarks2",remark)
    frappe.db.set_value("Education Checks",name,"na",1)
    frappe.db.set_value("Education Checks",name,"report_status","Not Applicable")

@frappe.whitelist()
def drop_value(name,date,remark):
    frappe.db.set_value("Reference Check",name,"workflow_state","Drop")
    frappe.db.set_value("Reference Check",name,"check_status","Drop")
    frappe.db.set_value("Reference Check",name,"drop_date",date)
    frappe.db.set_value("Reference Check",name,"remarks3",remark)
    frappe.db.set_value("Reference Check",name,"drop",1)
    frappe.db.set_value("Reference Check",name,"report_status","Drop")

@frappe.whitelist()
def na_value(name,date,remark):
    frappe.db.set_value("Reference Check",name,"workflow_state","Report Completed")
    frappe.db.set_value("Reference Check",name,"check_completion_date",today())
    frappe.db.set_value("Reference Check",name,"mark_na_on",date)
    frappe.db.set_value("Reference Check",name,"remarks2",remark)
    frappe.db.set_value("Reference Check",name,"na",1)
    frappe.db.set_value("Reference Check",name,"report_status","Not Applicable")




@frappe.whitelist()
def doc_mark(name,date,remark):
    frappe.db.set_value("Family",name,"workflow_state","Drop")
    frappe.db.set_value("Family",name,"check_status","Drop")
    frappe.db.set_value("Family",name,"drop_date",date)
    frappe.db.set_value("Family",name,"remarks3",remark)
    frappe.db.set_value("Family",name,"drop",1)
    frappe.db.set_value("Family",name,"report_status","Drop")

@frappe.whitelist()
def report_check(name,date,remark):
    frappe.db.set_value("Family",name,"workflow_state","Report Completed")
    frappe.db.set_value("Family",name,"check_completion_date",today())
    frappe.db.set_value("Family",name,"mark_na_on",date)
    frappe.db.set_value("Family",name,"remarks2",remark)
    frappe.db.set_value("Family",name,"na",1)
    frappe.db.set_value("Family",name,"report_status","Not Applicable")


@frappe.whitelist()
def status_mark(name,date,remark):
    frappe.db.set_value("Identity Aadhar",name,"workflow_state","Drop")
    frappe.db.set_value("Identity Aadhar",name,"check_status","Drop")
    frappe.db.set_value("Identity Aadhar",name,"drop_date",date)
    frappe.db.set_value("Identity Aadhar",name,"remarks3",remark)
    frappe.db.set_value("Identity Aadhar",name,"drop",1)
    frappe.db.set_value("Identity Aadhar",name,"report_status","Drop")

@frappe.whitelist()
def na_applicable(name,date,remark):
    frappe.db.set_value("Identity Aadhar",name,"workflow_state","Report Completed")
    frappe.db.set_value("Identity Aadhar",name,"check_completion_date",today())
    frappe.db.set_value("Identity Aadhar",name,"mark_na_on",date)
    frappe.db.set_value("Identity Aadhar",name,"remarks2",remark)
    frappe.db.set_value("Identity Aadhar",name,"na",1)
    frappe.db.set_value("Identity Aadhar",name,"report_status","Not Applicable")


@frappe.whitelist()
def document_state(name,date,remark):
    frappe.db.set_value("Employment",name,"workflow_state","Drop")
    frappe.db.set_value("Employment",name,"check_status","Drop")
    frappe.db.set_value("Employment",name,"drop_date",date)
    frappe.db.set_value("Employment",name,"remarks3",remark)
    frappe.db.set_value("Employment",name,"drop",1)
    frappe.db.set_value("Employment",name,"report_status","Drop")

@frappe.whitelist()
def state_report(name,date,remark):
    frappe.db.set_value("Employment",name,"workflow_state","Report Completed")
    frappe.db.set_value("Employment",name,"check_completion_date",today())
    frappe.db.set_value("Employment",name,"mark_na_on",date)
    frappe.db.set_value("Employment",name,"remarks2",remark)
    frappe.db.set_value("Employment",name,"na",1)
    frappe.db.set_value("Employment",name,"report_status","Not Applicable")


@frappe.whitelist()
def update_state(name,date,remark):
    frappe.db.set_value("Court",name,"workflow_state","Drop")
    frappe.db.set_value("Court",name,"check_status","Drop")
    frappe.db.set_value("Court",name,"drop_date",date)
    frappe.db.set_value("Court",name,"remarks3",remark)
    frappe.db.set_value("Court",name,"drop",1)
    frappe.db.set_value("Court",name,"report_status","Drop")

@frappe.whitelist()
def update_na(name,date,remark):
    frappe.db.set_value("Court",name,"workflow_state","Report Completed")
    frappe.db.set_value("Court",name,"check_completion_date",today())
    frappe.db.set_value("Court",name,"mark_na_on",date)
    frappe.db.set_value("Court",name,"remarks2",remark)
    frappe.db.set_value("Court",name,"na",1)
    frappe.db.set_value("Court",name,"report_status","Not Applicable")

@frappe.whitelist()
def set_status(name,date,remark):
    frappe.db.set_value("Criminal",name,"workflow_state","Drop")
    frappe.db.set_value("Criminal",name,"check_status","Drop")
    frappe.db.set_value("Criminal",name,"drop_date",date)
    frappe.db.set_value("Criminal",name,"remarks3",remark)
    frappe.db.set_value("Criminal",name,"drop",1)
    frappe.db.set_value("Criminal",name,"report_status","Drop")

@frappe.whitelist()
def set_na(name,date,remark):
    frappe.db.set_value("Criminal",name,"workflow_state","Report Completed")
    frappe.db.set_value("Criminal",name,"check_completion_date",today())
    frappe.db.set_value("Criminal",name,"mark_na_on",date)
    frappe.db.set_value("Criminal",name,"remarks2",remark)
    frappe.db.set_value("Criminal",name,"na",1)
    frappe.db.set_value("Criminal",name,"report_status","Not Applicable")


@frappe.whitelist()
def get_drop_status(name,date,remark):
    frappe.db.set_value("Address Check",name,"workflow_state","Drop")
    frappe.db.set_value("Address Check",name,"check_status","Drop")
    frappe.db.set_value("Address Check",name,"drop_date",date)
    frappe.db.set_value("Address Check",name,"remarks3",remark)
    frappe.db.set_value("Address Check",name,"drop",1)
    frappe.db.set_value("Address Check",name,"report_status","Drop")

@frappe.whitelist()
def get_ns_status(name,date,remark):
    frappe.db.set_value("Address Check",name,"workflow_state","Report Completed")
    frappe.db.set_value("Address Check",name,"check_completion_date",today())
    frappe.db.set_value("Address Check",name,"mark_na_on",date)
    frappe.db.set_value("Address Check",name,"remarks2",remark)
    frappe.db.set_value("Address Check",name,"na",1)
    frappe.db.set_value("Address Check",name,"report_status","Not Applicable")


@frappe.whitelist()
def update_status_issue():
    # ta = frappe.db.sql("""select count(*) as count from `tabIssue` where  task = '' and status = "Resolved" """,as_dict = True)
    ta = frappe.db.count("Issue",{'task':"",'status':"Resolved"})
    print(ta)
    # task = frappe.db.sql("""select * from `tabIssue` where  task = '' and status = "Resolved" """,as_dict = True)
    # for i in task:
    # 	print(i.name)
        # if not frappe.db.exists("Task",{'issue':i.name}):
        # 	frappe.db.set_value("Issue",i.issue,"status","Closed")
                # if i.status == "Pending Review":
                # 	frappe.db.set_value("Issue",i.issue,"status","Resolved")
                # if i.status == "Completed":
                # 	frappe.db.set_value("Issue",i.issue,"status","Closed")

@frappe.whitelist()
def meeting_mom_mail(meet):
    data = ''
    meet_doc = frappe.get_doc("Meeting",meet)
    data += 'Dear Sir,<br><br>Kindly Find the below List of MOM points against the meeting happened at %s' % formatdate(meet_doc.date)
    for i in meet_doc.minutes:
        data += '<tr><td>%s . %s </td></tr>' % (i.idx,i.description)
    for j in meet_doc.attendees:
        frappe.sendmail(
            recipients=[j.attendee],
            message=data,
            subject=_("Minutes of Meeting -  %s on %s " %(meet_doc.title,formatdate(meet_doc.date))),
        )
    for j in meet_doc.persons_to_be_informed:
        frappe.sendmail(
            recipients=[j.attendee],
            message=data,
            subject=_("Minutes of Meeting -  %s on %s " %(meet_doc.title,formatdate(meet_doc.date))),
        )

@frappe.whitelist()
def update_insuff_days(date1,date2):
    date=(date_diff(date2,date1))
    sql_query = f"""
        SELECT COUNT(*) 
        FROM `tabHoliday` 
        WHERE parent = 'TEAMPRO 2023 - Checkpro' 
        AND holiday_date BETWEEN '{date1}' AND '{date2}'
    """
    count = frappe.db.sql(sql_query, as_list=True)[0][0]
    date1 = (date-count)+1
    return date1

@frappe.whitelist()
def holidays(date1, date2):
    
    sql_query = f"""
        SELECT COUNT(*) 
        FROM `tabHoliday` 
        WHERE parent = 'TEAMPRO 2023 - Checkpro' 
        AND holiday_date BETWEEN '{date1}' AND '{date2}'
    """

    count = frappe.db.sql(sql_query, as_list=True)[0][0]

    return count
# @frappe.whitelist()
# def holidays(date1, date2,name):
#     doc=frappe.get_doc("Case",name)
#     if doc.date_of_initiating:
#         from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
#         holiday_list_name = 'TEAMPRO 2023 - Checkpro'
#         start_date = doc.date_of_initiating
#         working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
#         current_date = start_date
#         holiday = []
#         while working_days > 0:
#             if not is_holiday(holiday_list_name, current_date):
#                 holiday.append(current_date)
#                 working_days -= 1
#             current_date = add_days(current_date, 1)
#         sql_query = f"""
#             SELECT COUNT(*) 
#             FROM `tabHoliday` 
#             WHERE parent = 'TEAMPRO 2023 - Checkpro' 
#             AND holiday_date BETWEEN '{date1}' AND '{date2}'
#         """

#         count = frappe.db.sql(sql_query, as_list=True)[0][0]

#         # return count
#         return holiday[-1],count

@frappe.whitelist()
def tat_calculation(date1,date2,date3,date4,pac_tat):
    actual_tat=0
    tat = 0
    tat_monitor = ''
    date = 0
    dat = 0
    variation = 0
    doc=frappe.db.get_list("Case",["name","case_completion_date","case_status"])
    for i in doc:
        if i.case_completion_date:
            date=(date_diff(date2,date1))+1
            dat=(sum([int(date3),int(date4)]))
            actual_tat=date - dat
            variation = int(actual_tat)-int(pac_tat)

            if variation < 0:
                tat=0
                tat_monitor = "In TAT"
            else:
                tat=variation
                tat_monitor = "Out TAT"
    return actual_tat,tat,tat_monitor


# @frappe.whitelist()
# def update_att():
# 	att = frappe.db.sql("""delete from `tabAttendance` where name = "HR-ATT-2023-10801" """,as_dict = True)
# 	print(att)
# 	att = frappe.db.sql("""delete from `tabAttendance Request` where name = "HR-ARQ-23-11-00016" """,as_dict = True)
# 	print(att)
@frappe.whitelist()
def update_case_status(case):
    list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
    case_sts=[]
    case_status=''
    # insuff=""
    insuff=[]
    for i in list:
        doc=frappe.get_all(i,{"case_id":case},["name","workflow_state","detailsof_insufficiency","check_type"])        
        for j in doc:
            case_sts.append(j.workflow_state)
            if j.workflow_state =="Insufficient Data":
                insuff.append(j.detailsof_insufficiency)
            if any(status == "Insufficient Data" for status in case_sts):
                if insuff:
                    if any(k=="Entry" for k in insuff):
                        case_status="Entry-Insuff"
                    else:
                        case_status="Execution-Insuff"
                # if j.detailsof_insufficiency =="Entry":
                #     case_status="Entry-Insuff"
                # if j.detailsof_insufficiency =="Execution":
                #     case_status="Execution-Insuff"
            elif all(status == "Entry Completed" for status in case_sts):
                case_status = "Entry Completed"
                frappe.db.set_value("Case",case,"date_of_entry_completion",frappe.utils.now())
            elif all(status == "Entry QC Pending" for status in case_sts):
                case_status = "Entry-QC"
            elif all(status == "Execution Initiated" for status in case_sts):
                case_status = "Execution"
            elif all(status == "Execution Pending" for status in case_sts):
                case_status = "Execution"
            elif all(status == "Final QC Pending" for status in case_sts):
                case_status = "Final-QC"
            elif all(status == "Report Completed" for status in case_sts):
                case_status = "Generate Report"
                frappe.db.set_value("Case",case,"date_of_execution_completion",frappe.utils.now())
                frappe.db.set_value("Case",case,"date_of_final_qc_completion",frappe.utils.now())
                frappe.db.set_value("Case",case,"date_of_generate_report",frappe.utils.now())
            elif all(status == "Not Applicable" for status in case_sts):
                case_status = "Generate Report"
            elif all(status == "Drop" for status in case_sts):
                case_status = "Drop"
            elif any(status == "Draft" for status in case_sts):
                case_status = "Draft"
            elif any(status == "Entry Completed" for status in case_sts):
                case_status = "Entry Completed"
                frappe.db.set_value("Case",case,"date_of_entry_completion",frappe.utils.now())
            elif any(status == "Entry QC Pending" for status in case_sts):
                case_status = "Entry-QC"
            elif any(status == "Entry QC Completed" for status in case_sts):
                case_status = "Entry-QC"
                frappe.db.set_value("Case",case,"date_of_entry_completion",frappe.utils.now())
                frappe.db.set_value("Case",case,"date_of_entry_qc_completion",frappe.utils.now())
            elif any(status == "Execution Initiated" for status in case_sts):
                case_status = "Execution"
            elif any(status == "Execution Pending" for status in case_sts):
                case_status = "Execution"
            elif any(status == "Execution Completed" for status in case_sts):
                case_status = "Execution"
                frappe.db.set_value("Case",case,"date_of_execution_completion",frappe.utils.now())
            elif any(status == "Final QC Pending" for status in case_sts):
                case_status = "Final-QC"
            elif any(status == "Final QC Completed" for status in case_sts):
                case_status = "Final-QC"
                frappe.db.set_value("Case",case,"date_of_final_qc_completion",frappe.utils.now())
    frappe.db.set_value("Case",case,"case_status",case_status)
    frappe.db.set_value("Case",case,"custom_case_update_status",case_status)


# @frappe.whitelist()
# def update_query():
# 	frappe.db.sql("""update `tabCase` set case_status = 'Generate Report' where name = 'CS-005392'""")
# @frappe.whitelist()
# def update_query():
# 	frappe.db.sql("""update `tabEnergy Point And Non Conformity` set workflow_state = 'Rejected' where name = 'EPNC-24-10-00027-2'""")

@frappe.whitelist()
def update_query():
    frappe.db.sql("""update `tabCandidate` set pending_for = 'Linedup' where name = 'CD124210'""")

@frappe.whitelist()
def drop_case(case_id):
    list = ["Education Checks","Family","Reference Check","Court","Social Media","Criminal","Employment","Identity Aadhar","Address Check"]
    for i in list:
        checks=frappe.db.get_all(i,{"case_id":case_id},["name"])
        for j in checks:
            frappe.db.set_value(i,j.name,"workflow_state","Drop")
            frappe.db.set_value(i,j.name,"dropped_date",nowdate())

@frappe.whitelist()
def enqueue_check_status_update():
    from frappe.utils.background_jobs import enqueue
    filename = "5bba9ec645"
    enqueue(method=check_status_update, queue="long", timeout=6000, filename=filename)

def check_status_update(filename):
    from frappe.utils.file_manager import get_file
    # filename = frappe.get_value("File",{'file_url':fileurl},'name')
    filepath = get_file(filename)
    pps = read_csv_content(filepath[1])
    for pp in pps:
        if pp[0] == 'Check':
            pass
        else:
            if frappe.db.exists(pp[0],pp[1]):
                frappe.db.set_value(pp[0],pp[1],'workflow_state',pp[2])

@frappe.whitelist()
def case_report_update(filename):
    print(filename)
    from frappe.utils.file_manager import get_file
    filepath = get_file(filename)
    print(filepath)
    pps = read_csv_content(filepath[1])
    # for pp in pps:
    # 	if pp[0] == 'ID':
    # 		pass
    # 	else:
    # 		so=frappe.db.get_all("Closure",{'name':pp[0]},['project','customer','task',"given_name",'mobile','payment','candidate_owner','sa_id','billing_currency','territory','associate','passport_no','expected_doj','account_manager','service','sa_id','associate_si','client_si','candidate_si'])
    # 		for i in so:
    # 			create_sale_order(pp[0],i.project,i.customer,i.task,i.given_name,i.mobile,i.payment,"",i.candidate_owner,i.sa_id,'',i.billing_currency,i.territory,i.associate,i.passport_no,i.expected_doj,i.candidate_owner,i.account_manager,i.service,i.sa_id,i.associate_si,i.client_si,i.candidate_si)

@frappe.whitelist()
def case_status_update_from_csv(filename):
    print(filename)
    from frappe.utils.file_manager import get_file
    filepath = get_file(filename)
    print(filepath)
    pps = read_csv_content(filepath[1])
    i=1
    for pp in pps:
        if pp[0] != 'ID':
            i+=1
            print(pp[0])
            # frappe.db.sql("""update `tabCase` set case_status = %s where name = %s""",(pp[1],pp[0]))
            # frappe.db.sql("""delete from `tabEmployment` where name = %s""",(pp[0]),as_dict=True)

    print(i)

@frappe.whitelist()
def inactive_employee(doc,method):
    if doc.status=="Active":
        if doc.relieving_date:
            throw(_("Please remove the relieving date for the Active Employee."))

@frappe.whitelist()
def ops_mail(project):
    data = ''
    project_doc = frappe.get_doc("Project",project)
    data += 'New Project %s has been created  and Pending for Approval'%(project_doc.name)
    frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','sangeetha.a@groupteampro.com'],
        # recipients=["gifty.p@groupteampro.com"],
        message=data,
        subject=_("New Project Created -  %s" %(project_doc.project_name)),
    )

@frappe.whitelist()
def rns_mail(project):
    data = ''
    project_doc = frappe.get_doc("Project",project)
    data += 'Project %s has been returned to R&S'%(project_doc.name)
    created_by=project_doc.owner
    frappe.sendmail(
        recipients=created_by,
        message=data,
        subject=_("Project Returned -  %s" %(project_doc.project_name)),
    )

@frappe.whitelist()
def confirm_mail(project):
    data = ''
    project_doc = frappe.get_doc("Project",project)
    data += 'Project %s has been confirmed to OPS'%(project_doc.name)
    frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com',"sangeetha.s@groupteampro.com","sangeetha.a@groupteampro.com","sams@groupteampro.com","dm@groupteampro.com"],
        message=data,
        subject=_("Project Confirmed -  %s" %(project_doc.project_name)),
    )

# @frappe.whitelist()
# def update_lead():
# 	lead=frappe.get_all("Lead",{"website":('!=',''),'status':'Converted'},['*'])
# 	for i in lead:
# 		print(i.name)
# 		frappe.db.set_value("Lead",i.name,'website__',i.website)

@frappe.whitelist()
def update_batch_age():
    age=0
    tat_var=0
    tat_mon=''
    tat_sts=''
    doc=frappe.db.get_list("Batch",["name","expected_start_date","batch_status",'package_tat'])
    for i in doc:
        if i.batch_status not in ("Completed",'',"Drop",'Proposed SO'):
            if i.expected_start_date:
                date=(date_diff(nowdate(),i.expected_start_date))+1
                sql_query = f"""
                    SELECT COUNT(*) 
                    FROM `tabHoliday` 
                    WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                    AND holiday_date BETWEEN '{i.expected_start_date}' AND '{nowdate()}'
                """
                count = frappe.db.sql(sql_query, as_list=True)[0][0]
                print(count)
                # print(type(count))
                print(date)
                # print(type(date))
                if count==0:
                    age=date
                else:
                    age = date-count
                print(i.name)
                print(age)
                tat_var=i.package_tat-age
                if tat_var>0:
                    tat_mon='In TAT'
                else:
                    tat_mon='Out TAT'
                if age<(0.4*i.package_tat):
                    tat_sts='Regular'
                elif age<(0.65*i.package_tat):
                    tat_sts='Critical'
                else:
                    tat_sts='Most Critical'
                frappe.db.set_value("Batch",i.name,"actual_tat",age)
                frappe.db.set_value("Batch",i.name,"custom_holidays",count)
                frappe.db.set_value("Batch",i.name,"tat_variation",tat_var)
                frappe.db.set_value("Batch",i.name,"tat__monitor",tat_mon)
                frappe.db.set_value("Batch",i.name,"custom_tat_status",tat_sts)


@frappe.whitelist()
def update_case_age():
    age=0
    tat_var=0
    tat_mon=''
    tat_sts=''
    doc=frappe.db.get_list("Case",["name","date_of_initiating","case_status",'insufficiency_days','package_tat'],order_by='date_of_initiating ASC')
    for i in doc:
        if i.case_status not in ("Case Completed","Drop","Generate Report with Insuff",'',"Drop"):
            if i.date_of_initiating:
                date=(date_diff(nowdate(),i.date_of_initiating))+1
                sql_query = f"""
                    SELECT COUNT(*) 
                    FROM `tabHoliday` 
                    WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                    AND holiday_date BETWEEN '{i.date_of_initiating}' AND '{nowdate()}'
                """
                count = frappe.db.sql(sql_query, as_list=True)[0][0]
                if count==0:
                    age=date-i.insufficiency_days
                else:
                    age = (date-(count+i.insufficiency_days))
                print(i.name)
                print(age)
                if age > 15:
                    cl = '#f50f0f'
                elif age >10:
                    cl = '#EC864B'
                elif age >5:
                    
                    cl = '#449CF0'
                else:
                    cl = '#000000'
                tat_var=i.package_tat-age
                if tat_var>0:
                    tat_mon='In TAT'
                else:
                    tat_mon='Out TAT'
                if age<(0.4*i.package_tat):
                    tat_sts='Regular'
                elif age<(0.65*i.package_tat):
                    tat_sts='Critical'
                else:
                    tat_sts='Most Critical'
                frappe.db.set_value("Case",i.name,"actual_tat",age)
                print(cl)
                frappe.db.set_value("Case",i.name,"color",cl)
                frappe.db.set_value("Case",i.name,"holidays",count)
                frappe.db.set_value("Case",i.name,"tat_variation",tat_var)
                frappe.db.set_value("Case",i.name,"tat_monitor",tat_mon)
                frappe.db.set_value("Case",i.name,"custom_tat_status",tat_sts)


@frappe.whitelist()
def update_check_age():
    list = ["Education Checks","Employment","Address Check","Criminal","Reference Check","Court","Identity Aadhar","Family","Social Media"]
    
    age=0
    tat_var=0
    tat_mon=''
    tat_sts=''
    for i in list:
        doc=frappe.db.get_list(i,["name","check_creation_date","workflow_state",'package_tat','insufficiency_days'])
        for j in doc:
            if j.workflow_state not in ('Report Completed', '', 'Drop', 'Dropped', 'Not Applicable'):
                if j.check_creation_date:
                    date=(date_diff(nowdate(),j.check_creation_date))+1
                    sql_query = f"""
                        SELECT COUNT(*) 
                        FROM `tabHoliday` 
                        WHERE parent = 'TEAMPRO 2023 - Checkpro' 
                        AND holiday_date BETWEEN '{j.check_creation_date}' AND '{nowdate()}'
                    """
                    count = frappe.db.sql(sql_query, as_list=True)[0][0]
                    print(count)
                    # print(type(count))
                    print(date)
                    # print(type(date))
                    if count==0:
                        age=date-j.insufficiency_days
                    else:
                        age = date-(count+j.insufficiency_days)
                    tat_var=int(j.package_tat)-age
                    if tat_var>0:
                        tat_mon='In TAT'
                    else:
                        tat_mon='Out TAT'
                    if age<(0.4*int(j.package_tat)):
                        tat_sts='Regular'
                    elif age<(0.65*int(j.package_tat)):
                        tat_sts='Critical'
                    else:
                        tat_sts='Most Critical'
                    print(j.name)
                    print(i)
                    frappe.db.set_value(i,j.name,"actual_tat",age)
                    frappe.db.set_value(i,j.name,"holidays",count)
                    frappe.db.set_value(i,j.name,"tat_variation",tat_var)
                    frappe.db.set_value(i,j.name,"tat_monitor",tat_mon)
                    frappe.db.set_value(i,j.name,"custom_tat_status",tat_sts)

def update_cv_age():
    age=0
    ind=0
    cand=0
    doc=frappe.db.get_list("Candidate",{"pending_for":("not in",['IDB','Sourced','Proposed PSL'])},["name","source","submitted_date"],)
    for i in doc:
        cand+=1
        if i.submitted_date:
            ind+=1
            age=(date_diff(nowdate(),i.submitted_date))
            print(date)
            print(i.name)
            print(age)
            frappe.db.set_value("Candidate",i.name,"age_of_cv",age)



@frappe.whitelist()
def get_workflow_state(doctype):
    doc = frappe.get_doc("Workflow", doctype)
    states = [state.state for state in doc.states]
    return states

@frappe.whitelist()
def update_lead_status(doc,method):
    if doc.status=="Lost":
        frappe.db.set_value("Lead", doc.party_name, {"disabled": 0, "docstatus": 0})

@frappe.whitelist()
def update_att():
    att = frappe.db.sql("""update `tabAttendance` set docstatus = 0  where attendance_date between "2023-12-01" and "2023-12-31" """)
    print(att)

@frappe.whitelist()
def update_mandatory(doctype,status):
    settings = frappe.get_doc("Checkpro Settings")
    key_value_pairs_list = []

    for child in settings.get("settings"):
        if child.check_type == doctype and child.status == status:
            reqd_fields = child.mandatory_fields
            field_list = reqd_fields.split(',')
            for field in field_list:
                key, value = field.split('-')
                key = key.strip()
                value = value.strip()
                key_value_pairs_list.append({key: value})

    return key_value_pairs_list

@frappe.whitelist()
def update_meeting_id(meet):
    meeting=frappe.get_doc("Meeting",meet)
    for i in meeting.minutes:
        if i.custom_action == "Task":
            frappe.db.set_value("Task",i.custom_id,'custom_meeting_id',meet)
        elif i.custom_action == "To Do":
            frappe.db.set_value("ToDo",i.custom_id,'custom_meeting_id',meet)

from frappe import _

@frappe.whitelist()
def meeting_mail(meet):
    meet_doc = frappe.get_doc("Meeting",meet)
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
<tr style="background-color: #063970;">
    <td width="5%" style="text-align:center;">S.No</td>
    <td width="75%" style="text-align:center;">Description</td>
    <td width="10%" style="text-align:center;">Action</td>
    <td width="10%" style="text-align:center;">ID</td>
</tr>
'''

    ind = 1
    recipients = []
    cc = []
    for i in meet_doc.minutes:
        data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(ind, i.description, i.custom_action or '', i.custom_id)
        ind += 1
    data += '</table>'
    for j in meet_doc.attendees:
        if j.attended == '1':
            recipients.append(j.attendee)
        else:
            cc.append(j.attendee)
    frappe.sendmail(
        recipients=recipients,
        cc=cc,
        subject=_("Minutes of Meeting -  %s on %s " % (meet_doc.title, formatdate(meet_doc.date))),
        message="""
            Dear Sir/Madam,<br>Kindly Find the below List of MOM points against the meeting happened at {} {}<br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
            """.format(formatdate(meet_doc.date), data)
    )
    return "Ok"


@frappe.whitelist()
def get_task_details(name):
    data = ''
    data += "<table width=100% style='border:1px solid black' ><tr><td colspan=6 style='text-align:center;background-color:#009dd1;font-size:17px;border:1px solid black'><b>Task Details</b></td></tr>"
    data += "<tr style='font-size:17px;border:1px solid black'><td style='border:1px solid black'><b>Position</b></td><td style='border:1px solid black'><b># Vac</b></td><td style='border:1px solid black'><b># SP</b></td><td style='border:1px solid black'><b># FP</b></td><td style='border:1px solid black'><b>#SL</b></td><td style='border:1px solid black'><b>#PSL</b></td>"
    
    tasks = frappe.get_all("Task",{"project":name},['*'])
    for i in tasks:
        data += "<tr style='font-size:17px'><td style='border:1px solid black'>%s - %s</td><td style='border:1px solid black'>%s</td><td style='border:1px solid black'>%s</td><td style='border:1px solid black'>%s</td><td style='border:1px solid black'>%s</td><td style='border:1px solid black'>%s</td></tr>"%(i.name,i.subject,i.vac,i.sp,i.fp,i.sl,i.psl)
    data += "</table>"
    return data

# @frappe.whitelist()
# def update_att():
# 	att = frappe.db.sql("""
# 		UPDATE `tabClosure`
# 		SET visa_status = "Visa Received"
# 		WHERE name = "CL02458"
# 	""")
# 	print(att)




@frappe.whitelist(allow_guest=True)
def val_pass(passcode,email):	
    if frappe.db.exists("BG Entry Passcode",{'passcode':passcode,'email_id':email}):
        return "Yes"
    else:
        return "No"
    
@frappe.whitelist(allow_guest=True)
def submission_mail(email,name):
    frappe.sendmail(
        recipients=['anil.p@groupteampro.com'],
        subject=_("Proceed BGV"),
        message="""
            Dear Sir/madam,<br>Candidate %s with the Email ID %s has been applied for BGV. Kindly initate the BGV<br><br><br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
            """%(name, email)

        )
@frappe.whitelist(allow_guest=True)
def submission_mail2(email,name):
    frappe.sendmail(
        recipients=['anil.p@groupteampro.com'],
        subject=_("Applied for BGV"),
        message="""
            Dear ,<br>Candidate %s with the Email ID %s has submitted the documents successfully. BGV initiated for the Candidate<br><br>
            Thanks & Regards,<br>TEAM ERP<br>"This email has been automatically generated. Please do not reply"
            """%(name,email)
        )

@frappe.whitelist()
def update_pi_workflow(doc,method):
    pi=frappe.get_all("Purchase Invoice",{'sales_order':doc.name},['workflow_state','name'])
    for i in pi:
        frappe.db.sql("update `tabPurchase Invoice` set workflow_state='Cancelled' where name=%s",(i.name))

@frappe.whitelist()
def check_supplier(supplier):
    supp=frappe.get_doc("Supplier",supplier)
    if supp.custom_is_lead == 1:
        return "OK"
    
@frappe.whitelist()
def update_workflow_state(doc,method):
    if doc.workflow_state:
        frappe.db.sql("""update `tabPurchase Invoice` set custom_status = %s where name = %s""",(doc.workflow_state,doc.name))

@frappe.whitelist()
def update_bg_entry():
    
    frappe.db.sql("""update `tabEmployment` set entry_allocation_date = '2024-03-21' where name = 'Employment-2500'""")


def send_closure_report_with_table():
    filename = "Closure_" + today()
    xlsx_file = build_xlsx_response_closure(filename)
    html_table, total_count = closure_next_action()
    if total_count > 0:
        send_mail_with_attachment_and_html(filename, xlsx_file.getvalue(), html_table)

def send_mail_with_attachment_and_html(filename, file_content, html_table):
    subject = "DND DPR - %s" % nowdate()
    message = (
        "Dear Sir/Madam,<br>"
        "Please find attached the attached Report based on Next Action.<br><br>"
        + html_table +
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    frappe.sendmail(
        # recipients=['jeniba.a@groupteampro.com'],
        recipients=['dc@groupteampro.com'],
        cc=['sangeetha.a@groupteampro.com','sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
        sender=None,
        subject=subject,
        message=message,
        attachments=attachments,
    )

def build_xlsx_response_closure(filename):
    return make_xlsx_closure(filename)

def make_xlsx_closure(filename, sheet_name=None, wb=None, column_widths=None):
    action = add_days(nowdate(), 1)
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name or filename, 0)  
    default_column_widths = [15, 25, 25, 15, 25, 20]
    column_widths = column_widths or default_column_widths    
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width  
    header_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    ws.append(["ID", "Candidate Name", "Customer", "Status", "Next Action", "Remark", "Next Action Date"])
    for cell in ws[1]: 
        cell.fill = header_fill
    closures = frappe.get_all("Closure", {"custom_next_follow_up_on":action,'status':["Not In", ['Onboarded','Dropped']]}, ['*'])
    if closures:
        for closure in closures:
            ws.append([closure.name, closure.given_name, closure.customer, closure.status, closure.std_remarks, closure.remark, closure.custom_next_follow_up_on])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)    
    return xlsx_file

def closure_next_action():
    action_date = add_days(nowdate(), 1)
    closures = frappe.get_all("Closure", {"custom_next_follow_up_on": action_date,'status':("Not In", ['Onboarded','Dropped'])}, ["customer", "status"])
    customer_status_count = {}
    for closure in closures:
        customer = closure.customer
        status = closure.status
        if customer not in customer_status_count:
            customer_status_count[customer] = {}
        if status not in customer_status_count[customer]:
            customer_status_count[customer][status] = 0
        customer_status_count[customer][status] += 1
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">Customer</td><td style="width: 30%; font-weight: bold; text-align: center;">Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Count</td></tr>'
    for customer, statuses in customer_status_count.items():
        total_counts = sum(statuses.values())
        table += '<tr><td><b>%s</b></td><td></td><td><b>%s</b></td></tr>' % (customer, total_counts)        
        for status, count in statuses.items():
            table += '<tr><td></td><td>%s</td><td>%s</td></tr>' % (status, count)
    table += '</table>'
    total_count = sum(sum(status.values()) for status in customer_status_count.values())
    return table, total_count

def send_project_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename = "PTSR_" + posting_date
    xlsx_file = build_xlsx_response_project(filename)
    send_mail_with_attachment_project(filename, xlsx_file.getvalue())

def send_mail_with_attachment_project(filename, file_content):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    subject = "REC : Project – Task Status Report : - %s" % posting_date
    message = (
        "Dear Sir/Madam,<br>"
        "Please find attached the attached Project Report.<br><br>"
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    frappe.sendmail(
        recipients=["dineshbabu.k@groupteampro.com","sangeetha.a@groupteampro.com","sangeetha.s@groupteampro.com","annie.m@groupteampro.com","vijiyalakshmi.k@groupteampro.com",'keerthana.k@groupteampro.com','lokeshkumar.a@groupteampro.com','aruna.g@groupteampro.com','prabhu.m@groupteampro.com'],
        # recipients=['sangeetha.a@groupteampro.com'],
        cc='',
        sender=None,
        subject=subject,
        message=message,
        attachments=attachments,
    )
def build_xlsx_response_project(filename):
    return make_xlsx_project(filename)
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

def make_xlsx_project(filename, sheet_name=None, wb=None, column_widths=None):
    if wb is None:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name or filename, 0)
    default_column_widths = [8, 30, 15, 43, 43, 15, 15, 15, 15, 34, 18]
    column_widths = column_widths or default_column_widths
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    posting_date = datetime.now().strftime("%d-%m-%Y")
    ftitle = "REC : Project – Task Status Report : - " + posting_date
    ws.append([ftitle])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    header_font = Font(color="FFFFFF")
    headers = ["SI NO", "CUSTOMER/PROJECT NAME", "Project Priority", "AM Remark", "PM Remark",'SPOC Remark', 'Expected Value', 'Expected PSL', 'Sourcing Status', 'Territory', 'TASK', 'Task Priority', '#VAC', '#SP', '#FP', '#SL', '#PSL', '#LP']
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    ws.append(headers)
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = black_border
    cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
    row = 3
    serial_number = 1
    grand_totals = {'vac':0,'sp': 0,'fp': 0,'sl':0,'psl':0,'custom_lp':0}
    for c in cust:
        priority = {"High": 1, "Medium": 2, "Low": 3}
        pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'],order_by= "priority ASC")
        if not pname:
            continue
        task_totals = {'vac':0,'sp':0,'fp':0,'sl':0,'psl':0,'custom_lp':0}
        project_data = []      
        for p in pname:
            pdata = []
            print(p.project_name)        
            taskid = frappe.get_all("Task", {"status": ("in",('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'],order_by= "priority ASC")              
            # print(p['project_name'])
            # for tn in taskid:
                # print(tn.name)
            for t in taskid:
                pdata.append([p['project_name'] if p['project_name'] else "",p['priority'] if p['priority'] else "",p['remark'] if p['remark'] else "",p['account_manager_remark'] if p['account_manager_remark'] else "",p['custom_spoc_remark'] if p['custom_spoc_remark'] else "",p['expected_value'] if p['expected_value'] else "",p['expected_psl'] if p['expected_psl'] else "",p['sourcing_statu'] if p['sourcing_statu'] else "",p['territory'] if p['territory'] else "",t['subject'],t['priority'],t['vac'],t['sp'],t['fp'],t['sl'],t['psl'],t['custom_lp']])
                task_totals['vac'] +=t['vac']
                task_totals['sp'] +=t['sp']
                task_totals['fp']+= t['fp']
                task_totals['sl'] +=t['sl']
                task_totals['psl'] += t['psl']
                task_totals['custom_lp'] += t['custom_lp']
            project_data.append({
                'project_name': p['project_name'],'priority': p['priority'],
                'remark': p['remark'],'account_manager_remark': p['account_manager_remark'],'custom_spoc_remark':p['custom_spoc_remark'],'sourcing_statu': p['sourcing_statu'],'territory': p['territory'],
                'expected_value': p['expected_value'],'expected_psl': p['expected_psl'],'tasks': pdata})
        blue_fill = PatternFill(start_color="98d7f5", end_color="98d7f5", fill_type="solid")
        row_data = [serial_number, c['name']] + [""] * 10 + [task_totals['vac'], task_totals['sp'], task_totals['fp'], task_totals['sl'], task_totals['psl'],task_totals['custom_lp']]
        ws.append(row_data)
        row_to_fill = ws.max_row
        for col, cell in enumerate(ws[row_to_fill], start=1):
            cell.fill = blue_fill
            if col > 11:
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
            cell.border = black_border
        ws.merge_cells(start_row=row_to_fill, start_column=2, end_row=row_to_fill, end_column=3)
        serial_number += 1
        row += 1
        current_row_start = row
        for project in project_data:
            project_row_start = row
            for task_data in project['tasks']:
                ws.append([""] + task_data)
                for col in range(2, len(task_data) + 2):
                    cell = ws.cell(row=row, column=col)
                    if 2 <= col <= 11:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = black_border
                row += 1
            if project_row_start < row - 1:
                for col in range(2, 10):
                    ws.merge_cells(start_row=project_row_start, start_column=col, end_row=row-1, end_column=col)
                ws.merge_cells(start_row=project_row_start, start_column=1, end_row=row-1, end_column=1)
        grand_totals['vac'] += task_totals['vac']
        grand_totals['sp'] += task_totals['sp']
        grand_totals['fp'] += task_totals['fp']
        grand_totals['sl'] += task_totals['sl']
        grand_totals['psl'] += task_totals['psl']
        grand_totals['custom_lp'] +=task_totals['custom_lp']
    yellow_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    ws.append(['Total'] + [''] * 10 + [grand_totals['vac'], grand_totals['sp'], grand_totals['fp'], grand_totals['sl'], grand_totals['psl'],grand_totals['custom_lp']])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = black_border
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file



# @frappe.whitelist()
# def update_check_box(check,name):
#     candidate=frappe.get_all("Candidate",{"task":name},["name"])
#     for i in candidate:
#         if check=='1':
#             frappe.db.set_value("Candidate",i.name,"custom_experience_and_education_certificates_required",check)
#         else:
#             frappe.db.set_value("Candidate",i.name,"custom_experience_and_education_certificates_required",check)

# @frappe.whitelist()
# def update_check_box_mask(check,name):
#     candidate=frappe.get_all("Candidate",{"task":name},["name"])
#     for i in candidate:
#         if check=='1':
#             frappe.db.set_value("Candidate",i.name,"custom_non_masked_cvs",check)
#         else:
#             frappe.db.set_value("Candidate",i.name,"custom_non_masked_cvs",check)

# @frappe.whitelist()
# def update_check_box_in_task(check,name):
#     tasks=frappe.get_value("Task",{"project":name},["name"])
#     for i in tasks:
#         if check =='1':
#             frappe.db.set_value("Task",i.name,"custom_experience_and_education_certificates_required",check)
#         else:
#             frappe.db.set_value("Task",i.name,"custom_experience_and_education_certificates_required",'0')

# @frappe.whitelist()
# def update_check_box_task_mask(check,name):
#     candidate=frappe.get_all("Task",{"project":name},["name"])
#     for i in candidate:
#         if check=='1':
#             frappe.db.set_value("Task",i.name,"custom_non_masked_cvs",check)
#         else:
#             frappe.db.set_value("Task",i.name,"custom_non_masked_cvs",check)



# def validate_date1():
#     doc=frappe.get_doc("Purchase Invoice",'ACC-PINV-2024-00498')
#     current_date = datetime.now().date()
#     date_str = doc.posting_date
#     print(type(current_date))
#     date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
#     print(type(date_obj))
#     # if date_str < current_date:
#         throw(_("Submitting the Purchase Invoice on back date will have impact on the GST .Please click on the 'Edit Posting Date' and change the Date to Today."))


import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, add_days, date_diff,format_datetime
from datetime import date, timedelta, datetime, time
from frappe.utils import (
    add_days,
    cstr,
    flt,
    format_datetime,
    formatdate,
    get_datetime,
    get_first_day,
    get_last_day,
    get_link_to_form,
    get_number_format_info,
    getdate,
    nowdate,
    today,
)
def ep_mail():
    first_date=get_first_day(add_days(today(),-1))
    last_date=get_last_day(add_days(today(),-1))
    j=0
    ep = frappe.db.sql(
        """
        SELECT sum(total) as total,sum(total_nc) as total_nc, emp_name as emp_name,emp,employee_mail,reporting_manager_mail from `tabEnergy Point  Non Conformity` where date(creation) BETWEEN %s AND %s AND docstatus=1 group by emp""",(first_date, last_date), as_dict=True)
    s_no=0
    for i in ep:
        s_no+=1
        j+=1
        data = """<table class='table table-bordered' style='border-collapse: collapse; width: 100%;'><tr style='border: 1px solid black; background-color: #0f1568; color: white;'><th>S No</th><th>Employee ID</th><th>Employee Name</th><th>Energy Score</th><th>NC Score</th></tr>""" 

        data += """<tr style='border: 1px solid black;'><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" %( s_no,i.emp,i.emp_name,i['total'] or "-",i['total_nc'] or "-" )
        data += "</table>"
        subject = "Energy Point(EP) And Non Conformity(NC) List-  %s" % nowdate()
        message = """
        Dear Sir/Madam,<br><br>
        Kindly find the below Monthly Energy Point And Non Conformity List<br><br>{}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        <i>This email has been automatically generated. Please do not reply</i>
        """.format(data)
        if j>0:
            frappe.sendmail(
            recipients=[i.reporting_manager_mail,i.employee_mail],
            subject=subject,
            message=message,
        
        )


import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, add_days, date_diff,format_datetime
from datetime import date, timedelta, datetime, time
from frappe.utils import (
    add_days,
    cstr,
    flt,
    format_datetime,
    formatdate,
    get_datetime,
    get_first_day,
    get_last_day,
    get_link_to_form,
    get_number_format_info,
    getdate,
    nowdate,
    today,
)
def epnc_send_mail():
    first_date=get_first_day(add_days(today(),-1))
    last_date=get_last_day(add_days(today(),-1))
    j=0
    ep = frappe.db.sql(
        """
        SELECT sum(total) as total,sum(total_nc) as total_nc, emp_name as emp_name,emp,employee_mail,reporting_manager_mail from `tabEnergy Point  Non Conformity` where date(creation) BETWEEN %s AND %s AND docstatus=1 group by emp""",(first_date, last_date), as_dict=True)
    s_no=0
    data = """<table class='table table-bordered' style='border-collapse: collapse; width: 100%;'><tr style='border: 1px solid black; background-color: #0f1568; color: white;'><th>S No</th><th>Employee ID</th><th>Employee Name</th><th>Energy Score</th><th>NC Score</th></tr>""" 
    for i in ep:
        s_no+=1
        j+=1
    # data = """<table class='table table-bordered' style='border-collapse: collapse; width: 100%;'><tr style='border: 1px solid black; background-color: #0f1568; color: white;'><th>S No</th><th>Employee ID</th><th>Employee Name</th><th>Energy Score</th><th>NC Score</th></tr>""" 
        data += """<tr style='border: 1px solid black;'><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" %( s_no,i.emp,i.emp_name,i['total'] or "-",i['total_nc'] or "-" )
    data += "</table>"
    subject = "Energy Point(EP) And Non Conformity(NC) List-  %s" % nowdate()
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Monthly Energy Point And Non Conformity List<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(data)
    if j>0:
        frappe.sendmail(
        recipients=['dineshbabu.k@groupteampro.com','sangeetha.s@groupteampro.com','sangeetha.a@groupteampro.com'],
        subject=subject,
        message=message,

        )

@frappe.whitelist()
def closure_mail(subject,id,action_taken,live,et,at,revision,service,proof,allocated=None,project=None,issue=None,domain=None,spoc=None,reason=None,dev_spoc=None):
    if service=='IT-SW':
        percentage=et_at_calculation(id, et, at, allocated,subject)
        reports=frappe.db.get_value("Employee",{'user_id':allocated},['reports_to'])
        reports_to=frappe.db.get_value("Employee",{'name':reports},['user_id'])
        tl=frappe.db.get_value("Employee",{'user_id':allocated},["custom_tl"])
        tl_mail=frappe.db.get_value("Employee",{'name':tl},['user_id'])
        et_rate= 'ET : %s and AT : %s'%(et,round(percentage,2))
        if issue:
            raised_by=frappe.db.get_value("Issue",{'name':issue},['raised_by'])
        else:
            raised_by='None'
        data = ''
        data += f"<table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>\
        <tr><td colspan='2' style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black;'><b>Task / Issue Pending Review Note</b></td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Task ID</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/app/task/{id}' target='_blank'>{id}</a></td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Project</b></td><td style='border: 1px solid black;'>{project}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task Raised By</b></td><td style='border: 1px solid black;'>{reports_to}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue ID</b></td><td style='border: 1px solid black;'>{issue}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Raised By</b></td><td style='border: 1px solid black;'>{raised_by}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Statement</b></td><td style='border: 1px solid black;'>{subject}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Action Taken</b></td><td style='border: 1px solid black;'>{action_taken}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Live At</b></td><td style='border: 1px solid black;'>{live}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Domain</b></td><td style='border: 1px solid black;'>{domain}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Proof</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/{proof}' target='_blank'>Link to Proof</a></td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>ET & AT</b></td><td style='border: 1px solid black;'>{et_rate}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Re-Open Count</b></td><td style='border: 1px solid black;'>{revision}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Cause of Re-Open</b></td><td style='border: 1px solid black;'>{reason}</td></tr></table>"
        cc = [reports_to, allocated,spoc,'anil.p@groupteampro.com'] + ([dev_spoc] if dev_spoc else []) +([tl_mail] if tl_mail else [])
        frappe.sendmail(
            sender=allocated,
            # recipients=allocated,
            # recipients='divya.p@groupteampro.com',
            recipients=spoc,
            # cc=[reports_to,allocated,'anil.p@groupteampro.com'],
            cc=cc,
            subject='Task : %s Pending Review : Forward for Review to Mark Completion or Re-Open' % id,
            message = """
            <b>Dear Patron,<br><br>Greeting !!!</b><br><br>
           The attached Task has been completed by Development and forwarded for your kind review, please confirm if it satisfies all your requirement and Mark the Task Status as Client Review / Completed or if you feel it is still pending for some action please change the status as “OPEN” and give your remark for Re-open <br><br>
           {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(data)
        )
        


@frappe.whitelist()
def nc_for_reopen(subject,cause=None,id=None,allocated=None):
    if allocated:
        emp_id=frappe.db.get_value("Employee",{'user_id':allocated},['name'])
        reopen_cause='Task : %s - (%s) Re-Open ' % (id,subject)
        nc = frappe.new_doc('Energy Point And Non Conformity')
        nc.emp = emp_id
        nc.action='Non Conformity(NC)'
        # nc.nc_reported_by = 'Administrator'
        nc.class_proposed = 'Critical'
        nc.reason_of_ep = reopen_cause
        # nc.save(ignore_permissions=True)
        # frappe.db.commit()

@frappe.whitelist()
def et_at_calculation(id, et, at, allocated,subject):
    today_date = datetime.now().strftime('%Y-%m-%d')
    emp_id = frappe.db.get_value("Employee", {'user_id': allocated}, 'name')
    emp_name = frappe.db.get_value("Employee", {'name': emp_id}, 'employee_name')
    timesheets = frappe.get_all("Timesheet",{'employee': emp_id, 'docstatus': ['!=',2]},['name'])
    overall_at = 0.0
    if timesheets:
        for timesheet in timesheets:
            time_logs = frappe.get_all("Timesheet Detail", filters={'parent': timesheet['name'], 'task': id}, fields=['hours'])
            for log in time_logs:
                overall_at += log['hours']
    else:
        overall_at = 0.0

    # if overall_at > 1.5 * float(et):
    #     reopen_cause='Task : %s - (%s) AT is 150 times greater than its ET'% (id,subject)
    #     nc = frappe.new_doc('Energy Point And Non Conformity')
    #     nc.emp = emp_id
    #     # nc.nc_reported_by = 'Administrator'
    #     nc.class_proposed = 'Major'
    #     nc.action='Non Conformity(NC)'
    #     nc.reason_of_ep = reopen_cause
    #     nc.save(ignore_permissions=True)
    #     frappe.db.commit()
    # # else:
    #     total_hours = 0.0
    #     total_time_taken=float(at)
    #     total_time_taken_et=1.5 * float(et)
    #     if float(at)>total_time_taken_et:
    #         reopen_cause='Task : %s AT is 150 times greater than its ET'% id
    #         nc = frappe.new_doc('Energy Point And Non Conformity')
    #         nc.emp = emp_id
    #         # nc.nc_reported_by = 'Administrator'
    #         nc.class_proposed = 'Major'
    #         nc.action='Non Conformity(NC)'
    #         nc.reason_of_ep = reopen_cause
    #         nc.save(ignore_permissions=True)
            # frappe.db.commit()
    return overall_at


# @frappe.whitelist()
# def dpr_mail():
#     current_date = datetime.now().strftime("%Y-%m-%d")
    
#     data = '''
#     <html>
#     <body>
#         <table border="1" cellpadding="5" cellspacing="0">
#             <thead>
#                 <tr>
#                     <th>S.NO</th>
#                     <th>Task Name</th>
#                     <th>Project Name</th>
#                     <th>Subject</th>
#                     <th>CB</th>
#                     <th>Priority</th>
#                     <th>Status</th>
#                     <th>Revisions</th>
#                     <th>Actual Time</th>
#                     <th>Expected Time</th>
#                     <th>RT</th>
#                     <th>Custom Allocated On</th>
#                 </tr>
#             </thead>
#             <tbody>
#                 <tr>
#                     <td>DPR ({})</td>
#                     <td colspan="10"></td>
#                 </tr>
#     '''.format(current_date)
    
#     tasks = frappe.db.get_all("Task", filters={'status': 'Working', 'service': 'IT-SW'}, fields=["name", "project_name", "subject", "cb", "status", "revisions", "actual_time", "expected_time", "rt", "priority", "custom_allocated_on"])
    
#     # Sort tasks by cb, project_name, and priority
#     tasks_sorted = sorted(tasks, key=lambda x: (x.get('cb', ''), x.get('project_name', ''), x.get('priority', '')))
    
#     for idx, task in enumerate(tasks_sorted, start=2):
#         data += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#             idx, task.get('name', ''), task.get('project_name', ''), task.get('subject', ''), task.get('cb', ''), 
#             task.get('priority', ''), task.get('status', ''), task.get('revisions', ''), task.get('actual_time', ''), 
#             task.get('expected_time', ''), task.get('rt', ''), task.get('custom_allocated_on', '')
#         )
    
#     data += '''
#             </tbody>
#         </table>
#     </body>
#     </html>
#     '''
    
#     frappe.sendmail(
#         recipients='siva.m@groupteampro.com',
#         message=data
#     )


# @frappe.whitelist()
# def send_dsr_report():
#     total_open=0
#     total_working=0
#     total_pr=0
#     total_cr=0
#     total_overdue=0
#     mail = frappe.get_all("Project",{'spoc':"sivarenisha.m@groupteampro.com"},["name"])
#     for i in mail:
#         task_status=['Open','Working','Pending Review','Overdue','Client Review']
#         for a in task_status:
#             task_priority=['High','Low','Medium']


@frappe.whitelist()
def dpnd_excel_format():
    filename = "DND Details_" + today() +".xlsx"
    xlsx_file = build_xlsx_response(filename)
    dnd_report(filename, xlsx_file.getvalue())

def dnd_report(filename,file_content):
    task=frappe.db.get_all("Closure",{"custom_status_transition":nowdate()},["*"],order_by='project')
    count=0
    closure_status=["PSL","Sales Order","Client Offer Letter","Signed Offer Letter","Visa","Premedical","PCC","Certificate Attestation","Final Medical","Biometric","Visa Stamping","Emigration","Ticket","Onboarding","Arrived","Dropped","Waitlisted"]
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 45%; font-weight: bold; text-align: center;">Closure ID</td><td style="width: 30%; font-weight: bold; text-align: center;">Candidate Name</td><td style="width: 25%; font-weight: bold; text-align: center;">Project</td><td style="width: 45%; font-weight: bold; text-align: center;">Status</td><td style="width: 25%; font-weight: bold; text-align: center;">Current Status</td></tr>'
    for i in task:
        if i.status in closure_status:
            indx=closure_status.index(i.status)
            next_indx=closure_status[indx-1]
        table += """<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>""" % (i.name,i.given_name,i.project,next_indx,i.status)
        count+=1        
    table += '</table>'
    subject = "DND Transition -  %s" % nowdate()
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Transition :<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(table)
    if count>=1:
        frappe.sendmail(
            recipients=["dc@groupteampro.com","sangeetha.a@groupteampro.com"],
            # recipients=["divya.p@groupteampro.com"],
            subject=subject,
            message=message,
            # attachments=attachments,
            attachments=[{
                "fname": filename,
                "fcontent": file_content,
            }
            ]
        )


def build_xlsx_response(filename):
    xlsx_file = make_xlsx_dnd(filename)
    return xlsx_file

def make_xlsx_dnd(filename, sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    data_row=[]
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_row = ["Closure ID", "Candidate Name", "Project", "Status","Current Status"]
    ws.append(header_row)
    for cell in ws[1]:
        cell.fill = fill
        cell.border = thin_border
    # ws.append(["Closure ID","Candidate Name","Project","Status"])
    # ws.append([args.closure_id,args.candidate_name,args.project,args.status])
    closure_data=frappe.db.get_all("Closure",{"custom_status_transition":nowdate()},["*"],order_by='project')
    closure_status=["PSL","Sales Order","Client Offer Letter","Signed Offer Letter","Visa","Premedical","PCC","Certificate Attestation","Final Medical","Biometric","Visa Stamping","Emigration","Ticket","Onboarding","Arrived","Dropped","Waitlisted"]
    for i in closure_data:
        if i.status in closure_status:
            indx=closure_status.index(i.status)
            next_indx=closure_status[indx-1]
        data_row = [i.name or '', i.given_name or '', i.project or '',next_indx, i.status or '']
        ws.append(data_row)
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


@frappe.whitelist()
def closure_mail_client(client_email,subject,id,action_taken,live,et,at,revision,service,proof,allocated=None,project=None,issue=None,domain=None,spoc=None,reason=None):
    if service=='IT-SW':
        percentage=et_at_calculation(id, et, at, allocated,subject)
        reports=frappe.db.get_value("Employee",{'user_id':allocated},['reports_to'])
        reports_to=frappe.db.get_value("Employee",{'name':reports},['user_id'])
        if issue:
            raised_by=frappe.db.get_value("Issue",{'name':issue},['raised_by'])
        else:
            raised_by='None'
        data = ''
        data += f"<table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>\
        <tr><td colspan='2' style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black;'><b>Task / Issue Client Review Note</b></td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Task ID</b></td><td style='border: 1px solid black;'>{id}</td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Project</b></td><td style='border: 1px solid black;'>{project}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task Raised By</b></td><td style='border: 1px solid black;'>{reports_to}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue ID</b></td><td style='border: 1px solid black;'>{issue}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Raised By</b></td><td style='border: 1px solid black;'>{raised_by}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Statement</b></td><td style='border: 1px solid black;'>{subject}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Action Taken</b></td><td style='border: 1px solid black;'>{action_taken}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Live At</b></td><td style='border: 1px solid black;'>{live}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Domain</b></td><td style='border: 1px solid black;'>{domain}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Proof</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/{proof}' target='_blank'>Link to Proof</a></td></tr>\
        </table>"

        frappe.sendmail(
            sender=allocated,
            # recipients=client_email,
            recipients=client_email,
            cc=[spoc,allocated,reports_to,'anil.p@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject='Task : %s Client Review : Forward for Review to Mark Completion or Re-Open' % id,
            message = """
            <b>Dear Patron,<br><br>Greeting !!!</b><br><br>
           The attached Task has been completed by Development and forwarded for your kind review, please confirm if it satisfies all your requirement and Mark the Task Status as Completed or if you feel it is still pending for some action please give your remark for Re-open <br><br>
           {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(data)
        )
        

@frappe.whitelist()
def get_allocated_tasks(date,name,service,type):
    total=0
    percent=0
    if type:
        task_id=frappe.db.get_all("Task",{"custom_production_date":date,"service":service,"type":type},['*'],order_by='cb asc, project asc, priority asc')
    else:
        task_id=frappe.db.get_all("Task",{"custom_production_date":date,"service":service},['*'],order_by='cb asc, project asc, priority asc')

    task_det=frappe.db.get_all("Task",{"custom_production_date":date,"service":service},['*'],order_by='cb asc',group_by='custom_allocated_to asc')
    parent_doc = frappe.get_doc("Daily Monitor", name)
    parent_doc.task_details=[]
    parent_doc.dm_summary=[]
    issue_id=''
    for i in task_id:
        parent_doc.append("task_details", {"id": i.name,"a_task_type":i.type,"cb":i.cb})
        employee_id=frappe.db.get_value('Employee',{'short_code':i.cb},['user_id'])
        issue_id=frappe.db.get_all("Issue",{'assigned_to':employee_id,'status':'Open'},['*'])
    for j in issue_id:
        issue_list=frappe.db.get_value("Employee",{'user_id':j.assigned_to},['short_code'])
        parent_doc.append("task_details", {"id": j.name,"project_name":j.project,"subject":j.subject,"cb":issue_list,"status":j.status})
    for k in task_det:
        actual_aph=frappe.db.get_value('Employee',{'short_code':k.cb},['custom_aph'])
        sum_et=frappe.db.sql("""select sum(rt) as et from `tabTask` where custom_allocated_to=%s and custom_production_date=%s group by cb""",(k.custom_allocated_to,date), as_dict=True)
        emp_cb=frappe.db.get_value('Employee',{'user_id':k.custom_allocated_to},['short_code'])
        total+=sum_et[0].et
        if actual_aph is not None:
            percent=(float(sum_et[0].et)/float(actual_aph))*100
            parent_doc.append("dm_summary",{'d_cb':emp_cb,'d_aph':actual_aph or '8','d_rt':sum_et[0].et,'d_actual_time_taken':'','rt_vs_aph_':round(percent,2) or '0'})
        else:
            percent=(float(sum_et[0].et)/8)*100
            parent_doc.append("dm_summary",{'d_cb':emp_cb,'d_aph':'8','d_rt':sum_et[0].et,'d_actual_time_taken':'','rt_vs_aph_':round(percent,2) or '0'})
    parent_doc.save()
    frappe.db.commit()
    frappe.db.set_value("Daily Monitor",name,'dm_status',"DPR Pending")
    # frappe.db.set_value("Daily Monitor",name,'workflow_state',"DPR Pending")

    # return data



@frappe.whitelist()
def update_dsr(date, name,service,type):
    parent_doc = frappe.get_doc("Daily Monitor",name)
    parent_doc.task_details = []
    if type:
        task_id_list = frappe.db.get_all("Task", {"custom_production_date":date,"service":service,"type":type}, ['*'], order_by='custom_allocated_to asc, project asc, priority asc')
    else:
        task_id_list = frappe.db.get_all("Task", {"custom_production_date": date,"service":service}, ['*'], order_by='custom_allocated_to asc, project asc, priority asc')
       
    task_list = frappe.db.get_all("Task", {"custom_production_date":date}, ['*'], order_by='cb asc, project asc, priority asc',group_by='custom_allocated_to asc')
    issues = []
    meetings = []
    tasks = []
    appended_issues = set()
    appended_meetings = set()
    parent_doc.dm_summary=[]
    for task in task_id_list:
        emp_id = frappe.db.get_value("Employee", {'user_id': task.custom_allocated_to}, ['name'])
        emp_short_code= frappe.db.get_value("Employee", {'name': emp_id},['short_code'])
        timesheet = frappe.db.get_value("Timesheet", {'start_date':date, 'employee': emp_id},['name'])   
        task_hours_total = 0.0 
        if timesheet:
            issue_logs = frappe.get_all("Timesheet Detail", filters={'parent': timesheet, 'custom_issue': ['!=', '']}, fields=['*'])
            i_taken=0.0
            for issue in issue_logs:
                i_taken+=issue.hours
                if issue.custom_issue not in appended_issues:
                    short_code=frappe.db.get_value("Employee",{'name':emp_id},['short_code'])
                    prior = frappe.db.get_value("Issue", {'name': issue.custom_issue}, ['priority'])
                    i_id = frappe.db.get_value("Issue", {'name': issue.custom_issue}, ['status'])
                    issues.append({
                        "id": issue.custom_issue,
                        "at_taken": issue.hours,
                        'project_name': issue.project_name,
                        'subject': issue.custom_subject_issue,
                        'status': i_id,
                        'cb':short_code,
                        'priority':prior
                    })
                    appended_issues.add(issue.custom_issue)            
            meeting_logs = frappe.get_all("Timesheet Detail", filters={'parent': timesheet, 'custom_meeting': ['!=', '']}, fields=['*'])
            for meeting in meeting_logs:
                if meeting.custom_meeting not in appended_meetings:
                    m_id = frappe.db.get_value("Issue", {'name': meeting.custom_meeting}, ['status'])
                    short_code=frappe.db.get_value("Employee",{'name':emp_id},['short_code'])
                    meetings.append({
                        "id": meeting.custom_meeting,
                        "at_taken": meeting.hours,
                        'subject':meeting.custom_subject_meeting,
                        'cb':short_code
                    })
                    appended_meetings.add(meeting.custom_meeting)  
            # task_hours_total = 0.0
            task_logs = frappe.get_all("Timesheet Detail", filters={'parent': timesheet, 'task': task.name}, fields=['hours'])
            # timesheet = frappe.db.get_value("Timesheet", {'start_date': date, 'employee':emp_id}, ['total_hours'])      
            for log in task_logs:
                task_hours_total += log.hours
        tasks.append({
            "id": task.name,
            "at_taken": task_hours_total,
            "a_task_type":task.type,
            "cb":emp_short_code,
            "current_status":task.status
        })
    for d in issues:
        parent_doc.append("task_details", d)
    for meeting in meetings:
        parent_doc.append("task_details", meeting)
    for task in tasks:
        parent_doc.append("task_details", task)
    parent_doc.dsr_check = 1
    for j in task_list:
            employee_id=frappe.db.get_value('Employee',{'user_id':j.custom_allocated_to},['name'])
            emp_cb=frappe.db.get_value('Employee',{'user_id':j.custom_allocated_to},['short_code'])
            timesheet = frappe.db.get_value("Timesheet", {'start_date': date, 'employee':employee_id}, ['total_hours'])      
            actual_aph=frappe.db.get_value('Employee',{'name':employee_id},['custom_aph'])
            sum_et=frappe.db.sql("""select sum(rt) as et from `tabTask` where custom_allocated_to=%s and custom_production_date=%s group by custom_allocated_to""",(j.custom_allocated_to,date), as_dict=True)
            if timesheet and actual_aph is not None:
                percent=(float(timesheet)/float(actual_aph))*100
                parent_doc.append("dm_summary",{'d_cb':emp_cb,'d_aph':actual_aph,'d_rt':sum_et[0].et if sum_et else '0','d_actual_time_taken':round(timesheet,2),'rt_vs_aph_':round(percent,2)})
            else:
                parent_doc.append("dm_summary",{'d_cb':emp_cb,'d_aph':actual_aph,'d_rt':sum_et[0].et if sum_et else '0' ,'d_actual_time_taken':'0','rt_vs_aph_':'0'})
    parent_doc.save()
    frappe.db.commit()
    frappe.db.set_value('Daily Monitor',name,'dm_status','DSR Pending')

@frappe.whitelist()
def dpr_task_mail(name,date,service,task_type):
    total=0
    total_count=0
    percent=0
    or_count=0
    pr_count=0
    date_obj = datetime.strptime(date, '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00005'},['*'])
    recievers=[]
    for i in emp:
        recievers.append(i.user_id)
    recievers.append('abdulla.pi@groupteampro.com')
    recievers.append('dineshbabu.k@groupteampro.com')
    task_data=frappe.get_doc("Daily Monitor",name)
    # # priority = {"High": 1, "Medium": 2, "Low": 3}
    task=frappe.db.get_all("Task",{"custom_production_date":date},['*'],group_by='custom_allocated_to asc',order_by='cb asc, project asc, priority asc')
    total_at=0
    if task_data.dsr_check==1:
        count=1
        aph_totals=0
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '''
        <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
            <td style='width:4%'><b>SI NO</b></td>
            <td style='width:6%'><b>Task/Issue ID</b></td>
            <td style='width:12%'><b>Project </b></td>
            <td style='width:18%'><b>Subject</b></td>
            <td style='width:4%'><b>CB</b></td>
            <td style='width:7%'><b>Status</b></td>
            <td style='width:4%'><b>Revision</b></td>
            <td style='width:4%'><b>AT</b></td>
            <td style='width:4%'><b>ET</b></td>
            <td style='width:4%'><b>RT</b></td>
            <td style='width:6%'><b>Priority</b></td>
           <td style='width:8%'><b>Allocated On</b></td>
           <td style='width:4%'><b>Time Taken</b></td>
           <td style='width:10%'><b>Remarks</b></td>
           <td style='width:9%'><b>TL Remarks</b></td>
        </tr>
        '''
        table = '<table border="1" width="70%" style="border-collapse: collapse;text-align:center;">'
        table += '''
        <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
            <td style='width:1%'><b>CB</b></td>
            <td style='width:1%'><b>APH</b></td>
            <td style='width:1%'><b>RT</b></td>
            <td style='width:1%'><b>Actual Time Taken</b></td>
            <td style='width:1%'><b>RT Vs APH %</b></td>
            <td style='width:1%'><b>OR</b></td>
            <td style='width:1%'><b>PR</b></td>
        </tr>
        '''
        sorted_task_details = sorted(task_data.task_details, key=lambda i: i.cb)
        count = 1
        for i in sorted_task_details:
    #         # emp_cb=frappe.db.get_value('Employee',)
            vtaken = float(i.at)
            value_taken = round(vtaken, 3)
            if i.at_taken:
                t_taken = float(i.at_taken)
                today_taken = round(t_taken, 3)
            else:
                t_taken='0'
                today_taken='0'
            remark = '-' if i.remark is None else i.remark
            tl_remark = '-' if i.tl_remark is None else i.tl_remark
            if i.id is not None:
                id=i.id
            elif i.issue is not None:
                id=i.issue
            elif i.meeting is not None:
                id=i.meeting
            else:
                id='-'
            data += '<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'% (count,id, i.project_name, i.subject, i.cb, i.status, i.revisions, value_taken,i.et, i.rt, i.priority, i.allocated_on, today_taken, remark, tl_remark)
            count += 1
        data += '</table>'
        for j in task:
            employee_id=frappe.db.get_value('Employee',{'user_id':j.custom_allocated_to},['name'])
            emp_cb=frappe.db.get_value('Employee',{'user_id':j.custom_allocated_to},['short_code'])
            timesheet = frappe.db.get_value("Timesheet", {'start_date': date, 'employee':employee_id}, ['total_hours'])  
            actual_aph=frappe.db.get_value('Employee',{'short_code':emp_cb},['custom_aph'])
            sum_et=frappe.db.sql("""select sum(rt) as et from `tabTask` where custom_allocated_to=%s and custom_production_date=%s and service="IT-SW" group by custom_allocated_to""",(j.custom_allocated_to,date), as_dict=True)
            if sum_et:
                total+=sum_et[0].et
            if actual_aph is not None:
                aph_totals+=float(actual_aph)
            if timesheet is not None:
                total_at+=float(timesheet)
            if actual_aph and timesheet:
                percent=(float(timesheet)/float(actual_aph))*100
                value=actual_aph
                total_count=float(total)/float(aph_totals)*100
                or_count=float(timesheet)/float(value)*100
                pr_count=float(sum_et[0].et)/float(timesheet)*100
                or_total=float(total_at)/float(aph_totals)*100
                pr_total=float(total)/float(total_at)*100
            if percent:
                table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,actual_aph or '8',sum_et[0].et if sum_et else '0',round(timesheet,2) if timesheet is not None else '0',round(percent,2) if timesheet is not None else '0',round(or_count) if timesheet is not None else '0',round(pr_count) if timesheet is not None else '0')
            else:
                table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,actual_aph or '8',sum_et[0].et if sum_et else '0',round(timesheet,2) if timesheet is not None else'0','0',round(or_count,2) or '0',round(pr_count,2) or '0')
        table+='<tr style="font-size: 14px;" ><td colspan=1>Total</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(aph_totals,total,round(total_at,2),round(total_count,2),round(or_total),round(pr_total))
        table += "</table>"
           
        
        frappe.sendmail(
                # sender='abdulla.pi@groupteampro.com',
                # recipients='dineshbabu.k@groupteampro.com',
                # cc='abdulla.pi@groupteampro.com',
                # recipients='divya.p@groupteampro.com',
                subject='DSR %s -Reg' % formatted_date,
                message = """
               <b>Dear Team,</b><br><br>
                Please find the below DSR for {} for your kind reference.<br><br>
                {}<br><br>
                {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,table,data)
            )
        frappe.msgprint("DSR mail has been successfully sent.")
        frappe.db.set_value('Daily Monitor',name,'dm_status','Submitted')
        frappe.db.set_value("Daily Monitor",name,'dsr_submitted_on',today())
    else:
        count=1
        aph_total=0
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '''
        <tr style="background-color: #0f1568 ;text-align:center;color: white;"><b>
            <td style='width:5%'><b>SI NO</b></td>
            <td style='width:10%'><b>ID</b></td>
            <td style='width:15%'><b>Project </b></td>
            <td style='width:20%'><b>Subject</b></td>
            <td style='width:5%'><b>CB</b></td>
            <td style='width:10%'><b>Status</b></td>
            <td style='width:5%'><b>Revision</b></td>
            <td style='width:5%'><b>AT</b></td>
            <td style='width:5%'><b>ET</b></td>
            <td style='width:5%'><b>RT</b></td>
            <td style='width:7%'><b>Priority</b></td>
            <td style='width:13%'><b>Allocated On</b></td>
        </b></tr>
        '''
        table = '<table border="1" width="50%" style="border-collapse: collapse;text-align:center;">'
        table += '''
        <tr style="background-color: #0f1568 ;color: white;text-align:center;font-size: 12px;">
            <td style='width:1%'><b>CB</b></td>
            <td style='width:1%'><b>APH</b></td>
            <td style='width:1%'><b>RT</b></td>
            <td style='width:1%'><b>RT Vs APH%</b></td>
        </tr>
        '''
        for i in task_data.task_details:
            value_taken = round(i.at, 3)
            data+='<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td>'%(count,i.id or i.issue,i.project_name or '-',i.subject,i.cb,i.status,i.revisions,value_taken,i.et,i.rt,i.priority,i.allocated_on or '')
            count+=1
        data += '</table>'
        task_det=frappe.db.get_all("Task",{"custom_production_date":date,"type":task_type,"service":service},['*'],order_by='cb asc',group_by='custom_allocated_to asc')
        for k in task_det:
            employee_id=frappe.db.get_value('Employee',{'user_id':k.custom_allocated_to},['user_id'])
            emp_cb=frappe.db.get_value('Employee',{'user_id':k.custom_allocated_to},['short_code'])
            actual_aph=frappe.db.get_value('Employee',{'short_code':emp_cb},['custom_aph'])
            sum_et=frappe.db.sql("""select sum(rt) as et from `tabTask` where custom_allocated_to=%s and custom_production_date=%s and service="IT-SW" group by custom_allocated_to""",(k.custom_allocated_to,date), as_dict=True)
            if sum_et:
                total+=sum_et[0].et
            if actual_aph is not None and sum_et:
                percent=(float(sum_et[0].et)/float(actual_aph))*100
                value=actual_aph
                aph_total+=float(value)
            total_count=float(total)/float(aph_total)*100
            # print(employee_id)
            if percent:
                table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,value or '8',sum_et[0].et if sum_et else '0',round(percent,2) or '-')
            else:
                table+='<tr style="font-size: 14px;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td>' %(emp_cb,actual_aph or '',sum_et[0].et if sum_et else '0',round(percent,2) or '-')
        table+='<tr style="font-size: 14px;" ><td colspan=1>Total</td><td>%s</td><td>%s</td><td>%s</td>'%(aph_total,total,round(total_count,2))
        table+='</table>'
        frappe.sendmail(
                # sender='abdulla.pi@groupteampro.com',
                # recipients=recievers,
                # recipients='divya.p@groupteampro.com',
                subject='DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action, ensure all the Tasks allocated on time and as per the requirement, for each Revision and AT going beyond 150% there will be NC applied and accumulated NC will be reviewed every week and directly affects your Performance.<br><br>

            {}<br><br>
            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,table,data)
            )
        frappe.msgprint("DPR mail has been successfully sent")
        frappe.db.set_value('Daily Monitor',name,'dm_status','DPR Completed')
        frappe.db.set_value('Daily Monitor',name,'dpr_submitted_on',today())


@frappe.whitelist()
def reverse_revision(user):
    role = frappe.db.sql("""
        SELECT `tabUser`.name as name
        FROM `tabUser`
        LEFT JOIN `tabHas Role` ON `tabHas Role`.parent = `tabUser`.name
        WHERE `tabHas Role`.role = 'Customer Executive'
        AND `tabUser`.enabled = 1
        AND `tabUser`.name = %s
    """, (user,), as_dict=True)

    if role:
        return role[0].get('name')
    else:
        return None

@frappe.whitelist()
def reverse_revision_nc(name):
    reason = frappe.db.get_value("Energy Point And Non Conformity", {'reason_of_ep': ['like', '%'+name+'%']}, "name")

    if reason:
        nc = frappe.get_doc("Energy Point And Non Conformity", reason)
        nc.delete()

@frappe.whitelist()
def task_type_updates(name,id):
    cb=frappe.db.get_value("Employee",{"user_id":name},["short_code"])
    # frappe.db.set_value("Task",id,'cb',cb)
    return cb


@frappe.whitelist()
def dpr_send_alert():
    date = datetime.strptime(today(), '%Y-%m-%d')
    formatted_date = date.strftime('%d/%m/%Y')
    today_date=today()
    doc_name = frappe.db.get_value("Daily Monitor",{"date": today_date,"service":"IT-SW"},["name"])
    if doc_name:
        print("hi")
        parent_doc = frappe.get_doc("Daily Monitor", doc_name)
        if parent_doc.dpr_submitted_on is None:
            hod_mail=frappe.db.get_value("Services",{"name":"IT-SW"},['hod'])
            frappe.sendmail(
                recipients=[hod_mail,"divya.p@groupteampro.com"],
                # recipients="divya.p@groupteampro.com",
                subject= "DPR Reminder - %s" % formatted_date,
                message=""" Dear Sir,<br><br>
                The DPR is not updated today.Kindly update the DPR<br><br>
                Thanks & Regards,<br>TEAM ERP<br>"""
            )
    else:
        print("hello")
        hod_mail=frappe.db.get_value("Services",{"name":"IT-SW"},['hod'])
        frappe.sendmail(
            recipients=[hod_mail,"divya.p@groupteampro.com"],
            # recipients="divya.p@groupteampro.com",
            subject= "Daily Monitor Reminder - %s" % formatted_date,
            message=""" Dear Sir,<br><br>
            The Daily Monitor  is not created today.Kindly create the Daily Monitor<br><br>
            Thanks & Regards,<br>TEAM ERP<br>"""
        )

@frappe.whitelist()
def dsr_send_alert():
    date = datetime.strptime(add_days(today(),-1), '%Y-%m-%d')
    formatted_date = date.strftime('%d/%m/%Y')
    dates = datetime.strptime(today(), '%Y-%m-%d')
    formatted_dates = dates.strftime('%d/%m/%Y')
    previous_date = add_days(today(),-1)
    doc_name = frappe.db.get_value("Daily Monitor",{"date": previous_date,"service":"IT-SW"},["name"])
    print(doc_name)
    if doc_name:
        parent_doc = frappe.get_doc("Daily Monitor", doc_name)
        if parent_doc.dsr_submitted_on == "":
            hod_mail=frappe.db.get_value("Services",{"name":"IT-SW"},['hod'])
            frappe.sendmail(
                # recipients=[hod_mail,"divya.p@groupteampro.com"],
                recipients="divya.p@groupteampro.com",
                subject= "DSR Reminder - %s" % formatted_dates,
                message=""" Dear Sir,<br><br>
                The DSR is not updated for -%s.Kindly update the DSR.<br><br>
                Thanks & Regards,<br>TEAM ERP<br>""" % formatted_date
            )

@frappe.whitelist()
def update_tat_completion_date(name):
    doc=frappe.get_doc("Case",name)
    if doc.insufficiency_closed:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insufficiency_closed
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        sql_query = f"""
            SELECT COUNT(*) 
            FROM `tabHoliday` 
            WHERE parent = 'TEAMPRO 2023 - Checkpro' 
            AND holiday_date BETWEEN '{doc.insufficiency_closed}' AND '{holiday[-1]}'
        """
        count = frappe.db.sql(sql_query, as_list=True)[0][0]
        return holiday[-1],count

@frappe.whitelist()
def update_issue_type(doc,method):
    frappe.db.set_value("Issue",doc.issue,"custom_issue_status",doc.status)
    frappe.db.set_value("Issue",doc.issue,"issue_type",doc.custom_issue_type)
    if doc.status=="Open" or doc.status=="Overdue":
        frappe.db.set_value("Issue",doc.issue,"status","Open")
    elif doc.status=="Hold":
        frappe.db.set_value("Issue",doc.issue,"status","On Hold")
    elif doc.status=="Working":
        frappe.db.set_value("Issue",doc.issue,"status","Replied")
    elif doc.status=="Pending Review" or doc.status=="Client Review":
        frappe.db.set_value("Issue",doc.issue,"status","Resolved")
    elif doc.status=="Completed" or doc.status=="Cancelled":
        frappe.db.set_value("Issue",doc.issue,"status","Closed")


    # if doc.status in ["Open","Overdue","Hold"]:
    #     frappe.db.set_value("Issue",doc.issue,"status","Open")
    # elif doc.status in ["Working"]:
    #     frappe.db.set_value("Issue",doc.issue,"status","Replied")
    # elif doc.status in ["Pending Review","Client Review"]:
    #     frappe.db.set_value("Issue",doc.issue,"status","Resolved")
    # elif doc.status in ["Completed","Cancelled"]:
    #     frappe.db.set_value("Issue",doc.issue,"status","Closed")

       

@frappe.whitelist()
def update_issue_typein_issue(doc,method):
    frappe.db.set_value("Issue",doc.issue,"task",doc.name)  
    
@frappe.whitelist()
def update_country_flag(doc, method):
    mobile_no = frappe.db.get_value("Employee", {"user_id":doc.custom_allocated_to}, ["company_mobile_number"])
    if mobile_no:
        if doc.service in ["REC-I", "REC-D"]:
            if doc.territory:
                flag_url = frappe.db.get_value("Territory", {"name": doc.territory}, ["custom_country_flag"])
                doc.custom_country_flag = flag_url
            if doc.custom_allocated_to:
                doc.custom_recruiter_contact = mobile_no
                
from datetime import datetime
@frappe.whitelist()
def update_issue_status():
    issues=frappe.get_all("Issue",{'custom_issue_status':('in',['Open' ,'Working'])},['custom_expected_end_date','name',"custom_issue_status"])
    for issue in issues:
        if issue.custom_expected_end_date < datetime.now().date():
            frappe.db.set_value("Issue",issue.name,"custom_issue_status","Overdue")


@frappe.whitelist()
def client_reviwew_mail(client_email,subject,created_by,action_taken,issue,live,proof,project=None,domain=None):
    data = ''
    data += f"<table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>\
    <tr><td colspan='2' style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black;'><b>Issue Client Review Note</b></td></tr>\
    <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Issue ID</b></td><td style='border: 1px solid black;'>{issue}</td></tr>\
    <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Project</b></td><td style='border: 1px solid black;'>{project}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Raised By</b></td><td style='border: 1px solid black;'>{created_by}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Statement</b></td><td style='border: 1px solid black;'>{subject}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Action Taken</b></td><td style='border: 1px solid black;'>{action_taken}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Live At</b></td><td style='border: 1px solid black;'>{live}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Domain</b></td><td style='border: 1px solid black;'>{domain}</td></tr>\
    <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Proof</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/{proof}' target='_blank'>Link to Proof</a></td></tr>\
    </table>"

    frappe.sendmail(
        sender=created_by,
        # recipients=["divya.p@groupteampro.com"],
        recipients=client_email,
        # cc=[allocated,reports_to,'anil.p@groupteampro.com','dineshbabu.k@groupteampro.com'],
        subject='Task : %s Client Review : Forward for Review to Mark Completion' % issue,
        message = """
        <b>Dear Patron,<br><br>Greeting !!!</b><br><br>
        The attached Issue has been completed by Development and forwarded for your kind review, please confirm if it satisfies all your requirement and Mark the Issue Status as Completed  <br><br>
        {}<br><br>
        Thanks & Regards,<br>TEAM ERP<br>
        
        <i>This email has been automatically generated. Please do not reply</i>
        """.format(data)
    )



@frappe.whitelist()
def purchase_invoice_due_above():
    purchase_invoices = frappe.get_all("Purchase Invoice",{"status": ("in", ["Partly Paid", "Unpaid", "Overdue"]),"docstatus": ("!=",2),"due_date": (">=", nowdate())},["*"])
    # print(purchase_invoices)
    # purchase_invoices = frappe.get_all("Purchase Invoice",{"status": "Paid","docstatus": ("!=",2),"due_date": (">=", nowdate())},["*"])
    
    count = 1
    f_count=0
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
    <tr style="background-color: #0f1568; text-align:center; color: white;">
        <td style='width:5%'><b>SI NO</b></td>
        <td style='width:10%'><b>Invoice Number</b></td>
        <td style='width:15%'><b>Party Name</b></td>
        <td style='width:20%'><b>Bill Value</b></td>
        <td style='width:5%'><b>Outstanding Value</b></td>
        <td style='width:10%'><b>Age of the Bill</b></td>
        <td style='width:10%'><b>Remarks</b></td>
    </tr>
    '''
    current_date = getdate(today())
    for i in purchase_invoices:
        print(i.due_date)
        posting_date = getdate(i.posting_date)
        due=getdate(i.due_date)
        age_of_bill = (current_date - posting_date).days
        age=(due -current_date).days
        data += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>The Due date for this invoice will  approach in {} days</td></tr>'.format(count,i.name, i.supplier,i.grand_total,i.outstanding_amount,age_of_bill,age)
        count += 1
        f_count +=1
    print(count)

    data += '</table>'
    if f_count>=1:
        frappe.sendmail(
            recipients=['accounts@groupteampro.com','sangeetha.s@groupteampro.com'],
            subject='Purchase Invoice-Due Date Approaching',
            message="""
            <b>Dear Sir/Mam,</b><br><br>
            Please find the below purchase invoice list for your kind reference and action.<br><br>
            {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(data)
        )



@frappe.whitelist()
def purchase_invoice_beyond_duedate():
    purchase_invoices = frappe.db.get_all("Purchase Invoice",{"status": ["in", ["Partly Paid", "Unpaid", "Overdue"]],"docstatus": ["!=",2],"due_date": ["<=", nowdate()]},["*"])
    count = 1
    f_count=0
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '''
    <tr style="background-color: #0f1568; text-align:center; color: white;">
        <td style='width:5%'><b>SI NO</b></td>
        <td style='width:10%'><b>Invoice Number</b></td>
        <td style='width:15%'><b>Party Name</b></td>
        <td style='width:20%'><b>Bill Value</b></td>
        <td style='width:5%'><b>Outstanding Value</b></td>
        <td style='width:10%'><b>Age of the Bill</b></td>
    </tr>
    '''
    current_date = getdate(today())
    for i in purchase_invoices:
        posting_date = getdate(i.posting_date)
        age_of_bill = (current_date - posting_date).days
        data += '<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(count,i.name, i.supplier,i.grand_total,i.outstanding_amount,age_of_bill)
        count += 1
        f_count +=1
    print(count)

    data += '</table>'
    if f_count>=1:
        frappe.sendmail(
            recipients=['accounts@groupteampro.com','sangeetha.s@groupteampro.com'],
            subject='Purchase Invoice-Beyond Due Date',
            message="""
            <b>Dear Sir/Mam,</b><br><br>
            Please find the below purchase invoice list for your kind reference and action.<br><br>
            {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(data)
        )


@frappe.whitelist()
def sales_dpr():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
    recievers=[]
    custom_date = today()
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    data=[]
    for i in emp:
        recievers.append(i.user_id)
    recievers.append('anil.p@groupteampro.com')
    emp_list=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com'])},['*'])
    for j in emp_list:
        recievers.append(j.user_id)
    recievers.append('annie.m@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="8"><b>R&S DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Appointments</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Opportunity</b></td>
            <td style="width:13%;"><b>Customer</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    appointments = frappe.get_all("Appointment", filters={"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]}, fields=["*"])

    appointment_count = 0
    
    for user_email in recievers:
        for appointment in appointments:
        # Get the distinct user count for each appointment
            user_counts = frappe.db.sql("""
                SELECT DISTINCT c.user, COUNT(c.user) AS count 
                FROM `tabAppointment` p 
                INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
                WHERE c.user like %s and p.name = %s
            """, (user_email, appointment.name), as_dict=True)
            if user_counts:
                for user_count in user_counts:
                    print(user_count['user'], user_count['count'])
                appointment_count = user_count['count']
            else:
                appointment_count = 0
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        # appointment_count=frappe.db.count("Sales Follow Up",{"appointment_date":["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
    
    todo_list=frappe.db.get_all("ToDo",{"custom_production_date":formatted_next_date},["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        s_no=0
        for i in todo_list:
            s_no+=1
            short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.name,i.custom_subject)
    
    
    appointment_list = frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.scheduled_time between '%s' and '%s'""" %(formatted_next_date + " 00:00:00", formatted_next_date + " 23:59:59"),as_dict=1)
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Appointment</td></tr>
        '''
        for i in appointment_list:
            short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.time,i.name)
            
    data += '</table>'



    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['anil.p@groupteampro.com','annie.m@groupteampro.com'],
                cc='dineshbabu.k@groupteampro.com',
                subject='R&S DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    
    recievers = ['harish.g@groupteampro.com', 'aarthi.e@groupteampro.com', 'vijiyalakshmi.k@groupteampro.com']

    for user_email in recievers:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="8"><b>R&S DPR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appointments</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Opportunity</b></td>
                <td style="width:13%;"><b>Customer</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
            </tr>
        '''

        # Fetch appointment counts for each user
        appointments = frappe.get_all("Appointment", filters={"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]}, fields=["*"])
        appointment_count = frappe.db.count("Appointment", {"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]], "owner": user_email})

        # Other counts
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_next_date})

        # Populate table row for the specific user
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            short_code, appointment_count, lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count
        )

        # ToDo Section
        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": user_email}, ["*"])
        if todo_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
            '''
            for todo in todo_list:
                short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject)

        # Appointment Section
        appointment_list = frappe.db.sql("""
            SELECT p.scheduled_time AS time, p.name AS name, c.user AS user 
            FROM `tabAppointment` p 
            INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
            WHERE p.scheduled_time BETWEEN %s AND %s AND c.user = %s
        """, (formatted_next_date + " 00:00:00", formatted_next_date + " 23:59:59", user_email), as_dict=True)

        if appointment_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>Appointment</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Appointment</td></tr>
            '''
            for appointment in appointment_list:
                short_code = frappe.db.get_value("Employee", {"user_id": appointment.user}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, appointment.time, appointment.name)

        data += '</table>'

        # Send the email to the specific user
        frappe.sendmail(
            recipients=[user_email],
            subject='R&S DPR %s - Reg' % formatted_date,
            message="""
                <b>Dear {user},</b><br><br>
                Please find the below DPR for {date} for your kind reference and action.<br><br>
                {table}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """.format(user=user_email.split('@')[0], date=formatted_date, table=data)
        )

    
@frappe.whitelist()
def send_sales_dsr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date=nowdate()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    appointments = frappe.get_all("Appointment", filters={"custom_completed_date":formatted_next_date}, fields=["*"])

    for i in emp:
        emp_emails.append(i.user_id)
    emp_emails.append('anil.p@groupteampro.com')
    emp_list=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com'])},['*'])
    for j in emp_list:
        emp_emails.append(j.user_id)
    emp_emails.append('annie.m@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="8"><b>R&S DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td width="25%" colspan="1"><b>Exe</b></td>
            <td width="25%" colspan="1"><b>Effective</b></td>
            <td width="25%" colspan="1"><b>Non Effective</b></td>
            <td width="25%" colspan="1"><b>Appointment</b></td>    
            <td width="25%" colspan="1"><b>Total</b></td>
        </tr>
    '''
    for c in emp_emails:
        for appointment in appointments:
        # Get the distinct user count for each appointment
            user_counts = frappe.db.sql("""
                SELECT DISTINCT c.user, COUNT(c.user) AS count 
                FROM `tabAppointment` p 
                INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
                WHERE c.user like %s and p.name = %s
            """, (c, appointment.name), as_dict=True)
            if user_counts:
                for user_count in user_counts:
                    print(user_count['user'], user_count['count'])
                appointment_count = user_count['count']
            else:
                appointment_count = 0
        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
        non_effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Non Effective"})
        total_calls = effective_call + non_effective_call
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,effective_call if effective_call else '0' , non_effective_call if non_effective_call else '0',appointment_count if appointment_count else '0',total_calls if total_calls else '0'
            )
    appointment_list = frappe.db.sql("""select p.status as status, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s'""" %(formatted_next_date),as_dict=1)
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="2" style="text-align:center; ">Customer</td><td colspan="2">Status</td></tr>
        '''
        for i in appointment_list:
            short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td></tr>'.format(short_code,i.name,i.status)

    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": ["in", emp_emails], "status": ('not in',['Cancelled'])}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="1" style=" text-align: center;">Subject</td><td colspan="2">Current Status</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 10px;">{}</td><td colspan="2">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                # sender='sangeetha.a@groupteampro.com',
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['anil.p@groupteampro.com','annie.m@groupteampro.com'], 
                cc='dineshbabu.k@groupteampro.com',
                subject='R&S DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )

@frappe.whitelist()
def update_opportunity_age():
    opportunity = frappe.db.get_all("Opportunity",{"status":("not in", ["Lost", "Converted"])},["name","transaction_date"])
    current_date = getdate(today())
    ind = 0
    for i in opportunity:
        opportunity_date = getdate(i.transaction_date)
        age = (current_date - opportunity_date).days
        ind += 1
        frappe.db.set_value("Opportunity",i.name,"custom_opportunity_age",age)






from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import frappe


def send_project_spoc_report_daily():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    spoc_list = get_spoc_list()

    for spoc in spoc_list:
        filename = "DSR_" + spoc + "_" + posting_date
        xlsx_file = build_xlsx_response_spoc_project(spoc, filename)
        send_mail_with_attachment_spoc_project(spoc, filename, xlsx_file.getvalue())

def send_mail_with_attachment_spoc_project(spoc, filename, file_content):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    subject = f"DSR Report for {spoc} : - {posting_date}"
    message = (
        f"Dear {spoc},<br>"
        "Please find attached the DSR Report.<br><br>"
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    
    # Send the email for each SPOC
    frappe.sendmail(
        recipients=[spoc],  # Assuming spoc is the email ID of the SPOC
        # recipients=['jeniba.a@groupteampro.com'],
        sender=None,
        subject=subject,
        message=message,
        attachments=attachments,
    )

def build_xlsx_response_spoc_project(spoc, filename):
    return make_xlsx_spoc_project(spoc, filename)

def get_spoc_list():
    projects = frappe.get_all("Project", filters={'status': 'Open', 'service': 'IT-SW',"spoc":("not in",["abdulla.pi@groupteampro.com","sarath.v@groupteampro.com"])}, fields=['spoc'])
    spoc_set = {project['spoc'] for project in projects if project.get('spoc')}
    return list(spoc_set)

def make_xlsx_spoc_project(spoc, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "DSR Report"
    today = datetime.now().strftime('%Y-%m-%d')
    header_fill = PatternFill(start_color="A6CAF0", end_color="A6CAF0", fill_type="solid")
    head_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    row_fill = PatternFill(start_color="FFC1CC", end_color="FFC1CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=17):
        for cell in row:
            cell.border = thin_border
    posting_date = datetime.now().strftime("%d-%m-%Y")
    spoc_code = frappe.db.get_value("Employee", {"user_id": spoc}, ["short_code"])
    header_value = f"{spoc_code} DSR {posting_date}"

    # Manually place the header value in the first cell
    first_cell = ws.cell(row=1, column=1)
    first_cell.value = header_value

    # Apply styles to the first cell (where the value is placed)
    first_cell.fill = header_fill
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")
    first_cell.border = thin_border

    # Merge cells from column 1 to 17
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

    # Apply styles to the entire merged range (though value only goes into the first cell)
    for col in range(1, 18):  # Merged range is from column 1 to 17
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    headers = ["S.NO", "Project", "Priority", "New", "", "Open", "", "Working", "", "Overdue", "", "PR", "", "CR", "", "Total", ""]
    sub_headers = ["", "", "", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue"]
    ws.append(headers)
    ws.append(sub_headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = head_fill
        cell.font = Font(bold=True, color="FFFFFF")  
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for col in range(1, len(sub_headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = row_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    merged_cell = ws.cell(row=3, column=1)  # This is the top-left cell of the merged area

    # Create a new fill color
    new_fill_color = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    # Apply the new fill color to the merged cell
    merged_cell.fill = new_fill_color

    for col in range(4, 17, 2):
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
    serial_number = 1
    priority_rows = [("High",), ("Medium",), ("Low",)]

    cust = frappe.db.get_all("Project", {"status":"Open","spoc": spoc, "service": "IT-SW"}, ["*"])
    total_new_tasks=0
    total_new_issues=0
    total_open_tasks=0
    total_open_issues=0
    total_working_tasks=0
    total_working_issues=0
    total_overdue_tasks=0
    total_overdue_issues=0
    total_pr_tasks=0
    total_pr_issues=0
    total_cr_tasks=0
    total_cr_issues=0
    total_all_tasks=0
    total_all_issues=0
    current_row = 4 
    s_row=4
    for c in cust:
        total_task_count = 0
        total_issue_count = 0
        priority_levels = ["High", "Medium", "Low"]
        h_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "High","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
        h_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "High","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
        h_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "High"})
        h_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "High"})
        h_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"High"})
        h_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"High"})
        h_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"High"})
        h_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"High"})
        high_task_count = h_open_taskcount + h_working_taskcount + h_overdue_taskcount + h_pr_taskcount + h_cr_taskcount 
        high_issue_count = h_open_issuecount + h_working_issuecount + h_overdue_issuecount + h_pr_issuecount + h_cr_issuecount
        e_high_task_count = h_open_taskcount + h_working_taskcount + h_overdue_taskcount + h_pr_taskcount + h_cr_taskcount
        e_high_issue_count = h_open_issuecount + h_working_issuecount + h_overdue_issuecount + h_pr_issuecount + h_cr_issuecount

        total_new_tasks += h_new_taskcount 
        total_new_issues += h_new_issuecount 
        total_open_tasks += h_open_taskcount 
        total_open_issues += h_open_issuecount 
        total_working_tasks += h_working_taskcount 
        total_working_issues += h_working_issuecount 
        total_overdue_tasks += h_overdue_taskcount 
        total_overdue_issues += h_overdue_issuecount 
        total_pr_tasks += h_pr_taskcount 
        total_pr_issues += h_pr_issuecount
        total_cr_tasks += h_cr_taskcount 
        total_cr_issues += h_cr_issuecount
        total_all_tasks += e_high_task_count
        total_all_issues += e_high_issue_count
        # Add to total task/issue counts
        total_task_count += high_task_count
        total_issue_count += high_issue_count
        # Initialize row data
        row_data = [serial_number, c['project_name'], "High"] + [""] * 14
        
        # Prepare the row data, excluding the project name
        row_data[3] = '' if h_new_taskcount == 0 else h_new_taskcount  # Open Task Count (High)
        row_data[4] = '' if h_new_issuecount == 0 else h_new_issuecount  # Open Issue Count (High)
        row_data[5] = '' if h_open_taskcount == 0 else h_open_taskcount  # Open Task Count (High)
        row_data[6] = '' if h_open_issuecount == 0 else h_open_issuecount  # Open Issue Count (High)
        row_data[7] = '' if h_working_taskcount == 0 else h_working_taskcount  # Working Task Count (High)
        row_data[8] = '' if h_working_issuecount == 0 else h_working_issuecount  # Working Issue Count (High)
        row_data[9] = '' if h_overdue_taskcount == 0 else h_overdue_taskcount  # Overdue Task Count (High)
        row_data[10] = '' if h_overdue_issuecount == 0 else h_overdue_issuecount  # Overdue Issue Count (High)
        row_data[11] = '' if h_pr_taskcount == 0 else h_pr_taskcount  # Pending Review Task Count (High)
        row_data[12] = '' if h_pr_issuecount == 0 else h_pr_issuecount  # Pending Review Issue Count (High)
        row_data[13] = '' if h_cr_taskcount == 0 else h_cr_taskcount  # Client Review Task Count (High)
        row_data[14] = '' if h_cr_issuecount == 0 else h_cr_issuecount  # Client Review Issue Count (High)
        row_data[15] = '' if total_task_count == 0 else total_task_count  # Total Task Count (High)
        row_data[16] = '' if total_issue_count == 0 else total_issue_count  # Total Issue Count (High)


        ws.append(row_data)
 
        priority_cell = ws.cell(row=ws.max_row, column=3)  # Column C for "High"
        priority_cell.font = Font(color="FF0000")
        for idx in [3,4,5, 6, 7, 8, 9, 10, 11, 12, 13,14, 15, 16,17]:
            cell = ws.cell(row=ws.max_row, column=idx)
            cell.font = Font(color="FF0000")

        # # Now check the third column for "High" and color it red
        for row in range(4, ws.max_row+1):  # Adjust based on where your actual data starts
            cell = ws.cell(row=row, column=3)  # Third column
            if cell.value and cell.value.strip() == "High":  # Check for "High"
                cell.font = Font(color="FF0000")  # Set the font color to red
 # Change font color to red
        for priority in priority_rows[1:]: 
            total_mediumtask_count = 0
            total_mediumissue_count = 0
     # Start from Medium to avoid duplicating 'High'
            priority_row_data = ["", c['project_name'], priority[0]] + [""] * 14  # Priority in column 3, rest as blanks
            # priority_row_data[1] = c['project_name']
            priority_cell = ws.cell(row=ws.max_row, column=3)  # Column C for Medium/Low
            priority_cell.font = Font(color="000000")  # Set the font color to black
            # Add counts for Medium and Low priority
            if priority[0] == "Medium":
                m_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "Medium","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                m_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "Medium","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                m_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "Medium"})
                m_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "Medium"})
                m_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"Medium"})
                m_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"Medium"})
                m_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"Medium"})
                m_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"Medium"})
                medium_task_count = m_open_taskcount + m_working_taskcount + m_overdue_taskcount + m_pr_taskcount + m_cr_taskcount 
                medium_issue_count = m_open_issuecount + m_working_issuecount + m_overdue_issuecount + m_pr_issuecount + m_cr_issuecount
                e_medium_task_count = m_open_taskcount + m_working_taskcount + m_overdue_taskcount + m_pr_taskcount + m_cr_taskcount 
                e_medium_issue_count = m_open_issuecount + m_working_issuecount + m_overdue_issuecount + m_pr_issuecount + m_cr_issuecount 
                total_new_tasks +=  m_new_taskcount 
                total_new_issues +=m_new_issuecount 
                total_open_tasks += m_open_taskcount 
                total_open_issues +=  m_open_issuecount
                total_working_tasks +=  m_working_taskcount 
                total_working_issues +=  m_working_issuecount 
                total_overdue_tasks +=  m_overdue_taskcount 
                total_overdue_issues +=  m_overdue_issuecount 
                total_pr_tasks += m_pr_taskcount 
                total_pr_issues +=m_pr_issuecount 
                total_cr_tasks += m_cr_taskcount 
                total_cr_issues += m_cr_issuecount             # Accumulate to total task/issue counts
                total_mediumtask_count += medium_task_count
                total_mediumissue_count += medium_issue_count
                total_all_tasks += e_medium_task_count
                total_all_issues += e_medium_issue_count 
                priority_row_data[3] = '' if m_new_taskcount == 0 else m_new_taskcount  # Open Task Count (High)
                priority_row_data[4] = '' if m_new_issuecount == 0 else m_new_issuecount  # Open Issue Count (High)
                priority_row_data[5] = '' if m_open_taskcount == 0 else m_open_taskcount  # Open Task Count (Medium)
                priority_row_data[6] = '' if m_open_issuecount == 0 else m_open_issuecount  # Open Issue Count (Medium)
                priority_row_data[7] = '' if m_working_taskcount == 0 else m_working_taskcount  # Working Task Count (Medium)
                priority_row_data[8] = '' if m_working_issuecount == 0 else m_working_issuecount  # Working Issue Count (Medium)
                priority_row_data[9] = '' if m_overdue_taskcount == 0 else m_overdue_taskcount  # Overdue Task Count (Medium)
                priority_row_data[10] = '' if m_overdue_issuecount == 0 else m_overdue_issuecount  # Overdue Issue Count (Medium)
                priority_row_data[11] = '' if m_pr_taskcount == 0 else m_pr_taskcount  # Pending Review Task Count (Medium)
                priority_row_data[12] = '' if m_pr_issuecount == 0 else m_pr_issuecount  # Pending Review Issue Count (Medium)
                priority_row_data[13] = '' if m_cr_taskcount == 0 else m_cr_taskcount  # Client Review Task Count (Medium)
                priority_row_data[14] = '' if m_cr_issuecount == 0 else m_cr_issuecount  # Client Review Issue Count (Medium)
                priority_row_data[15] = '' if total_mediumtask_count == 0 else total_mediumtask_count  # Total Task Count
                priority_row_data[16] = '' if total_mediumissue_count == 0 else total_mediumissue_count  # Total Issue Count


            elif priority[0] == "Low":
                l_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "Low","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                l_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "Low","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                l_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "Low"})
                l_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "Low"})
                l_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"Low"})
                l_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"Low"})
                l_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"Low"})
                l_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"Low"})
                low_task_count = l_open_taskcount + l_working_taskcount + l_overdue_taskcount + l_pr_taskcount + l_cr_taskcount 
                low_issue_count = l_open_issuecount + l_working_issuecount + l_overdue_issuecount + l_pr_issuecount + l_cr_issuecount 
                e_low_task_count = l_open_taskcount + l_working_taskcount + l_overdue_taskcount + l_pr_taskcount + l_cr_taskcount 
                e_low_issue_count = l_open_issuecount + l_working_issuecount + l_overdue_issuecount + l_pr_issuecount + l_cr_issuecount
                total_new_tasks += l_new_taskcount
                total_new_issues += l_new_issuecount
                total_open_tasks += l_open_taskcount
                total_open_issues += l_open_issuecount
                total_working_tasks += l_working_taskcount
                total_working_issues += l_working_issuecount
                total_overdue_tasks += l_overdue_taskcount
                total_overdue_issues += l_overdue_issuecount
                total_pr_tasks += l_pr_taskcount
                total_pr_issues +=l_pr_issuecount
                total_cr_tasks += l_cr_taskcount
                total_cr_issues += l_cr_issuecount
                # Accumulate to total task/issue counts
                total_mediumtask_count += low_task_count
                total_mediumissue_count += low_issue_count
                total_all_tasks += e_low_task_count
                total_all_issues += e_low_issue_count 
                priority_row_data[3] = '' if l_new_taskcount == 0 else l_new_taskcount  # Open Task Count (Low)
                priority_row_data[4] = '' if l_new_issuecount == 0 else l_new_issuecount  # Open Issue Count (Low)
                priority_row_data[5] = '' if l_open_taskcount == 0 else l_open_taskcount  # Open Task Count (Low)
                priority_row_data[6] = '' if l_open_issuecount == 0 else l_open_issuecount  # Open Issue Count (Low)
                priority_row_data[7] = '' if l_working_taskcount == 0 else l_working_taskcount  # Working Task Count (Low)
                priority_row_data[8] = '' if l_working_issuecount == 0 else l_working_issuecount  # Working Issue Count (Low)
                priority_row_data[9] = '' if l_overdue_taskcount == 0 else l_overdue_taskcount  # Overdue Task Count (Low)
                priority_row_data[10] = '' if l_overdue_issuecount == 0 else l_overdue_issuecount  # Overdue Issue Count (Low)
                priority_row_data[11] = '' if l_pr_taskcount == 0 else l_pr_taskcount  # Pending Review Task Count (Low)
                priority_row_data[12] = '' if l_pr_issuecount == 0 else l_pr_issuecount  # Pending Review Issue Count (Low)
                priority_row_data[13] = '' if l_cr_taskcount == 0 else l_cr_taskcount  # Client Review Task Count (Low)
                priority_row_data[14] = '' if l_cr_issuecount == 0 else l_cr_issuecount  # Client Review Issue Count (Low)
                priority_row_data[15] = '' if total_mediumtask_count == 0 else total_mediumtask_count  # Total Task Count
                priority_row_data[16] = '' if total_mediumissue_count == 0 else total_mediumissue_count  # Total Issue Count

            ws.append(priority_row_data)  
            # ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
 
            for idx in [3,4,5, 6, 7, 8, 9, 10, 11, 12, 13,14, 15, 16,17]:
                cell = ws.cell(row=ws.max_row, column=idx)
                cell.font = Font(color="000000")
            for row in range(4, ws.max_row+1):  # Adjust based on where your actual data starts
                cell = ws.cell(row=row, column=3)  # Third column
                if cell.value and cell.value.strip() == "High":  # Check for "High"
                    cell.font = Font(color="FF0000")

        serial_number += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
        ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row+2, end_column=1)
        merged_cell = ws.cell(row=current_row, column=2)
        merges_cell = ws.cell(row=s_row, column=1)  # Get the first cell of the merged range
        merged_cell.alignment = Alignment(horizontal="center",vertical='center')
        merges_cell.alignment = Alignment(horizontal="center",vertical='center')


        current_row += 3
        s_row +=3
    total_row = ["", "TOTAL", "",
        total_new_tasks, total_new_issues,
        total_open_tasks, total_open_issues,
        total_working_tasks, total_working_issues,
        total_overdue_tasks, total_overdue_issues,
        total_pr_tasks, total_pr_issues,
        total_cr_tasks, total_cr_issues,
        total_all_tasks, total_all_issues
    ]
    ws.append(total_row)
    total_row_index = ws.max_row
# Merge the first three cells in the total row
    ws.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=3)
    merged_cell = ws.cell(row=total_row_index, column=1)
    merged_cell.value = "TOTAL"
    for col in range(1, len(total_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = row_fill
        cell.font = Font(bold=True, color="000000")  # Bold black font for the total row
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=17):
        for cell in row:
            cell.border = thin_border
    # Save the workbook to a BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    
    return xlsx_file

@frappe.whitelist()
def rec_update_dsr_test():
    parent_doc = frappe.get_doc("Daily Monitor", "DM-00141")
    task_ids = set()

    # Iterate through existing task details
    for j in parent_doc.dm_rec_task_details:
        candidate_count = frappe.db.sql("""
        SELECT COUNT(cs.status) AS status_count
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE DATE(cs.sourced_date) = %s
        AND c.candidate_created_by = %s
        AND c.task = %s
        AND cs.status IN (%s, %s, %s, %s, %s, %s)
        """, ("2024-10-07", j.allocated_to, j.id, 
            "Submitted(Internal)", "Submitted(Client)", 
            "Submit(SPOC)", "Linedup", "QC Cleared", "Shortlisted"))

        # Ensure to fetch the count correctly
        j.actual_count = candidate_count[0][0] if candidate_count else 0  # Get the count value
        j.current_status = frappe.db.get_value("Task", {"name": j.id}, "status")
        task_ids.add((j.id, j.allocated_to))  # Store both task ID and allocated_to

    # Fetch candidates with pending status for the given date
    candidate_tasks = frappe.db.sql("""
        SELECT c.task, c.candidate_created_by
        FROM `tabCandidate` c
        INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
        WHERE DATE(cs.sourced_date) = %s
        AND cs.status IN (%s, %s, %s, %s, %s, %s)
    """, ("2024-10-07", 
            "Submitted(Internal)", "Submitted(Client)", 
            "Submit(SPOC)", "Linedup", "QC Cleared", "Shortlisted"))

    for i in candidate_tasks:
        task_id = i[0]  # Access the first element of the tuple (task)
        owner = i[1]    # Access the second element of the tuple (candidate_created_by)

        # Check if the task with the same owner is already in the task_ids
        if (task_id, owner) not in task_ids:
            add_count = frappe.db.sql("""
                SELECT COUNT(cs.status) AS status_count
                FROM `tabCandidate` c
                INNER JOIN `tabCandidate status` cs ON c.name = cs.parent
                WHERE DATE(cs.sourced_date) = %s
                AND c.candidate_created_by = %s
                AND c.task = %s
                AND cs.status IN (%s, %s, %s, %s, %s, %s)
            """, ("2024-10-07", owner, task_id, 
                "Submitted(Internal)", "Submitted(Client)", 
                "Submit(SPOC)", "Linedup", "QC Cleared", "Shortlisted"))

            # Get the count value from the add_count query
            add_count_value = add_count[0][0] if add_count else 0
            
            # Append the details to the child table
            parent_doc.append("dm_rec_task_details", {
                "id": task_id,
                "allocated_to": owner,
                "actual_count": add_count_value
            })
            task_ids.add((task_id, owner))  # Add to task_ids after appending

    parent_doc.dsr_check = 1
    parent_doc.save()
    frappe.db.commit()


# @frappe.whitelist()
# def update_so_st():
#     frappe.db.set_value("Education Checks",{"name":"Education Checks-20179"},"check_status","Draft")
#     frappe.db.set_value("Education Checks",{"name":"Education Checks-20179"},"workflow_state","Draft")


# @frappe.whitelist()
# def update_service_in_payment_entries():
#     payment_entries = frappe.get_all("Payment Entry", filters={"docstatus": 1},fields=["name"])
#     ind=0
#     for entry in payment_entries:
#         ind+=1
#         doc = frappe.get_doc("Payment Entry", entry.name)
#         print(doc)
#         for i in doc.references:
#             reference_doctype = i.reference_doctype
#             reference_name = i.reference_name
#             if reference_doctype=="Sales Order":
#                 print("Inside of sales order")
#                 if reference_doctype and reference_name:
#                     reference_doc = frappe.get_doc(reference_doctype, reference_name)
#                     if hasattr(reference_doc, 'service'):
#                         print(reference_doc.service)
#                         service_value = reference_doc.service
#                         i.service = service_value
#             elif reference_doctype=="Sales Invoice":
#                 print("Inside of sales invoice")
#                 if reference_doctype and reference_name:
#                     reference_doc = frappe.get_doc(reference_doctype, reference_name)
#                     if hasattr(reference_doc, 'services'):
#                         print(reference_doc.services)
#                         service_value = reference_doc.services
#                         i.service = service_value
#             elif reference_doctype=="Purchase Invoice":
#                 print("Inside of Purchase invoice")
#                 if reference_doctype and reference_name:
#                     reference_doc = frappe.get_doc(reference_doctype, reference_name)
#                     if hasattr(reference_doc, 'services'):
#                         print(reference_doc.services)
#                         service_value = reference_doc.services
#                         i.service = service_value
#         doc.save()
#     print(ind)

@frappe.whitelist()
def update_service_in_payment_entries(ref_doc,ref_name):
    if ref_doc=="Sales Order":
        if ref_doc and ref_name:
            reference_doc = frappe.get_doc(ref_doc, ref_name)
            if hasattr(reference_doc, 'service'):
                service_value = reference_doc.service
    elif ref_doc=="Sales Invoice":
        if ref_doc and ref_name:
            reference_doc = frappe.get_doc(ref_doc, ref_name)
            if hasattr(reference_doc, 'services'):
                service_value = reference_doc.services
    elif ref_doc=="Purchase Invoice":
        if ref_doc and ref_name:
            reference_doc = frappe.get_doc(ref_doc, ref_name)
            if hasattr(reference_doc, 'services'):
                service_value = reference_doc.services
    elif ref_doc=="Purchase Order":
        if ref_doc and ref_name:
            reference_doc = frappe.get_doc(ref_doc, ref_name)
            if hasattr(reference_doc, 'custom_service'):
                service_value = reference_doc.custom_service

    return service_value



@frappe.whitelist()
def update_attachement_from_issue(issue):
    spoc=frappe.db.get_value("Issue",{"name":issue},["custom_spoc"])
    attach=frappe.db.get_value("Issue",{"name":issue},["custom_attachement"])
    frappe.db.set_value("Task",{"issue":issue},"spoc",spoc)
    frappe.db.set_value("Task",{"issue":issue},"custom_issue_attachment",attach)

@frappe.whitelist()
def update_lead_as_converted(doc,method):
    if doc.lead_name:
        if frappe.db.exists("Lead",doc.lead_name):
            lead=frappe.db.get_doc("Lead",doc.lead_name)
            lead.status='Converted'
            lead.save(ignore_permissions=True)

@frappe.whitelist()
def update_lead_as_qualified(doc,method):
    if doc.lead_name:
        if frappe.db.exists("Lead",doc.lead_name):
            lead=frappe.get_doc("Lead",doc.lead_name)
            lead.status='Converted'
            lead.save(ignore_permissions=True)

@frappe.whitelist()
def update_existing_lead(doc,method):
    la=frappe.new_doc("Existing Leads")
    la.lead_name=doc.company_name
    la.lead_id=doc.name
    la.insert()
    la.save(ignore_permissions=True)
    frappe.db.commit()

from frappe.utils import flt, fmt_money
@frappe.whitelist()
def dashboard_data_receivable(from_date, to_date):
    query = """
    SELECT 
        SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN so.base_total ELSE 0 END, 0)) AS `THIS`,
        SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO Food Products' THEN so.base_total ELSE 0 END, 0)) AS `TFP`,
        SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN so.base_total ELSE 0 END, 0)) AS `TGTP`,
        SUM(COALESCE(CASE WHEN so.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN so.base_total ELSE 0 END, 0)) AS `MFAL`,
        SUM(COALESCE(so.base_total, 0)) AS Total
    FROM `tabSales Order` so
    WHERE so.status NOT IN ('On Hold', 'Cancelled', 'Closed')
    AND so.transaction_date >= %s
    AND so.transaction_date <= %s;
    """
    result = frappe.db.sql(query, (from_date, to_date), as_dict=True)
    
    query2 = """SELECT 
            SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN si.base_total ELSE 0 END, 0)) AS `THIS`,
            SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO Food Products' THEN si.base_total ELSE 0 END, 0)) AS `TFP`,
            SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN si.base_total ELSE 0 END, 0)) AS `TGTP`,
            SUM(COALESCE(CASE WHEN si.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN si.base_total ELSE 0 END, 0)) AS `MFAL`,
            SUM(COALESCE(si.base_total, 0)) AS Total
        FROM `tabSales Invoice` si
        WHERE si.status NOT IN ('Cancelled', 'Return', 'Credit Note Issued')
        AND si.posting_date >= %s
        AND si.posting_date <= %s;"""
    result2 = frappe.db.sql(query2, (from_date, to_date), as_dict=True)
    query3 = """SELECT
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN pi.paid_amount ELSE 0 END, 0)) AS `THIS`,
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO Food Products' THEN pi.paid_amount ELSE 0 END, 0)) AS `TFP`,
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN pi.paid_amount ELSE 0 END, 0)) AS `TGTP`,
        SUM(COALESCE(CASE WHEN pi.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN pi.paid_amount ELSE 0 END, 0)) AS `MFAL`,
        SUM(COALESCE(pi.paid_amount, 0)) AS Total
    FROM `tabPayment Entry` pi
    WHERE pi.status NOT IN ('Draft', 'Cancelled')
    AND pi.payment_type = 'Receive'
    AND pi.posting_date >= %s
    AND pi.posting_date <= %s"""
    result3 = frappe.db.sql(query3, (from_date, to_date), as_dict=True)

    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="background-color: #063970;color: white;text-align:center"><td width=20% >Title</td><td width=15% >THIS</td><td width=20% >TFP</td><td width=20% >TGTP</td><td width=20% >MFAL</td><td width=20% >Total</td></tr>'
    data += f'<tr><td width=20% >Sales Order(Net Total)</td><td width=20% >{fmt_money(result[0]["THIS"])}</td><td width=20% >{fmt_money(result[0]["TFP"])}</td><td width=20% >{fmt_money(result[0]["TGTP"])}</td><td width=20% >{fmt_money(result[0]["MFAL"])}</td><td width=20% >{fmt_money(result[0]["Total"])}</td></tr>'
    data += f'<tr><td width=20% >Sales Invoice(SC)</td><td width=20% >{fmt_money(result2[0]["THIS"])}</td><td width=20% >{fmt_money(result2[0]["TFP"])}</td><td width=20% >{fmt_money(result2[0]["TGTP"])}</td><td width=20% >{fmt_money(result2[0]["MFAL"])}</td><td width=20% >{fmt_money(result2[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Collection(Grand Total)</td><td width=20% >{fmt_money(result3[0]["THIS"])}</td><td width=20% >{fmt_money(result3[0]["TFP"])}</td><td width=20% >{fmt_money(result3[0]["TGTP"])}</td><td width=20% >{fmt_money(result3[0]["MFAL"])}</td><td width=20% >{fmt_money(result3[0]["Total"])}</td></tr>'
    data += '</table>'

    
    return data

@frappe.whitelist()
def count_sfp():
    import frappe
    count = frappe.db.sql("""
        SELECT COUNT(*) 
        FROM `tabSales Follow Up` AS sf
        WHERE sf.lead IS NOT NULL
        AND NOT EXISTS (
            SELECT 1 
            FROM `tabLead Contacts` AS lc 
            WHERE lc.parent = sf.name
        )
    """)[0][0]

    print(count)


@frappe.whitelist()
def dashboard_data_payable(from_date, to_date):
    query = """
    SELECT 
       SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN so.grand_total ELSE 0 END, 0)) AS `THIS`,
       SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO Food Products' THEN so.grand_total ELSE 0 END, 0)) AS `TFP`,
       SUM(COALESCE(CASE WHEN so.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN so.grand_total ELSE 0 END, 0)) AS `TGTP`,
       SUM(COALESCE(CASE WHEN so.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN so.grand_total ELSE 0 END, 0)) AS `MFAL`,
       SUM(COALESCE(so.grand_total, 0)) AS Total
  FROM `tabPurchase Order` so
   WHERE so.status NOT IN ('On Hold', 'Cancelled', 'Closed')
   AND so.transaction_date >= %s
   AND so.transaction_date <= %s
    """
    result = frappe.db.sql(query, (from_date, to_date), as_dict=True)
    
    query2 = """SELECT
        SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN si.grand_total ELSE 0 END, 0)) AS `THIS`,
        SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO Food Products' THEN si.grand_total ELSE 0 END, 0)) AS `TFP`,
        SUM(COALESCE(CASE WHEN si.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN si.grand_total ELSE 0 END, 0)) AS `TGTP`,
        SUM(COALESCE(CASE WHEN si.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN si.grand_total ELSE 0 END, 0)) AS `MFAL`,
        SUM(COALESCE(si.grand_total, 0)) AS Total
    FROM `tabPurchase Invoice` si
    WHERE si.status NOT IN ('Cancelled', 'Return', 'Credit Note Issued')
    AND si.posting_date >= %s
    AND si.posting_date <= %s """
    result2 = frappe.db.sql(query2, (from_date, to_date), as_dict=True)
    query3 = """SELECT
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN pi.paid_amount ELSE 0 END, 0)) AS `THIS`,
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO Food Products' THEN pi.paid_amount ELSE 0 END, 0)) AS `TFP`,
        SUM(COALESCE(CASE WHEN pi.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN pi.paid_amount ELSE 0 END, 0)) AS `TGTP`,
        SUM(COALESCE(CASE WHEN pi.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN pi.paid_amount ELSE 0 END, 0)) AS `MFAL`,
        SUM(COALESCE(pi.paid_amount, 0)) AS Total
    FROM `tabPayment Entry` pi
    WHERE pi.status NOT IN ('Draft', 'Cancelled')
    AND pi.payment_type = 'Pay'
    AND pi.posting_date >= %s
    AND pi.posting_date <= %s"""
    result3 = frappe.db.sql(query3, (from_date, to_date), as_dict=True)

    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="background-color: #063970;color: white;text-align:center"><td width=20% >Title</td><td width=15% >THIS</td><td width=20% >TFP</td><td width=20% >TGTP</td><td width=20% >MFAL</td><td width=20% >Total</td></tr>'
    data += f'<tr><td width=20% >Purchase Order</td><td width=20% >{fmt_money(result[0]["THIS"])}</td><td width=20% >{fmt_money(result[0]["TFP"])}</td><td width=20% >{fmt_money(result[0]["TGTP"])}</td><td width=20% >{fmt_money(result[0]["MFAL"])}</td><td width=20% >{fmt_money(result[0]["Total"])}</td></tr>'
    data += f'<tr><td width=20% >Purchase Invoice</td><td width=20% >{fmt_money(result2[0]["THIS"])}</td><td width=20% >{fmt_money(result2[0]["TFP"])}</td><td width=20% >{fmt_money(result2[0]["TGTP"])}</td><td width=20% >{fmt_money(result2[0]["MFAL"])}</td><td width=20% >{fmt_money(result2[0]["Total"])}</td></tr>'
    data += f'<tr><td width=20% >Payment</td><td width=20% >{fmt_money(result3[0]["THIS"])}</td><td width=20% >{fmt_money(result3[0]["TFP"])}</td><td width=20% >{fmt_money(result3[0]["TGTP"])}</td><td width=20% >{fmt_money(result3[0]["MFAL"])}</td><td width=20% >{fmt_money(result3[0]["Total"])}</td></tr>'
    data += '</table>'

    
    return data


@frappe.whitelist()
def dashboard_data_financial_status(from_date, to_date):
    query = """
    SELECT 
       CAST(FORMAT(SUM(CASE WHEN si.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN si.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT(SUM(CASE WHEN si.company = 'TEAMPRO Food Products' THEN si.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT(SUM(CASE WHEN si.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN si.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT(SUM(CASE WHEN si.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN si.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(SUM(si.outstanding_amount), 2) AS CHAR) AS Total
    FROM `tabSales Invoice` si
    WHERE si.status NOT IN ('Paid', 'Cancelled', 'Return', 'Credit Note Issued')
    """
    result = frappe.db.sql(query, as_dict=True)
    
    query2 = """SELECT 
        CAST(FORMAT(SUM(CASE WHEN so.company = 'TEAMPRO Food Products' THEN so.base_grand_total - ((so.base_grand_total * so.amount_billed) + so.advance_paid) ELSE 0 END), 2) AS CHAR) AS `TFP`,
        CAST(FORMAT(SUM(CASE WHEN so.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN so.base_grand_total - ((so.base_grand_total * so.amount_billed) + so.advance_paid) ELSE 0 END), 2) AS CHAR) AS `THIS`,
        CAST(FORMAT(SUM(CASE WHEN so.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN so.base_grand_total - ((so.base_grand_total * so.amount_billed) + so.advance_paid) ELSE 0 END), 2) AS CHAR) AS `TGTP`,
        CAST(FORMAT(SUM(CASE WHEN so.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN so.base_grand_total - ((so.base_grand_total * so.amount_billed) + so.advance_paid) ELSE 0 END), 2) AS CHAR) AS `MFAL`,
        CAST(FORMAT(SUM(so.base_grand_total - ((so.base_grand_total * so.amount_billed) + so.advance_paid)), 2) AS CHAR) AS Total
    FROM `tabSales Order` so
     WHERE so.status NOT IN ('On Hold', 'To Deliver', 'Closed', 'Cancelled', 'Completed')
    """
    result2 = frappe.db.sql(query2, as_dict=True)
    query3 = """SELECT 
       CAST(FORMAT(SUM(CASE WHEN pi.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN pi.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT(SUM(CASE WHEN pi.company = 'TEAMPRO Food Products' THEN pi.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT(SUM(CASE WHEN pi.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN pi.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT(SUM(CASE WHEN pi.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN pi.outstanding_amount ELSE 0 END), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(SUM(pi.outstanding_amount), 2) AS CHAR) AS Total
    FROM `tabPurchase Invoice` pi
    WHERE pi.status NOT IN ('Paid', 'Cancelled', 'Debit Note Issued','Return')
    """
    result3 = frappe.db.sql(query3, as_dict=True)

    query4 = """SELECT
       CAST(FORMAT(SUM(CASE WHEN po.company = 'TEAMPRO HR & IT Services Pvt. Ltd.' THEN (po.grand_total - po.advance_paid) ELSE 0 END), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT(SUM(CASE WHEN po.company = 'TEAMPRO Food Products' THEN (po.grand_total - po.advance_paid) ELSE 0 END), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT(SUM(CASE WHEN po.company = 'TEAMPRO General Trading Pvt. Ltd.' THEN (po.grand_total - po.advance_paid) ELSE 0 END), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT(SUM(CASE WHEN po.company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT' THEN (po.grand_total - po.advance_paid) ELSE 0 END), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(SUM(po.grand_total - po.advance_paid), 2) AS CHAR) AS Total
    FROM `tabPurchase Order` po
    WHERE po.status IN ('To Receive and Bill', 'To Bill')
    """
    result4 = frappe.db.sql(query4,  as_dict=True)

    query5 = """
    SELECT 
       CAST(FORMAT((IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT((IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT((IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0)), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)) + (IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)) +(IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)) + (IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0))), 2) AS CHAR) AS Total
  FROM `tabCompany` c
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g2
    ON c.name = g2.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq2
    ON c.name = sq2.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - TFP'
           AND company = 'TEAMPRO Food Products'
           AND is_cancelled = 0
         GROUP BY company
       ) g
    ON c.name = g.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE  account = 'Cash - TFP'
           AND company = 'TEAMPRO Food Products'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq
    ON c.name = sq.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - TGT'
           AND company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g3
    ON c.name = g3.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - TGT'
           AND company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq3
    ON c.name = sq3.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - MAHTARVIN'
           AND company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_cancelled = 0
         GROUP BY company
       ) g4
    ON c.name = g4.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Cash - MAHTARVIN'
           AND company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq4
    ON c.name = sq4.company
 WHERE c.name IN ('TEAMPRO HR & IT Services Pvt. Ltd.','TEAMPRO Food Products', 'TEAMPRO General Trading Pvt. Ltd.', 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT')    """
    result5 = frappe.db.sql(query5,  as_dict=True)

    query6 = """SELECT
        CAST(FORMAT((IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)), 2) AS CHAR) AS `THIS`,
        CAST(FORMAT((IFNULL(g1.opening_debit, 0) + IFNULL(SUM(sq1.debit), 0)) - (IFNULL(g1.opening_credit, 0) + IFNULL(SUM(sq1.credit), 0)), 2) AS CHAR) AS `TFP`,
        CAST(FORMAT((IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)), 2) AS CHAR) AS `TGTP`,
        CAST(FORMAT((IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0)), 2) AS CHAR) AS `MFAL`,
        CAST(FORMAT(((IFNULL(g1.opening_debit, 0) + IFNULL(SUM(sq1.debit), 0)) - (IFNULL(g1.opening_credit, 0) + IFNULL(SUM(sq1.credit), 0)) + (IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)) + (IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)) + (IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0))), 2) AS CHAR) AS Total
    FROM `tabCompany` c
    LEFT JOIN (
            SELECT company,
                SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
                SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
            AND is_cancelled = 0
            GROUP BY company
        ) g2
        ON c.name = g2.company
    LEFT JOIN (
            SELECT company,
                SUM(debit_in_account_currency) AS debit,
                SUM(credit_in_account_currency) AS credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
            AND is_opening = 'No'
            AND is_cancelled = 0
            GROUP BY company
        ) sq2
        ON c.name = sq2.company
    LEFT JOIN (
            SELECT company,
                SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
                SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND is_cancelled = 0
            AND company = 'TEAMPRO Food Products'
            GROUP BY company
        ) g1
        ON c.name = g1.company
    LEFT JOIN (
            SELECT company,
                SUM(debit_in_account_currency) AS debit,
                SUM(credit_in_account_currency) AS credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'TEAMPRO Food Products'
            AND is_opening = 'No'
            AND is_cancelled = 0
            GROUP BY company
        ) sq1
        ON c.name = sq1.company
    LEFT JOIN (
            SELECT company,
                SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
                SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'TEAMPRO General Trading Pvt. Ltd.'
            AND is_cancelled = 0
            GROUP BY company
        ) g3
        ON c.name = g3.company
    LEFT JOIN (
            SELECT company,
                SUM(debit_in_account_currency) AS debit,
                SUM(credit_in_account_currency) AS credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'TEAMPRO General Trading Pvt. Ltd.'
            AND is_opening = 'No'
            AND is_cancelled = 0
            GROUP BY company
        ) sq3
        ON c.name = sq3.company
    LEFT JOIN (
            SELECT company,
                SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
                SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
            AND is_cancelled = 0
            GROUP BY company
        ) g4
        ON c.name = g4.company
    LEFT JOIN (
            SELECT company,
                SUM(debit_in_account_currency) AS debit,
                SUM(credit_in_account_currency) AS credit
            FROM `tabGL Entry`
            WHERE account IN ('777705160983 - ICICI Bank - THIS', '50200054611436 - HDFC - THIS','50200082906246-HDFC - This','50200050787897 - HDFC Account - TGT','50200059117831 - HDFC Bank - TFP')
            AND company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
            AND is_opening = 'No'
            AND is_cancelled = 0
            GROUP BY company
        ) sq4
        ON c.name = sq4.company
    WHERE c.name IN ('TEAMPRO HR & IT Services Pvt. Ltd.','TEAMPRO Food Products', 'TEAMPRO General Trading Pvt. Ltd.', 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT')"""
    result6 = frappe.db.sql(query6,  as_dict=True)
        
    query7 = """SELECT
       CAST(FORMAT((IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT((IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT((IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0)), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)) + (IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)) +(IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)) + (IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0))), 2) AS CHAR) AS Total
  FROM `tabCompany` c
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g2
    ON c.name = g2.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq2
    ON c.name = sq2.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO Food Products'
           AND is_cancelled = 0
         GROUP BY company
       ) g
    ON c.name = g.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO Food Products'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq
    ON c.name = sq.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits - TGTP'
           AND company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g3
    ON c.name = g3.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits - TGTP'
           AND company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq3
    ON c.name = sq3.company
  LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_cancelled = 0
         GROUP BY company
       ) g4
    ON c.name = g4.company
  LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq4
    ON c.name = sq4.company
    WHERE c.name IN ('TEAMPRO HR & IT Services Pvt. Ltd.','TEAMPRO Food Products', 'TEAMPRO General Trading Pvt. Ltd.', 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT')"""
    result7 = frappe.db.sql(query7, as_dict=True)

    query8 = """SELECT
       CAST(FORMAT((IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)), 2) AS CHAR) AS `THIS`,
       CAST(FORMAT((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)), 2) AS CHAR) AS `TFP`,
       CAST(FORMAT((IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)), 2) AS CHAR) AS `TGTP`,
       CAST(FORMAT((IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0)), 2) AS CHAR) AS `MFAL`,
       CAST(FORMAT(((IFNULL(g.opening_debit, 0) + IFNULL(SUM(sq.debit), 0)) - (IFNULL(g.opening_credit, 0) + IFNULL(SUM(sq.credit), 0)) + (IFNULL(g2.opening_debit, 0) + IFNULL(SUM(sq2.debit), 0)) - (IFNULL(g2.opening_credit, 0) + IFNULL(SUM(sq2.credit), 0)) +(IFNULL(g3.opening_debit, 0) + IFNULL(SUM(sq3.debit), 0)) - (IFNULL(g3.opening_credit, 0) + IFNULL(SUM(sq3.credit), 0)) + (IFNULL(g4.opening_debit, 0) + IFNULL(SUM(sq4.debit), 0)) - (IFNULL(g4.opening_credit, 0) + IFNULL(SUM(sq4.credit), 0))), 2) AS CHAR) AS Total
    FROM `tabCompany` c
    LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits(Res) - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g2
    ON c.name = g2.company
    LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE account = 'Fixed Deposits(Res) - THIS'
           AND company = 'TEAMPRO HR & IT Services Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq2
    ON c.name = sq2.company
    LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO Food Products'
           AND is_cancelled = 0
         GROUP BY company
       ) g
    ON c.name = g.company
    LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO Food Products'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq
    ON c.name = sq.company
    LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_cancelled = 0
         GROUP BY company
       ) g3
    ON c.name = g3.company
    LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE company = 'TEAMPRO General Trading Pvt. Ltd.'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq3
    ON c.name = sq3.company
    LEFT JOIN (
        SELECT company,
               SUM(CASE WHEN is_opening = 'Yes' THEN debit ELSE 0 END) AS opening_debit,
               SUM(CASE WHEN is_opening = 'Yes' THEN credit ELSE 0 END) AS opening_credit
          FROM `tabGL Entry`
         WHERE company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_cancelled = 0
         GROUP BY company
       ) g4
    ON c.name = g4.company
    LEFT JOIN (
        SELECT company,
               SUM(debit_in_account_currency) AS debit,
               SUM(credit_in_account_currency) AS credit
          FROM `tabGL Entry`
         WHERE company = 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT'
           AND is_opening = 'No'
           AND is_cancelled = 0
         GROUP BY company
       ) sq4
    ON c.name = sq4.company
    WHERE c.name IN ('TEAMPRO HR & IT Services Pvt. Ltd.','TEAMPRO Food Products', 'TEAMPRO General Trading Pvt. Ltd.', 'MUSISA FERIK AL-MAHTARVIN LALAMAGALAT')"""
    result8 = frappe.db.sql(query8, as_dict=True)

    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="background-color: #063970;color: white;text-align:center"><td width=20% >Title</td><td width=15% >THIS</td><td width=20% >TFP</td><td width=20% >TGTP</td><td width=20% >MFAL</td><td width=20% >Total</td></tr>'
    data += f'<tr><td width=20% >Collection Outstanding</td><td width=20% >{fmt_money(result[0]["THIS"])}</td><td width=20% >{fmt_money(result[0]["TFP"])}</td><td width=20% >{fmt_money(result[0]["TGTP"])}</td><td width=20% >{fmt_money(result[0]["MFAL"])}</td><td width=20% >{fmt_money(result[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Bill Outstanding</td><td width=20% >{fmt_money(result2[0]["THIS"])}</td><td width=20% >{fmt_money(result2[0]["TFP"])}</td><td width=20% >{fmt_money(result2[0]["TGTP"])}</td><td width=20% >{fmt_money(result2[0]["MFAL"])}</td><td width=20% >{fmt_money(result2[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Purchase Invoice Outstanding</td><td width=20% >{fmt_money(result3[0]["THIS"])}</td><td width=20% >{fmt_money(result3[0]["TFP"])}</td><td width=20% >{fmt_money(result3[0]["TGTP"])}</td><td width=20% >{fmt_money(result3[0]["MFAL"])}</td><td width=20% >{fmt_money(result3[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Purchase Order Outstanding</td><td width=20% >{fmt_money(result4[0]["THIS"])}</td><td width=20% >{fmt_money(result4[0]["TFP"])}</td><td width=20% >{fmt_money(result4[0]["TGTP"])}</td><td width=20% >{fmt_money(result4[0]["MFAL"])}</td><td width=20% >{fmt_money(result4[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Cash in Hand</td><td width=20% >{fmt_money(result5[0]["THIS"])}</td><td width=20% >{fmt_money(result5[0]["TFP"])}</td><td width=20% >{fmt_money(result5[0]["TGTP"])}</td><td width=20% >{fmt_money(result5[0]["MFAL"])}</td><td width=20% >{fmt_money(result5[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >Cash in Bank</td><td width=20% >{fmt_money(result6[0]["THIS"])}</td><td width=20% >{fmt_money(result6[0]["TFP"])}</td><td width=20% >{fmt_money(result6[0]["TGTP"])}</td><td width=20% >{fmt_money(result6[0]["MFAL"])}</td><td width=20% >{fmt_money(result6[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >FD (In-Hand)</td><td width=20% >{fmt_money(result7[0]["THIS"])}</td><td width=20% >{fmt_money(result7[0]["TFP"])}</td><td width=20% >{fmt_money(result7[0]["TGTP"])}</td><td width=20% >{fmt_money(result7[0]["MFAL"])}</td><td width=20% >{fmt_money(result7[0]["Total"])}</td></tr>'

    data += f'<tr><td width=20% >FD (Reserved)</td><td width=20% >{fmt_money(result8[0]["THIS"])}</td><td width=20% >{fmt_money(result8[0]["TFP"])}</td><td width=20% >{fmt_money(result8[0]["TGTP"])}</td><td width=20% >{fmt_money(result8[0]["MFAL"])}</td><td width=20% >{fmt_money(result8[0]["Total"])}</td></tr>'


    data += '</table>'

    
    return data

def move_to_sfp():
    sfp=frappe.db.get_all("Sales Follow Up",{"lead":['!=',''],'name':'SFP-29098'},['*'])
    for i in sfp:
        sales=frappe.get_doc("Sales Follow Up",{'name':'SFP-29098'})
        count=0
        for s in sales.contacts:
            count+=1
        if count <= 0:
            lead=frappe.get_doc("Lead",{'name':i.lead})
            for k in lead.lead_contacts:
                print(k)
                sales.append('contacts',{
                'person_name': k.person_name,
                'mobile':k.mobile,
                'is_primary':k.is_primary,
                'email_id':k.email_id,
                'is_primaryemail':k.is_primaryemail,
                'has_whatsapp':k.has_whatsapp,
                'service':k.service
                })
            sales.save()

@frappe.whitelist()
def calculate_distance(docname):
    doc = frappe.get_doc('Appointment', docname)

    origin_lat = float(doc.custom_checkin_latitude)
    origin_lon = float(doc.custom_checkin_longitude)
    destination_lat = float(doc.custom_checkout_latitude)
    destination_lon = float(doc.custom_checkout_longitude)
    if destination_lat == origin_lat and destination_lon == origin_lon:
        distance = 0
    else:
        distance = get_road_distance_graphhopper(origin_lat, origin_lon, destination_lat, destination_lon)
    distance = float(distance)
    return distance

import requests
def get_road_distance_graphhopper(origin_lat, origin_lon, destination_lat, destination_lon):
    api_key = "f0801316-68c4-43b8-8c19-788ebff443df"
    url = f"https://graphhopper.com/api/1/route?point={origin_lat},{origin_lon}&point={destination_lat},{destination_lon}&vehicle=car&locale=en&key={api_key}&calc_points=false"

    response = requests.get(url)
    data = response.json()

    if 'paths' in data:
        distance_in_meters = data['paths'][0]['distance']
        distance_in_km = distance_in_meters / 1000
        return distance_in_km
    else:
        return "Error calculating distance"


@frappe.whitelist()
def set_quotation(doc,method):
    # opportunity = frappe.get_all("Quotation",{"docstatus":("!=",2)},["name","opportunity"])
    # for i in opportunity:
    #     if i.opportunity:
    frappe.db.set_value("Opportunity",doc.opportunity,"custom_quotation",doc.name)

# @frappe.whitelist()
# def dpr_for_ani_team():
#     job = frappe.db.exists('Scheduled Job Type', 'sales_ss_team_dsr_daily')
#     if not job:
#         sjt = frappe.new_doc("Scheduled Job Type")
#         sjt.update({
#             "method": 'teampro.custom.sales_ss_team_dsr_daily',
#             "frequency": 'Cron',
#             "cron_format": '30 20 * * *'
#         })
#         sjt.save(ignore_permissions=True)

@frappe.whitelist()
def sales_app_team_dpr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
    recievers=[]
    reciever=[]
    custom_date = add_days(today(),1)
    # custom_date=today()
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= add_days(today(),1)
    # next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    print(formatted_next_date)
    data=[]
    for j in emp:
        reciever.append(j.user_id)
    for i in emp:
        recievers.append(i.user_id)
    recievers.append('anil.p@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="9"><b>APP & Team DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''

    appointment_count = 0
    appointment_list = []
    app_individual=[]
    for user_email in recievers:
        print(user_email)
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
        appt = frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where date(p.scheduled_time) between '%s' and '%s' and c.user='%s'""" %(formatted_next_date, formatted_next_date,user_email),as_dict=1)
        appointment_list.append(appt)
    todo_list=frappe.db.get_all("ToDo",{"custom_production_date":formatted_before_date,"allocated_to":user_email},["*"])
    if todo_list:
        data += '''
              <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        s_no=0
        for i in todo_list:
            s_no+=1
            short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="6" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td></tr>'.format(short_code,i.name,i.custom_subject,i.status)
    print(appointment_list[1])
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user in recievers:
        appointment_day_beforecount = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (user, f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"), as_dict=True)[0].count or 0
        short_code = frappe.db.get_value("Employee", {"user_id": user}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":user,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_day_beforecount if appointment_day_beforecount else '0' , lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
    if appointment_list:
        data += f'''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment Schedule {formatted_date}</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
        '''
        for appt_group in appointment_list:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                user_email = i['user']
                short_code = frappe.db.get_value("Employee", {"user_id":i.user }, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.time,i.name)
            
    data += '</table>'



    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['anil.p@groupteampro.com'],
                cc='dineshbabu.k@groupteampro.com',
                subject='APP & Team DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for user_email in reciever:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="8"><b>APP & Team DPR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>APT</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
            </tr>
        '''
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.scheduled_time BETWEEN %s AND %s
    """, (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        # Other counts
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_next_date})

        # Populate table row for the specific user
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            short_code, appointment_count, lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count
        )
        app_i= frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where c.user='%s' and date(p.scheduled_time) between '%s' and '%s'""" %(user_email,formatted_next_date, formatted_next_date),as_dict=1)
        app_individual.append(app_i)

        # ToDo Section
    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": user_email}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject)

        # Appointment Section
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for users in reciever:
        appointment_day_beforecount = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (users, f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"), as_dict=True)[0].count or 0
        short_code = frappe.db.get_value("Employee", {"user_id": users}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":users,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_day_beforecount if appointment_day_beforecount else '0' , lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
        if app_individual:
            data += f'''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>Appointment Schedule {formatted_date}</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
            '''
            for appt_group in app_individual:
                for i in appt_group:  # each 'i' is a dictionary with appointment details
                    user_email = i['user']
                    short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, i.time, i.name)

        data += '</table>'
        frappe.sendmail(
            recipients=[user_email],
            # recipients=['divya.p@groupteampro.com'],
            subject='APP & Team DPR %s - Reg' % formatted_date,
            message="""
                <b>Dear {user},</b><br><br>
                Please find the below DPR for {date} for your kind reference and action.<br><br>
                {table}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """.format(user=user_email.split('@')[0], date=formatted_date, table=data)
        )


@frappe.whitelist()
def sales_ami_team_dpr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com'])},['*'])
    recievers=[]
    reciever=[]
    # custom_date=today()
    custom_date = add_days(today(),1)
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= add_days(today(),1)
    # next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    data=[]
    for i in emp:
        recievers.append(i.user_id) 

    for j in emp:
        reciever.append(j.user_id)
    recievers.append('annie.m@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="9"><b>ANI & Team DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user_email in recievers:
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,'3' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
        
    for todo in  recievers:  
        todo_list=frappe.db.get_all("ToDo",{"custom_production_date":formatted_next_date,"allocated_to":todo},["*"])
        if todo_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
            '''
            s_no=0
            for i in todo_list:
                s_no+=1
                print(i.allocated_to)

                short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.name,i.custom_subject)
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user in recievers:
        short_code = frappe.db.get_value("Employee", {"user_id": user}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":user,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,'3', lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
    data += '</table>'



    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                # recipients=['vijiyalakshmi.k@groupteampro.com'],
                recipients=['annie.m@groupteampro.com'],
                cc='dineshbabu.k@groupteampro.com',
                subject='ANI & Team DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for user_email in reciever:
        # Start of the table
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        
        # Header for ANI & Team DPR
        data += '<tr style="text-align:center;"><td colspan="9"><b>ANI & Team DPR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></td>
            </tr>
        '''
        
        # Populate ANI & Team DPR rows
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_next_date})

        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            short_code, '3', lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count
        )

        # Header for Non Updated Followup
        data += '''
            <tr style="text-align:center;">
                <td colspan="9"><b>Non Updated Followup</b></td>
            </tr>
        '''
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></td>
            </tr>
        '''

        # Populate Non Updated Followup rows
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_day_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Lead", "status": "Lead"})
        open_day_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Lead", "status": "Open"})
        replied_day_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_day_before_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_before_date, "follow_up_to": "Customer"})
        todo_day_before_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_before_date})

        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
            short_code or '-', '3', lead_day_before_count or '0', open_day_before_count or '0', replied_day_before_count or '0',
            interested_before_count or '0', opportunity_before_count or '0', customer_day_before_count or '0', todo_day_before_count or '0'
        )

        # Close the table
        data += '</table>'
        frappe.sendmail(
            recipients=[user_email],
            # recipients=['vijiyalakshmi.k@groupteampro.com'],
            # recipients=['divya.p@groupteampro.com'],
            subject='ANI & Team DPR %s - Reg' % formatted_date,
            message="""
                <b>Dear {user},</b><br><br>
                Please find the below DPR for {date} for your kind reference and action.<br><br>
                {table}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """.format(user=user_email.split('@')[0], date=formatted_date, table=data)
        )


@frappe.whitelist()
def sales_jss_team_dpr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00191'},['*'])
    recievers=[]
    reciever=[]
    custom_date = add_days(today(),1)
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= add_days(today(),1)
    # next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    print(formatted_next_date)
    data=[]
    for j in emp:
        reciever.append(j.user_id)
    for i in emp:
        recievers.append(i.user_id)
    recievers.append('jayaraman.s@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="9"><b>JSS & Team DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''

    appointment_count = 0
    appointment_list = []
    app_individual=[]
    for user_email in recievers:
        print(user_email)
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
        appt = frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where date(p.scheduled_time) between '%s' and '%s' and c.user='%s'""" %(formatted_next_date, formatted_next_date,user_email),as_dict=1)
        appointment_list.append(appt)
    todo_list=frappe.db.get_all("ToDo",{"custom_production_date":formatted_before_date,"allocated_to":user_email},["*"])
    if todo_list:
        data += '''
              <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        s_no=0
        for i in todo_list:
            s_no+=1
            short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="6" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td></tr>'.format(short_code,i.name,i.custom_subject,i.status)
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user in recievers:
        appointment_day_beforecount = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (user, f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"), as_dict=True)[0].count or 0
        short_code = frappe.db.get_value("Employee", {"user_id": user}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":user,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_day_beforecount if appointment_day_beforecount else '0' , lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
    if appointment_list:
        data += f'''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment Schedule {formatted_date}</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
        '''
        for appt_group in appointment_list:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                user_email = i['user']
                short_code = frappe.db.get_value("Employee", {"user_id":i.user }, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.time,i.name)
            
    data += '</table>'



    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['jayaraman.s@groupteampro.com'],
                cc='dineshbabu.k@groupteampro.com',
                subject='JSS & Team DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for user_email in reciever:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="8"><b>JSS & Team DPR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>APT</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
            </tr>
        '''
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.scheduled_time BETWEEN %s AND %s
    """, (user_email, f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"), as_dict=True)[0].count or 0
        # Other counts
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Lead"})
        open_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Open"})
        replied_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Replied"})
        interested_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Interested"})
        opportunity_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Lead", "status": "Opportunity"})
        customer_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email, "next_contact_date": formatted_next_date, "follow_up_to": "Customer"})
        todo_count = frappe.db.count("ToDo", {"allocated_to": user_email, "custom_production_date": formatted_next_date})

        # Populate table row for the specific user
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
            short_code, appointment_count, lead_count, open_count, replied_count, interested_count, opportunity_count, customer_count, todo_count
        )
        app_i= frappe.db.sql("""select p.scheduled_time as time, p.name as name, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where c.user='%s' and date(p.scheduled_time) between '%s' and '%s'""" %(user_email,formatted_next_date, formatted_next_date),as_dict=1)
        app_individual.append(app_i)

        # ToDo Section
    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": user_email}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject)

        # Appointment Section
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for users in reciever:
        appointment_day_beforecount = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user=%s AND p.scheduled_time BETWEEN %s AND %s
        """, (users, f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"), as_dict=True)[0].count or 0
        short_code = frappe.db.get_value("Employee", {"user_id": users}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":users,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_day_beforecount if appointment_day_beforecount else '0' , lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
        if app_individual:
            data += f'''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>Appointment Schedule {formatted_date}</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Customer</td></tr>
            '''
            for appt_group in app_individual:
                for i in appt_group:  # each 'i' is a dictionary with appointment details
                    user_email = i['user']
                    short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code, i.time, i.name)

        data += '</table>'
        frappe.sendmail(
            recipients=[user_email],
            # recipients=['divya.p@groupteampro.com'],
            subject='JSS & Team DPR %s - Reg' % formatted_date,
            message="""
                <b>Dear {user},</b><br><br>
                Please find the below DPR for {date} for your kind reference and action.<br><br>
                {table}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                <i>This email has been automatically generated. Please do not reply</i>
            """.format(user=user_email.split('@')[0], date=formatted_date, table=data)
        )

# @frappe.whitelist()
# def sales_app_team_dsr_daily():
#     emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
#     emp_emails=[]
#     date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
#     formatted_date = date_obj.strftime('%d/%m/%Y')
#     next_date=today()
#     # next_date="2024-11-16"
#     # next_date=add_days(nowdate(),1)
#     next_dates=datetime.strptime(next_date, '%Y-%m-%d')
#     formatted_next_date=next_dates.strftime('%Y-%m-%d')
#     print(formatted_next_date)
#     appointments = frappe.get_all("Appointment", filters={"custom_completed_date":formatted_next_date}, fields=["*"])
#     user_mails=[]
#     for i in emp:
#         emp_emails.append(i.user_id)
#     for j in emp:
#         user_mails.append(j.user_id)

#     emp_emails.append('anil.p@groupteampro.com')
#     # emp_list=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com'])},['*'])
#     # for j in emp_list:
#     #     emp_emails.append(j.user_id)
#     # emp_emails.append('annie.m@groupteampro.com') 
#     data = '<table border="1" width="100%" style="border-collapse: collapse;">'
#     data += '<tr style="text-align:center;"><td colspan="9"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td width="10%" colspan="1"><b>Exe</b></td>
#             <td width="25%" colspan="1"><b>Effective</b></td>
#             <td width="25%" colspan="1"><b>Non Effective</b></td>
#             <td width="25%" colspan="1"><b>Appointment</b></td>
#             <td width="25%" colspan="1"><b>TODO</b></td>     
#             <td width="25%" colspan="1"><b>Total</b></td>
#         </tr>
#     '''
#     appointment_lists = []
#     app_individual=[]
#     todo_lists=[]
#     for c in emp_emails:
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (c,formatted_next_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

#         short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
#         effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
#         non_effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Non Effective"})
#         total_calls = effective_call + non_effective_call
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,effective_call if effective_call else '0' , non_effective_call if non_effective_call else '0',appointment_count if appointment_count else '0',todo_count if todo_count else '0',total_calls if total_calls else '0'
#             )
#         app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,c),as_dict=1)
#         todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": c, "status": ('not in',['Cancelled'])}, ["*"])
#         if app:
#             appointment_lists.append(app)
#         if todo_list:
#             todo_lists.append(todo_list)
#     if appointment_lists:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="9";"><b>Appointment</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="1">Status</td><td colspan="3">Remarks</td></tr>
#         '''
#         for appt_group in appointment_lists:
#             for i in appt_group:  # each 'i' is a dictionary with appointment details
#                 user_email = i['user']
#                 short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
#                 data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td><td colspan="3">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
#     if todo_lists:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="9";"><b>ToDo</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="1" style=" text-align: center;">Subject</td><td colspan="3">Remarks</td></tr>
#         '''
#         for k in todo_lists:
#             for m in k:
#                 short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
#                 data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 10px;">{}</td><td colspan="3">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
#     data += '</table>'
#     frappe.sendmail(
#                 # recipients=recievers,
#                 # recipients=['divya.p@groupteampro.com'],
#                 recipients=['anil.p@groupteampro.com'], 
#                 cc='dineshbabu.k@groupteampro.com',
#                 subject='APP & Team DSR %s -Reg' % formatted_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DSR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data)
#             )
#     for c in user_mails:
#         data = '<table border="1" width="100%" style="border-collapse: collapse;">'
#         data += '<tr style="text-align:center;"><td colspan="9"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
#         data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td width="25%" colspan="1"><b>Exe</b></td>
#             <td width="25%" colspan="1"><b>Effective</b></td>
#             <td width="25%" colspan="1"><b>Non Effective</b></td>
#             <td width="25%" colspan="1"><b>Appointment</b></td>
#             <td width="25%" colspan="1"><b>TODO</b></td>     
#             <td width="25%" colspan="1"><b>Total</b></td>
#         </tr>
#         '''
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (c,formatted_next_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

#         short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
#         effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
#         non_effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Non Effective"})
#         total_calls = effective_call + non_effective_call
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,effective_call if effective_call else '0' , non_effective_call if non_effective_call else '0',appointment_count if appointment_count else '0',todo_count if todo_count else '0',total_calls if total_calls else '0'
#             )
#         app_i = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,c),as_dict=1)
#         todo_lists = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])}, ["*"])

#         if app_i:
#             app_individual.append(app_i)
#     if app_individual:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="9";"><b>Appointment</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="1">Status</td><td colspan="3">Remarks</td></tr>
#         '''
#         for appt_group in app_individual:
#             for i in appt_group:  # each 'i' is a dictionary with appointment details
#                 user_email = i['user']
#                 short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
#                 data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td><td colspan="3">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)

#     if todo_lists:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="9";"><b>ToDo</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="1" style=" text-align: center;">Subject</td><td colspan="3">Remarks</td></tr>
#         '''
#         for todo in todo_lists:
#             short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
#             data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 10px;">{}</td><td colspan="3">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
#     data += '</table>'
#     frappe.sendmail(
#                 recipients=[c],
#                 # recipients=recievers,
#                 # recipients=['divya.p@groupteampro.com'],
#                 # recipients=['anil.p@groupteampro.com'], 
#                 # cc='dineshbabu.k@groupteampro.com',
#                 subject='APP & Team DSR  %s -Reg' % formatted_next_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DSR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data)
#             )
# @frappe.whitelist()
# def sales_app_team_dsr_daily():
#     emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
#     emp_emails=[]
#     date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
#     # date_obj = datetime.strptime(add_days(today(), -1), '%Y-%m-%d')
#     formatted_date = date_obj.strftime('%d/%m/%Y')
#     # next_date=add_days(today(),-1)
#     next_date=today()
#     next_dates=datetime.strptime(next_date, '%Y-%m-%d')
#     formatted_next_date=next_dates.strftime('%Y-%m-%d')
#     before_date=add_days(today(),-1)
#     before_dates=datetime.strptime(before_date, '%Y-%m-%d')
#     formatted_before_date=before_dates.strftime('%Y-%m-%d')
#     user_mails=[]
#     for i in emp:
#         emp_emails.append(i.user_id)
#     for j in emp:
#         user_mails.append(j.user_id)

#     emp_emails.append('anil.p@groupteampro.com') 
#     data = '<table border="1" width="100%" style="border-collapse: collapse;">'
#     data += '<tr style="text-align:center;"><td colspan="11"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td style="width:10%;"><b>Exe</b></td>
#             <td style="width:15%;"><b>Apt</b></td>
#             <td style="width:20%;"><b>Lead</b></td>
#             <td style="width:13%;"><b>Open</b></td>
#             <td style="width:10%;"><b>Replied</b></td>
#             <td style="width:7%;"><b>Interested</b></td>
#             <td style="width:13%;"><b>Oppr</b></td>
#             <td style="width:13%;"><b>Cust</b></td>
#             <td style="width:10%;"><b>ToDo</b></b></td>
#             <td style="width:10%;"><b>OR%</b></b></td>
#             <td style="width:10%;"><b>PR%</b></b></td>
#         </tr>
#     '''
#     appointment_lists = []
#     app_individual=[]
#     todo_lists=[]
    
#     for c in emp_emails:
#         appointment = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]},["*"])
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (c,formatted_next_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )
#         app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,c),as_dict=1)

#         todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": c, "status": ('not in',['Cancelled'])}, ["*"])
#         if app:
#             appointment_lists.append(app)
#         if todo_list:
#             todo_lists.append(todo_list)
#     if appointment_lists:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>Appointment Taken</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
#         '''
#         for appt_group in appointment_lists:
#             for i in appt_group:  # each 'i' is a dictionary with appointment details
#                 user_email = i['user']
#                 short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
#                 data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
#     data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>Appointment Fixed</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="3" style="text-align:center; ">Customer</td><td colspan="1">Status</td><td colspan="7">Remarks</td></tr>
#         '''
#     for j in appointment:
#         data+='<tr style="text-align:center;"><td colspan="3" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td><td colspan="7">{}</td></tr>'.format(j.name,j.status,j.custom_remarks)
#     data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td style="width:10%;"><b>Exe</b></td>
#             <td style="width:15%;"><b>Apt</b></td>
#             <td style="width:20%;"><b>Lead</b></td>
#             <td style="width:13%;"><b>Open</b></td>
#             <td style="width:10%;"><b>Replied</b></td>
#             <td style="width:7%;"><b>Interested</b></td>
#             <td style="width:13%;"><b>Oppr</b></td>
#             <td style="width:13%;"><b>Cust</b></td>
#             <td style="width:10%;"><b>ToDo</b></b></td>
#             <td style="width:10%;"><b>OR%</b></b></td>
#             <td style="width:10%;"><b>PR%</b></b></td>
#         </tr>
#     '''
#     for user in emp_emails:
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (user,formatted_before_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )

#     if todo_lists:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>ToDo</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
#         '''
#         for k in todo_lists:
#             for m in k:
#                 short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
#                 data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
#     data += '</table>'
#     frappe.sendmail(
#                 # recipients=recievers,
#                 # recipients=['divya.p@groupteampro.com'],
#                 recipients=['anil.p@groupteampro.com'], 
#                 cc='dineshbabu.k@groupteampro.com',
#                 subject='APP & Team DSR %s -Reg' % formatted_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DSR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data)
#             )
#     for d in user_mails:
#         print(d)
#         data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        
#         data += '<tr style="text-align:center;"><td colspan="11"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td style="width:10%;"><b>Exe</b></td>
#                 <td style="width:15%;"><b>Appt</b></td>
#                 <td style="width:20%;"><b>Lead</b></td>
#                 <td style="width:13%;"><b>Open</b></td>
#                 <td style="width:10%;"><b>Replied</b></td>
#                 <td style="width:7%;"><b>Interested</b></td>
#                 <td style="width:13%;"><b>Oppr</b></td>
#                 <td style="width:13%;"><b>Cust</b></td>
#                 <td style="width:10%;"><b>ToDo</b></td>
#                 <td style="width:10%;"><b>OR%</b></b></td>
#                 <td style="width:10%;"><b>PR%</b></b></td>
#             </tr>
#         '''
#         appointment_ind = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]},["*"])
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (d,formatted_next_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":d,"custom_production_date":formatted_next_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )
#         app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,d),as_dict=1)

#         todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
#         if app:
#             app_individual.append(app)
#         if todo_list:
#             todo_lists.append(todo_list)
#     if app_individual:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";><b>Appointment Taken</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
#         '''
#         for appt_group in app_individual:
#             for i in appt_group:  # each 'i' is a dictionary with appointment details
#                 user_email = i['user']
#                 short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
#                 data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
#     data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>Appointment Fixed</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="3" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
#         '''
#     for j in appointment_ind:
#         data+='<tr style="text-align:center;"><td colspan="3" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(j.name,j.status,j.custom_remarks)

#     data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td style="width:10%;"><b>Exe</b></td>
#             <td style="width:15%;"><b>Apt</b></td>
#             <td style="width:20%;"><b>Lead</b></td>
#             <td style="width:13%;"><b>Open</b></td>
#             <td style="width:10%;"><b>Replied</b></td>
#             <td style="width:7%;"><b>Interested</b></td>
#             <td style="width:13%;"><b>Oppr</b></td>
#             <td style="width:13%;"><b>Cust</b></td>
#             <td style="width:10%;"><b>ToDo</b></b></td>
#             <td style="width:10%;"><b>OR%</b></b></td>
#             <td style="width:10%;"><b>PR%</b></b></td>
#         </tr>
#     '''
#     for users in user_mails:
#         appointment_count = frappe.db.sql("""
#         SELECT COUNT(DISTINCT p.name) AS count 
#         FROM `tabAppointment` p 
#         INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
#         WHERE c.user = %s AND p.custom_completed_date=%s
#     """, (users,formatted_before_date), as_dict=True)[0].count or 0
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":users, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":users},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":users,"custom_production_date":formatted_before_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )

#         if todo_lists:
#             data += '''
#                 <tr style="background-color: #0f1568; color: white; text-align:center;">
#                     <td colspan="11";><b>ToDo</b></b></td>
#                 </tr>
#                 <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
#             '''
#             for k in todo_lists:
#                 for m in k:
#                     short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
#                     data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
#         data += '</table>'
#         frappe.sendmail(
#                     recipients=[d],
#                     # recipients=recievers,
#                     # recipients=['divya.p@groupteampro.com'],
#                     subject='APP & Team DSR  %s -Reg' % formatted_next_date,
#                     message = """
#                     <b>Dear Team,</b><br><br>
#     Please find the below DSR for {} for your kind reference and action.<br><br>

#                 {}<br><br>
#                     Thanks & Regards,<br>TEAM ERP<br>
                    
#                     <i>This email has been automatically generated. Please do not reply</i>
#                     """.format(formatted_date,data)
#                 )
@frappe.whitelist()
def sales_app_team_dsr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00023','user_id':('not in',['sivarenisha.m@groupteampro.com','jeniba.a@groupteampro.com'])},['*'])
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    # date_obj = datetime.strptime(add_days(today(), -1), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    # next_date=add_days(today(),-1)
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    user_mails=[]
    for i in emp:
        emp_emails.append(i.user_id)
    for j in emp:
        user_mails.append(j.user_id)

    emp_emails.append('anil.p@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="11"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    appointment_lists = []
    app_individual=[]
    todo_lists=[]
    
    for c in emp_emails:
        appointment = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabSales Follow Up` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE (c.user = %s OR p.visted_by = %s) AND p.visted_date=%s
    """, (c,c,formatted_next_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        # app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,c),as_dict=1)
        app = frappe.db.sql("""select p.app_status as status, p.organization_name as name,p.appointment_remarks as custom_remarks, c.user as user,p.visted_by as visted_by from `tabSales Follow Up` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.visted_date = '%s' and (c.user = '%s' OR p.visted_by = '%s')""" %(formatted_next_date,c,c),as_dict=1)

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": c, "status": ('not in',['Cancelled'])}, ["*"])
        if app:
            appointment_lists.append(app)
        if todo_list:
            todo_lists.append(todo_list)
    if appointment_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Taken</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        processed_users = set()  # Track processed user and visited_by combinations

        for appt_group in appointment_lists:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                # Handle 'user'
                if i['user'] and i['user'] not in processed_users:
                    short_code = frappe.db.get_value("Employee", {"user_id": i['user']}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(
                        short_code or "N/A", i['name'], i['status'], i['custom_remarks']
                    )
                    processed_users.add(i['user'])  # Mark 'user' as processed

                # Handle 'visited_by'
                if i['visted_by'] and i['visted_by'] not in processed_users:
                    short_code = frappe.db.get_value("Employee", {"user_id": i['visted_by']}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(
                        short_code or "N/A", i['name'], i['status'], i['custom_remarks']
                    )
                    processed_users.add(i['visted_by'])  # Mark 'visited_by' as processed
        # for appt_group in appointment_lists:
        #     for i in appt_group:  # each 'i' is a dictionary with appointment details
        #         # if i['user']  or i['visted_by']:
        #             # user_email = i['user'] or i['visted_by']
        #         if i['user']:
        #             short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
        #             data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)

        #         if i['visted_by']:
        #             short_code = frappe.db.get_value("Employee", {"user_id": i.visted_by}, "short_code")
        #         #     short_code = frappe.db.get_value("Employee", {"user_id": i.visted_by}, "short_code")
        #             data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
    data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="3" style="text-align:center; ">Customer</td><td colspan="1">Status</td><td colspan="7">Remarks</td></tr>
        '''
    for j in appointment:
        data+='<tr style="text-align:center;"><td colspan="3" style="text-align: left; padding-left: 50px;">{}</td><td colspan="1">{}</td><td colspan="7">{}</td></tr>'.format(j.organization_name,j.app_status,j.custom_details1)
    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for user in emp_emails:
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabSales Follow Up` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE (c.user = %s OR p.visted_by = %s) AND p.visted_date=%s
    """, (user,user,formatted_before_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    if todo_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for k in todo_lists:
            for m in k:
                short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['anil.p@groupteampro.com'], 
                cc='dineshbabu.k@groupteampro.com',
                subject='APP & Team DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for d in user_mails:
        print(d)
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        
        data += '<tr style="text-align:center;"><td colspan="11"><b>APP & Team DSR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        appointment_ind = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabSales Follow Up` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE (c.user = %s OR p.visted_by = %s) AND p.visted_date=%s
    """, (d,d,formatted_next_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":d,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        app = frappe.db.sql("""select p.app_status as status, p.organization_name as name,p.appointment_remarks as custom_remarks, c.user as user from `tabSales Follow Up` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.visted_date = '%s' and (c.user = '%s' OR p.visted_by = '%s')""" %(formatted_next_date,d,d),as_dict=1)

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
        if app:
            app_individual.append(app)
        if todo_list:
            todo_lists.append(todo_list)
    if app_individual:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";><b>Appointment Taken</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        processed_user = set()  # Track processed user and visited_by combinations

        for appt_group in app_individual:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                # Handle 'user'
                if i['user'] and i['user'] not in processed_users:
                    short_code = frappe.db.get_value("Employee", {"user_id": i['user']}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(
                        short_code or "N/A", i['name'], i['status'], i['custom_remarks']
                    )
                    processed_user.add(i['user'])  # Mark 'user' as processed

                # Handle 'visited_by'
                if i['visted_by'] and i['visted_by'] not in processed_users:
                    short_code = frappe.db.get_value("Employee", {"user_id": i['visted_by']}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(
                        short_code or "N/A", i['name'], i['status'], i['custom_remarks']
                    )
                    processed_user.add(i['visted_by'])  # Mark 'visited_by' as processed
        # for appt_group in app_individual:
        #     for i in appt_group:  # each 'i' is a dictionary with appointment details
        #         user_email = i['user']
        #         short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
        #         data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
    data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="3" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
    for j in appointment_ind:
        data+='<tr style="text-align:center;"><td colspan="3" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(j.organization_name,j.app_status,j.custom_details1)

    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for users in user_mails:
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabSales Follow Up` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE (c.user = %s OR p.visted_by = %s) AND p.visted_date=%s
    """, (users,users,formatted_before_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":users, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":users},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":users,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

        if todo_lists:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="11";><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
            '''
            for k in todo_lists:
                for m in k:
                    short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
                    data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
        data += '</table>'
        frappe.sendmail(
                    recipients=[d],
                    # recipients=recievers,
                    # recipients=['divya.p@groupteampro.com'],
                    subject='APP & Team DSR  %s -Reg' % formatted_next_date,
                    message = """
                    <b>Dear Team,</b><br><br>
    Please find the below DSR for {} for your kind reference and action.<br><br>

                {}<br><br>
                    Thanks & Regards,<br>TEAM ERP<br>
                    
                    <i>This email has been automatically generated. Please do not reply</i>
                    """.format(formatted_date,data)
                )
# @frappe.whitelist()
# def sales_ami_team_dsr_daily():
#     emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com']),"name":('!=',("TC00058"))},['*'])
#     emp_emails=[]
#     date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
#     formatted_date = date_obj.strftime('%d/%m/%Y')
#     next_date=nowdate()
#     next_dates=datetime.strptime(next_date, '%Y-%m-%d')
#     # next_date=add_days(nowdate(),1)
#     formatted_next_date=next_dates.strftime('%Y-%m-%d')
#     before_date=add_days(today(),-1)
#     before_dates=datetime.strptime(before_date, '%Y-%m-%d')
#     formatted_before_date=before_dates.strftime('%Y-%m-%d')
#     user_mails=[]
#     for i in emp:
#         emp_emails.append(i.user_id)
#     for j in emp:
#         user_mails.append(j.user_id)

#     emp_emails.append('annie.m@groupteampro.com')
#     data = '<table border="1" width="100%" style="border-collapse: collapse;">'
#     data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td style="width:10%;"><b>Exe</b></td>
#             <td style="width:15%;"><b>Apt</b></td>
#             <td style="width:20%;"><b>Lead</b></td>
#             <td style="width:13%;"><b>Open</b></td>
#             <td style="width:10%;"><b>Replied</b></td>
#             <td style="width:7%;"><b>Interested</b></td>
#             <td style="width:13%;"><b>Oppr</b></td>
#             <td style="width:13%;"><b>Cust</b></td>
#             <td style="width:10%;"><b>ToDo</b></b></td>
#             <td style="width:10%;"><b>OR%</b></b></td>
#             <td style="width:10%;"><b>PR%</b></b></td>
#         </tr>
#     '''
#     for c in emp_emails:
        
#         # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":c})
#         appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": c,"appointment_created_on":formatted_next_date})
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

#         short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )
#     appointment_list = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":["in",emp_emails]},["*"])
#     if appointment_list:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>Appointment Fixed</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="2" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
#         '''
#         for i in appointment_list:
#             short_code = frappe.db.get_value("Employee", {"user_id": i.owner}, "short_code")
#             data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
#     data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
#     data += '''
#         <tr style="background-color: #0f1568; color: white; text-align:center;">
#             <td style="width:10%;"><b>Exe</b></td>
#             <td style="width:15%;"><b>Apt</b></td>
#             <td style="width:20%;"><b>Lead</b></td>
#             <td style="width:13%;"><b>Open</b></td>
#             <td style="width:10%;"><b>Replied</b></td>
#             <td style="width:7%;"><b>Interested</b></td>
#             <td style="width:13%;"><b>Oppr</b></td>
#             <td style="width:13%;"><b>Cust</b></td>
#             <td style="width:10%;"><b>ToDo</b></b></td>
#             <td style="width:10%;"><b>OR%</b></b></td>
#             <td style="width:10%;"><b>PR%</b></b></td>
            
#         </tr>
#     '''
#     for user in emp_emails:
#         appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":user})
#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else '0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )

#     todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": ["in", emp_emails], "status": ('not in',['Cancelled'])}, ["*"])
#     if todo_list:
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td colspan="11";"><b>ToDo</b></b></td>
#             </tr>
#             <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
#         '''
#         for todo in todo_list:
#             short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
#             data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
#     data += '</table>'
#     frappe.sendmail(
#                 # recipients=['divya.p@groupteampro.com'],
#                 recipients=['annie.m@groupteampro.com'], 
#                 cc='dineshbabu.k@groupteampro.com',
#                 subject='ANI & Team DSR %s -Reg' % formatted_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DSR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data)
#             )
#     for d in user_mails:
#         data = '<table border="1" width="100%" style="border-collapse: collapse;">'
#         data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td style="width:10%;"><b>Exe</b></td>
#                 <td style="width:15%;"><b>Apt</b></td>
#                 <td style="width:20%;"><b>Lead</b></td>
#                 <td style="width:13%;"><b>Open</b></td>
#                 <td style="width:10%;"><b>Replied</b></td>
#                 <td style="width:7%;"><b>Interested</b></td>
#                 <td style="width:13%;"><b>Oppr</b></td>
#                 <td style="width:13%;"><b>Cust</b></td>
#                 <td style="width:10%;"><b>ToDo</b></b></td>
#                 <td style="width:10%;"><b>OR%</b></b></td>
#                 <td style="width:10%;"><b>PR%</b></b></td>
#             </tr>
#         '''
#         appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d})

#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})

#         short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' ,effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )
#         appointment_list = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d},["*"])
#         if appointment_list:
#             data += '''
#                 <tr style="background-color: #0f1568; color: white; text-align:center;">
#                     <td colspan="11";><b>Appointment Fixed</b></b></td>
#                 </tr>
#                 <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="1" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
#             '''
#         for i in appointment_list:
#             short_code = frappe.db.get_value("Employee", {"user_id": i.owner}, "short_code")
#             data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
#         data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
#         data += '''
#             <tr style="background-color: #0f1568; color: white; text-align:center;">
#                 <td style="width:10%;"><b>Exe</b></td>
#                 <td style="width:15%;"><b>Apt</b></td>
#                 <td style="width:20%;"><b>Lead</b></td>
#                 <td style="width:13%;"><b>Open</b></td>
#                 <td style="width:10%;"><b>Replied</b></td>
#                 <td style="width:7%;"><b>Interested</b></td>
#                 <td style="width:13%;"><b>Oppr</b></td>
#                 <td style="width:13%;"><b>Cust</b></td>
#                 <td style="width:10%;"><b>ToDo</b></b></td>
#                 <td style="width:10%;"><b>OR%</b></b></td>
#                 <td style="width:10%;"><b>PR%</b></b></td>
#             </tr>
#         '''
#         appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":d})

#         todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
#         short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
#         # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
#         lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
#         effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
#         open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
#         effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
#         replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
#         effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
#         interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
#         effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
#         opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
#         effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
#         customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
#         effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
#         todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
#         data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
#                 short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
#             )

#         todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
#         if todo_list:
#             data += '''
#                 <tr style="background-color: #0f1568; color: white; text-align:center;">
#                     <td colspan="11";"><b>ToDo</b></b></td>
#                 </tr>
#                 <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
#             '''
#             for todo in todo_list:
#                 short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
#                 data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
#         data += '</table>'

#         frappe.sendmail(
#                 recipients=[d],
#                 # recipients=['divya.p@groupteampro.com'],
#                 subject='ANI & Team DSR  %s -Reg' % formatted_date,
#                 message = """
#                 <b>Dear Team,</b><br><br>
# Please find the below DSR for {} for your kind reference and action.<br><br>

#             {}<br><br>
#                 Thanks & Regards,<br>TEAM ERP<br>
                
#                 <i>This email has been automatically generated. Please do not reply</i>
#                 """.format(formatted_date,data)
#             )
@frappe.whitelist()
def sales_ami_team_dsr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00007','user_id':('not in',['dm@groupteampro.com']),"name":('!=',("TC00058"))},['*'])
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date=nowdate()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    user_mails=[]
    for i in emp:
        emp_emails.append(i.user_id)
    for j in emp:
        user_mails.append(j.user_id)

    emp_emails.append('annie.m@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for c in emp_emails:
        
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":c})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": c,"appointment_created_on":formatted_next_date})
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
    appointment_list = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="2" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        for i in appointment_list:
            short_code = frappe.db.get_value("Employee", {"user_id": i.next_contact_by}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.organization_name,i.app_status,i.custom_details1)
    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
            
        </tr>
    '''
    for user in emp_emails:
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": c,"appointment_created_on":formatted_next_date})
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":user})
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else '0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": ["in", emp_emails], "status": ('not in',['Cancelled'])}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                # recipients=['divya.p@groupteampro.com'],
                recipients=['annie.m@groupteampro.com'], 
                cc='dineshbabu.k@groupteampro.com',
                subject='ANI & Team DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for d in user_mails:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        data += '<tr style="text-align:center;"><td colspan="11"><b>ANI & Team DSR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Apt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": d,"appointment_created_on":formatted_next_date})

        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})

        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' ,effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        # appointment_list = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":d},["*"])
        appointment_list = frappe.db.get_all("Sales Follow Up",{"appointment_created_on":formatted_next_date},["*"])
        if appointment_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="11";><b>Appointment Fixed</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="1" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
            '''
        for i in appointment_list:
            # short_code = frappe.db.get_value("Employee", {"user_id": i.owner}, "short_code")
            # data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
            short_code = frappe.db.get_value("Employee", {"user_id": i.next_contact_by}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.organization_name,i.app_status,i.custom_details1)

        data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Apt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        # appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":d})
        appointment_count = frappe.db.count("Sales Follow Up",{"next_contact_by": d,"appointment_created_on":formatted_next_date})

        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
        if todo_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="11";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
            '''
            for todo in todo_list:
                short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
        data += '</table>'

        frappe.sendmail(
                recipients=[d],
                # recipients=['divya.p@groupteampro.com'],
                subject='ANI & Team DSR  %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )



@frappe.whitelist()
def sales_ss_team_dsr_daily():
    emp=frappe.db.get_all("Employee",{'status':'Active','reports_to':'TI00191'},['*'])
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    user_mails=[]
    for i in emp:
        emp_emails.append(i.user_id)
    for j in emp:
        user_mails.append(j.user_id)

    emp_emails.append('jayaraman.s@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="11"><b>JSS & Team DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    appointment_lists = []
    app_individual=[]
    todo_lists=[]
    appointment_ind = [] 
    for c in emp_emails:
        appointment = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]},["*"])
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.custom_completed_date=%s
    """, (c,formatted_next_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,c),as_dict=1)

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": c, "status": ('not in',['Cancelled'])}, ["*"])
        if app:
            appointment_lists.append(app)
        if todo_list:
            todo_lists.append(todo_list)
    if appointment_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";><b>Appointment Taken</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        for appt_group in appointment_lists:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                user_email = i['user']
                short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
    data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="2" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
    for j in appointment:
        data+='<tr style="text-align:center;"><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(j.name,j.status,j.custom_remarks)
    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for user in emp_emails:
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.custom_completed_date=%s
    """, (user,formatted_before_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    if todo_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for k in todo_lists:
            for m in k:
                short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['jayaraman.s@groupteampro.com'], 
                cc='dineshbabu.k@groupteampro.com',
                subject='JSS & Team DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )
    for d in user_mails:
        data = '<table border="1" width="100%" style="border-collapse: collapse;">'
        
        data += '<tr style="text-align:center;"><td colspan="11"><b>JSS & Team DSR, {}</b></td></tr>'.format(formatted_date)
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td style="width:10%;"><b>Exe</b></td>
                <td style="width:15%;"><b>Appt</b></td>
                <td style="width:20%;"><b>Lead</b></td>
                <td style="width:13%;"><b>Open</b></td>
                <td style="width:10%;"><b>Replied</b></td>
                <td style="width:7%;"><b>Interested</b></td>
                <td style="width:13%;"><b>Oppr</b></td>
                <td style="width:13%;"><b>Cust</b></td>
                <td style="width:10%;"><b>ToDo</b></td>
                <td style="width:10%;"><b>OR%</b></b></td>
                <td style="width:10%;"><b>PR%</b></b></td>
            </tr>
        '''
        appointment_ind = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]]},["*"]) or []
        
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.custom_completed_date=%s
    """, (d,formatted_next_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":d, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":d},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": d,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":d,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
        app = frappe.db.sql("""select p.status as status, p.name as name,p.custom_remarks as custom_remarks, c.user as user from `tabAppointment` p inner join `tabDPR Mail Users` c on c.parent = p.name where p.custom_completed_date = '%s' and c.user='%s'""" %(formatted_next_date,d),as_dict=1)

        todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": d, "status": ('not in',['Cancelled'])}, ["*"])
        if app:
            appointment_lists.append(app)
        if todo_list:
            todo_lists.append(todo_list)
    if appointment_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Taken</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td style="text-align:left;"colspan="1">Exe</td><td colspan="1" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        for appt_group in appointment_lists:
            for i in appt_group:  # each 'i' is a dictionary with appointment details
                user_email = i['user']
                short_code = frappe.db.get_value("Employee", {"user_id": i.user}, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
    data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="9";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="2" style="text-align:center; ">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
    if appointment_ind:
        for j in appointment_ind:
            data+='<tr style="text-align:center;"><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(j.name,j.status,j.custom_remarks)

    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for users in user_mails:
        appointment_count = frappe.db.sql("""
        SELECT COUNT(DISTINCT p.name) AS count 
        FROM `tabAppointment` p 
        INNER JOIN `tabDPR Mail Users` c ON c.parent = p.name 
        WHERE c.user = %s AND p.custom_completed_date=%s
    """, (users,formatted_before_date), as_dict=True)[0].count or 0
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":users, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":users},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": users,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":users,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    if todo_lists:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td  style="text-align:left;" colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for k in todo_lists:
            for m in k:
                short_code = frappe.db.get_value("Employee", {"user_id": m.allocated_to}, "short_code")
                data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, m.name, m.custom_subject, m.current_status_ or '')

    
        data += '</table>'
        frappe.sendmail(
                    recipients=[d],
                    # recipients=recievers,
                    # recipients=['divya.p@groupteampro.com'],
                    subject='JSS & Team DSR  %s -Reg' % formatted_next_date,
                    message = """
                    <b>Dear Team,</b><br><br>
    Please find the below DSR for {} for your kind reference and action.<br><br>

                {}<br><br>
                    Thanks & Regards,<br>TEAM ERP<br>
                    
                    <i>This email has been automatically generated. Please do not reply</i>
                    """.format(formatted_date,data)
                )
    
@frappe.whitelist()
def todo_report(allocated_to):
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr><td colspan="18" style="text-align:center; font-weight:bold;">TODO Report</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Subject</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Current Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Date</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Status</td>' \
            '</tr>'
    todo_data = frappe.db.get_all("ToDo",{"allocated_to":allocated_to,"status":"Open"}, ["*"])
    for i in todo_data:
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.custom_subject}</td>' \
                f'<td style="text-align:center;">{i.current_status_}</td>' \
                f'<td style="text-align:center;">{i.created_on}</td>' \
                f'<td style="text-align:center;">{i.status}</td>' \
                '</tr>'
        s_no += 1

    data += '</table>'
    return data

from datetime import datetime
from openpyxl.styles import PatternFill
@frappe.whitelist()
def download_todo_report():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    filename_todo = "ToDo Report"+ posting_date
    build_xlsx_response_todo(filename_todo)
    
def build_xlsx_response_todo(filename_todo):
    xlsx_file = make_xlsx_todo(filename_todo)
    frappe.response['filename'] = filename_todo + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
def make_xlsx_todo(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    fill_color = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  # White color font
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20 
    black_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    header = ["ToDo Report"]
    ws.append(header) 

    ws.append(["S NO", "ID", "Subject", "Current Status", "Date", "Status"])

    for cell in ws[2]: 
        cell.fill = fill_color
        cell.font = header_font  # Apply white font to each header cell 
        cell.border = black_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)



    data1= get_data_of_todo(args)
    for row in data1:
        ws.append(row)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file

def get_data_of_todo(args):
    data = []
    s_no=1
    todo_data = frappe.db.get_all("ToDo",{"allocated_to":args.allocated_to,"status":"Open"}, ["*"])
    for i in todo_data:
        data.append([s_no, i.name, i.custom_subject, i.current_status_, i.created_on, i.status])
        s_no+=1
    return data


@frappe.whitelist()
def so_creation_mail_it_sw(doc,method):
    if doc.service=="IT-SW":
        subject = f"New Sales Order Created: {doc.name}"
        message = f"""
            <p>Dear Sir/Mam,</p>
            <p>A new Sales Order <strong>{doc.name}</strong> has been created for the service <strong>IT-SW</strong>.</p>
            """
        frappe.sendmail(
            # recipients=['divya.p@groupteampro.com'],
            recipients=['sarath.v@groupteampro.com'],
            cc=['sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject=subject,
            message=message
        )

@frappe.whitelist()
def create_hooks_expense():
    job = frappe.db.exists('Scheduled Job Type', 'send_mail_for_expenseapproval_weekly_md')
    if not job:
        exp = frappe.new_doc("Scheduled Job Type")
        exp.update({
            "method": 'teampro.custom.send_mail_for_expenseapproval_weekly_md',
            "frequency": 'Cron',
            "cron_format": '00 00 * * 0'
        })
        exp.save(ignore_permissions=True)


@frappe.whitelist()
def send_mail_for_expenseapproval_weekly_hod():
    expenses=frappe.db.get_all("Expense Claim",{"workflow_state":"Pending for HOD"},["*"])
    # Group expense claims by Expense Approver
    approver_expenses = {}
    for expense in expenses:
        approver = expense.get("expense_approver")
        if approver not in approver_expenses:
            approver_expenses[approver] = []
        approver_expenses[approver].append(expense)
    for approver, approver_expense_list in approver_expenses.items():
        s_no = 1
        data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
        data += '<tr style="background-color: #002060; color: white;">' \
                '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
                '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
                '<td style="text-align:center; font-weight:bold; color:white;">Employee</td>' \
                '<td style="text-align:center; font-weight:bold; color:white;">Total Claimed Amount</td>' \
                '</tr>'
        
        # Add each expense claim to the email content
        for expense in approver_expense_list:
            data += f'<tr>' \
                    f'<td style="text-align:center;">{s_no}</td>' \
                    f'<td style="text-align:center;">{expense["name"]}</td>' \
                    f'<td style="text-align:center;">{expense["employee_name"]}</td>' \
                    f'<td style="text-align:center;">{expense["total_claimed_amount"]}</td>' \
                    '</tr>'
            s_no += 1
        data += '</table>'
        
        # Send the email
        frappe.sendmail(
            # recipients=["divya.p@groupteampro.com"],
            recipients=[approver],  # Send to the respective approver
            subject='Expense Claim Waiting for Approval',
            message=f"""
                <b>Dear {approver},</b><br><br>
                Please find below the list of expense claims waiting for your approval:<br><br>
                {data}<br><br>
                Thanks & Regards,<br>TEAM ERP<br><br>
                <i>This email has been automatically generated. Please do not reply.</i>
            """
        )

@frappe.whitelist()
def send_mail_for_expenseapproval_weekly_ceo():
    expenses=frappe.db.get_all("Expense Claim",{"workflow_state":"Pending for CEO"},["*"])
    s_no=1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Employee</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Total Claimed Amount</td>' \
            '</tr>'
    for i in expenses:
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.employee_name}</td>' \
                f'<td style="text-align:center;">{i.total_claimed_amount}</td>' \
                '</tr>'
        s_no+=1
    data += '</table>'
    frappe.sendmail(
            recipients=['sangeetha.s@groupteampro.com'],
            subject='Expense Claim Waiting for Approval' ,
            message="""
            <b>Dear Sir/Madam,</b><br><br>
            Please find below the list of expense claims waiting for your approval:<br><br>
            {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br><br>
            <i>This email has been automatically generated. Please do not reply.</i>
        """.format(data)
        )

@frappe.whitelist()
def send_mail_for_expenseapproval_weekly_md():
    expenses=frappe.db.get_all("Expense Claim",{"workflow_state":"Pending for MD"},["*"])
    s_no=1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">ID</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Employee</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Total Claimed Amount</td>' \
            '</tr>'
    for i in expenses:
        data += f'<tr>' \
                f'<td style="text-align:center;">{s_no}</td>' \
                f'<td style="text-align:center;">{i.name}</td>' \
                f'<td style="text-align:center;">{i.employee_name}</td>' \
                f'<td style="text-align:center;">{i.total_claimed_amount}</td>' \
                '</tr>'
        s_no+=1
    data += '</table>'
    frappe.sendmail(
            recipients=['dineshbabu.k@groupteampro.com'],
            subject='Expense Claim Waiting for Approval' ,
            message="""
            <b>Dear Sir/Madam,</b><br><br>
            Please find below the list of expense claims waiting for your approval:<br><br>
            {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br><br>
            <i>This email has been automatically generated. Please do not reply.</i>
        """.format(data)
        )

def send_project_spoc_report_weekly():
    posting_date = datetime.now().strftime("%d-%m-%Y")
    spoc_list = get_spoc_list_weekly()

    for spoc in spoc_list:
        filename = "DSR_" + spoc + "_" + posting_date
        xlsx_file = build_xlsx_response_spoc_project_even_weekly(spoc, filename)
        send_mail_with_attachment_spoc_project_even_weekly(spoc, filename, xlsx_file.getvalue())

def send_mail_with_attachment_spoc_project_even_weekly(spoc, filename, file_content):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    subject = f"DSR Report for {spoc} : - {posting_date}"
    message = (
        f"Dear {spoc},<br>"
        "Please find attached the DSR Report.<br><br>"
        "<br>Thanks & Regards,<br>TEAM ERP<br>"
        "This email has been automatically generated. Please do not reply"
    )
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    week_date=today()
    day_of_week=week_date.weekday()
    week_number = (week_date.day - 1) // 7 + 1
    if day_of_week == 5 and (week_number == 2 or week_number == 4):
        # Send the email for each SPOC
        frappe.sendmail(
            recipients=[spoc],  # Assuming spoc is the email ID of the SPOC
            # recipients=['jeniba.a@groupteampro.com'],
            sender=None,
            subject=subject,
            message=message,
            attachments=attachments,
        )

def build_xlsx_response_spoc_project_even_weekly(spoc, filename):
    return make_xlsx_spoc_project_weekly(spoc, filename)

def get_spoc_list_weekly():
    projects = frappe.get_all("Project", filters={'status': 'Open', 'service': 'IT-SW',"spoc":("not in",["abdulla.pi@groupteampro.com","sarath.v@groupteampro.com"])}, fields=['spoc'])
    spoc_set = {project['spoc'] for project in projects if project.get('spoc')}
    return list(spoc_set)

def make_xlsx_spoc_project_weekly(spoc, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "DSR Report"
    today = datetime.now().strftime('%Y-%m-%d')
    header_fill = PatternFill(start_color="A6CAF0", end_color="A6CAF0", fill_type="solid")
    head_fill = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")
    row_fill = PatternFill(start_color="FFC1CC", end_color="FFC1CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=17):
        for cell in row:
            cell.border = thin_border
    posting_date = datetime.now().strftime("%d-%m-%Y")
    spoc_code = frappe.db.get_value("Employee", {"user_id": spoc}, ["short_code"])
    header_value = f"{spoc_code} DSR {posting_date}"

    # Manually place the header value in the first cell
    first_cell = ws.cell(row=1, column=1)
    first_cell.value = header_value

    # Apply styles to the first cell (where the value is placed)
    first_cell.fill = header_fill
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")
    first_cell.border = thin_border

    # Merge cells from column 1 to 17
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)

    # Apply styles to the entire merged range (though value only goes into the first cell)
    for col in range(1, 18):  # Merged range is from column 1 to 17
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    headers = ["S.NO", "Project", "Priority", "New", "", "Open", "", "Working", "", "Overdue", "", "PR", "", "CR", "", "Total", ""]
    sub_headers = ["", "", "", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue", "Task", "Issue"]
    ws.append(headers)
    ws.append(sub_headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = head_fill
        cell.font = Font(bold=True, color="FFFFFF")  
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for col in range(1, len(sub_headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = row_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    merged_cell = ws.cell(row=3, column=1)  # This is the top-left cell of the merged area

    # Create a new fill color
    new_fill_color = PatternFill(start_color="0f1568", end_color="0f1568", fill_type="solid")

    # Apply the new fill color to the merged cell
    merged_cell.fill = new_fill_color

    for col in range(4, 17, 2):
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
    serial_number = 1
    priority_rows = [("High",), ("Medium",), ("Low",)]

    cust = frappe.db.get_all("Project", {"status":"Open","spoc": spoc, "service": "IT-SW"}, ["*"])
    total_new_tasks=0
    total_new_issues=0
    total_open_tasks=0
    total_open_issues=0
    total_working_tasks=0
    total_working_issues=0
    total_overdue_tasks=0
    total_overdue_issues=0
    total_pr_tasks=0
    total_pr_issues=0
    total_cr_tasks=0
    total_cr_issues=0
    total_all_tasks=0
    total_all_issues=0
    current_row = 4 
    s_row=4
    for c in cust:
        total_task_count = 0
        total_issue_count = 0
        priority_levels = ["High", "Medium", "Low"]
        h_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "High","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
        h_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "High","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
        h_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "High"})
        h_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "High"})
        h_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"High"})
        h_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"High"})
        h_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"High"})
        h_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"High"})
        h_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"High"})
        high_task_count = h_open_taskcount + h_working_taskcount + h_overdue_taskcount + h_pr_taskcount + h_cr_taskcount 
        high_issue_count = h_open_issuecount + h_working_issuecount + h_overdue_issuecount + h_pr_issuecount + h_cr_issuecount
        e_high_task_count = h_open_taskcount + h_working_taskcount + h_overdue_taskcount + h_pr_taskcount + h_cr_taskcount
        e_high_issue_count = h_open_issuecount + h_working_issuecount + h_overdue_issuecount + h_pr_issuecount + h_cr_issuecount

        total_new_tasks += h_new_taskcount 
        total_new_issues += h_new_issuecount 
        total_open_tasks += h_open_taskcount 
        total_open_issues += h_open_issuecount 
        total_working_tasks += h_working_taskcount 
        total_working_issues += h_working_issuecount 
        total_overdue_tasks += h_overdue_taskcount 
        total_overdue_issues += h_overdue_issuecount 
        total_pr_tasks += h_pr_taskcount 
        total_pr_issues += h_pr_issuecount
        total_cr_tasks += h_cr_taskcount 
        total_cr_issues += h_cr_issuecount
        total_all_tasks += e_high_task_count
        total_all_issues += e_high_issue_count
        # Add to total task/issue counts
        total_task_count += high_task_count
        total_issue_count += high_issue_count
        # Initialize row data
        row_data = [serial_number, c['project_name'], "High"] + [""] * 14
        
        # Prepare the row data, excluding the project name
        row_data[3] = '' if h_new_taskcount == 0 else h_new_taskcount  # Open Task Count (High)
        row_data[4] = '' if h_new_issuecount == 0 else h_new_issuecount  # Open Issue Count (High)
        row_data[5] = '' if h_open_taskcount == 0 else h_open_taskcount  # Open Task Count (High)
        row_data[6] = '' if h_open_issuecount == 0 else h_open_issuecount  # Open Issue Count (High)
        row_data[7] = '' if h_working_taskcount == 0 else h_working_taskcount  # Working Task Count (High)
        row_data[8] = '' if h_working_issuecount == 0 else h_working_issuecount  # Working Issue Count (High)
        row_data[9] = '' if h_overdue_taskcount == 0 else h_overdue_taskcount  # Overdue Task Count (High)
        row_data[10] = '' if h_overdue_issuecount == 0 else h_overdue_issuecount  # Overdue Issue Count (High)
        row_data[11] = '' if h_pr_taskcount == 0 else h_pr_taskcount  # Pending Review Task Count (High)
        row_data[12] = '' if h_pr_issuecount == 0 else h_pr_issuecount  # Pending Review Issue Count (High)
        row_data[13] = '' if h_cr_taskcount == 0 else h_cr_taskcount  # Client Review Task Count (High)
        row_data[14] = '' if h_cr_issuecount == 0 else h_cr_issuecount  # Client Review Issue Count (High)
        row_data[15] = '' if total_task_count == 0 else total_task_count  # Total Task Count (High)
        row_data[16] = '' if total_issue_count == 0 else total_issue_count  # Total Issue Count (High)


        ws.append(row_data)
 
        priority_cell = ws.cell(row=ws.max_row, column=3)  # Column C for "High"
        priority_cell.font = Font(color="FF0000")
        for idx in [3,4,5, 6, 7, 8, 9, 10, 11, 12, 13,14, 15, 16,17]:
            cell = ws.cell(row=ws.max_row, column=idx)
            cell.font = Font(color="FF0000")

        # # Now check the third column for "High" and color it red
        for row in range(4, ws.max_row+1):  # Adjust based on where your actual data starts
            cell = ws.cell(row=row, column=3)  # Third column
            if cell.value and cell.value.strip() == "High":  # Check for "High"
                cell.font = Font(color="FF0000")  # Set the font color to red
 # Change font color to red
        for priority in priority_rows[1:]: 
            total_mediumtask_count = 0
            total_mediumissue_count = 0
     # Start from Medium to avoid duplicating 'High'
            priority_row_data = ["", c['project_name'], priority[0]] + [""] * 14  # Priority in column 3, rest as blanks
            # priority_row_data[1] = c['project_name']
            priority_cell = ws.cell(row=ws.max_row, column=3)  # Column C for Medium/Low
            priority_cell.font = Font(color="000000")  # Set the font color to black
            # Add counts for Medium and Low priority
            if priority[0] == "Medium":
                m_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "Medium","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                m_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "Medium","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                m_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "Medium"})
                m_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "Medium"})
                m_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"Medium"})
                m_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"Medium"})
                m_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"Medium"})
                m_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"Medium"})
                m_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"Medium"})
                medium_task_count = m_open_taskcount + m_working_taskcount + m_overdue_taskcount + m_pr_taskcount + m_cr_taskcount 
                medium_issue_count = m_open_issuecount + m_working_issuecount + m_overdue_issuecount + m_pr_issuecount + m_cr_issuecount
                e_medium_task_count = m_open_taskcount + m_working_taskcount + m_overdue_taskcount + m_pr_taskcount + m_cr_taskcount 
                e_medium_issue_count = m_open_issuecount + m_working_issuecount + m_overdue_issuecount + m_pr_issuecount + m_cr_issuecount 
                total_new_tasks +=  m_new_taskcount 
                total_new_issues +=m_new_issuecount 
                total_open_tasks += m_open_taskcount 
                total_open_issues +=  m_open_issuecount
                total_working_tasks +=  m_working_taskcount 
                total_working_issues +=  m_working_issuecount 
                total_overdue_tasks +=  m_overdue_taskcount 
                total_overdue_issues +=  m_overdue_issuecount 
                total_pr_tasks += m_pr_taskcount 
                total_pr_issues +=m_pr_issuecount 
                total_cr_tasks += m_cr_taskcount 
                total_cr_issues += m_cr_issuecount             # Accumulate to total task/issue counts
                total_mediumtask_count += medium_task_count
                total_mediumissue_count += medium_issue_count
                total_all_tasks += e_medium_task_count
                total_all_issues += e_medium_issue_count 
                priority_row_data[3] = '' if m_new_taskcount == 0 else m_new_taskcount  # Open Task Count (High)
                priority_row_data[4] = '' if m_new_issuecount == 0 else m_new_issuecount  # Open Issue Count (High)
                priority_row_data[5] = '' if m_open_taskcount == 0 else m_open_taskcount  # Open Task Count (Medium)
                priority_row_data[6] = '' if m_open_issuecount == 0 else m_open_issuecount  # Open Issue Count (Medium)
                priority_row_data[7] = '' if m_working_taskcount == 0 else m_working_taskcount  # Working Task Count (Medium)
                priority_row_data[8] = '' if m_working_issuecount == 0 else m_working_issuecount  # Working Issue Count (Medium)
                priority_row_data[9] = '' if m_overdue_taskcount == 0 else m_overdue_taskcount  # Overdue Task Count (Medium)
                priority_row_data[10] = '' if m_overdue_issuecount == 0 else m_overdue_issuecount  # Overdue Issue Count (Medium)
                priority_row_data[11] = '' if m_pr_taskcount == 0 else m_pr_taskcount  # Pending Review Task Count (Medium)
                priority_row_data[12] = '' if m_pr_issuecount == 0 else m_pr_issuecount  # Pending Review Issue Count (Medium)
                priority_row_data[13] = '' if m_cr_taskcount == 0 else m_cr_taskcount  # Client Review Task Count (Medium)
                priority_row_data[14] = '' if m_cr_issuecount == 0 else m_cr_issuecount  # Client Review Issue Count (Medium)
                priority_row_data[15] = '' if total_mediumtask_count == 0 else total_mediumtask_count  # Total Task Count
                priority_row_data[16] = '' if total_mediumissue_count == 0 else total_mediumissue_count  # Total Issue Count


            elif priority[0] == "Low":
                l_new_taskcount = frappe.db.count("Task", {"spoc": spoc, "project_name": c.project_name, "priority": "Low","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                l_new_issuecount = frappe.db.count("Issue", {"project": c.project_name, "priority": "Low","creation": ["between", [today + " 00:00:00", today + " 23:59:59"]]})
                l_open_taskcount = frappe.db.count("Task", {"status": "Open", "spoc": spoc, "project_name": c.project_name, "priority": "Low"})
                l_open_issuecount = frappe.db.count("Issue", {"status": "Open", "project": c.project_name, "priority": "Low"})
                l_working_taskcount=frappe.db.count("Task",{"status":"Working","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_working_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Working","project":c.project_name,"priority":"Low"})
                l_overdue_taskcount=frappe.db.count("Task",{"status":"Overdue","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_overdue_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Overdue","project":c.project_name,"priority":"Low"})
                l_pr_taskcount=frappe.db.count("Task",{"status":"Pending Review","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_pr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Pending Review","project":c.project_name,"priority":"Low"})
                l_cr_taskcount=frappe.db.count("Task",{"status":"Client Review","spoc":spoc,"project_name":c.project_name,"priority":"Low"})
                l_cr_issuecount=frappe.db.count("Issue",{"custom_issue_status":"Client Review","project":c.project_name,"priority":"Low"})
                low_task_count = l_open_taskcount + l_working_taskcount + l_overdue_taskcount + l_pr_taskcount + l_cr_taskcount 
                low_issue_count = l_open_issuecount + l_working_issuecount + l_overdue_issuecount + l_pr_issuecount + l_cr_issuecount 
                e_low_task_count = l_open_taskcount + l_working_taskcount + l_overdue_taskcount + l_pr_taskcount + l_cr_taskcount 
                e_low_issue_count = l_open_issuecount + l_working_issuecount + l_overdue_issuecount + l_pr_issuecount + l_cr_issuecount
                total_new_tasks += l_new_taskcount
                total_new_issues += l_new_issuecount
                total_open_tasks += l_open_taskcount
                total_open_issues += l_open_issuecount
                total_working_tasks += l_working_taskcount
                total_working_issues += l_working_issuecount
                total_overdue_tasks += l_overdue_taskcount
                total_overdue_issues += l_overdue_issuecount
                total_pr_tasks += l_pr_taskcount
                total_pr_issues +=l_pr_issuecount
                total_cr_tasks += l_cr_taskcount
                total_cr_issues += l_cr_issuecount
                # Accumulate to total task/issue counts
                total_mediumtask_count += low_task_count
                total_mediumissue_count += low_issue_count
                total_all_tasks += e_low_task_count
                total_all_issues += e_low_issue_count 
                priority_row_data[3] = '' if l_new_taskcount == 0 else l_new_taskcount  # Open Task Count (Low)
                priority_row_data[4] = '' if l_new_issuecount == 0 else l_new_issuecount  # Open Issue Count (Low)
                priority_row_data[5] = '' if l_open_taskcount == 0 else l_open_taskcount  # Open Task Count (Low)
                priority_row_data[6] = '' if l_open_issuecount == 0 else l_open_issuecount  # Open Issue Count (Low)
                priority_row_data[7] = '' if l_working_taskcount == 0 else l_working_taskcount  # Working Task Count (Low)
                priority_row_data[8] = '' if l_working_issuecount == 0 else l_working_issuecount  # Working Issue Count (Low)
                priority_row_data[9] = '' if l_overdue_taskcount == 0 else l_overdue_taskcount  # Overdue Task Count (Low)
                priority_row_data[10] = '' if l_overdue_issuecount == 0 else l_overdue_issuecount  # Overdue Issue Count (Low)
                priority_row_data[11] = '' if l_pr_taskcount == 0 else l_pr_taskcount  # Pending Review Task Count (Low)
                priority_row_data[12] = '' if l_pr_issuecount == 0 else l_pr_issuecount  # Pending Review Issue Count (Low)
                priority_row_data[13] = '' if l_cr_taskcount == 0 else l_cr_taskcount  # Client Review Task Count (Low)
                priority_row_data[14] = '' if l_cr_issuecount == 0 else l_cr_issuecount  # Client Review Issue Count (Low)
                priority_row_data[15] = '' if total_mediumtask_count == 0 else total_mediumtask_count  # Total Task Count
                priority_row_data[16] = '' if total_mediumissue_count == 0 else total_mediumissue_count  # Total Issue Count

            ws.append(priority_row_data)  
            # ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
 
            for idx in [3,4,5, 6, 7, 8, 9, 10, 11, 12, 13,14, 15, 16,17]:
                cell = ws.cell(row=ws.max_row, column=idx)
                cell.font = Font(color="000000")
            for row in range(4, ws.max_row+1):  # Adjust based on where your actual data starts
                cell = ws.cell(row=row, column=3)  # Third column
                if cell.value and cell.value.strip() == "High":  # Check for "High"
                    cell.font = Font(color="FF0000")

        serial_number += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+2, end_column=2)
        ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row+2, end_column=1)
        merged_cell = ws.cell(row=current_row, column=2)
        merges_cell = ws.cell(row=s_row, column=1)  # Get the first cell of the merged range
        merged_cell.alignment = Alignment(horizontal="center",vertical='center')
        merges_cell.alignment = Alignment(horizontal="center",vertical='center')


        current_row += 3
        s_row +=3
    total_row = ["", "TOTAL", "",
        total_new_tasks, total_new_issues,
        total_open_tasks, total_open_issues,
        total_working_tasks, total_working_issues,
        total_overdue_tasks, total_overdue_issues,
        total_pr_tasks, total_pr_issues,
        total_cr_tasks, total_cr_issues,
        total_all_tasks, total_all_issues
    ]
    ws.append(total_row)
    total_row_index = ws.max_row
# Merge the first three cells in the total row
    ws.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=3)
    merged_cell = ws.cell(row=total_row_index, column=1)
    merged_cell.value = "TOTAL"
    for col in range(1, len(total_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = row_fill
        cell.font = Font(bold=True, color="000000")  # Bold black font for the total row
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=17):
        for cell in row:
            cell.border = thin_border
    # Save the workbook to a BytesIO object
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    
    return xlsx_file

@frappe.whitelist()
def kk_dpr_daily():
    recievers=[]
    # custom_date=today()
    custom_date = add_days(today(),1)
    date_obj = datetime.strptime(str(custom_date), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date= add_days(today(),1)
    # next_date=today()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    data=[]
    recievers.append('keerthana.k@groupteampro.com') 
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="9"><b>KK DPR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user_email in recievers:
        short_code = frappe.db.get_value("Employee", {"user_id": user_email}, "short_code")
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user_email,"next_contact_date": formatted_next_date,"follow_up_to":"Customer"})
        
        # appointment_count = frappe.db.count("Appointment",{"scheduled_time": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":user_email})
        todo_count=frappe.db.count("ToDo",{"allocated_to":user_email,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,'3' , lead_count if lead_count else '0' , open_count if open_count else '0', replied_count if replied_count else '0', interested_count if interested_count else '0', opportunity_count if opportunity_count else '0', customer_count if customer_count else'0',todo_count if todo_count else '0'
            )
        
    for todo in  recievers:  
        todo_list=frappe.db.get_all("ToDo",{"custom_production_date":formatted_next_date,"allocated_to":todo},["*"])
        if todo_list:
            data += '''
                <tr style="background-color: #0f1568; color: white; text-align:center;">
                    <td colspan="9";"><b>ToDo</b></b></td>
                </tr>
                <tr style="text-align:center; font-weight: 500;"><td colspan="1">Exe</td><td colspan="1">ID</td><td colspan="7" style="padding-right: 300px;">Todo</td></tr>
            '''
            s_no=0
            for i in todo_list:
                s_no+=1
                print(i.allocated_to)

                short_code = frappe.db.get_value("Employee", {"user_id": i.allocated_to}, "short_code")
                data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="7" style="text-align: left; padding-left: 50px;">{}</td></tr>'.format(short_code,i.name,i.custom_subject)
    data += '<tr style="text-align:center;"><td colspan="9"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
        </tr>
    '''
    for user in recievers:
        short_code = frappe.db.get_value("Employee", {"user_id": user}, "short_code")
        lead_day_bforecount = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        open_day_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        replied_day_beforecount= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        interested_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        opportunity_before_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        customer_day_befor_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"next_contact_date": formatted_before_date,"follow_up_to":"Customer"})
        todo_day_before_count=frappe.db.count("ToDo",{"allocated_to":user,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,'3', lead_day_bforecount if lead_day_bforecount else '0' , open_day_before_count if open_day_before_count else '0', replied_day_beforecount if replied_day_beforecount else '0', interested_before_count if interested_before_count else '0', opportunity_before_count if opportunity_before_count else '0', customer_day_befor_count if customer_day_befor_count else'0',todo_day_before_count if todo_day_before_count else '0'
            )
    data += '</table>'



    frappe.sendmail(
                # recipients=recievers,
                # recipients=['divya.p@groupteampro.com'],
                recipients=['keerthana.k@groupteampro.com'],
                cc=['dineshbabu.k@groupteampro.com','sangeetha.a@groupteampro.com'],
                subject='KK DPR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DPR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )

@frappe.whitelist()
def kk_dsr_daily():
    emp_emails=[]
    date_obj = datetime.strptime(str(date.today()), '%Y-%m-%d')
    formatted_date = date_obj.strftime('%d/%m/%Y')
    next_date=nowdate()
    next_dates=datetime.strptime(next_date, '%Y-%m-%d')
    # next_date=add_days(nowdate(),1)
    formatted_next_date=next_dates.strftime('%Y-%m-%d')
    before_date=add_days(today(),-1)
    before_dates=datetime.strptime(before_date, '%Y-%m-%d')
    formatted_before_date=before_dates.strftime('%Y-%m-%d')
    emp_emails.append('keerthana.k@groupteampro.com')
    data = '<table border="1" width="100%" style="border-collapse: collapse;">'
    data += '<tr style="text-align:center;"><td colspan="11"><b>KK DSR, {}</b></td></tr>'.format(formatted_date)
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
        </tr>
    '''
    for c in emp_emails:
        appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":c})

        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_next_date, "allocated_to":c, "status": ('not in',['Cancelled'])})

        short_code=frappe.db.get_value("Employee",{"user_id":c},["short_code"])
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_next_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_next_date})
        data += '<tr style="text-align:center;"><td>{}</td><td>{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}/{}</td><td>{}</td><td>{}</td><td>{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else'0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )
    appointment_list = frappe.db.get_all("Appointment",{"creation": ["between", [f"{formatted_next_date} 00:00:00", f"{formatted_next_date} 23:59:59"]],"owner":["in",emp_emails]},["*"])
    if appointment_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>Appointment Fixed</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1">Exe</td><td colspan="2" style="text-align:center;">Customer</td><td colspan="2">Status</td><td colspan="7">Remarks</td></tr>
        '''
        for i in appointment_list:
            short_code = frappe.db.get_value("Employee", {"user_id": i.owner}, "short_code")
            data+='<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 50px;">{}</td><td colspan="2">{}</td><td colspan="7">{}</td></tr>'.format(short_code,i.name,i.status,i.custom_remarks)
    data += '<tr style="text-align:center;"><td colspan="11"><b>Non Updated Followup</b></td></tr>'
    data += '''
        <tr style="background-color: #0f1568; color: white; text-align:center;">
            <td style="width:10%;"><b>Exe</b></td>
            <td style="width:15%;"><b>Apt</b></td>
            <td style="width:20%;"><b>Lead</b></td>
            <td style="width:13%;"><b>Open</b></td>
            <td style="width:10%;"><b>Replied</b></td>
            <td style="width:7%;"><b>Interested</b></td>
            <td style="width:13%;"><b>Oppr</b></td>
            <td style="width:13%;"><b>Cust</b></td>
            <td style="width:10%;"><b>ToDo</b></b></td>
            <td style="width:10%;"><b>OR%</b></b></td>
            <td style="width:10%;"><b>PR%</b></b></td>
            
        </tr>
    '''
    for user in emp_emails:
        appointment_count = frappe.db.count("Appointment",{"creation": ["between", [f"{formatted_before_date} 00:00:00", f"{formatted_before_date} 23:59:59"]],"owner":user})
        todo_count = frappe.db.count("ToDo", {"custom_production_date": formatted_before_date, "allocated_to":user, "status": ('not in',['Cancelled'])})
        short_code=frappe.db.get_value("Employee",{"user_id":user},["short_code"])
        # effective_call=frappe.db.count("Sales Follow Up", {"next_contact_by": c,"last_contacted_on": formatted_before_date,"call_status": "Effective"})
        lead_count = frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead"})
        effective_call_lead=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Lead","call_status": "Effective"})
        open_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open"})
        effective_call_open=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Open","call_status": "Effective"})
        replied_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied"})       
        effective_call_replied=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Replied","call_status": "Effective"})
        interested_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested"})               
        effective_call_interested=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Interested","call_status": "Effective"})
        opportunity_count= frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity"})   
        effective_call_oppr=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Lead","status": "Opportunity","call_status": "Effective"})
        customer_count=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer"})
        effective_call_cust=frappe.db.count("Sales Follow Up", {"next_contact_by": user,"last_contacted_on": formatted_before_date,"follow_up_to":"Customer","call_status": "Effective"})
        todo_count=frappe.db.count("ToDo",{"allocated_to":c,"custom_production_date":formatted_before_date})
        data += '<tr style="text-align:center;"><td>{}</td><td style="color: red;">{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}/{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td><td style="color: red;">{}</td></tr>'.format(
                short_code,appointment_count if appointment_count else '0' , effective_call_lead if effective_call_lead else '0',lead_count if lead_count else '0', effective_call_open if effective_call_open else '0',open_count if open_count else '0', effective_call_replied if effective_call_replied else '0',replied_count if replied_count else '0', effective_call_interested if effective_call_interested else '0',interested_count if interested_count else '0', effective_call_oppr if effective_call_oppr else '0',opportunity_count if opportunity_count else '0',effective_call_cust if effective_call_cust else '0',customer_count if customer_count else '0',todo_count if todo_count else '0','',''
            )

    todo_list = frappe.db.get_all("ToDo", {"custom_production_date": formatted_next_date, "allocated_to": ["in", emp_emails], "status": ('not in',['Cancelled'])}, ["*"])
    if todo_list:
        data += '''
            <tr style="background-color: #0f1568; color: white; text-align:center;">
                <td colspan="11";"><b>ToDo</b></b></td>
            </tr>
            <tr style="text-align:center; font-weight: 500;background-color: #87CEEB;"><td colspan="1" width="">Exe</td><td colspan="1">ID</td><td colspan="2" style=" text-align: center;">Subject</td><td colspan="7">Remarks</td></tr>
        '''
        for todo in todo_list:
            short_code = frappe.db.get_value("Employee", {"user_id": todo.allocated_to}, "short_code")
            data += '<tr style="text-align:center;"><td colspan="1">{}</td><td colspan="1">{}</td><td colspan="2" style="text-align: left; padding-left: 10px;">{}</td><td colspan="7">{}</td></tr>'.format(short_code, todo.name, todo.custom_subject, todo.current_status_ or '')

    
    data += '</table>'
    frappe.sendmail(
                # recipients=['divya.p@groupteampro.com'],
                recipients=['keerthana.k@groupteampro.com'], 
                cc=['dineshbabu.k@groupteampro.com','sangeetha.a@groupteampro.com'],
                subject='KK DSR %s -Reg' % formatted_date,
                message = """
                <b>Dear Team,</b><br><br>
Please find the below DSR for {} for your kind reference and action.<br><br>

            {}<br><br>
                Thanks & Regards,<br>TEAM ERP<br>
                
                <i>This email has been automatically generated. Please do not reply</i>
                """.format(formatted_date,data)
            )

@frappe.whitelist()
def get_so_details(so):
    sales_order = frappe.get_doc("Sales Order",so)
    return sales_order.items



@frappe.whitelist()
#return the last execution time of attendance cron
def update_last_execution():
    doc=frappe.db.get_value("Scheduled Job Log",{"scheduled_job_type":"mark_attendance.mark_att","status":"Complete"},["creation"])
    if doc:
        return doc

@frappe.whitelist()
def update_tat_completion_date_inchecks(name):
    doc=frappe.get_doc("Criminal",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_inchecks_criminal(name):
    doc=frappe.get_doc("Criminal",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_ed(name):
    doc=frappe.get_doc("Education Checks",name)
    if doc.clear_insufficiency:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.clear_insufficiency
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_education_ch(name):
    doc=frappe.get_doc("Education Checks",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_employment(name):
    doc=frappe.get_doc("Employment",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)

        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_employment_ch(name):
    doc=frappe.get_doc("Employment",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_address(name):
    doc=frappe.get_doc("Address Check",name)
    if doc.clear_insufficiency:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.clear_insufficiency
        if doc.check_package:
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_address_ch(name):
    doc=frappe.get_doc("Address Check",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        if doc.check_package:
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_court(name):
    doc=frappe.get_doc("Court",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        if doc.check_package:
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Court",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_court_ch(name):
    doc=frappe.get_doc("Court",name)
    if doc.check_creation_date and not doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        if doc.check_package:
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Court",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_reference(name):
    doc=frappe.get_doc("Reference Check",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Reference Check",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_reference_ch(name):
    doc=frappe.get_doc("Reference Check",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        if doc.check_package:
            working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Reference Check",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_id(name):
    doc=frappe.get_doc("Identity Aadhar",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Identity Aadhar",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_id_ch(name):
    doc=frappe.get_doc("Identity Aadhar",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Identity Aadhar",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_sm(name):
    doc=frappe.get_doc("Social Media",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
    
@frappe.whitelist()
def update_tat_completion_date_sm_ch(name):
    doc=frappe.get_doc("Social Media",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]

@frappe.whitelist()
def update_tat_completion_date_family(name):
    doc=frappe.get_doc("Family",name)
    if doc.insuff_cleared_on:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.insuff_cleared_on
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        
        return holiday[-1]
        # frappe.db.set_value("Family",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def update_tat_completion_date_family_ch(name):
    doc=frappe.get_doc("Family",name)
    if doc.check_creation_date:
        from erpnext.setup.doctype.holiday_list.holiday_list import is_holiday
        holiday_list_name = 'TEAMPRO 2023 - Checkpro'
        start_date = doc.check_creation_date
        working_days = int(frappe.db.get_value("Check Package",{'name':doc.check_package},['package_tat']))
        current_date = start_date
        holiday = []
        while working_days > 0:
            if not is_holiday(holiday_list_name, current_date):
                holiday.append(current_date)
                working_days -= 1
            current_date = add_days(current_date, 1)
        return holiday[-1]
        # frappe.db.set_value("Family",doc.name,"tat_completion_date",holiday[-1])

@frappe.whitelist()
def validate_permission_request(doc,method):
    start_date = get_first_day(doc.permission_date) 
    end_date = get_last_day(doc.permission_date) 
    permission=0
    permission_list = frappe.db.get_all("Attendance Permission",{"name": ("!=", doc.name),"employee": doc.employee,"permission_date": ("between", [start_date, end_date]),"docstatus":("!=",2)},["*"])    
    if permission_list:
        for i in permission_list:
            permission+=int(i.total_time)
            if permission>=2:
                frappe.throw("Only 2 hours Permission is allowed for the month")
            elif permission<2:
                permission+=int(doc.total_time)
                if permission>2:
                    frappe.throw("Already applied for 1 hour permission.You are only  allow to apply additionaly 1 hour")


@frappe.whitelist()
def update_permission_req_in_att(doc,method):
    attendance=frappe.get_doc("Attendance",{"attendance_date":doc.permission_date,"docstatus":("!=",2),"employee":doc.employee})
    hours=0
    if attendance:
        attendance.custom_attendance_permission=doc.name
        if attendance.bt_difference:
            diff=(attendance.bt_difference)
            hours = diff+int(doc.total_time)
        if hours>=8:
            attendance.status="Present"
        elif hours>=4 and hours < 8:
            attendance.status="Half Day"
        else:
            attendance.status="Absent"
    attendance.save()
    frappe.db.commit()

@frappe.whitelist()
def update_permission_req_in_att_cancel(doc,method):
    attendance=frappe.get_doc("Attendance",{"attendance_date":doc.permission_date,"docstatus":("!=",2),"employee":doc.employee})
    hours=0
    if attendance:
        attendance.custom_attendance_permission=""
        if attendance.bt_difference:
            diff=(attendance.bt_difference)
            hours = diff
        if hours>=8:
            attendance.status="Present"
        elif hours>=4 and hours < 8:
            attendance.status="Half Day"
        else:
            attendance.status="Absent"
    attendance.save()
    frappe.db.commit()


@frappe.whitelist()
def update_sfp_remarks(doc,method):
    if doc.status in ['Open','Overdue','Enquiry']:
        if doc.customer:
            status=frappe.db.get_value('Customer',{'name':doc.customer},['disabled'])
            sfp=frappe.db.get_all("Sales Follow Up",{'party_from':'Customer','party_name':doc.customer,'service':doc.service},['active','name'])
            if status==0 and sfp:
                for s in sfp:
                    if s.active==0:
                        frappe.db.set_value("Sales Follow Up",s.name,'active',True)   
             

@frappe.whitelist()
def update_sfp_status():
    sfp=frappe.db.get_all("Sales Follow Up",{'party_from':'Customer'},['party_name','service','name'])
    for s in sfp:
        status=frappe.db.get_value('Customer',{'name':s.party_name},['disabled'])
        if status==0:
            if frappe.db.exists("Project",{'Customer':s.party_name,'status':('in',('Open','Overdue','Enquiry')),'service':s.service}):
                frappe.db.set_value('Sales Follow Up',s.name,'active',1)
            else:
                frappe.db.set_value('Sales Follow Up',s.name,'active',0)
            
            
@frappe.whitelist()
def update_profile_submission_project(project):
    tasks=frappe.db.get_all("Task",{"project":project},["name"])
    return tasks

@frappe.whitelist()
def send_mail_for_profile_submission(project):
    pro=frappe.get_doc("Project",{"name":project})
    posting_date = datetime.now().strftime("%d-%m-%Y")
    spoc=pro.spoc
    table = '<table text-align="center" border="1" width="75%" style="border-collapse: collapse;">'
    table += '<tr style="background-color: #87CEFA"><td style= width="1%;font-weight: bold;"><b>Project</b></td><td style= width="1%";font-weight: bold;"><b>Task ID</b></td><td style="width:1%; font-weight: bold;">Position</td><td style="width: 1%; font-weight: bold;">#Profiles</td><td style="width:2%; font-weight: bold;">Date</td></tr>'
    for i in pro.custom_profile_submission:
        formatted_date = frappe.utils.formatdate(i.date, 'dd-mm-yyyy')
        table+="""<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>"""%(i.project,i.task,i.position,i.profiles,formatted_date)
    table+='</table>'
    subject="New Project Batch Profile Submission Plan-  %s" % posting_date
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Project Details:<p><b>Project Name:</b>{}<p><b>Project ID:</b>{}<p><b>Project Manager</b>{}<p> A new project has been opened for your further action.<br>Kindly find the 1 st batch profile submission plan<br>{}<br></p><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(pro.project_name,pro.name,pro.project_manager,table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["sangeetha.a@groupteampro.com","dineshbabu.k@groupteampro.com","sangeetha.s@groupteampro.com","aruna.g@groupteampro.com",spoc],
        subject=subject,
        message=message
    )


@frappe.whitelist()
def project_send_mail_to_creation_adv(name):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    pro = frappe.get_doc("Project",name)
    tasks = frappe.get_all("Task",{'project': pro.name},['*'])
    t=frappe.db.get_all("Task",{'project': pro.name},['food'],group_by='food')
    food_count=len(t)
    qualification=frappe.get_all("Task",{'project': pro.name},['qualification_type'],group_by='qualification_type')
    qual_count=len(qualification)
    experience=frappe.get_all("Task",{'project': pro.name},['total_experience'],group_by='total_experience')
    exp_count=len(experience)
    g_experience=frappe.get_all("Task",{'project': pro.name},['gulf_experience'],group_by='gulf_experience')
    g_exp=len(g_experience)
    interview=frappe.get_all("Task",{'project': pro.name},['mode_of_interview'],group_by='mode_of_interview')
    int_count=len(interview)
    acc=frappe.get_all("Task",{'project': pro.name},['accommodation'],group_by='accommodation')
    a_count=len(acc)
    transport=frappe.get_all("Task",{'project': pro.name},['transportation'],group_by='transportation')
    trans_count=len(transport)
    visa=frappe.get_all("Task",{'project': pro.name},['visa_type'],group_by='visa_type')
    v_count=len(visa)
    con=frappe.get_all("Task",{'project': pro.name},['contract_period_year'],group_by='contract_period_year')
    con_count=len(con)
    categorys=frappe.get_all("Task",{'project': pro.name},['category'],group_by='category')
    ca_count=len(categorys)
    keys=frappe.get_all("Task",{'project': pro.name},['custom_major_key_skills'],group_by='custom_major_key_skills')
    key_count=len(keys)
    rec=frappe.get_all("Task",{'project': pro.name},['custom_free_recruitment'],group_by='custom_free_recruitment')
    rec_count=len(rec)
    task_count=(frappe.db.count("Task",{'project': pro.name}))
    serial_no = 1
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 10%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">Title</td><td style="width: 60%; font-weight: bold; text-align: center;">Details</td></tr>'
    table += """<tr><td>1</td><td>Project ID</td><td>{}</td></tr>""".format(pro.name or '')
    table += """<tr><td>2</td><td>Date</td><td>{}</td></tr>""".format(pro.custom_actionconfirmed_datetime or '')
    table += """<tr><td>3</td><td>Country</td><td>{}</td></tr>""".format(pro.territory or '')
    table += """<tr><td>4</td><td>Client</td><td>{}</td></tr>""".format(pro.customer or '')
    for i in tasks:
        if tasks.index(i)==0:
            table += """<tr><td rowspan={}>5</td><td rowspan={}>Positions</td><td>{}</td></tr>""".format(task_count,task_count,i.subject or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(i.subject or '')
    for k in keys:
        if keys.index(k)==0:
            table += """<tr><td rowspan={}>6</td><td rowspan={}>Major Key Skills</td><td>{}</td></tr>""".format(key_count,key_count,k.custom_major_key_skills or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(k.custom_major_key_skills or '')
    for d in qualification:
        if qualification.index(d)==0:
            table += """<tr><td rowspan={}>7</td><td rowspan={}>Qualification</td><td>{}</td></tr>""".format(qual_count,qual_count,d.qualification_type or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(d.qualification_type or '')
    for e in experience:
        if experience.index(e)==0:
            table += """<tr><td rowspan={}>8</td><td rowspan={}>Experience</td><td>{}</td></tr>""".format(exp_count,exp_count,e.total_experience or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(e.total_experience or '')
    for g in g_experience:
        if g_experience.index(g)==0:
            table += """<tr><td rowspan={}>9</td><td rowspan={}>GCC Experience</td><td>{}</td></tr>""".format(g_exp,g_exp,g.gulf_experience or '')
        else:
            table +="""<tr><td>{}</td></tr>""".format(g.gulf_experience or '')
    for r in rec:
        if rec.index(r)==0:
            table += """<tr><td rowspan={}>10</td><td rowspan={}>Free Recruitment</td><td>{}</td></tr>""".format(rec_count,rec_count,r.custom_free_recruitment or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(r.custom_free_recruitment or '')
    for m in interview:
        if interview.index(m)==0:
            table += """<tr><td rowspan={}>11</td><td rowspan={}>Mode Of Interview</td><td>{}</td></tr>""".format(int_count,int_count,m.mode_of_interview or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(m.mode_of_interview or '')
    table += """<tr><td>12</td><td>If Direct Client Interview - Location & Date</td><td></td></tr>"""
    table += """<tr><td>13</td><td>Contact Number</td><td>+91 75502 24400,+91 73050 56204</td></tr>"""
    table += """<tr><td>14</td><td>Mail ID</td><td>aruna.g@groupteampro.com</td></tr>"""
    for v in visa:
            if visa.index(v)==0:
                table += """<tr><td rowspan={}>15</td><td rowspan={}>Visa Type</td><td>{}</td></tr>""".format(v_count,v_count,v.visa_type or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(v.visa_type or '')
    for cons in con:
            if con.index(cons)==0:
                table += """<tr><td rowspan={}>16</td><td rowspan={}>Contract</td><td>{}</td></tr>""".format(con_count,con_count,cons.contract_period_year or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cons.contract_period_year or '')
    for cat in categorys:
            if categorys.index(cat)==0:
                table += """<tr><td rowspan={}>17</td><td rowspan={}>ECR/ECNR</td><td>{}</td></tr>""".format(ca_count,ca_count,cat.category or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cat.category or '')
    table += """<tr><td>18</td><td>Special Remarks</td><td>Attractive Salary</td></tr>"""
    table += """<tr><td>19</td><td>Common Version 1.0</td><td>Company Name + Logo + RA Licence + Location + Website + Common Number (7305056202) + Common Mail ID</td></tr>"""
    table += '</table>'
    subject = "Advertisement Details - %s" %posting_date
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Advertisement Confirmed Details:<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["dineshbabu.k@groupteampro.com","dm@groupteampro.com","annie.m@groupteampro.com"],
        subject=subject,
        message=message
    )

@frappe.whitelist()     
def new_project_creation_mail(name):
    create=frappe.get_doc("Project",name)
    posting_date = datetime.now().strftime("%d-%m-%Y")
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 10%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">Title</td><td style="width: 60%; font-weight: bold; text-align: center;">Details</td></tr>'
    table += """<tr><td>1</td><td>Project ID</td><td>{}</td></tr>""".format(create.name or '')
    table += """<tr><td>2</td><td>Project Name</td><td>{}</td></tr>""".format(create.project_name or '')
    table += """<tr><td>3</td><td>Customer</td><td>{}</td></tr>""".format(create.customer or '')
    table += """<tr><td>4</td><td>Mode Of Interview</td><td>{}</td></tr>""".format(create.mode_of_interview or '')
    table += """<tr><td>5</td><td>#Positions</td><td>{}</td></tr>""".format(create.task or '')
    table += """<tr><td>6</td><td>#vacancies</td><td>{}</td></tr>""".format(create.tvac or '')
    table+='</table>'
    subject="Project Created-  %s" % posting_date
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below project Created Details:<p><b>Project Name:</b>{}<p><b>Project ID:</b>{}– a new project has been created for your further action.<br>{}<br></p><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(create.project_name,create.name,table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["sangeetha.a@groupteampro.com","dineshbabu.k@groupteampro.com","sangeetha.s@groupteampro.com","aruna.g@groupteampro.com"],
        subject=subject,
        message=message
    )

@frappe.whitelist()
def get_to_date(from_date):
    return get_last_day(from_date)


@frappe.whitelist()
def create_job_fail():
    job = frappe.db.exists('Scheduled Job Type', 'cron_failed')
    if not job:
        emc = frappe.new_doc("Scheduled Job Type")
        emc.update({
            "method": 'teampro.custom.cron_failed_method',
            "frequency": 'Cron',
            "cron_format": '*/5 * * * *'
        })
        emc.save(ignore_permissions=True)


@frappe.whitelist()
def cron_failed_method():
    cutoff_time = datetime.now() - timedelta(minutes=5)
    failed_jobs = frappe.get_all(
        "Scheduled Job Log",
        filters={
            "status": "Failed",
            "creation": [">=", cutoff_time]
        },
        fields=["scheduled_job_type"]
    )
    unique_job_types = set()
    for job in failed_jobs:
        unique_job_types.add(job['scheduled_job_type'])

    for job_type in unique_job_types:
        frappe.sendmail(
            recipients = ["erp@groupteampro.com","jenisha.p@groupteampro.com","pavithra.s@groupteampro.com","gifty.p@groupteampro.com"],
            subject = 'Failed Cron List - Internal',
            message = 'Dear Sir / Mam <br> Kindly find the below failed Scheduled Job  %s'%(job_type)
        )


@frappe.whitelist()
def create_new_epnc_review():
    employee=frappe.db.get_all("Employee",{"status":"Active"},["*"])
    today_date=today()
    start_date=get_first_day(today_date)
    end_date=get_last_day(today_date)
    for i in employee:
        doc=frappe.new_doc("Monthly EP NC Review")
        doc.start_date=start_date
        doc.end_date=end_date
        doc.employee=i.name
        doc.total_score='100'
        doc.save()
    frappe.db.commit()



@frappe.whitelist()
def update_score_in_epnc_review(doc, method):
    from datetime import datetime
    creation_date_result = frappe.db.sql(
        """
        SELECT date(creation) as creation from `tabEnergy Point And Non Conformity` where creation=%s AND name=%s
        """, (doc.creation,doc.name), as_dict=True)
    creation_date = creation_date_result[0]['creation'] if creation_date_result else None
    first_date=get_first_day(creation_date)
    documents=frappe.db.get_all("Monthly EP NC Review",{"start_date":first_date,'employee':doc.emp},["name"])
    for i in documents:
        doc=frappe.get_doc("Monthly EP NC Review",i)
        ep = frappe.db.sql(
            """
            SELECT sum(total) as total from `tabEnergy Point And Non Conformity` where emp=%s AND action='Energy Point(EP)' AND docstatus=1 AND date(creation) BETWEEN %s AND %s
            """, (doc.employee, doc.start_date, doc.end_date), as_dict=True)
        nc = frappe.db.sql(
            """
            SELECT sum(total_nc) as total_nc from `tabEnergy Point And Non Conformity` where emp=%s AND action='Non Conformity(NC)' AND docstatus=1 AND date(creation) BETWEEN %s AND %s
            """, (doc.employee, doc.start_date, doc.end_date), as_dict=True)
        doc.total_ep = ep[0]['total'] if ep and ep[0]['total'] is not None else 0
        doc.total_nc = nc[0]['total_nc'] if nc and nc[0]['total_nc'] is not None else 0
        total = 100 - doc.total_nc + doc.total_ep
        doc.total_score = total
        doc.save(ignore_permissions=True)

    frappe.db.commit()


@frappe.whitelist()
def update_service(name):
    doc = frappe.get_doc("Target Manager",name)
    total_ct = 0
    total_ft = 0
    doc.target_child=[]
    doc.monthly_ft_allocation=[]
    if doc.service_list:
        for i in doc.service_list:
            total_ct += i.ct
            total_ft += i.ft
    doc.annual_ct = total_ct
    doc.annual_ft = total_ft
    months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
    month_no = {'Apr':'12','May':'11','Jun':'10','Jul':'9','Aug':'8','Sep':'7','Oct':'6','Nov':'5','Dec':'4','Jan':'3','Feb':'2','Mar':'1'}
    value =doc.annual_ct / 12
    value_ft =doc.annual_ft / 12
    for month in months:
        doc.append("target_child", {
            'month': month,
            'month_nos': month_no[month],
            'ct': value
        })
        doc.append("monthly_ft_allocation", {
            'month': month,
            'month_nos': month_no[month],
            'ft': value_ft
        })

    doc.save()
    frappe.db.commit()

@frappe.whitelist()
def update_service_tm(doc,method):
    doc = frappe.get_doc("Target Manager",doc.name)
    total_ct = 0
    total_ft = 0
    doc.target_child=[]
    doc.monthly_ft_allocation=[]
    if doc.service_list:
        for i in doc.service_list:
            total_ct += i.ct
            total_ft += i.ft
    doc.annual_ct = total_ct
    doc.annual_ft = total_ft
    months = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
    month_no = {'Apr':'12','May':'11','Jun':'10','Jul':'9','Aug':'8','Sep':'7','Oct':'6','Nov':'5','Dec':'4','Jan':'3','Feb':'2','Mar':'1'}
    value =doc.annual_ct / 12
    value_ft =doc.annual_ft / 12
    for month in months:
        doc.append("target_child", {
            'month': month,
            'month_nos': month_no[month],
            'ct': value
        })
        doc.append("monthly_ft_allocation", {
            'month': month,
            'month_nos': month_no[month],
            'ft': value_ft
        })

    doc.save()
    frappe.db.commit()

@frappe.whitelist()
def organization_update_sp(customer):
    customer=frappe.db.get_value("Customer",{"name":customer},["name"])
    return customer

@frappe.whitelist()
def organization_update_sp_lead(lead):
    company_name=frappe.db.get_value("Lead",{"name":lead},["company_name"])
    return company_name

# @frappe.whitelist()
# def organization_update_sp_test():
#     customer=frappe.db.get_all("Sales Follow Up",{"organization_name":("=",''),"party_from":"Customer"},["*"])
#     count=0
#     for i in customer:
#         organization_name=frappe.db.get_value("Customer",{"name":i.customer},["name"])
#         print(i.name)
#         print(organization_name)
#         frappe.db.set_value("Sales Follow Up",i.name,"organization_name",organization_name)
#         count+=1
#     print(count)

import frappe
from frappe.utils import today
from datetime import datetime

@frappe.whitelist()
def send_mail_for_update_checkpro_holiday():
    holiday_list = frappe.get_doc("Holiday List", {"name": "TEAMPRO 2023 - Checkpro"})
    if holiday_list.holidays:
        current_date = today()
        current_year = datetime.now().year
        current_month = datetime.now().month

        if current_month == 12:  # Perform the check only in December
            next_year = current_year + 1  # Calculate the next year
            holiday_for_next_year = any(
                holiday.holiday_date.year == next_year for holiday in holiday_list.holidays
            )
            
            if not holiday_for_next_year:
                subject = f"Add Holidays for TEAMPRO 2023 - Checkpro{next_year}"
                message = (
                    f"Dear Sir/Mam, <br><br>"
                    f"The holiday list for the year {next_year} is missing in the 'TEAMPRO 2023 - Checkpro' holiday list. "
                    f"Please update the holiday list for {next_year} to avoid any disruptions. <br><br>"
                    f"Regards,<br>Team"
                )
                recipients = ["divya.p@groupteampro.com","chitra.g@groupteampro.com","sangeetha.a@groupteampro.com","sangeetha.s@groupteampro.com"] 
                frappe.sendmail(
                    recipients=recipients, 
                    subject=subject, 
                    message=message)
                
# def dsr_for_project_spoc_daily():
#     job = frappe.db.exists('Scheduled Job Type', 'send_mail_for_update_checkpro_holiday')
#     if not job:
#         sjt = frappe.new_doc("Scheduled Job Type")
#         sjt.update({
#             "method": 'teampro.custom.send_mail_for_update_checkpro_holiday',
#             "frequency": 'Cron',
#             "cron_format": '00 00 1 * *'
#         })
#         sjt.save(ignore_permissions=True)


@frappe.whitelist()
def update_currency_amount(currency, amount):
    if amount:
        if currency == "INR" or not currency:
            amount_value = amount
        else:
            conversion = get_exchange_rate(currency, "INR")
            conversion_amt= float (conversion) * float(amount)
            amount_value =conversion_amt
        return amount_value


@frappe.whitelist()
def update_sla_details(name, service, sla_from_date, sla_to_date, sla_type, status,attach):
    try:
        customer = frappe.get_doc("Customer", name)
        if customer.custom_sla_details:
            for i in customer.custom_sla_details:
                customer.append("custom_sla_history", {
                    "service": i.service,
                    "sla_from_date": i.sla_from_date,
                    "sla_to_date": i.sla_to_date,
                    "sla_type": i.sla_type,
                    "status": i.status,
                    "attach":i.attach,
                })
            customer.set("custom_sla_details", [])
        customer.append("custom_sla_details", {
            "service": service,
            "sla_from_date": sla_from_date,
            "sla_to_date": sla_to_date,
            "sla_type": sla_type,
            "status": status,
            "attach":attach,
        })

        customer.save(ignore_permissions=True)
        frappe.db.commit()
        return {"status": "success", "message": "SLA details updated successfully"}
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Update SLA Details")
        return {"status": "error", "message": str(e)}



# @frappe.whitelist()
# def update_sla_status_and_notify():
# 	customers = frappe.db.get_all("Customer", {"disabled": 0,"name":"Divya"}, ["*"])
# 	expired_sla_data = []

# 	for customer_info in customers:
# 		customer = frappe.get_doc("Customer", customer_info.name)
# 		expired_services = []
        
# 		if customer.custom_sla_details:
# 			for sla in customer.custom_sla_details:
# 				print(sla.service)
# 				if sla.sla_to_date and getdate(sla.sla_to_date) < getdate(nowdate()):
# 					expired_services.append(sla)
# 					if expired_services:
# 						for sla in expired_services:
# 							print(sla.service)
# 							expired_sla_data.append({
# 								"customer_name": customer.customer_name,
# 								"service": sla.service,
# 								"sla_to_date": sla.sla_to_date
# 							})
# 			customer.save(ignore_permissions=True)
# 			frappe.db.commit()

# 	print(expired_sla_data)
# 	# Send a single email if there are expired SLAs
# 	if expired_sla_data:
# 		table_rows = "".join(
# 			f"<tr><td>{entry['customer_name']}</td><td>{entry['service']}</td><td>{entry['sla_to_date']}</td></tr>"
# 			for entry in expired_sla_data
# 		)

# 		table_html = f"""
# 			<table border="1" style="border-collapse: collapse; width: 100%;">
# 				<thead>
# 					<tr>
# 						<th style="padding: 8px; text-align: left;">Customer Name</th>
# 						<th style="padding: 8px; text-align: left;">Service</th>
# 						<th style="padding: 8px; text-align: left;">SLA To Date</th>
# 					</tr>
# 				</thead>
# 				<tbody>
# 					{table_rows}
# 				</tbody>
# 			</table>
# 		"""

# 		subject = "Alert: Expired SLA Details"
# 		message = f"""
# 			Dear Sir/Mam,<br><br>
# 			The following SLA services have expired:<br><br>
# 			{table_html}
# 			<br><br>
# 			Please take necessary actions to renew these services.<br><br>
# 			Best Regards,<br>Your Team
# 		"""
# 		recipients = ["divya.p@groupteampro.com"]

# 		frappe.sendmail(
# 			recipients=recipients,
# 			subject=subject,
# 			message=message,
# 		)

# import frappe
# from frappe.utils import getdate, nowdate, add_days

# @frappe.whitelist()
# def update_sla_status_and_notify():
# 	customers = frappe.db.get_all("Customer", {"disabled": 0}, ["*"])
# 	expired_sla_data = []
# 	nearing_expiry_sla_data = []

# 	for customer_info in customers:
# 		customer = frappe.get_doc("Customer", customer_info.name)
        
# 		if customer.custom_sla_details:
# 			for sla in customer.custom_sla_details:
# 				if sla.sla_to_date:
# 					sla_date = getdate(sla.sla_to_date)
# 					formatted_sla_date = formatdate(sla.sla_to_date, "dd-MM-yyyy")
# 					today = getdate(nowdate())

# 					# Expired SLAs (already past)
# 					if sla_date < today:
# 						expired_sla_data.append({
# 							"customer_name": customer.customer_name,
# 							"service": sla.service,
# 							"sla_to_date": formatted_sla_date
# 						})

# 					# Nearing expiry SLAs (expiring within the next 60 days)
# 					elif 0 <= (sla_date - today).days <= 60:
# 						nearing_expiry_sla_data.append({
# 							"customer_name": customer.customer_name,
# 							"service": sla.service,
# 							"sla_to_date": formatted_sla_date
# 						})

# 	# If there's any data to notify
# 	if expired_sla_data or nearing_expiry_sla_data:
        
# 		expired_table_rows = "".join(
# 			f"<tr><td>{entry['customer_name']}</td><td>{entry['service']}</td><td>{entry['sla_to_date']}</td></tr>"
# 			for entry in expired_sla_data
# 		)
# 		nearing_expiry_table_rows = "".join(
# 			f"<tr><td>{entry['customer_name']}</td><td>{entry['service']}</td><td>{entry['sla_to_date']}</td></tr>"
# 			for entry in nearing_expiry_sla_data
# 		)

# 		# Tables for both cases
# 		expired_table_html = f"""
# 			<h3>Expired SLAs</h3>
# 			<table border="1" style="border-collapse: collapse; width: 100%;">
# 				<thead>
# 					<tr>
# 						<th style="padding: 8px; text-align: left;">Customer Name</th>
# 						<th style="padding: 8px; text-align: left;">Service</th>
# 						<th style="padding: 8px; text-align: left;">SLA To Date</th>
# 					</tr>
# 				</thead>
# 				<tbody>
# 					{expired_table_rows}
# 				</tbody>
# 			</table>
# 		""" if expired_sla_data else ""

# 		nearing_expiry_table_html = f"""
# 			<h3>SLAs Nearing Expiry (Next 60 Days)</h3>
# 			<table border="1" style="border-collapse: collapse; width: 100%;">
# 				<thead>
# 					<tr>
# 						<th style="padding: 8px; text-align: left;">Customer Name</th>
# 						<th style="padding: 8px; text-align: left;">Service</th>
# 						<th style="padding: 8px; text-align: left;">SLA To Date</th>
# 					</tr>
# 				</thead>
# 				<tbody>
# 					{nearing_expiry_table_rows}
# 				</tbody>
# 			</table>
# 		""" if nearing_expiry_sla_data else ""

# 		# Construct final email message
# 		subject = "Alert: SLA Expiry Notifications"
# 		message = f"""
# 			Dear Sir/Mam,<br><br>
# 			Please find the SLA expiry details below:<br><br>
# 			{expired_table_html}
# 			<br>
# 			{nearing_expiry_table_html}
# 			<br><br>
# 			Please take necessary actions.<br><br>
# 			Best Regards,<br>Teampro
# 		"""

# 		recipients = ["divya.p@groupteampro.com"]

# 		frappe.sendmail(
# 			recipients=recipients,
# 			subject=subject,
# 			message=message,
# 		)
import frappe
from frappe.utils import getdate, nowdate, add_days, formatdate

@frappe.whitelist()
def update_sla_status_and_notify():
    customers = frappe.db.get_all("Customer", {"disabled": 0}, ["*"])
    expired_sla_data = []
    nearing_expiry_sla_data = []

    today = getdate(nowdate())

    for customer_info in customers:
        customer = frappe.get_doc("Customer", customer_info.name)
        
        if customer.custom_sla_details:
            for sla in customer.custom_sla_details:
                if sla.sla_to_date:
                    sla_date = getdate(sla.sla_to_date)
                    formatted_sla_date = formatdate(sla.sla_to_date, "dd-MM-yyyy")

                    # Expired SLAs (already past)
                    if sla_date < today:
                        expired_sla_data.append({
                            "customer_name": customer.customer_name,
                            "service": sla.service,
                            "sla_to_date": formatted_sla_date,
                            "description":sla.description,
                            "sla_type":sla.sla_type
                        })

                    # Nearing Expiry SLAs (expiring within the next 60 days)
                    elif 0 <= (sla_date - today).days <= 60:
                        nearing_expiry_sla_data.append({
                            "customer_name": customer.customer_name,
                            "service": sla.service,
                            "sla_to_date": formatted_sla_date,
                            "days_remaining": (sla_date - today).days,
                            "description":sla.description,
                            "sla_type":sla.sla_type
                        })

    # If there's any data to notify
    if expired_sla_data or nearing_expiry_sla_data:
        expired_table_rows = "".join(
            f"<tr><td>{entry['customer_name']}</td><td>{entry['service']}</td><td>{entry['sla_type']}</td><td>{entry['sla_to_date']}</td><td>{entry['description']}</td></tr>"
            for entry in expired_sla_data
        )
        nearing_expiry_table_rows = "".join(
            f"<tr><td>{entry['customer_name']}</td><td>{entry['service']}</td><td>{entry['sla_type']}</td><td>{entry['sla_to_date']}</td><td>{entry['days_remaining']} days</td><td>{entry['description']}</td></tr>"
            for entry in nearing_expiry_sla_data
        )

        # Tables with equal widths
        expired_table_html = f"""
            <div>
                <h3>Expired SLAs</h3>
                <table border="1" width="100%" style="border-collapse: collapse;">
                    <thead>
                        <tr style="background-color: #f2f2f2;">
                            <th style="width: 40%;">Customer Name</th>
                            <th style="width: 20%;">Service</th>
                            <th style="width: 20%;">Type</th>
                            <th style="width: 20%;">SLA To Date</th>
                            <th style="width: 40%;">Description</th>
                        </tr>
                    </thead>
                    <tbody>
                        {expired_table_rows}
                    </tbody>
                </table>
            </div>
        """ if expired_sla_data else ""

        nearing_expiry_table_html = f"""
            <div>
                <h3>SLAs Nearing Expiry (Next 60 Days)</h3>
                <table border="1" width="100%" style="border-collapse: collapse;">
                    <thead>
                        <tr style="background-color: #f2f2f2;">
                            <th style="width: 30%;">Customer Name</th>
                            <th style="width: 10%;">Service</th>
                            <th style="width: 10%;">Type</th>
                            <th style="width: 20%;">SLA To Date</th>
                            <th style="width: 20%;">Days Remaining</th>
                            <th style="width: 40%;">Description</th>
                        </tr>
                    </thead>
                    <tbody>
                        {nearing_expiry_table_rows}
                    </tbody>
                </table>
            </div>
        """ if nearing_expiry_sla_data else ""

        # Construct final email message
        subject = "Alert: SLA Expiry Notifications"
        message = f"""
            Dear Sir/Mam,<br><br>
            Please find the SLA expiry details below:<br><br>
            {expired_table_html}
            {nearing_expiry_table_html}
            <br><br>
            These alerts will be sent daily until the SLA expires.<br><br>
            Best Regards,<br>Teampro
        """


        recipients = ["annie.m@groupteampro.com","dineshbabu.k@groupteampro.com"]
        # recipients = ["divya.p@groupteampro.com"]

        frappe.sendmail(
            recipients=recipients,
            subject=subject,
            message=message,
        )


@frappe.whitelist()
def update_sla_status():
    customers = frappe.db.get_all("Customer", {"disabled": 0}, ["*"])

    for customer_info in customers:
        customer = frappe.get_doc("Customer", customer_info.name)
        if customer.custom_sla_details:
            for sla in customer.custom_sla_details:
                if sla.sla_to_date and getdate(sla.sla_to_date) < getdate(nowdate()):
                    sla.status = "Expired"
            customer.save(ignore_permissions=True)
            frappe.db.commit()

# @frappe.whitelist()
# def task_mail_notification_status ():
#     job = frappe.db.exists('Scheduled Job Type','send_daily_psr_report')
#     if not job:
#         task = frappe.new_doc("Scheduled Job Type")
#         task.update({
#             "method": 'teampro.teampro.doctype.psr_report_dashboard.psr_report_dashboard.send_daily_psr_report',
#             "frequency": 'Cron',
#             "cron_format": '0 9 * * *'
#         })
#         task.save(ignore_permissions=True)

# @frappe.whitelist()
# def update_case_age_test():
#     age=0
#     tat_var=0
#     tat_mon=''
#     tat_sts=''
#     doc=frappe.db.get_list("Case",{"name":"SHF-130125-15473-00001"},["name","date_of_initiating","case_status",'insufficiency_days','package_tat','insufficiency_closed'],order_by='date_of_initiating ASC')
#     for i in doc:
#         if i.case_status not in ("Case Completed","Drop","Generate Report with Insuff",'',"Drop"):
#             if i.date_of_initiating and not i.insufficiency_closed:
#                 date=(date_diff(nowdate(),i.date_of_initiating))+1
#                 sql_query = f"""
#                     SELECT COUNT(*) 
#                     FROM `tabHoliday` 
#                     WHERE parent = 'TEAMPRO 2023 - Checkpro' 
#                     AND holiday_date BETWEEN '{i.date_of_initiating}' AND '{nowdate()}'
#                 """
#                 count = frappe.db.sql(sql_query, as_list=True)[0][0]
#                 print(count)
#                 print(date)
#                 if count==0:
#                     age=date-i.insufficiency_days
#                 else:
#                     age = (date-(count+i.insufficiency_days))
#                 print(i.name)
#                 print(age)
#             else:
#                 if i.insufficiency_closed:
#                     date=(date_diff(nowdate(),i.insufficiency_closed))+1
#                     sql_query = f"""
#                         SELECT COUNT(*) 
#                         FROM `tabHoliday` 
#                         WHERE parent = 'TEAMPRO 2023 - Checkpro' 
#                         AND holiday_date BETWEEN '{i.date_of_initiating}' AND '{nowdate()}'
#                     """
#                     count = frappe.db.sql(sql_query, as_list=True)[0][0]
#                     print(count)
#                     print(date)
#                     if count==0:
#                         age=date-i.insufficiency_days
#                     else:
#                         age = (date-(count))
#                     print(i.name)
#                     print(age)

frappe.whitelist()
def update_wh_att(doc,method):
    if frappe.db.exists("Attendance",{"attendance_request":doc.name,"employee":doc.employee,'docstatus':['!=',2]}):
        att=frappe.db.get_value("Attendance",{"attendance_request":doc.name,"employee":doc.employee,'docstatus':['!=',2]},['name'])
        if att:
            frappe.db.set_value("Attendance",att,'attendance_request','')          

# frappe.whitelist()
# def att_od():
#     frappe.db.set_value("Attendance",'HR-ATT-2025-01291','attendance_request','') 

@frappe.whitelist()
def validate_timesheet(doc,method):
    if doc.department=="IT. Development - THIS" and doc.total_hours <6 and not doc.custom_or_remarks:
        frappe.throw("Your Total hours is less than 6 hours.Kindly fill the OR Remarks")
    if doc.timesheet_summary:
        for i in doc.timesheet_summary:
            if i.status=="Working" and not i.remarks:
                frappe.throw(
                        f"Row #{i.idx}:Kindly fill the Working Remarks."
                    )

@frappe.whitelist()
def update_working_remarks(doc,method):
    if doc.timesheet_summary:
        for i in doc.timesheet_summary:
            if i.status=="Working" and i.remarks:
                frappe.db.set_value("Task",i.id,"custom_remarks",i.remarks)
# import frappe

# def validate_et_vs_at(task):

#     if not task.employee:
#         frappe.throw("No employee is linked to this task. Please set an employee.")

   
#     timesheets = frappe.get_all("Timesheet Detail",
#         filters={"task": task.name},
#         fields=["parent", "hours"]
#     )

#     if not timesheets:
#         return  

    
#     total_actual_time = sum(ts["hours"] for ts in timesheets)

    
#     if total_actual_time > task.expected_time and not task.et_vs_at_remarks:
#         frappe.throw(f"ET vs AT Remarks is mandatory when Actual Time ({total_actual_time} hours) exceeds Estimated Time ({task.expected_time} hours).")

# @frappe.whitelist()
# def validate(doc, method):
#     validate_et_vs_at(doc)


# import frappe
# @frappe.whitelist()
# def et_vs_at(task, employee):
#     if not task or not employee:
#         frappe.throw("Task or Employee is missing.")

#     # Fetch actual time from timesheets
#     timesheet_data = frappe.db.sql("""
#         SELECT SUM(td.hours) AS total_hours
#         FROM `tabTimesheet Detail` td
#         JOIN `tabTimesheet` ti ON td.parent = ti.name
#         WHERE td.task = %s AND ti.employee = %s
#     """, (task, employee), as_dict=True)

#     actual_time = float(timesheet_data[0].total_hours) if timesheet_data and timesheet_data[0].total_hours else 0.0

#     # Fetch estimated time and ensure it's a float
#     estimated_time = frappe.db.get_value("Task", task, "expected_time") or 0.0
#     estimated_time = float(estimated_time)  # Convert to float if needed

#     # Fetch task document
#     task_doc = frappe.get_doc("Task", task)

#     # Check if ET vs AT remarks are required
#     if actual_time > estimated_time and not task_doc.custom_et_vs_at_remark:
#         frappe.throw(f"ET vs AT Remarks is mandatory when Actual Time ({actual_time} hours) exceeds Estimated Time ({estimated_time} hours).")

#     return {
#         "task": task,
#         "employee": employee,
#         "estimated_time": estimated_time,
#         "actual_time": actual_time
#     }

import frappe

import frappe

def validate_et_vs_at(doc, method):
    if doc.service=="IT-SW":
        employee = frappe.get_value("Employee", {"user_id": frappe.session.user}, "name")

        if not employee:
            frappe.throw(f"Employee record for the current user ({frappe.session.user}) not found.")

        timesheet_data = frappe.db.sql("""
            SELECT SUM(td.hours) AS total_hours
            FROM `tabTimesheet Detail` td
            INNER JOIN `tabTimesheet` ti ON td.parent = ti.name
            WHERE td.task = %s AND ti.employee = %s
        """, (doc.name, employee), as_dict=True)

        actual_time = timesheet_data[0].total_hours if timesheet_data and timesheet_data[0].total_hours else 0.0
        estimated_time = frappe.db.get_value(
        "Task", 
        {"name": doc.name, "status": "Working"},  
        "expected_time"
        ) or 0.0

        if not isinstance(actual_time, (int, float)) or not isinstance(estimated_time, (int, float)):
            frappe.throw("Invalid time values detected.")

        if actual_time > estimated_time and not doc.custom_et_vs_at_remark:
            frappe.throw(f"ET vs AT Remark field is mandatory when Actual Time ({actual_time} hours) exceeds Estimated Time ({estimated_time} hours).")

@frappe.whitelist()
def get_per_billed_in_so():
    so = frappe.get_doc("Sales Order",'SAL-ORD-2021-00004')
    so.per_billed = 100
    so.save(ignore_permissions=True)
    

@frappe.whitelist()
def update_cost_center(doc, method):
    for row in doc.accounts:
        row.cost_center = doc.custom_cost_center


# import frappe
# import pypdf
# from frappe.utils.pdf import get_pdf
# from frappe.utils.file_manager import get_file_path
# from frappe import _
# from io import BytesIO
# from pypdf import PdfReader, PdfMerger
# from PIL import Image
# import os

# def is_valid_pdf(file_path):
#     """Check if a file is a valid PDF."""
#     try:
#         with open(file_path, "rb") as f:
#             PdfReader(f, strict=False)  # Try loading the PDF
#         return True
#     except Exception:
#         return False

# def convert_image_to_pdf(image_path):
#     """Convert an image (JPG, PNG) to PDF."""
#     img = Image.open(image_path)
#     pdf_path = f"/tmp/{os.path.basename(image_path)}.pdf"
    
#     img = img.convert("RGB")  # Ensure it's in RGB mode
#     img.save(pdf_path, "PDF")
    
#     return pdf_path

# def remove_blank_pages(pdf_reader):
#     """Remove blank pages from a PdfReader object."""
#     valid_pages = []
#     for page in pdf_reader.pages:
#         text = page.extract_text()
#         if text and not text.isspace():  # Keep only pages with actual content
#             valid_pages.append(page)
#     return valid_pages

# @frappe.whitelist()
# def merge_and_download_pdf(docname):
#     doc = frappe.get_doc("Candidate", docname)

#     print_format_pdf = get_pdf(frappe.get_print("Candidate", docname, "Interview form Masked"))
#     print_pdf_file = BytesIO(print_format_pdf)  # Convert to file-like object
#     print_pdf_file.seek(0)  # Reset cursor

#     reader1 = PdfReader(print_pdf_file, strict=False)
#     pages1 = remove_blank_pages(reader1)
#     file_doc = frappe.db.get_value("File", 
#                                    {"attached_to_doctype": "Candidate", "attached_to_name": docname}, 
#                                    ["file_url"])
    
#     if not file_doc:
#         frappe.throw(_("No attached file found!"))

#     file_path = get_file_path(file_doc)

#     if is_valid_pdf(file_path):
#         pdf_to_merge = file_path  # It's already a PDF
#     else:
#         try:
#             pdf_to_merge = convert_image_to_pdf(file_path)  # Convert image to PDF
#         except Exception as e:
#             frappe.throw(_("Error converting image to PDF: {0}").format(str(e)))

#     reader2 = PdfReader(pdf_to_merge, strict=False)
#     pages2 = remove_blank_pages(reader2)

#     merger = PdfMerger()
#     print_pdf_path = f"/tmp/clean_print_{docname}.pdf"
#     with open(print_pdf_path, "wb") as temp_pdf:
#         writer = pypdf.PdfWriter()
#         for page in pages1:
#             writer.add_page(page)
#         writer.write(temp_pdf)
#     clean_attached_pdf_path = f"/tmp/clean_attached_{docname}.pdf"
#     with open(clean_attached_pdf_path, "wb") as temp_pdf:
#         writer = pypdf.PdfWriter()
#         for page in pages2:
#             writer.add_page(page)
#         writer.write(temp_pdf)
#     merger.append(print_pdf_path)
#     merger.append(clean_attached_pdf_path)
#     merged_pdf_path = f"/tmp/merged_{docname}.pdf"
#     with open(merged_pdf_path, "wb") as output_file:
#         merger.write(output_file)

#     merger.close()
#     with open(merged_pdf_path, "rb") as file:
#         frappe.local.response.filename = f"Merged_{docname}.pdf"
#         frappe.local.response.filecontent = file.read()
#         frappe.local.response.type = "download"

import frappe
import pypdf
from frappe.utils.pdf import get_pdf
from frappe.utils.file_manager import get_file
from frappe import _
from io import BytesIO
from pypdf import PdfReader, PdfMerger
from PIL import Image
import os
import requests

def is_valid_pdf(file_path):
    """Check if a file is a valid PDF."""
    try:
        with open(file_path, "rb") as f:
            PdfReader(f, strict=False)
        return True
    except Exception:
        return False


def convert_image_to_pdf(image_path):
    """Convert an image (JPG, PNG) to a proper A4-sized PDF."""
    img = Image.open(image_path)
    pdf_path = f"/tmp/{os.path.basename(image_path)}.pdf"

    # Convert to RGB and resize to A4 dimensions
    img = img.convert("RGB")
    a4_width, a4_height = (595, 842)  # A4 size in points

    img.thumbnail((a4_width, a4_height))  # Resize while maintaining aspect ratio
    new_img = Image.new("RGB", (a4_width, a4_height), (255, 255, 255))  # White background
    new_img.paste(img, ((a4_width - img.width) // 2, (a4_height - img.height) // 2))

    # Save as PDF
    new_img.save(pdf_path, "PDF", quality=75)
    return pdf_path

def compress_pdf_with_fitz(input_path, output_path, quality=60):
    """Compress PDF using PyMuPDF."""
    doc = fitz.open(input_path)
    doc.save(output_path, garbage=4, deflate=True, clean=True)
    doc.close()


def remove_blank_pages(pdf_reader):
    """Remove blank pages from a PdfReader object."""
    valid_pages = []
    for page in pdf_reader.pages:
        text = page.extract_text()
        if text and not text.isspace():
            valid_pages.append(page)
    return valid_pages

def download_external_file(file_url):
    """Download file from external storage (dfp_external_storage, S3, etc.)."""
    if not file_url.startswith("http"):
        file_url = frappe.utils.get_url() + file_url  # Convert relative URL to full URL
    
    response = requests.get(file_url, stream=True)
    if response.status_code != 200:
        frappe.throw(_("Failed to download file from external storage."))

    file_path = f"/tmp/{os.path.basename(file_url)}"
    with open(file_path, "wb") as f:
        f.write(response.content)

    return file_path

import fitz  # PyMuPDF

def add_logo_to_pdf_with_fitz(input_pdf_path, output_pdf_path, logo_path):
    doc = fitz.open(input_pdf_path)
    logo = fitz.open(logo_path)
    
    logo_width = 100
    logo_height = 70  # Slightly smaller height
    top_margin = 1    # Closer to the top

    for page in doc:
        page_width = page.rect.width
        logo_rect = fitz.Rect(
            page_width - logo_width - 30,
            top_margin,
            page_width - 30,
            top_margin + logo_height
        )
        page.insert_image(logo_rect, filename=logo_path)

    doc.save(output_pdf_path)
    doc.close()



@frappe.whitelist()
def merge_and_download_pdf(docname):
    doc = frappe.get_doc("Candidate", docname)

    # Generate the masked interview form PDF
    print_format_pdf = get_pdf(frappe.get_print("Candidate", docname, "Interview form Masked"))
    print_pdf_file = BytesIO(print_format_pdf)
    print_pdf_file.seek(0)

    reader1 = PdfReader(print_pdf_file, strict=False)
    pages1 = remove_blank_pages(reader1)

    # Fetch the attached file URL
    file_doc = doc.updated__masked_cv
    # file_doc = frappe.db.get_value("File", {"attached_to_doctype": "Candidate", "attached_to_name": docname}, "file_url")
    
    if not file_doc:
        frappe.throw(_("No attached file found!"))

    # Determine if the file is stored externally and download it
    try:
        file_path = download_external_file(file_doc)
    except Exception as e:
        frappe.throw(_("Error retrieving the file: {0}").format(str(e)))

    # Check if it's a valid PDF; otherwise, convert it
    if is_valid_pdf(file_path):
        pdf_to_merge = file_path
    else:
        try:
            pdf_to_merge = convert_image_to_pdf(file_path)
        except Exception as e:
            frappe.throw(_("Error converting image to PDF: {0}").format(str(e)))

    reader2 = PdfReader(pdf_to_merge, strict=False)
    pages2 = remove_blank_pages(reader2)
    if not pages2:  # Ensure images are included
        pages2 = [reader2.pages[i] for i in range(len(reader2.pages))]

    # Merge PDFs
    merger = PdfMerger()
    print_pdf_path = f"/tmp/clean_print_{docname}.pdf"
    with open(print_pdf_path, "wb") as temp_pdf:
        writer = pypdf.PdfWriter()
        for page in pages1:
            writer.add_page(page)
        writer.write(temp_pdf)

    clean_attached_pdf_path = f"/tmp/clean_attached_{docname}.pdf"
    with open(clean_attached_pdf_path, "wb") as temp_pdf:
        writer = pypdf.PdfWriter()
        for page in pages2:
            writer.add_page(page)
        writer.write(temp_pdf)

    merger.append(print_pdf_path)
    merger.append(clean_attached_pdf_path)
    merged_pdf_path = f"/tmp/merged_{docname}.pdf"
    with open(merged_pdf_path, "wb") as output_file:
        merger.write(output_file)
    
    merger.close()
    # Add logo to each page
    file_doc = frappe.get_doc("File", {"file_name": "27a5425e7c98ba2cb891TEAMPRO STROKE 3.png"})
    file_url = file_doc.file_url

    # Converts file_url (e.g., /private/files/logo.png) to full file system path
    logo_path = frappe.get_site_path(file_url.strip("/"))
    # logo_path = frappe.get_site_path("private", "files", "27a5425e7c98ba2cb891TEAMPRO STROKE 3.png")
    final_pdf_path = f"/tmp/final_with_logo_{docname}.pdf"
    add_logo_to_pdf_with_fitz(merged_pdf_path, final_pdf_path, logo_path)
    compressed_pdf_path = f"/tmp/compressed_final_{docname}.pdf"
    compress_pdf_with_fitz(final_pdf_path, compressed_pdf_path)
    # Send final PDF with logo to user
    # with open(final_pdf_path, "rb") as file:
    # 	frappe.local.response.filename = f"Merged_{docname}.pdf"
    # 	frappe.local.response.filecontent = file.read()
    # 	frappe.local.response.type = "download"
    with open(compressed_pdf_path, "rb") as file:
        frappe.local.response.filename = f"Merged_{docname}.pdf"
        frappe.local.response.filecontent = file.read()
        frappe.local.response.type = "download"



@frappe.whitelist()
def project_send_mail_to_creation_adv_button(name):
    posting_date = datetime.now().strftime("%d-%m-%Y")
    pro = frappe.get_doc("Project",name)
    tasks = frappe.get_all("Task",{'project': pro.name},['*'])
    t=frappe.db.get_all("Task",{'project': pro.name},['food'],group_by='food')
    food_count=len(t)
    qualification=frappe.get_all("Task",{'project': pro.name},['qualification_type'],group_by='qualification_type')
    qual_count=len(qualification)
    experience=frappe.get_all("Task",{'project': pro.name},['total_experience'],group_by='total_experience')
    exp_count=len(experience)
    g_experience=frappe.get_all("Task",{'project': pro.name},['gulf_experience'],group_by='gulf_experience')
    g_exp=len(g_experience)
    interview=frappe.get_all("Task",{'project': pro.name},['mode_of_interview'],group_by='mode_of_interview')
    int_count=len(interview)
    acc=frappe.get_all("Task",{'project': pro.name},['accommodation'],group_by='accommodation')
    a_count=len(acc)
    transport=frappe.get_all("Task",{'project': pro.name},['transportation'],group_by='transportation')
    trans_count=len(transport)
    visa=frappe.get_all("Task",{'project': pro.name},['visa_type'],group_by='visa_type')
    v_count=len(visa)
    con=frappe.get_all("Task",{'project': pro.name},['contract_period_year'],group_by='contract_period_year')
    con_count=len(con)
    categorys=frappe.get_all("Task",{'project': pro.name},['category'],group_by='category')
    ca_count=len(categorys)
    keys=frappe.get_all("Task",{'project': pro.name},['custom_major_key_skills'],group_by='custom_major_key_skills')
    key_count=len(keys)
    rec=frappe.get_all("Task",{'project': pro.name},['custom_free_recruitment'],group_by='custom_free_recruitment')
    rec_count=len(rec)
    task_count=(frappe.db.count("Task",{'project': pro.name}))
    serial_no = 1
    table = '<table text-align="center" border="1" width="100%" style="border-collapse: collapse;text-align: center;">'
    table += '<tr style="background-color: #87CEFA"><td style="width: 10%; font-weight: bold; text-align: center;">S.NO</td><td style="width: 30%; font-weight: bold; text-align: center;">Title</td><td style="width: 60%; font-weight: bold; text-align: center;">Details</td></tr>'
    table += """<tr><td>1</td><td>Project ID</td><td>{}</td></tr>""".format(pro.name or '')
    table += """<tr><td>2</td><td>Date</td><td>{}</td></tr>""".format(pro.custom_actionconfirmed_datetime or '')
    table += """<tr><td>3</td><td>Country</td><td>{}</td></tr>""".format(pro.territory or '')
    table += """<tr><td>4</td><td>Client</td><td>{}</td></tr>""".format(pro.customer or '')
    for i in tasks:
        if tasks.index(i)==0:
            table += """<tr><td rowspan={}>5</td><td rowspan={}>Positions</td><td>{}</td></tr>""".format(task_count,task_count,i.subject or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(i.subject or '')
    for k in keys:
        if keys.index(k)==0:
            table += """<tr><td rowspan={}>6</td><td rowspan={}>Major Key Skills</td><td>{}</td></tr>""".format(key_count,key_count,k.custom_major_key_skills or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(k.custom_major_key_skills or '')
    for d in qualification:
        if qualification.index(d)==0:
            table += """<tr><td rowspan={}>7</td><td rowspan={}>Qualification</td><td>{}</td></tr>""".format(qual_count,qual_count,d.qualification_type or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(d.qualification_type or '')
    for e in experience:
        if experience.index(e)==0:
            table += """<tr><td rowspan={}>8</td><td rowspan={}>Experience</td><td>{}</td></tr>""".format(exp_count,exp_count,e.total_experience or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(e.total_experience or '')
    for g in g_experience:
        if g_experience.index(g)==0:
            table += """<tr><td rowspan={}>9</td><td rowspan={}>GCC Experience</td><td>{}</td></tr>""".format(g_exp,g_exp,g.gulf_experience or '')
        else:
            table +="""<tr><td>{}</td></tr>""".format(g.gulf_experience or '')
    for r in rec:
        if rec.index(r)==0:
            table += """<tr><td rowspan={}>10</td><td rowspan={}>Free Recruitment</td><td>{}</td></tr>""".format(rec_count,rec_count,r.custom_free_recruitment or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(r.custom_free_recruitment or '')
    for m in interview:
        if interview.index(m)==0:
            table += """<tr><td rowspan={}>11</td><td rowspan={}>Mode Of Interview</td><td>{}</td></tr>""".format(int_count,int_count,m.mode_of_interview or '')
        else:
            table += """<tr><td>{}</td></tr>""".format(m.mode_of_interview or '')
    table += """<tr><td>12</td><td>If Direct Client Interview - Location & Date</td><td></td></tr>"""
    table += """<tr><td>13</td><td>Contact Number</td><td>+91 75502 24400,+91 73050 56204</td></tr>"""
    table += """<tr><td>14</td><td>Mail ID</td><td>aruna.g@groupteampro.com</td></tr>"""
    for v in visa:
            if visa.index(v)==0:
                table += """<tr><td rowspan={}>15</td><td rowspan={}>Visa Type</td><td>{}</td></tr>""".format(v_count,v_count,v.visa_type or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(v.visa_type or '')
    for cons in con:
            if con.index(cons)==0:
                table += """<tr><td rowspan={}>16</td><td rowspan={}>Contract</td><td>{}</td></tr>""".format(con_count,con_count,cons.contract_period_year or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cons.contract_period_year or '')
    for cat in categorys:
            if categorys.index(cat)==0:
                table += """<tr><td rowspan={}>17</td><td rowspan={}>ECR/ECNR</td><td>{}</td></tr>""".format(ca_count,ca_count,cat.category or '')
            else:
                table += """<tr><td>{}</td></tr>""".format(cat.category or '')
    table += """<tr><td>18</td><td>Special Remarks</td><td>Attractive Salary</td></tr>"""
    table += """<tr><td>19</td><td>Common</td><td>Company Name + Logo + RA Licence + Location + Website + Common Number (7305056202) + Common Mail ID</td></tr>"""
    table += '</table>'
    subject = "Advertisement Details - %s" %posting_date
    message = """
    Dear Sir/Madam,<br><br>
    Kindly find the below Advertisement Confirmed Details:<br><br>{}<br><br>
    Thanks & Regards,<br>TEAM ERP<br>
    <i>This email has been automatically generated. Please do not reply</i>
    """.format(table)
    frappe.sendmail(
        # recipients=["divya.p@groupteampro.com"],
        recipients=["dineshbabu.k@groupteampro.com","dm@groupteampro.com","annie.m@groupteampro.com"],
        subject=subject,
        message=message
    )


def ptsr_report():
    from datetime import datetime
    posting_date = datetime.now().strftime("%d-%m-%Y")
    s_no = 1
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += f'<tr><td colspan="17" style="text-align:center; font-weight:bold;">REC : Project - Task Status Report - {posting_date}</td></tr>'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; font-weight:bold; color:white;">S.NO</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Customer / Project Name</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Project Priority</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">AM Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">PM Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Spoc Remark</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Expected Value</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Expected PSL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Sourcing Status</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">Task</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border-right-style: hidden;">Task Priority</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;border-left-style: hidden;">#VAC</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#SP</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#FP</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#SL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#PSL</td>' \
            '<td style="text-align:center; font-weight:bold; color:white;">#LP</td>' \
            '</tr>'
    
    cust = frappe.db.sql("""SELECT * FROM `tabCustomer` WHERE `disabled` = 0 AND service IN ('REC-I','REC-D') ORDER BY `customer_name` ASC""", as_dict=True)
    
    ev_total = 0
    grand_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}

    for c in cust:
        pname = frappe.get_all("Project", {"status": ("in", ['Open', 'Enquiry']), "customer": c['name'], "service": ("in", ['REC-I', 'REC-D'])}, ['*'], order_by="priority ASC")
        task_totals = {'vac': 0, 'sp': 0, 'fp': 0, 'sl': 0, 'psl': 0, 'custom_lp': 0}

        for p in pname:
            try:
                ev_total += float(p.get('expected_value', 0) or 0)
            except ValueError:
                frappe.log_error(f"Expected value is not a number for project {p.get('project_name')}", "Data Error in PSR Report")
            
            taskid = frappe.get_all("Task", {"status": ("in", ('Working', 'Open', 'Overdue', 'Pending Review')), "project": p.name}, ['*'], order_by="priority ASC")
            
            # Project row
            data += f'<tr style="background-color: #98d7f5;">' \
                    f'<td style="text-align:center;">{s_no}</td>' \
                    f'<td colspan=4 style="text-align:left;">{c.name} - {p.territory}</td>' \
                    f'<td colspan=13></td>' \
                    '</tr>'
            
            # Each task row within a project
            for t in taskid:
                data += f'<tr>' \
                        f'<td style="text-align:center;"></td>' \
                        f'<td style="text-align:left;">{p.project_name}</td>' \
                        f'<td style="text-align:center;">{p.priority}</td>' \
                        f'<td style="text-align:left;">{p.remark}</td>' \
                        f'<td style="text-align:left;">{p.account_manager_remark}</td>' \
                        f'<td style="text-align:left;">{p.custom_spoc_remark}</td>' \
                        f'<td style="text-align:center;">{p.expected_value}</td>' \
                        f'<td style="text-align:center;">{p.expected_psl}</td>' \
                        f'<td style="text-align:center;">{p.sourcing_statu}</td>' \
                        f'<td style="text-align:left;">{t.subject}</td>' \
                        f'<td style="text-align:center;border-right-style: hidden;">{t.priority}</td>' \
                        f'<td style="text-align:center;border-left-style: hidden;">{t.vac}</td>' \
                        f'<td style="text-align:center;">{t.sp}</td>' \
                        f'<td style="text-align:center;">{t.fp}</td>' \
                        f'<td style="text-align:center;">{t.sl}</td>' \
                        f'<td style="text-align:center;">{t.pl}</td>' \
                        f'<td style="text-align:center;">{t.custom_lp}</td>' \
                        '</tr>'
                
                # Update task totals
                task_totals['vac'] += t.get('vac', 0)
                task_totals['sp'] += t.get('sp', 0)
                task_totals['fp'] += t.get('fp', 0)
                task_totals['sl'] += t.get('sl', 0)
                task_totals['psl'] += t.get('psl', 0)
                task_totals['custom_lp'] += t.get('custom_lp', 0)
                s_no += 1

    # Grand total row
    data += f'<tr style="background-color: #002060;">' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">Total</td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{ev_total}</td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td></td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["vac"]}</td>' \
            f'<td style="text-align:center; font-weight: bold;color: #ffffff;">{grand_totals["sp"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["fp"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["sl"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["psl"]}</td>' \
            f'<td style="text-align:center; font-weight: bold; color: #ffffff;">{grand_totals["custom_lp"]}</td>' \
            '</tr>'
    
    data += '</table>'
    return data

@frappe.whitelist()
def update_qualification_status(name,qualification_status):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name": name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.qualification_status = qualification_status
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 

@frappe.whitelist()
def update_visit_status_sfp(name,visit_status):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name": name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.visit_status = visit_status
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 
    
@frappe.whitelist()
def update_spf_status(doc,method):
    if doc.lead_name:
        sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name": doc.lead_name}, "name")
        if sfp_name:
            for i in sfp_name:
                sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
                sfp_doc.status = "Converted"
                sfp_doc.party_from="Customer"
                sfp_doc.party_name=doc.name
                sfp_doc.save()  
                sfp_doc.reload() 
                frappe.db.commit() 

@frappe.whitelist()
def add_custom_appointment_details_in_lead(name, lead, visted_date, visted_by, appointment_remarks):
    lead_doc = frappe.get_doc("Lead", lead)

    if lead_doc:
        if lead_doc.visit_status1:  # Ensure it exists
            lead_doc.append("visit_status1", {
                "sales_follow_up": name,
                "visit_date": visted_date,
                "visit_by": visted_by,
                "remarks": appointment_remarks
            })
                
        else:
            # If the child table is empty, add a new entry
            lead_doc.append("visit_status1", {
                "sales_follow_up": name,
                "visit_date": visted_date,
                "visit_by": visted_by,
                "remarks": appointment_remarks
            })

        lead_doc.save()
        frappe.db.commit()  # Not necessary, but keeping it here if needed

@frappe.whitelist()
def update_sfp_opportunity(doc,method):
    if doc.custom_sales_follow_up and doc.custom_quotation and doc.status=="Lost":
        frappe.db.set_value("Sales Follow Up",doc.custom_sales_follow_up,"status","Replied")

@frappe.whitelist()
def update_dnc(name):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name":name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.status = "Do Not Contact"
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 

@frappe.whitelist()
def update_dnc_converted(name):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name":name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.status = "Converted"
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 

@frappe.whitelist()
def update_territory_sfp(name,territory):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name": name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.territory = territory
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 

@frappe.whitelist()
def update_market_segment_sfp(name,market_segment):
    sfp_name = frappe.db.get_all("Sales Follow Up", {"party_name": name}, "name")
    if sfp_name:
        for i in sfp_name:
            sfp_doc = frappe.get_doc("Sales Follow Up", i.name) 
            sfp_doc.market_segment = market_segment
            sfp_doc.save()  
            sfp_doc.reload() 
            frappe.db.commit() 

@frappe.whitelist()
def update_check_existing_lead(doc,method):
    if doc.custom_check_existing:
        # frappe.errprint(doc.custom_check_existing)
        # if frappe.db.exists("Lead",{'name':['!=',doc.name],'custom_existing_lead':doc.custom_existing_lead}):
        #     lead_name=frappe.get_doc("Lead",{'name':['!=',doc.name],'custom_existing_lead':doc.custom_existing_lead})
        #     form_link = get_link_to_form("Lead", lead_name.name)
        #     msg = _("Already another lead found {0}").format(form_link)
        #     frappe.throw(msg)
        
        leads=frappe.get_doc("Existing Leads",doc.custom_check_existing)
        leads.lead_id=doc.name
        leads.save()  
        leads.reload() 
        frappe.db.commit() 

@frappe.whitelist()
def update_company_name(name):
    organiation_name=frappe.db.get_value("Existing Leads",name,"lead_name")
    return organiation_name

@frappe.whitelist()
def update_project_dates(doc, method):
    if doc.custom_sla_details:
        for row in doc.custom_sla_details:
            if row.project and row.sla_from_date and row.sla_to_date:
                frappe.db.set_value("Project", row.project, {
                    "expected_start_date": row.sla_from_date,
                    "expected_end_date": row.sla_to_date
                })
                frappe.db.commit()

@frappe.whitelist()
def employee_chc_print(doc):
    chc = frappe.get_doc("Employee Onboarding", {"employee": doc.name})
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:left; font-weight:bold; color:white;font-size: 18px;width: 5%">S NO</td>' \
            '<td style="text-align:left; font-weight:bold; color:white;font-size: 18px;width: 100%">Task</td>' \
            '<td style="text-align:left; font-weight:bold; color:white;font-size: 18px;width: 100%">Status</td>' \
            '</tr>'
    s_no = 1
    
    if chc:
        for i in chc.custom_employee_chc:
            data += f'<tr>' \
                    f'<td style="text-align:left;font-size: 18px;">{s_no}</td>' \
                    f'<td style="text-align:left;font-size: 18px;">{i.activity_name}</td>' \
                    f'<td style="text-align:left;font-size: 18px;">{i.status}</td>' \
                    '</tr>'
            s_no += 1
    data += '</table>'
    return data

@frappe.whitelist()
def employee_joining_print(doc):
    chc = frappe.get_doc("Employee Onboarding", {"employee": doc.name})
    data = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    data += '<tr style="background-color: #002060; color: white;">' \
            '<td style="text-align:center; width: 5%;background-color: #0f1568 !important; color: white">S.NO</td>' \
            '<td style="text-align:center; width: 50%;background-color: #0f1568 !important; color: white">Task</td>' \
            '<td style="text-align:center; width: 30%;background-color: #0f1568 !important; color: white">Resp.</td>' \
            '<td style="text-align:center; width: 30%;background-color: #0f1568 !important; color: white">Completed ON</td>' \
            '<td style="text-align:center; width: 30%;background-color: #0f1568 !important; color: white">Completed by Name / Sign</td>' \
            '<td style="text-align:center; width: 30%;background-color: #0f1568 !important; color: white">Status</td>' \
            '</tr>'
    s_no = 1
    
    if chc:
        for i in chc.activities:
            data += f'<tr>' \
                    f'<td style="text-align:left;">{s_no}</td>' \
                    f'<td style="text-align:left;">{i.activity_name if i.activity_name else ""}</td>' \
                    f'<td style="text-align:left;">{i.user if i.user else ""}</td>' \
                    f'<td style="text-align:left;">{i.completed_on_date if i.completed_on_date else ""}</td>' \
                    f'<td style="text-align:left;"></td>' \
                    f'<td style="text-align:left;">{i.status if i.status else ""}</td>' \
                    '</tr>'
            s_no += 1
    data += '</table>'
    return data

@frappe.whitelist()
def update_lead_contacts_sfp(doc,method):
    if doc.party_from=="Lead" and doc.party_name:
        lead_contact = frappe.get_doc("Lead", doc.party_name)
        for i in lead_contact.lead_contacts:
            doc.append("contacts", {
                'person_name': i.person_name,
                'mobile': i.mobile,
                'is_primary': i.is_primary,
                'has_whatsapp': i.has_whatsapp,
                'email_id': i.email_id,
                'is_primaryemail': i.is_primaryemail,
                'service':i.service
            })
            doc.append("custom_contact_details", {
                'person_name': i.person_name,
                'mobile': i.mobile,
                'is_primary': i.is_primary,
                'has_whatsapp': i.has_whatsapp,
                'email_id': i.email_id,
                'is_primaryemail': i.is_primaryemail,
                'service':i.service
            })
    elif doc.party_from=="Customer" and doc.party_name:
        lead_contact = frappe.get_doc("Customer", doc.party_name)
        for i in lead_contact.customer_contact:
            doc.append("customer_contacts", {
                "person_name": i.person_name or '',
                "mobile": i.mobile or '',
                "is_primary": i.is_primary or False,
                "has_whatsapp": i.has_whatsapp or False,
                "email_id": i.email_id or '',
                "is_primaryemail": i.is_primaryemail or False,
                "service":i.service or ''
            })
    doc.save()
    frappe.db.commit()  
            
# @frappe.whitelist()
# def update_project_type():
#     frappe.db.set_value("Sales Follow Up","SFP-30292","status","Opportunity")

@frappe.whitelist()
def clear_payment_table_si(doc,method):
    doc.payment_schedule=[]

@frappe.whitelist()
def validate_maintain_stok_si(doc,method):
    if doc.pos_profile not in ["Main Store","VM1_Precision"]:
    # if doc.pos_profile!="Main Store":
        if doc.items:
            for i in doc.items:
                stock=frappe.db.get_value("Item",i.item_code,"is_stock_item")
                if stock and not i.delivery_note:
                    frappe.throw(_("Row {0}:Stock Item '{1}' requires a Delivery Note.Kindly create Invoice from Delivery Note.").format(i.idx, i.item_code))

# @frappe.whitelist()
# def update_status_sfp():
#     frappe.db.set_value("Sales Follow Up","SFP-28768","status","Converted")

@frappe.whitelist()
def update_month_cycle(doc,method):
    from_date_obj = datetime.strptime(doc.start_date, "%Y-%m-%d")
    month_year = from_date_obj.strftime("%b %Y") 
    frappe.db.set_value("Appraisal Cycle",doc.name,"custom_cycle_month",month_year)

@frappe.whitelist()
def update_spf_details_lead(doc,method):
    created_on=now_datetime()
    if doc.party_from=="Lead":
        lead=frappe.get_doc("Lead",doc.party_name)
        lead.append("custom_sfp_details", {"sfp_id": doc.name,"sfp_owner":doc.account_manager_lead_owner,"created_on":created_on,"service":doc.service})
        lead.save()
        frappe.db.commit()  
    if doc.party_from=="Customer":
        customer=frappe.get_doc("Customer",doc.party_name)
        customer.append("custom_sfp_details", {"sfp_id": doc.name,"sfp_owner":doc.account_manager_lead_owner,"created_on":created_on,"service":doc.service})
        customer.save()
        frappe.db.commit() 

# @frappe.whitelist()
# def update_exixting_spf_details_lead():
#     count=0
#     sfp=frappe.db.get_all("Sales Follow Up",{"party_from":"Lead"},["party_name","account_manager_lead_owner","service","name","creation"])
#     for i in sfp:
#         if i.party_name:
#             lead=frappe.get_doc("Lead",i.party_name)
#             # Check if the sfp_id is already in the child table
#             existing_sfp_ids = [row.sfp_id for row in lead.custom_sfp_details]
#             if i.name not in existing_sfp_ids:  # Avoid duplicate entries
#                 lead.append("custom_sfp_details", {
#                     "sfp_id": i.name,
#                     "sfp_owner": i.account_manager_lead_owner,
#                     "created_on": i.creation,
#                     "service": i.service
#                 })
#                 lead.save()
#                 frappe.db.commit()
#                 count += 1
#     print(count)

# @frappe.whitelist()
# def enqueue_sfp():
#     from frappe.utils.background_jobs import enqueue
#     enqueue(method=update_exixting_spf_details_lead, queue="long", timeout=96000)



import frappe
from frappe import _
import requests
import json

# Replace this with your actual Mattermost incoming webhook URL
MATTERMOST_WEBHOOK_URL = "https://pm.teamproit.com/hooks/8rm94z3knfdptf8phf6cmrszpe"


@frappe.whitelist(allow_guest=True)
def create_issue_from_mattermost_new():
    data = frappe.local.form_dict
    user_name = data.get("user_name")
    text = data.get("text")  # whatever comes after the slash command
    channel = data.get("channel_name")
    # Proceed if token is OK
    user_name = frappe.form_dict.get("user_name") or "unknown"
    text = frappe.form_dict.get("text") or "No message"
    project_name = frappe.db.get_value(
        "Project",
        {"name": ["like", f"%{channel}%"]},
        "name"
    )
    issue=frappe.new_doc("Issue")
    issue.subject=f"Issue from Mattermost by {user_name}"
    issue.description=text
    issue.project=project_name
    issue.save()
    frappe.db.commit()
    message = f"Your query has been registered successfully. Please refer to Ticket Number:*{issue.name}* for any future communication.\n> {text}"
    try:
        response = requests.post(
            MATTERMOST_WEBHOOK_URL,
            headers={"Content-Type": "application/json"},
            data=json.dumps({"text": message})
        )
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        frappe.log_error(f"Failed to post to Mattermost: {e}", "Mattermost Webhook")

    return "OK"

@frappe.whitelist()
def send_amil_for_working(subject,id,action_taken,live,et,at,code_review_revision,service,proof,allocated=None,project=None,issue=None,domain=None,spoc=None,reason=None,dev_spoc=None,code_review_comment=None):
    if service=='IT-SW':
        percentage=et_at_calculation(id, et, at, allocated,subject)
        reports=frappe.db.get_value("Employee",{'user_id':allocated},['reports_to'])
        reports_to=frappe.db.get_value("Employee",{'name':reports},['user_id'])
        tl=frappe.db.get_value("Employee",{'user_id':allocated},["custom_tl"])
        tl_mail=frappe.db.get_value("Employee",{'name':tl},['user_id'])
        et_rate= 'ET : %s and AT : %s'%(et,round(percentage,2))
        if issue:
            raised_by=frappe.db.get_value("Issue",{'name':issue},['raised_by'])
        else:
            raised_by='None'
        data = ''
        data += f"<table width='100%' style='border-collapse: collapse; border: 1px solid black; text-align: center;'>\
        <tr><td colspan='2' style='text-align: center; background-color: #0f1568;color: white; font-size: 17px; border: 1px solid black;'><b>Task / Issue Pending Review Note</b></td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Task ID</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/app/task/{id}' target='_blank'>{id}</a></td></tr>\
        <tr style='text-align: left;'><td width='25%'style='border: 1px solid black;'><b>Project</b></td><td style='border: 1px solid black;'>{project}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task Raised By</b></td><td style='border: 1px solid black;'>{reports_to}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue ID</b></td><td style='border: 1px solid black;'>{issue}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Issue Raised By</b></td><td style='border: 1px solid black;'>{raised_by}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Statement</b></td><td style='border: 1px solid black;'>{subject}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Task / Issue Action Taken</b></td><td style='border: 1px solid black;'>{action_taken}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Live At</b></td><td style='border: 1px solid black;'>{live}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Domain</b></td><td style='border: 1px solid black;'>{domain}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Proof</b></td><td style='border: 1px solid black;'><a href='https://erp.teamproit.com/{proof}' target='_blank'>Link to Proof</a></td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>ET & AT</b></td><td style='border: 1px solid black;'>{et_rate}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Code Review to Working Count</b></td><td style='border: 1px solid black;'>{code_review_revision}</td></tr>\
        <tr style='text-align: left;'><td style='border: 1px solid black;'><b>Reason:</b></td><td style='border: 1px solid black;'>{code_review_comment}</td></tr></table>"
        cc = [reports_to, allocated,spoc,'anil.p@groupteampro.com'] + ([dev_spoc] if dev_spoc else []) +([tl_mail] if tl_mail else [])
        frappe.sendmail(
            sender=allocated,
            recipients='divya.p@groupteampro.com',
            # recipients=spoc,
            # cc=cc,
            subject='Task : %s Status Changed to Working from Code Review' % id,
            message = """
            <b>Dear Patron,<br><br>Greeting !!!</b><br><br>
           The referenced task has been moved back to <b>Working</b> from <b>Code Review</b> for further development or changes kindly find the Code Reviewer Comment.<br><br>
            Please find the task details below:<br><br>
           {}<br><br>
            Thanks & Regards,<br>TEAM ERP<br>
            
            <i>This email has been automatically generated. Please do not reply</i>
            """.format(data)
        )

@frappe.whitelist()
def update_project_issue(doc,method):
    if not doc.project:
        frappe.db.set_value("Issue",doc.name,"project","Internal ERP - TEAMPRO V15")

# @frappe.whitelist()
# def update_status_lead_today():
#     filename='9940a8111b2bae3Lead.csv'
#     from frappe.utils.file_manager import get_file
#     filepath = get_file(filename)
#     pps = read_csv_content(filepath[1])
#     ind=0
#     for pp in pps:
#         if pp[0]!="ID":
#             frappe.db.set_value("Lead",{"name":pp[0]},"lead_owner",pp[1])
#             ind+=1
#             print(pp[0])
#             print(pp[1])
#     print(ind)


# method to update the criteria table during the task creation
@frappe.whitelist()
def update_criteria_table(doc, method):
    pass
    # if doc.service == 'REC-I':
    # 	proj = frappe.get_doc("Project", doc.project)        
    # 	doc.set("custom_criteria_table", [])
    # 	for row in proj.custom_criteria_table:
    # 		doc.append("custom_criteria_table", {
    # 			"scheduling_criteria": row.scheduling_criteria,
    # 			"scheduling_parameter": row.scheduling_parameter
    # 		}) 
    # 	doc.save()


import frappe
import requests
import json
from datetime import datetime

@frappe.whitelist()
def update_issue_wonjin(doc,method):
    if not doc.is_new():
        doc_task = frappe.get_doc("Task",doc.name)
        if doc.project == 'Wonjin_ERP_19.12.2023' and not doc.is_new():
            subject = f"{doc.subject} - {doc_task.creation.strftime('%d-%m-%Y')}"
            creation = doc_task.creation.strftime('%d-%m-%Y')

            params = {
                'creation': creation,
                'name': doc.name,
                'subject': subject,
                'description': doc.description,
                'priority': doc.priority,
                'status': doc.status,
                'pr_remarks': doc.custom_taskissue_action_taken,
                'proof': doc.custom_proof_of_closure_review,
                'issue_id': doc.issue,
                'allocated_to': doc.custom_allocated_to,

            }

            url = "https://erp.onegeneindia.in/api/method/onegene.www.update_issue.update_issue_from_teampro"
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'token 7503af112f2692c:812bd60c48b22ed'
            }


            try:
                response = requests.post(url, headers=headers, json=params, verify=False)
                response.raise_for_status()  # raises exception for 4xx/5xx errors

                res = response.json()
                return res

            except requests.exceptions.RequestException as e:
                frappe.throw(f"HTTP error: {str(e)}")
            except json.JSONDecodeError:
                frappe.throw("Failed to decode JSON response from server")

        return "No matching task found or it's new"

@frappe.whitelist()
def create_issue_wonjin(doc,method):
    doc_task = frappe.get_doc("Task",doc.name)
    if doc.project == 'Wonjin_ERP_19.12.2023':
        subject = f"{doc.subject} - {doc_task.creation.strftime('%d-%m-%Y')}"
        creation = doc_task.creation.strftime('%d-%m-%Y')

        params = {
            'creation': creation,
            'name': doc.name,
            'subject': subject,
            'description': doc.description,
            'priority': doc.priority,
            'status': doc.status,
            'pr_remarks': doc.custom_taskissue_action_taken,
            'proof': doc.custom_proof_of_closure_review,
            'issue_id': doc.issue,
        }

        url = "https://erp.onegeneindia.in/api/method/onegene.www.update_issue.update_issue_from_teampro"
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'token 7503af112f2692c:812bd60c48b22ed'
        }


        try:
            response = requests.post(url, headers=headers, json=params, verify=False)
            response.raise_for_status()  # raises exception for 4xx/5xx errors

            res = response.json()
            return res

        except requests.exceptions.RequestException as e:
            frappe.throw(f"HTTP error: {str(e)}")
        except json.JSONDecodeError:
            frappe.throw("Failed to decode JSON response from server")

    return "No matching task found or it's new"


@frappe.whitelist()
def update_issueid_wonjin(doc,method):
    # doc = frappe.get_doc("Issue",doc.name)
    if doc.raised_by == 'wonjin_corporate@onegeneindia.in':
        params = {
            'name': doc.name,
            'subject': doc.subject,
            'status': doc.status,
        }

        url = "https://erp.onegeneindia.in/api/method/onegene.www.update_issue.update_issueid_from_teampro"
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'token 7503af112f2692c:812bd60c48b22ed'
        }


        try:
            response = requests.post(url, headers=headers, json=params, verify=False)
            response.raise_for_status()  # raises exception for 4xx/5xx errors

            res = response.json()
            return res

        except requests.exceptions.RequestException as e:
            frappe.throw(f"HTTP error: {str(e)}")
        except json.JSONDecodeError:
            frappe.throw("Failed to decode JSON response from server")

    return "No matching task found or it's new"

@frappe.whitelist()
def update_criteria_changes(name):
    project = frappe.get_doc("Project", name)
    # if not project.custom_criteria_table:
    # 	return
    tasks = frappe.get_all("Task", {"project": name}, ["name"])
    for i in tasks:
        frappe.errprint(i.name)
        task = frappe.get_doc("Task", i.name)
        for row in project.custom_criteria_table:
            frappe.errprint('name')
            if not row.updated:
                frappe.errprint('name1')
                task.append("custom_criteria_table", {
                    "scheduling_criteria": row.scheduling_criteria,  
                    "scheduling_parameter": row.scheduling_parameter         
                })
        task.save()
    return "Ok"


# @frappe.whitelist()
# def get_service():
#     count =0
#     project = frappe.get_all('Project',filters={"service":"IT-SW"},fields=["name","customer"])
#     for i in project:
#         # i.customer
#         # i.name
#         task = frappe.get_all('Task',filters={'project':i.name,"service":"IT-SW",'customer':["!=",i.customer],'creation': ['between', ['2025-04-20', '2025-05-09']]},fields=['name','creation'])
#         for j in task:
#             count += 1
#             if(i.customer != j.customer):
#                 # if j.name == 'TS10119':
#                 print(j.name)
#                 # frappe.db.set_value('Task',j.name,'customer',i.customer)
#     print(count)

@frappe.whitelist()
def merge_material_request_items(material_request):
    childtab = frappe.db.sql(""" select `tabMaterial Request Item`.item_code,`tabMaterial Request Item`.warehouse,
    `tabMaterial Request Item`.item_name,`tabMaterial Request Item`.conversion_factor,
    sum(`tabMaterial Request Item`.qty) as qty, sum(`tabMaterial Request Item`.stock_qty) as stock_qty, `tabMaterial Request Item`.actual_qty as actual_stock_qty,
    `tabMaterial Request Item`.stock_uom,
    `tabMaterial Request Item`.uom,`tabMaterial Request Item`.warehouse,
    `tabMaterial Request Item`.from_warehouse,`tabMaterial Request Item`.stb
    from `tabMaterial Request` 
    left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent where `tabMaterial Request`.name = '%s' group by `tabMaterial Request Item`.item_code order by `tabMaterial Request Item`.idx """%(material_request),as_dict = 1)
    return childtab

@frappe.whitelist()
def update_ordered_qty(doc,method):
    material_request = frappe.db.get_value("Purchase Order Item", {"parent": doc.name}, "material_request")
    if frappe.db.exists("Material Request", material_request):
        for i in doc.items:
            mr_name = frappe.db.get_value("Material Request Clubbed Item",{'parent':i.material_request,'item_code':i.item_code},['name'])
            if mr_name:
                frappe.db.set_value("Material Request Clubbed Item",mr_name,'po_qty',i.qty)

@frappe.whitelist()
def update_ordered_qty_on_cancel(doc,method):
    material_request = frappe.db.get_value("Purchase Order Item", {"parent": doc.name}, "material_request")
    if frappe.db.exists("Material Request", material_request):
        for i in doc.items:
            mr_name = frappe.db.get_value("Material Request Clubbed Item",{'parent':i.material_request,'item_code':i.item_code},['name'])
            if mr_name:
                po_qty = frappe.db.get_value("Material Request Clubbed Item",{'parent':i.material_request,'item_code':i.item_code},['po_qty'])
                if po_qty and po_qty > 0:
                    frappe.db.set_value("Material Request Clubbed Item",mr_name,'po_qty',po_qty - i.qty)
        doc = frappe.get_doc("Material Request", material_request)
        row_count = 0
        conditions_satisfied_count = 0
        for row in doc.custom_merged_items:
            row_count += 1
            if row.po_qty >= row.purchase_qty:
                conditions_satisfied_count += 1
        if row_count > 0 and  row_count == conditions_satisfied_count:
            frappe.db.set_value("Material Request", material_request, "status", "Pending")

@frappe.whitelist()
def update_material_request_status_on_submit(doc,method):
    material_request = frappe.db.get_value("Purchase Order Item", {"parent": doc.name}, "material_request")
    if frappe.db.exists("Material Request", material_request):
        doc = frappe.get_doc("Material Request", material_request)
        row_count = 0
        conditions_satisfied_count = 0
        for row in doc.custom_merged_items:
            row_count += 1
            if row.po_qty >= row.purchase_qty:
                conditions_satisfied_count += 1
        if row_count > 0 and row_count == conditions_satisfied_count:
            frappe.db.set_value("Material Request", material_request, "status", "Ordered")
        
        
import frappe
from frappe.utils import add_days

@frappe.whitelist()
def get_previous_count(name,date):
    prev_date=add_days(date,-1)
    result = frappe.db.sql("""
        SELECT SUM(sd.count)
        FROM `tabStock Counting` sc
        JOIN `tabStock Counting Details` sd ON sc.name = sd.parent
        WHERE sc.date = %s AND sd.item = %s AND DATE(sd.date_and_time)=%s
    """, (prev_date, name,prev_date))

    return result[0][0] or 0

@frappe.whitelist()
def calc_cost_prize_po(docname):
    doc = frappe.get_doc("Purchase Order", docname)  # change doctype if needed
    warnings = []
    for f in doc.items:
        tfp_item = frappe.db.get_value("Item", f.item_code, "tfp")
        if tfp_item == 1:
            price = frappe.db.get_value("Item Price", {
                "price_list": "Cost Price TFP",
                "item_code": f.item_code
            }, "price_list_rate")
            if price:
                item_price = price / 1000 if f.uom == "Gram" else price
                item_rate = round(item_price, 2)
                if f.rate > item_rate:
                    warnings.append(f"{f.item_name}")

    return {"warnings": warnings}

@frappe.whitelist()
def validate_stock_counting(doc,method):
    if frappe.db.exists("Stock Counting",{"date":doc.date,"name": ["!=", doc.name]}):
        frappe.throw("Document already created for this date")
    item_codes = []
    if doc.details:
        for row in doc.details:
            if row.item in item_codes:
                frappe.throw(f"Item '{row.item}' has already been scanned. Duplicate entries are not allowed.")
            item_codes.append(row.item)

@frappe.whitelist()
def get_tfp_item():
    items_with_tfp = frappe.get_all(
        "Item",
        filters={
            "disabled": 0,
            "item_group": ["in", ["Savouries", "Food Products", "Bakery"]],
        },
        fields=["name", "item_name", "item_code"]
    )


    html = """
    <html>
    <head>
      <meta name="pdfkit-orientation" content="Portrait"/>
      <meta name="page-size" content="A4"/>
      <style>
        body {
          font-family: Arial, sans-serif;
          margin: 0;
          padding: 0;
        }
        .page {
          display: flex;
          flex-wrap: wrap;
          width: 100%;
          page-break-after: always;
        }
        .item-block {
  border: 1px solid black;
  box-sizing: border-box;
  margin: 2px;
  padding: 30px 10px 10px 20px; /* 30px top padding */
  width: 48%;
  height: 500px;
  display: flex;
  flex-direction: column;
  justify-content: flex-start;
  align-items: center;
}

        .item-name {
        font-size: 30px;
        font-weight: bold;
        text-transform: uppercase;
        text-align: center;
        margin-top: 0;
        whitespace:nowrap;
        }
        .item-code {
          font-size: 19px;
          margin-top: 0px;
        }
        .qr-code {
          margin-top: 60px;
        }
        .qr-code img {
          width: 250px;
          height: 250px;
        }
      </style>
    </head>
    <body>
    """

    for i, item in enumerate(items_with_tfp):
        if i % 4 == 0:
            if i != 0:
                html += "</div>" 
            html += '<div class="page">'

        html += f"""
        <div class="item-block">
          <div class="item-name">{item['item_name']}</div>
          <div class="item-code">({item['item_code']})</div>
          <div class="qr-code">
            <img src="https://api.qrserver.com/v1/create-qr-code/?data={item['name']}" alt="QR Code">
          </div>
        </div>
        """

    html += "</div></body></html>" 

    return html

from frappe.utils import formatdate

import frappe

@frappe.whitelist()
def get_tasks_by_date_and_employee(employee, date):
    emp=frappe.db.get_value("Employee",employee,["user_id"])
    tasks = frappe.get_all(
        "Task",
        filters={
            "custom_production_date": date,
            "custom_allocated_to": emp
        },
        fields=["name", "status", "rt","project","subject"]
    )

    if not tasks:
        return "<div>No tasks found for the selected date and employee.</div>"
    html = """
    <style>
        .task-table, .task-table th, .task-table td {
            border: 1px solid black;
            border-collapse: collapse;
        }
    </style>
    <table class="table table-bordered">
        <thead>
            <tr style = 'background-color:#0f1568;color:white;text-align:center'>
                <th>Task ID</th>
                <th>Project</th>
                <th>Subject</th>
                <th>Status</th>
                <th>RT</th>
            </tr>
        </thead>
        <tbody>
    """
    for task in tasks:
        html += f"""
            <tr>
                <td>{task.name}</td>
                <td>{task.project}</td>
                <td>{task.subject}</td>
                <td>{task.status}</td>
                <td style='text-align:right'>{task.rt or ''}</td>
            </tr>
        """
    html += "</tbody></table>"
    return html

@frappe.whitelist()
def update_app_visit_status(lead,visit):
    doc=frappe.get_doc('Lead',lead)
    doc.visit_status=visit
    doc.save()
    doc.reload()

@frappe.whitelist()
def update_table_in_task(doc,method):
   if doc.service == 'REC-I':
        # doc.reload()
        frappe.errprint("Hello")
        frappe.enqueue(
            update_criteria_to_tasks, 
            queue="long",
            timeout=36000,
            is_async=True, 
            now=False, 
            job_name='Update Tasks',
            enqueue_after_commit=False,
            name=doc.name,
        )
    
@frappe.whitelist()
def update_criteria_to_tasks(name):
    frappe.log_error(title='Task Update', message='Starting update_criteria_to_tasks')

    project = frappe.get_doc("Project", name)
    if not project.custom_criteria_table:
        return
    criteria_list = project.custom_criteria_table
    if not criteria_list:
        return
    tasks = frappe.get_all("Task", {"project": name}, ["name"])
    for t in tasks:
        task = frappe.get_doc("Task", t.name)
        for project_criteria in criteria_list:
            already_exists = any(
                task_criteria.scheduling_criteria == project_criteria.scheduling_criteria and
                task_criteria.scheduling_parameter == project_criteria.scheduling_parameter
                for task_criteria in task.custom_criteria_table
            )
            if not already_exists:
                task.append("custom_criteria_table", {
                    "scheduling_criteria": project_criteria.scheduling_criteria,
                    "scheduling_parameter": project_criteria.scheduling_parameter
                })
                frappe.log_error(title=f'Task Update - {t.name}', message='Criteria appended')
        task.save()
        frappe.db.commit()

    frappe.log_error(title='Task Update', message='Completed update_criteria_to_tasks')

@frappe.whitelist()
def update_existing_cust():
    cust=frappe.db.get_all("Customer",{'customer_name':['!=','']},['customer_name','name'])
    for c in cust:
        exist=frappe.new_doc("Existing Customer")
        exist.customer_name=c.customer_name
        exist.customer_id=c.name
        # exist.id=c.name
        exist.insert()
        exist.save()
        frappe.db.commit()
@frappe.whitelist()
def create_cust(name):
    if not frappe.db.exists("Existing Customer",name):
        cust=frappe.new_doc("Existing Customer")
        cust.customer_id=name
        cust.save()
        frappe.db.commit()

@frappe.whitelist()
def get_dn_packing_details(doc):
    dn = frappe.get_doc("Delivery Note", doc.name)
    so_no = ''
    data = ''

    if dn.items:
        data = '<table border="1" style="border-collapse: collapse; width: 100%; table-layout: auto;">'
        data += '''
        <tr style="background-color: #002060; color: white;">
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">Sr</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">SO ID</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">Item Name</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">QTY</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">UOM</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">St.QTY</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">UOM</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">Cover type</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">MRP</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">Mnfg. On</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;"># Covers</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">2P Type</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;">Per 2P</td>
            <td style="text-align:center; font-weight:bold; border:1px solid black;background-color: #002060; color: white;"># 2P</td>
        </tr>
        '''



        for idx, i in enumerate(dn.items, 1):
            so_id = i.against_sales_order or ''
            so_item_doc = None
            cover_type = mrp = manuf_date = no_of_covers = two_p_type = per_two_p = no_of_two_p = ''

            if i.so_detail:
                so_item_doc = frappe.get_doc("Sales Order Item", i.so_detail)

            if so_item_doc:
                cover_type = so_item_doc.get("custom_cover_type") or ''
                mrp = so_item_doc.get("mrp") or ''
                manuf_date = so_item_doc.get("custom_mfg_on") or ''
                no_of_covers = so_item_doc.get("custom_covers") or ''
                two_p_type = so_item_doc.get("custom_packing_type") or ''
                per_two_p = so_item_doc.get("custom_per_2p") or ''
                no_of_two_p = so_item_doc.get("custom_2nd_packing") or ''
                # if manuf_date:
                #     if isinstance(manuf_date, str):
                #         try:
                #             manuf_date = datetime.datetime.strptime(manuf_date, "%Y-%m-%d").strftime("%d-%m-%Y")
                #         except ValueError:
                #             manuf_date = manuf_date  # Leave as-is if format is unexpected
                #     elif isinstance(manuf_date, datetime.date) or isinstance(manuf_date, datetime.datetime):
                #         manuf_date = manuf_date.strftime("%d-%m-%Y")

            data += f'<tr>' \
                    f'<td style="text-align:left;">{idx}</td>' \
                    f'<td style="text-align:left;">{so_id}</td>' \
                    f'<td style="text-align:left;">{i.item_name}</td>' \
                    f'<td style="text-align:right;">{i.qty}</td>' \
                    f'<td style="text-align:center;">{i.uom}</td>' \
                    f'<td style="text-align:right;">{i.stock_qty}</td>' \
                    f'<td style="text-align:center;">{i.stock_uom}</td>' \
                    f'<td style="text-align:left;">{cover_type}</td>' \
                    f'<td style="text-align:right;">{mrp}</td>' \
                    f'<td style="text-align:center;">{formatdate(manuf_date) if manuf_date else ""}</td>' \
                    f'<td style="text-align:right;">{no_of_covers}</td>' \
                    f'<td style="text-align:left;">{two_p_type}</td>' \
                    f'<td style="text-align:right;">{per_two_p}</td>' \
                    f'<td style="text-align:right;">{no_of_two_p}</td>' \
                    f'</tr>'

        data += '</table>'
    return data

@frappe.whitelist()
def get_so_from_dn(doc):
    dn = frappe.get_doc("Delivery Note", doc.name)
    for item in dn.items:
        if item.against_sales_order:
            return item.against_sales_order
    return ""

# def validate_custom_cover_type(doc, method):
#     if doc.company=="TEAMPRO Food Products" and doc.name=="MAT-DN-2025-00534":
#         if doc.items:
#             for row in doc.items:
#                 if row.against_sales_order:
#                     so_item = frappe.db.get_value(
#                         "Sales Order Item",
#                         {
#                             "parent": row.against_sales_order,
#                             "item_code": row.item_code
#                         },
#                         "custom_cover_type",
#                     )
#                     if so_item:
#                         row.custom_cover_type = so_item

import frappe
from frappe.model.document import Document
from frappe.utils import flt

@frappe.whitelist()
def create_material_isse(doc, method):
    doc = frappe.get_doc("Delivery Note", doc.name) if isinstance(doc, str) else doc

    if doc.company != "TEAMPRO Food Products":
        return

    stock_entry = None
    for item in doc.items:
        items_to_issue = []

        if flt(item.custom_covers) > 0 and item.custom_cover_type:
            items_to_issue.append({
                "item_code": item.custom_cover_type,
                "qty": item.custom_covers
            })

        if flt(item.custom_bag) > 0 and item.custom_packing_type:
            items_to_issue.append({
                "item_code": item.custom_packing_type,
                "qty": item.custom_bag
            })

        if flt(item.custom_box) > 0 and item.custom_tertiary_packingbox:
            items_to_issue.append({
                "item_code": item.custom_tertiary_packingbox,
                "qty": item.custom_box
            })

        if items_to_issue:
            if not stock_entry:
                stock_entry = frappe.new_doc("Stock Entry")
                stock_entry.stock_entry_type = "Material Issue"
                stock_entry.from_warehouse = "Stores - TFP"
                stock_entry.company = doc.company
                # stock_entry.set_posting_time = 1
                stock_entry.posting_date = doc.posting_date
                stock_entry.posting_time = doc.posting_time
                stock_entry.custom_delivery_note=doc.name
            for i in items_to_issue:
                stock_entry.append("items", {
                    "item_code": i["item_code"],
                    "qty": i["qty"],
                    "uom":"Nos",
                    "s_warehouse": stock_entry.from_warehouse,
                    "cost_center": item.cost_center or frappe.db.get_value("Company", doc.company, "cost_center"),
                    "allow_zero_valuation_rate":1
                })

    if stock_entry:
        stock_entry.insert()
        stock_entry.submit()
        frappe.msgprint(f"Material Issue created: {stock_entry.name}")

@frappe.whitelist()
def cancel_material_isse(doc, method):
    stock_entries = frappe.get_all("Stock Entry", filters={"custom_delivery_note": doc.name, "docstatus": 1}, pluck="name")
    for stock_name in stock_entries:
        stock = frappe.get_doc("Stock Entry", stock_name)
        stock.cancel()



@frappe.whitelist()
def set_totals_in_delivery_note(doc, method):
    doc = frappe.get_doc("Delivery Note", doc.name) if isinstance(doc, str) else doc

    total_covers = 0
    total_bag = 0
    total_box = 0
    if doc.items:
        for item in doc.items:
            total_covers += flt(item.custom_covers)
            total_bag += flt(item.custom_bag)
            total_box += flt(item.custom_box)

        doc.custom_total_covers = total_covers
        doc.custom_total_bag = total_bag
        doc.custom_total_box = total_box


import frappe
from frappe.utils import formatdate

def get_packing_slip_table(doc):
    table_html = """
    <table style="width: 100%; border-collapse: collapse; font-size: 13px;" border="1">
        <thead>
            <tr>
                {header_row_1}
            </tr>
            <tr>
                {header_row_2}
            </tr>
        </thead>
        <tbody>
    """.format(
        header_row_1=''.join([
            '<th rowspan="2" style="background-color: #002060; color:white;">Sr</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">SO ID</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Item Name</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Qty</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">UOM</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">ST.Qty</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">UOM</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">MRP</th>',
            '<th colspan="2" style="background-color: #002060; color:white;">Packing</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Weight</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Rate</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Packing On</th>',
            '<th rowspan="2" style="background-color: #002060; color:white;">Mfg.On</th>',
        ]),
        header_row_2=''.join([
            '<th style="background-color: #002060; color:white;">Cover Type</th>',
            '<th style="background-color: #002060; color:white;">Count</th>',
        ])
    )

    for idx, item in enumerate(doc.items):
        packing_on = formatdate(item.custom_packing_on, "dd-MM-yyyy") if item.custom_packing_on else ""
        mfg_on = formatdate(item.custom_mfg_on, "dd-MM-yyyy") if item.custom_mfg_on else ""

        table_html += f"""
        <tr>
            <td rowspan="3">{idx + 1}</td>
            <td rowspan="3">{item.against_sales_order or ''}</td>
            <td rowspan="3">{item.item_name}</td>
            <td rowspan="3">{item.qty}</td>
            <td rowspan="3">{item.uom}</td>
            <td rowspan="3">{item.stock_qty}</td>
            <td rowspan="3">{item.stock_uom}</td>
            <td rowspan="3">{item.mrp or ''}</td>

            <td>Primary Packing (Cover)</td>
            <td>{item.custom_covers or ''}</td>
            <td rowspan="3">{item.custom_wrd_uom or ''}</td>
            <td rowspan="3">{item.custom_wrd_rate or ''}</td>
            <td rowspan="3">{packing_on}</td>
            <td rowspan="3">{mfg_on}</td>
        </tr>
        <tr>
            <td>Secondary Packing (Bag)</td>
            <td>{item.custom_bag or ''}</td>
        </tr>
        <tr>
            <td>Tertiary Packing (Box)</td>
            <td>{item.custom_box or ''}</td>
        </tr>
        """

    table_html += "</tbody></table>"
    return table_html

import frappe
from frappe.model.naming import make_autoname
from frappe.utils import now_datetime

@frappe.whitelist()
def set_customer_id(doc, method):
    if doc.is_new():
        if not doc.customer_id:
            last_customer = frappe.db.sql("""
                SELECT customer_id FROM `tabCustomer`
                WHERE customer_id REGEXP '^CUST-[0-9]{6}$'
                ORDER BY CAST(SUBSTRING(customer_id, 6) AS UNSIGNED) DESC
                LIMIT 1
            """, as_dict=True)

            if last_customer:
                last_id_num = int(last_customer[0]["customer_id"].split("-")[1])
                new_id_num = last_id_num + 1
            else:
                new_id_num = 1
            new_customer_id = f"CUST-{new_id_num:06d}"
            doc.customer_id = new_customer_id

def update_cover_count(doc, method):
    if doc.items:
        for i in doc.items:
            if i.custom_covers and i.custom_per_2p:
                if i.custom_per_2p != 0:
                    i.custom_bag = i.custom_covers / i.custom_per_2p
            if i.custom_covers and i.custom_per_3p:
                if i.custom_per_3p!=0:
                    i.custom_box=i.custom_covers/i.custom_per_3p

def update_cover_count_in_dn(doc, method):
    if doc.items:
        for i in doc.items:
            if i.custom_covers and i.custom_per_2p:
                if i.custom_per_2p != 0:
                    i.custom_bag = i.custom_covers / i.custom_per_2p
            if i.custom_covers and i.custom_per_3p:
                if i.custom_per_3p!=0:
                    i.custom_box=i.custom_covers/i.custom_per_3p

@frappe.whitelist()
def set_totals_in_sales_order(doc, method):
    doc = frappe.get_doc("Sales Order", doc.name) if isinstance(doc, str) else doc
    total_covers = 0
    total_bag = 0
    total_box = 0
    if doc.items:
        for item in doc.items:
            total_covers += flt(item.custom_covers)
            total_bag += flt(item.custom_bag)
            total_box += flt(item.custom_box)

        doc.custom_total_covers = total_covers
        doc.custom_total_bag = total_bag
        doc.custom_total_box = total_box

import frappe

@frappe.whitelist()
def send_mail_so_submission(doc,method):
    if doc.service == "TFP" and doc.company == "TEAMPRO Food Products":
        table = '''
            <table border="1" style="border-collapse: collapse; width: 50%; text-align: center;">
                <tr style="background-color: #87CEFA;">
                    <th>SO ID</th>
                    <th>Customer</th>
                </tr>
                <tr>
                    <td style="text-align:left">{so_id}</td>
                    <td style="text-align:left">{customer}</td>
                </tr>
            </table>
        '''.format(so_id=doc.name, customer=doc.customer)
        subject = "New Sales Order created for TFP"
        message = f"<p>Dear Team,</p><p>A new Sales Order has been created. Details are as follows:</p>{table}<p>Regards,<br>ERP System</p>"
        recipients = ["nishanthi.p@groupteampro.com"]
        frappe.sendmail(
            recipients=recipients,
            cc=['tfp@groupteampro.com','sangeetha.s@groupteampro.com','dineshbabu.k@groupteampro.com'],
            subject=subject,
            message=message,
        )

@frappe.whitelist()
def get_so_delivery_data(docname):
    from frappe.utils import flt
    from frappe import _
    from collections import defaultdict

    if not docname:
        return "<p>No Sales Order selected.</p>"

    if not frappe.db.exists("Sales Order", docname):
        return f"<p>Sales Order {docname} does not exist.</p>"

    # Fetch DN items
    dn_items = frappe.db.sql("""
        SELECT 
            dni.parent AS dn_name,
            dni.item_code,
            dni.item_name,
            dni.qty,
            dni.uom
        FROM `tabDelivery Note Item` dni
        JOIN `tabDelivery Note` dn ON dn.name = dni.parent
        WHERE dn.docstatus = 1 AND dn.sales_order = %s
        ORDER BY dni.parent, dni.item_code
    """, (docname,), as_dict=True)

    if not dn_items:
        return "<p>No Delivery Notes found for this Sales Order.</p>"

    # Group by DN name
    dn_map = defaultdict(list)
    for row in dn_items:
        dn_map[row.dn_name].append(row)

    # Build HTML table
    data = """
    <table width="100%" style="border-collapse: collapse; font-size: 13px;" border="1">
        <thead>
            <tr style="background-color: #d71920; color: white; text-align: center;">
                <td style="padding: 6px;background-color: #d71920; color: white;">S#</td>
                <td style="padding: 6px;background-color: #d71920; color: white;">Delivery Note</td>
                <td style="padding: 6px;background-color: #d71920; color: white;">Item Code</td>
                <td style="padding: 6px;background-color: #d71920; color: white;">Item Name</td>
                <td style="padding: 6px;background-color: #d71920; color: white;">Qty</td>
                <td style="padding: 6px;background-color: #d71920; color: white;">UOM</td>
            </tr>
        </thead>
        <tbody>
    """

    sr_no = 1
    for dn_name, items in dn_map.items():
        rowspan = len(items)
        for idx, item in enumerate(items):
            data += "<tr>"
            if idx == 0:
                data += f"""
                    <td style='text-align: left; vertical-align: middle;' rowspan='{rowspan}'>{sr_no}</td>
                    <td style='text-align: left; vertical-align: middle;' rowspan='{rowspan}'>{dn_name}</td>
                """
            data += f"""
                <td style="padding: 6px; text-align: left;">{item.item_code}</td>
                <td style="padding: 6px; text-align: left;">{item.item_name}</td>
                <td style="padding: 6px; text-align: center;">{flt(item.qty)}</td>
                <td style="padding: 6px; text-align: center;">{item.uom}</td>
            </tr>
            """
        sr_no += 1

    data += "</tbody></table><br><br>"
    return data




import frappe
from frappe.utils import nowdate, flt, today
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from io import BytesIO
import base64

@frappe.whitelist()
def stock_counting_report_excel():
    formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MM-yyyy")
    filename = "Physical_Vs_ERP_Stock_balance_" + formatted_date
    xlsx_file = build_xlsx_response_stock(filename)
    send_mail_with_attachment_stock(filename, xlsx_file.getvalue())
    

def send_mail_with_attachment_stock(filename,file_content):
    formatted_date = frappe.utils.format_datetime(frappe.utils.nowdate(), "dd-MM-yyyy")
    attachments = [{"fname": filename + '.xlsx', "fcontent": file_content}]
    frappe.sendmail(
        recipients=["divya.p@groupteampro.com","dineshbabu.k@groupteampro.com"],  # Change to real recipient
        cc=["sangeetha.s@groupteampro.com","tfp@groupteampro.com"],
        subject="Physical Vs ERP Stock Balance - " + formatted_date,
        message="Please find attached the Physical Vs ERP Stock Balance report.",
        attachments=attachments,

    )

def build_xlsx_response_stock(filename):
    return make_xlsx_physical_stock(filename)

def make_xlsx_physical_stock(filename, sheet_name=None, wb=None, column_widths=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Physical Vs ERP Stock Balance"

    # Styles
    text_wrap_left = Alignment(vertical="center", horizontal="center", wrap_text=True)
    text_wrap_center = Alignment(vertical="center", horizontal="center", wrap_text=True)
    bold_white_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF002060", end_color="FF002060", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths (A to F now)
    # for col in range(2, 8):  # A to F
    #     ws.column_dimensions[chr(64 + col)].width = 20

    ws.column_dimensions['A'].width = 10   # S.No
    ws.column_dimensions['B'].width = 20   # Item Code
    ws.column_dimensions['C'].width = 40   # Item Name
    ws.column_dimensions['D'].width = 10   # Stock Qty
    ws.column_dimensions['E'].width = 10   # Physical Qty
    ws.column_dimensions['F'].width = 10   # Difference
    ws.column_dimensions['G'].width = 10   # Status


    # Headers with status
    headers = ["S.No", "Item","Item Name", "Stock Qty", "Physical Qty", "Difference", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold_white_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = text_wrap_left

    # Data
    data = get_stock_data()
    for i, row in enumerate(data, start=1):
        status = "Match" if row["difference"] == 0 else "Variance"
        ws.append([
            i,
            row["item"],
            row["item_name"],
            row["stock_qty"],
            row["physical_qty"],
            row["difference"],
            status
        ])
        for j in range(1, 8):  # columns A to F
            cell = ws.cell(row=i+1, column=j)
            # cell.alignment = text_wrap_center
            # cell.border = border
            if j in [1, 2, 3, 7]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cell.border = border
            # Apply conditional font color for Status column
            if j == 7:  # Status column
                if status == "Match":
                    cell.font = Font(bold=True, color="008000")  # Green
                else:
                    cell.font = Font(bold=True, color="FF0000")  # Red

    # Save to in-memory file
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_stock_data():
    today = nowdate()
    latest_stock_counting = frappe.get_all(
        "Stock Counting",
        filters={"docstatus": 1},
        fields=["name", "date"],
        order_by="date desc",
        limit=1
    )

    if not latest_stock_counting:
        return []

    latest_date = latest_stock_counting[0].date

    items = frappe.get_all("Item", filters={"tfp": 1, "disabled": 0}, fields=["*"])
    data = []

    for item in items:
        item_code = item.name
        result = frappe.db.sql("""
            SELECT SUM(sd.count) AS physical_qty
            FROM `tabStock Counting Details` sd
            JOIN `tabStock Counting` sc ON sd.parent = sc.name
            WHERE sd.item = %(item_code)s
            AND sc.date = %(date)s
            AND sc.docstatus = 1
        """, {"item_code": item_code, "date": latest_date}, as_dict=True)

        physical_qty = flt(result[0].physical_qty) if result and result[0].physical_qty else 0
        stock_qty = frappe.db.get_value("Bin", {"item_code": item_code, "warehouse": "Stores - TFP"}, "actual_qty") or 0
        diff = flt(stock_qty) - flt(physical_qty)
        if physical_qty > 0 or stock_qty > 0:
            data.append({
                "item": item_code,
                "item_name":item.item_name,
                "stock_qty": round(stock_qty,2),
                "physical_qty": round(physical_qty,2),
                "difference": round(diff,2),
            })

    return data

@frappe.whitelist()
def get_physical_vs_erp_stock_data():
    from frappe.utils import flt

    # Step 1: Get latest submitted Stock Counting document
    latest_stock_counting = frappe.get_all(
        "Stock Counting",
        filters={"docstatus": 1},
        fields=["name", "date"],
        order_by="date desc",
        limit=1
    )

    if not latest_stock_counting:
        return []

    latest_date = latest_stock_counting[0].date

    # Step 2: Get items marked as TFP
    items = frappe.get_all("Item", filters={"tfp": 1, "disabled": 0}, fields=["name", "item_name"])
    data = []

    for item in items:
        item_code = item.name

        # Step 3: Get physical quantity from Stock Counting Details using latest date
        result = frappe.db.sql("""
            SELECT SUM(sd.count) AS physical_qty
            FROM `tabStock Counting Details` sd
            JOIN `tabStock Counting` sc ON sd.parent = sc.name
            WHERE sd.item = %(item_code)s
            AND sc.date = %(date)s
            AND sc.docstatus = 1
        """, {"item_code": item_code, "date": latest_date}, as_dict=True)

        physical_qty = flt(result[0].physical_qty) if result and result[0].physical_qty else 0

        # Step 4: Get stock qty from Bin
        stock_qty = frappe.db.get_value("Bin", {"item_code": item_code, "warehouse": "Stores - TFP"}, "actual_qty") or 0

        diff = flt(stock_qty) - flt(physical_qty)
        if physical_qty > 0 or stock_qty > 0:
            status = "Match" if diff == 0 else "Variance"
            data.append({
                "item": item_code,
                "item_name": item.item_name,
                "stock_qty": stock_qty,
                "physical_qty": physical_qty,
                "difference": diff,
                "status": status
            })

    return {
    "date": latest_date,
    "data": data
}


@frappe.whitelist()
def update_so_priority(doc,method):
    so=''
    for i in doc.items:
        if i.against_sales_order:
            so=i.against_sales_order
            break
    frappe.db.set_value("Sales Order",so,'custom_priority','')



@frappe.whitelist()
def get_so_item_details(doc,method):
    if not doc.sales_order:
        return

    for item in doc.items:
        if not item.item_code:
            continue

        so_item = frappe.db.get_value(
            "Sales Order Item",
            {
                "parent": item.against_sales_order or doc.sales_order,
                "item_code": item.item_code
            },
            [
                "custom_cover_type",
                "custom_packing_type",
                "custom_tertiary_packingbox",
                "custom_bag",
                "custom_covers",
                "custom_box",
                "custom_per_2p",
                "custom_per_3p",
                "custom_wrd_uom",
                "custom_wrd_rate"
            ],
            as_dict=True
        )

        if so_item:
            item.custom_cover_type = so_item.custom_cover_type
            item.custom_packing_type = so_item.custom_packing_type
            item.custom_tertiary_packingbox = so_item.custom_tertiary_packingbox
            item.custom_bag = so_item.custom_bag
            item.custom_covers = so_item.custom_covers
            item.custom_box = so_item.custom_box
            item.custom_per_2p = so_item.custom_per_2p
            item.custom_per_3p = so_item.custom_per_3p
            item.custom_wrd_uom = so_item.custom_wrd_uom
            item.custom_wrd_rate = so_item.custom_wrd_rate

    # Save updated child table values
    doc.save(ignore_permissions=True)

@frappe.whitelist()
def get_tfp_plan_html_plan_new():
    from frappe.utils import formatdate, flt

    headers = ["Sr", "SO ID", "PRT", "Customer Name", "Packing", "Delivery"]

    html = '''
    <div class="tfp-table-wrapper" style="max-height: 600px; overflow: auto; display: block; border: 1px solid #ccc;">
    <style>
        .tfp-table-wrapper td, .tfp-table-wrapper th {
            padding: 6px;
            vertical-align: middle;
            border: 1px solid #ccc;
        }
    </style>
    <table class="table table-bordered" style="border-collapse: collapse; width: 100%; table-layout: auto;">
    '''
    html += f''' <tr style="background-color: #002060; color: white;">
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="1">Sr</th>
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="3">SO</th>
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="1">PRT</th>
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="3">Customer</th>
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="1">Packing</th>
            <th style="position: sticky; top: 0; text-align: center;background-color: #002060; color: white;" colspan="2">Delivery</th>
        </tr>'''
    
    html += '</tr></thead><tbody>'

    s_no = 1
    grand_total_qty = 0
    grand_total_stock_qty = 0
    grand_total_covers = 0
    grand_total_2p = 0
    grand_total_bag = 0
    grand_total_box = 0

    so_list = frappe.db.get_all("Sales Order", {
        "service": "TFP",
        "status": "To Deliver and Bill"
    }, ["name", "customer", "custom_packing_on", "delivery_date"])

    if not so_list:
        html += '<tr><td colspan="6" style="text-align:center;">Nothing to show</td></tr>'
        html += '</tbody></table></div>'
        return html

    for so in so_list:
        items = frappe.db.get_all("Sales Order Item", {"parent": so.name}, [
            "item_code", "item_name", "qty", "uom", "stock_qty", "stock_uom", "custom_cover_type", "mrp",
            "custom_mfg_on", "custom_covers", "custom_packing_type", "custom_per_2p",
            "custom_2nd_packing", "custom_name_print", "custom_tertiary_packingbox", "custom_bag", "custom_box",
            "custom_wrd_uom", "custom_wrd_rate", "custom_packing_on"
        ])

        html += f'''
        <tr style="font-weight:bold; background-color:#f2f2f2;">
            <td style="text-align:center;"colspan="1">
                <button class="toggle-btn" data-sos="{so.name}" style="background:none; border:none; font-weight:bold; cursor:pointer;">+</button> {s_no}
            </td>
            <td colspan="3">{so.name}</td>
            <td colspan="1"></td>
            <td colspan="3">{so.customer}</td>
            <td style="text-align:center;"colspan="1">{formatdate(so.custom_packing_on)}</td>
            <td style="text-align:center;"colspan="2">{formatdate(so.delivery_date)}</td>
        </tr>
        <tr class="details-row sos-{so.name}" style="display:none; background-color: #d9e1f2; font-weight: bold;">
            <td style="text-align:center;">Item</td>
            <td style="text-align:center;">Qty</td>
            <td style="text-align:center;">UOM</td>
            <td style="text-align:center;">Stock Qty</td>
            <td style="text-align:center;background-color: #C00000; color: white;">CR. Stock</td>
            <td style="text-align:center;background-color: #C00000; color: white;">Stock Status</td>
            <td style="text-align:center;">UOM</td>
            <td style="text-align:center;">MRP</td>
            <td style="text-align:center;">Packing Details</td>
            <td style="text-align:center;">WRD Details</td>
            <td style="text-align:center;">Name Details</td>
        </tr>
        '''

        total_qty = total_stock_qty = total_covers = total_2p = total_bag = total_box = 0
        for item in items:
            total_qty += flt(item.qty)
            total_stock_qty += flt(item.stock_qty)
            total_covers += flt(item.custom_covers)
            total_2p += flt(item.custom_2nd_packing)
            total_bag += flt(item.custom_bag)
            total_box += flt(item.custom_box)

            primary = frappe.db.get_value("Item", item.custom_cover_type, "item_name") or ''
            secondary = frappe.db.get_value("Item", item.custom_packing_type, "item_name") or ''
            tertiary = frappe.db.get_value("Item", item.custom_tertiary_packingbox, "item_name") or ''

            cr_stock = frappe.db.get_value("Bin", {
                "item_code": item.item_code,
                "warehouse": "Stores - TFP"
            }, "actual_qty") or 0

            stock_status = '<span style="color: green; font-weight: bold;">In Stock</span>' \
                if flt(item.stock_qty) <= cr_stock else \
                '<span style="color: red; font-weight: bold;">Out of Stock</span>'
            if item.custom_wrd_rate:
                item_rate=f"{float(item.custom_wrd_rate):.2f}"
            else:
                item_rate=''

            html += f'''
            <tr class="details-row sos-{so.name}" style="display:none;">
                <td style="text-align:left;">{item.item_name}</td>
                <td style="text-align:center;">{item.qty}</td>
                <td style="text-align:center;">{item.uom}</td>
                <td style="text-align:right;">{item.stock_qty}</td>
                <td style="text-align:right;">{cr_stock}</td>
                <td style="text-align:center;">{stock_status}</td>
                <td style="text-align:center;">{item.stock_uom}</td>
                <td style="text-align:right;">{item.mrp}</td>
                <td style="text-align:left;">(C): {primary or "None"}: {item.custom_covers or "0"}<br>(B): {secondary or "None"}: {item.custom_bag or "0"}<br>(BX): {tertiary or "None"}: {item.custom_box or "0"}</td>
                <td style="text-align:left;">(W): {item.custom_wrd_uom or ""}<br>(R): {item_rate or ""}<br>(D): {formatdate(item.custom_mfg_on) if item.custom_mfg_on else ""}</td>
                <td style="text-align:center;">{item.custom_name_print or ""}</td>
            </tr>
            '''

        go_status = "CREATE DN" if all(
            flt(it.stock_qty) <= (frappe.db.get_value("Bin", {"item_code": it.item_code, "warehouse": "Stores - TFP"}, "actual_qty") or 0)
            for it in items
        ) else "CREATE MR"

        html += f'''
        <tr class="details-row sos-{so.name}" style="display:none; font-weight:bold; background-color: #d9e1f2;">
        <td colspan="1" style="text-align:right; border: 1px solid #ccc;">Total</td>
                <td style="text-align:center; border: 1px solid #ccc;">{total_qty}</td>
                <td></td>
                <td style="text-align:right; border: 1px solid #ccc;">{total_stock_qty}</td>
                <td colspan="2" style="text-align:center; border: 1px solid #ccc;vertical-align: middle;">{go_status}</td>
                <td></td>
            <td colspan="5" style="text-align:right;"></td>
        </tr>
        '''

        grand_total_qty += total_qty
        grand_total_stock_qty += total_stock_qty
        grand_total_covers += total_covers
        grand_total_2p += total_2p
        grand_total_bag += total_bag
        grand_total_box += total_box
        s_no += 1

    html += f'''
        <tr style="background-color: #002060; font-weight: bold; color: white;">
         <td colspan="1" style="text-align:center; border: 1px solid #ccc;">Grand Total</td>
            <td style="text-align:center; border: 1px solid #ccc;">{grand_total_qty}</td>
            <td style="border: 1px solid #ccc;"></td>
            <td style="text-align:right; border: 1px solid #ccc;">{grand_total_stock_qty:.2f}</td>
            <td colspan="7" style="border: 1px solid #ccc;"></td>
        </tr>
    </tbody></table></div>
    <script>
        document.querySelectorAll(".toggle-btn").forEach(btn => {{
            btn.addEventListener("click", function() {{
                const sos = this.dataset.sos;
                const rows = document.querySelectorAll(".sos-" + sos);
                const isVisible = rows[0].style.display === "table-row";
                rows.forEach(row => row.style.display = isVisible ? "none" : "table-row");
                this.textContent = isVisible ? "+" : "-";
            }});
        }});
    </script>
    '''
    

    return html
    
@frappe.whitelist()
def create_dn_from_so(sales_order):
    import frappe
    from frappe.model.mapper import get_mapped_doc

    if not frappe.db.exists("Sales Order", sales_order):
        frappe.throw("Invalid Sales Order")

    def set_missing_values(source, target):
        target.custom_delivery_date = nowdate()
        target.run_method("set_missing_values")
        target.run_method("calculate_taxes_and_totals")

    dn = get_mapped_doc(
        "Sales Order",
        sales_order,
        {
            "Sales Order": {
                "doctype": "Delivery Note",
                "validation": {
                    "docstatus": ["=", 1]
                }
            },
            "Sales Order Item": {
                "doctype": "Delivery Note Item",
                "field_map": {
                    "name": "so_detail",
                    "parent": "against_sales_order"
                },
                "condition": lambda doc: doc.delivered_qty < doc.qty
            },
            "Sales Taxes and Charges": {
                "doctype": "Sales Taxes and Charges",
                "add_if_empty": True
            }
        },
        target_doc=None,
        postprocess=set_missing_values
    )

    dn.insert(ignore_permissions=True)
    # dn.submit()  # Uncomment to auto-submit
    frappe.db.commit()
    return dn.name


import frappe
from frappe.utils import nowdate

@frappe.whitelist()
def create_mr_from_so(sales_order):
    so = frappe.get_doc("Sales Order", sales_order)
    if not so:
        frappe.throw("Sales Order not found")

    mr = frappe.new_doc("Material Request")
    mr.material_request_type = "Purchase"
    mr.transaction_date = nowdate()
    mr.schedule_date = nowdate()
    mr.company = so.company
    mr.set_warehouse = "Stores - TFP"
    mr.customer = so.customer
    mr.sales_order = so.name

    item_added = False  # Track if any items are added

    for item in so.items:
        current_stock = frappe.db.get_value("Bin", {
            "item_code": item.item_code,
            "warehouse": "Stores - TFP"
        }, "actual_qty") or 0

        shortage_qty = float(item.stock_qty) - float(current_stock)

        if shortage_qty > 0:
            mr.append("items", {
                "item_code": item.item_code,
                "qty": shortage_qty,
                "uom": item.uom,
                "schedule_date": nowdate(),
                "warehouse": "Stores - TFP"
            })
            item_added = True

    if not item_added:
        frappe.throw("All items are in stock. No Material Request created.")

    mr.save()
    # mr.submit()

    return mr.name

@frappe.whitelist()
def update_so_priority_on_submit(doc, method):
    # Fetch all TFP SOs with 'To Deliver and Bill' status
    sales_orders = frappe.get_all(
        "Sales Order",
        filters={
            "service": "TFP",
            "status": "To Deliver and Bill"
        },
        fields=["name", "custom_packing_on","custom_priority"]
    )

    # Group SOs by packing date
    from collections import defaultdict
    date_groups = defaultdict(list)

    for so in sales_orders:
        if so.custom_packing_on:
            date_groups[so.custom_packing_on].append(so.name)

    # Sort packing dates and assign priorities
    for priority, packing_date in enumerate(sorted(date_groups.keys()), start=1):
        for so_name in date_groups[packing_date]:
            # frappe.log_error(title=so_name,message=priority)
            frappe.db.set_value("Sales Order", so_name, "custom_priority", priority)
